import hashlib
import json
import struct
import uuid
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from test.fuxi.test_table_semantic_rows import _load_table_module

import rag.app.tabular_structure as tabular_structure

from rag.app.tabular_structure import (
    PRODUCER_SCHEMA_VERSION,
    PROJECTION_ROW_FIELDS,
    _ordered_fields,
    _formula_cached_result_kinds_from_biff_stream,
    _formula_coordinates_from_biff_stream,
    build_tabular_structure_projection,
    partition_tabular_structure_projection,
    store_tabular_structure_projection,
    validate_tabular_structure_projection,
)


def test_utf8_bounded_context_cannot_end_with_truncated_whitespace():
    value = "A" * 126 + "  suffix"

    bounded = tabular_structure._truncate_utf8(value, 128)

    assert len(bounded.encode("utf-8")) <= 128
    assert bounded == bounded.strip()


def test_current_producer_versions_invalidate_pre_enumeration_generations():
    assert tabular_structure.TABULAR_STRUCTURE_VERSION == "tabular-row/v2"
    assert PRODUCER_SCHEMA_VERSION == "table-producer/v6"
    assert tabular_structure.PROJECTION_VERSION == "tabular-structure-projection/v6"
    assert tabular_structure.PROJECTION_PART_VERSION == "tabular-structure-part/v3"
    assert tabular_structure.STRUCTURE_PRODUCER_ALGORITHM_VERSION == "region-producer/v21"
    assert tabular_structure.ENUMERATION_RULE_VERSION == "enumeration-rules/v9"


class _AxisClosureCell:
    def __init__(self, value):
        self.value = value


class _AxisClosureWorksheet:
    class _MergedCells:
        ranges = ()

    def __init__(self, values):
        self._values = values
        self.merged_cells = self._MergedCells()

    def cell(self, row, column):
        return _AxisClosureCell(self._values.get((row, column)))


class _AxisClosureParser:
    @staticmethod
    def _get_merged_cell_value(_worksheet, _row, _column, _merged_ranges):
        return None


def _axis_closure_item(
    *,
    required_offsets=(0, 1, 2),
    key_axis=True,
    source_column=1,
    width=3,
    row_values=(),
    row_role="data",
    data_row_count=None,
    structure_evidence=True,
):
    rows = []
    members = set()
    for row_ordinal, values in row_values:
        rows.append(
            {
                "row_ordinal_int": row_ordinal,
                "row_role_kwd": row_role,
            }
        )
        members.update(
            (row_ordinal, source_column + offset)
            for offset, value in enumerate(values)
            if value is not None and str(value).strip()
        )
    axis = {
        "record_row_ordinals": tuple(row for row, _values in row_values),
        "required_offsets": tuple(required_offsets),
        "record_key_axis_proven": key_axis,
        "single_record_axis_proven": False,
    }
    evidence = (
        {
            "record_axis_evidence": axis,
            "record_axis_source_column": source_column,
            "record_axis_width": width,
            "body_row_ordinals": [row for row, _values in row_values],
        }
        if structure_evidence
        else None
    )
    return {
        "bbox": (
            min((row for row, _values in row_values), default=1),
            source_column,
            max((row for row, _values in row_values), default=1),
            source_column + width - 1,
        ),
        "members": members,
        "member_columns": {column for _row, column in members},
        "rows": rows,
        "source_values": tuple(row_values),
        "source_column": source_column,
        "structure_evidence": evidence,
        "table": {
            "data_row_count": (
                len(row_values) if data_row_count is None else data_row_count
            )
        },
    }


def _axis_closure_worksheet(*items):
    values = {}
    for item in items:
        for row, row_values in item["source_values"]:
            evidence = item.get("structure_evidence") or {}
            source_column = evidence.get(
                "record_axis_source_column",
                item["source_column"],
            )
            for offset, value in enumerate(row_values):
                values[(row, source_column + offset)] = value
    return _AxisClosureWorksheet(values)


def test_axis_closure_accepts_a_value_free_footer_without_geometric_distance():
    earlier = _axis_closure_item(
        row_values=((2, (1, "A", 10)), (3, (2, "B", 11))),
    )
    later = _axis_closure_item(
        row_values=((5, (None, None, None)), (6, (None, None, None))),
        row_role="unknown",
        data_row_count=0,
    )

    assert tabular_structure._axis_closure_proven(
        parser=_AxisClosureParser(),
        worksheet=_axis_closure_worksheet(earlier, later),
        earlier=earlier,
        later=later,
    )


def test_axis_closure_does_not_depend_on_footer_values_in_the_key_column():
    earlier = _axis_closure_item(
        row_values=((2, (1, "A", 10)), (3, (2, "B", 11))),
    )
    later = _axis_closure_item(
        row_values=((5, (99, "approval", None)), (6, (None, "sign", None))),
        row_role="unknown",
        data_row_count=0,
    )

    assert tabular_structure._axis_closure_proven(
        parser=_AxisClosureParser(),
        worksheet=_axis_closure_worksheet(earlier, later),
        earlier=earlier,
        later=later,
    )


def test_axis_closure_rejects_a_candidate_with_the_earlier_record_arity():
    earlier = _axis_closure_item(
        row_values=((2, (1, "A", 10)), (3, (2, "B", 11))),
    )
    later = _axis_closure_item(
        row_values=((5, (3, "C", 12)),),
        row_role="unknown",
        data_row_count=0,
    )

    assert not tabular_structure._axis_closure_proven(
        parser=_AxisClosureParser(),
        worksheet=_axis_closure_worksheet(earlier, later),
        earlier=earlier,
        later=later,
    )


def test_axis_closure_rebases_candidate_columns_to_earlier_source_columns():
    earlier = _axis_closure_item(
        required_offsets=(0, 1, 2),
        source_column=3,
        row_values=((2, (1, "A", 10)), (3, (2, "B", 11))),
    )
    later = _axis_closure_item(
        source_column=4,
        row_values=((5, (None, "footer", None)),),
        row_role="unknown",
        data_row_count=0,
    )

    assert tabular_structure._axis_closure_proven(
        parser=_AxisClosureParser(),
        worksheet=_axis_closure_worksheet(earlier, later),
        earlier=earlier,
        later=later,
    )


def test_axis_closure_accepts_a_shifted_candidate_that_lacks_the_main_key_column():
    earlier = _axis_closure_item(
        source_column=3,
        row_values=((2, (1, "A", 10)), (3, (2, "B", 11))),
    )
    later = _axis_closure_item(
        source_column=4,
        row_values=((5, (3, "C", 12)),),
        row_role="unknown",
        data_row_count=0,
    )

    assert tabular_structure._axis_closure_proven(
        parser=_AxisClosureParser(),
        worksheet=_axis_closure_worksheet(earlier, later),
        earlier=earlier,
        later=later,
    )


def test_axis_closure_rebases_a_non_numeric_anchor_to_the_main_key_column():
    earlier = _axis_closure_item(
        key_axis=False,
        source_column=3,
        row_values=((2, ("A-1", "A", 10)), (3, ("A-2", "B", 11))),
    )
    later = _axis_closure_item(
        key_axis=False,
        source_column=4,
        row_values=((5, ("A-3", "approval", 12)),),
        row_role="unknown",
        data_row_count=0,
    )

    assert tabular_structure._axis_closure_proven(
        parser=_AxisClosureParser(),
        worksheet=_axis_closure_worksheet(earlier, later),
        earlier=earlier,
        later=later,
    )


def test_axis_closure_accepts_a_nonrecord_footer_without_structure_evidence():
    earlier = _axis_closure_item(
        row_values=((2, (1, "A", 10)), (3, (2, "B", 11))),
    )
    later = _axis_closure_item(
        row_values=((5, (None, None, None)),),
        row_role="unknown",
        data_row_count=0,
        structure_evidence=False,
    )

    assert tabular_structure._axis_closure_proven(
        parser=_AxisClosureParser(),
        worksheet=_axis_closure_worksheet(earlier, later),
        earlier=earlier,
        later=later,
    )


def test_axis_closure_rejects_a_complete_record_without_structure_evidence():
    earlier = _axis_closure_item(
        row_values=((2, (1, "A", 10)), (3, (2, "B", 11))),
    )
    later = _axis_closure_item(
        row_values=((5, (3, "C", 12)),),
        row_role="unknown",
        data_row_count=0,
        structure_evidence=False,
    )

    assert not tabular_structure._axis_closure_proven(
        parser=_AxisClosureParser(),
        worksheet=_axis_closure_worksheet(earlier, later),
        earlier=earlier,
        later=later,
    )


def test_axis_closure_accepts_a_partial_context_without_structure_evidence():
    earlier = _axis_closure_item(
        row_values=((2, (1, "A", 10)), (3, (2, "B", 11))),
    )
    later = _axis_closure_item(
        row_values=((5, (3, None, None)),),
        row_role="unknown",
        data_row_count=0,
        structure_evidence=False,
    )

    assert tabular_structure._axis_closure_proven(
        parser=_AxisClosureParser(),
        worksheet=_axis_closure_worksheet(earlier, later),
        earlier=earlier,
        later=later,
    )


def test_v20_projection_contract_remains_available_for_backfill():
    assert (
        "table-producer/v6",
        "tabular-structure-projection/v6",
        "region-producer/v20",
        "enumeration-rules/v9",
    ) in tabular_structure._KNOWN_BACKFILL_PROJECTION_CONTRACTS


def test_structurally_closed_header_only_table_proves_an_empty_record_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous empty register"
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "Anonymous register"
    sheet.append(
        [
            "Sequence",
            "Reference",
            "Received",
            "Issuer",
            "Issued",
            "Reason",
        ]
    )

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["rows"] == []
    assert len(projection["tables"]) == 1
    assert projection["tables"][0] == {
        "table_ref": projection["tables"][0]["table_ref"],
        "sheet_ordinal": 1,
        "table_ordinal": 1,
        "row_count": 0,
        "data_row_count": 0,
        "source_total_count": 0,
        "table_label": "Anonymous empty register",
        "table_context": [
            {"name": "context", "value": "Anonymous register"},
            *(
                {"name": "field", "value": name}
                for name in (
                    "Sequence",
                    "Reference",
                    "Received",
                    "Issuer",
                    "Issued",
                    "Reason",
                )
            ),
        ],
        "ordered_columns": [
            {
                "column_id": f"col_v1:1:{ordinal}",
                "column_ordinal": ordinal,
                "header_path": [name],
                "name": name,
            }
            for ordinal, name in enumerate(
                (
                    "Sequence",
                    "Reference",
                    "Received",
                    "Issuer",
                    "Issued",
                    "Reason",
                ),
                start=1,
            )
        ],
        "enumeration_status": "supported_complete",
        "enumeration_reason": "empty_record_axis_proven",
        "matched_rule": "L1-08",
    }


def test_empty_record_axis_ignores_a_disjoint_sidecar_outside_the_table_columns(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous empty register"
    sheet.merge_cells("A1:G1")
    sheet["A1"] = "Anonymous register"
    headers = [
        "Sequence",
        "Reference",
        "Received",
        "Issuer",
        "Issued",
        "Reason",
        "Category",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=column, value=header)
    sheet["H2"] = "Resource type"
    sheet["I2"] = "Related resource"
    sheet["I2"].hyperlink = "https://example.invalid/resource"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["matched_rule"] == "L1-08"
    assert complete[0]["source_total_count"] == 0
    assert [
        column["column_id"] for column in complete[0]["ordered_columns"]
    ] == [f"col_v1:1:{ordinal}" for ordinal in range(1, 8)]
    assert all(
        column["name"] != "Related resource"
        for column in complete[0]["ordered_columns"]
    )
    expected_membership = tabular_structure._region_membership_sha256(
        1,
        {
            (row, column)
            for row in (1, 2)
            for column in range(1, 8)
        },
    )
    assert complete[0]["table_ref"].startswith(
        f"tbl_v2_{expected_membership}_"
    )
    assert all(
        row["table_ref_kwd"] != complete[0]["table_ref"]
        for row in projection["rows"]
    )


def test_merged_title_with_trailing_blank_header_and_link_sidecar_proves_empty_record_axis(
    table_parser,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous header-only register"
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "Anonymous register"
    for column, header in enumerate(
        (
            "Sequence",
            "Reference",
            "Received",
            "Issuer",
            "Issued",
            "Reason",
            "Category",
        ),
        start=1,
    ):
        sheet.cell(row=2, column=column, value=header)
    sheet["I2"] = "Related resource"
    sheet["I2"].hyperlink = "https://example.invalid/resource"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["matched_rule"] == "L1-08"
    assert complete[0]["source_total_count"] == 0
    assert [column["name"] for column in complete[0]["ordered_columns"]] == [
        "Sequence",
        "Reference",
        "Received",
        "Issuer",
        "Issued",
        "Reason",
        "Category",
    ]


def test_unmerged_wide_single_row_is_not_promoted_to_an_empty_list(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Prepared by",
            "Approved by",
            "Reviewed by",
            "Issued on",
            "Status",
            "Revision",
            "Comment",
        ]
    )

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert all(table["matched_rule"] != "L1-08" for table in projection["tables"])


def test_trailing_blank_does_not_hide_content_inside_the_title_span(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:I1")
    sheet["A1"] = "Anonymous register"
    for column, header in enumerate(
        ("Sequence", "Reference", "Received", "Issuer", "Issued", "Reason", "Category"),
        start=1,
    ):
        sheet.cell(row=2, column=column, value=header)
    sheet["I2"] = "Ambiguous content"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert all(table["matched_rule"] != "L1-08" for table in projection["tables"])


def test_empty_record_axis_does_not_ignore_content_below_its_table_columns(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:G1")
    sheet["A1"] = "Anonymous register"
    for column, header in enumerate(
        ("Sequence", "Reference", "Received", "Issuer", "Issued", "Reason", "Category"),
        start=1,
    ):
        sheet.cell(row=2, column=column, value=header)
    sheet["H2"] = "Resource type"
    sheet["I2"] = "Related resource"
    sheet["I2"].hyperlink = "https://example.invalid/resource"
    sheet["C4"] = "Unresolved content"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert all(table["matched_rule"] != "L1-08" for table in projection["tables"])


def test_two_cell_signoff_row_is_not_promoted_to_an_empty_list(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Prepared by", "Approved by"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert all(table["matched_rule"] != "L1-08" for table in projection["tables"])


def test_delayed_structural_header_proves_an_empty_record_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:G1")
    sheet["A1"] = "Anonymous register"
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "Context A"
    sheet.merge_cells("C2:G2")
    sheet["C2"] = "Value A"
    sheet.merge_cells("A3:B3")
    sheet["A3"] = "Context B"
    sheet.merge_cells("C3:G3")
    sheet["C3"] = "Value B"
    for column, header in enumerate(
        (
            "Sequence",
            "Reference",
            "Received",
            "Issuer",
            "Issued",
            "Reason",
            "Category",
        ),
        start=1,
    ):
        sheet.cell(row=5, column=column, value=header)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 0
    assert complete[0]["matched_rule"] == "L1-08"
    assert [column["name"] for column in complete[0]["ordered_columns"]] == [
        "Sequence",
        "Reference",
        "Received",
        "Issuer",
        "Issued",
        "Reason",
        "Category",
    ]


def test_trailing_multilevel_header_after_context_proves_an_empty_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "Anonymous register"
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "Context A"
    sheet.merge_cells("C2:F2")
    sheet["C2"] = "Value A"
    sheet.merge_cells("A4:A5")
    sheet["A4"] = "Sequence"
    sheet.merge_cells("B4:C4")
    sheet["B4"] = "Identity"
    sheet["B5"] = "Code"
    sheet["C5"] = "Revision"
    sheet.merge_cells("D4:F4")
    sheet["D4"] = "Evidence"
    sheet["D5"] = "Received"
    sheet["E5"] = "Issuer"
    sheet["F5"] = "Reason"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 0
    assert complete[0]["matched_rule"] == "L1-08"


@pytest.mark.parametrize("header_row", (3, 20, 47, 121))
def test_trailing_header_uses_physical_region_coordinates_at_any_offset(
    table_parser,
    header_row,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Delayed header"
    for row in range(1, max(2, header_row - 1)):
        sheet.cell(row, 1, f"Context {row}")
    sheet.merge_cells(
        start_row=header_row,
        start_column=1,
        end_row=header_row,
        end_column=2,
    )
    sheet.cell(header_row, 1, "Identity")
    sheet.cell(header_row + 1, 1, "Code")
    sheet.cell(header_row + 1, 2, "Description")
    sheet.cell(header_row + 2, 1, "Type")
    sheet.cell(header_row + 2, 2, "Requirement")

    worksheet = table_parser._load_excel_to_workbook(
        BytesIO(_save_workbook(workbook))
    ).active
    rows, populated_rows, unresolved_rows = tabular_structure._complete_worksheet_rows(
        worksheet
    )

    assert worksheet.max_row == header_row + 2
    assert len(rows) == header_row + 2
    assert [row[0].row for row in rows] == list(range(1, header_row + 3))
    assert populated_rows == [
        *range(1, max(2, header_row - 1)),
        header_row,
        header_row + 1,
        header_row + 2,
    ]
    structure = tabular_structure._trailing_empty_record_axis_structure(
        table_parser,
        worksheet,
        rows,
        populated_rows,
        unresolved_rows,
    )

    assert structure is not None
    headers, header_paths, header_start, data_start = structure
    assert headers == ["Identity-Code-Type", "Identity-Description-Requirement"]
    assert header_paths == [
        ["Identity", "Code", "Type"],
        ["Identity", "Description", "Requirement"],
    ]
    assert (header_start, data_start) == (header_row - 1, header_row + 2)


@pytest.mark.parametrize("header_row", (3, 20, 47, 121))
def test_projection_handles_trailing_header_at_any_offset(table_parser, header_row):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Delayed header projection"
    for row in range(1, max(2, header_row - 1)):
        sheet.cell(row, 1, f"Context {row}")
    sheet.merge_cells(
        start_row=header_row,
        start_column=1,
        end_row=header_row,
        end_column=2,
    )
    sheet.cell(header_row, 1, "Identity")
    sheet.cell(header_row + 1, 1, "Code")
    sheet.cell(header_row + 1, 2, "Description")
    sheet.cell(header_row + 2, 1, "Type")
    sheet.cell(header_row + 2, 2, "Requirement")

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 0
    assert complete[0]["matched_rule"] == "L1-08"
    assert [column["header_path"] for column in complete[0]["ordered_columns"]] == [
        ["Identity", "Code", "Type"],
        ["Identity", "Description", "Requirement"],
    ]


def test_delayed_signoff_form_is_not_promoted_to_an_empty_record_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Anonymous approval form"
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "Prepared by"
    sheet.merge_cells("C2:D2")
    sheet["C2"] = "Reviewed by"
    sheet.merge_cells("A3:B3")
    sheet["A3"] = "Issued on"
    sheet.merge_cells("C3:D3")
    sheet["C3"] = "Revision"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert all(table["matched_rule"] != "L1-08" for table in projection["tables"])


def test_header_with_unresolved_body_content_is_not_promoted_to_an_empty_list(
    table_parser,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "Anonymous register"
    sheet.append(["Sequence", "Reference", "Reason"])
    sheet.merge_cells("A3:C3")
    sheet["A3"] = "Pending note"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert all(table["matched_rule"] != "L1-08" for table in projection["tables"])


def test_table_ref_identity_binds_all_versions_and_exact_membership(monkeypatch):
    source_sha256 = hashlib.sha256(b"same source").hexdigest()
    first = tabular_structure._table_ref(source_sha256, 1, 1, "a" * 64)
    changed_members = tabular_structure._table_ref(source_sha256, 1, 1, "b" * 64)

    monkeypatch.setattr(tabular_structure, "PRODUCER_SCHEMA_VERSION", "table-producer/test-next")
    changed_schema = tabular_structure._table_ref(source_sha256, 1, 1, "a" * 64)
    monkeypatch.setattr(tabular_structure, "PRODUCER_SCHEMA_VERSION", PRODUCER_SCHEMA_VERSION)
    monkeypatch.setattr(
        tabular_structure,
        "PROJECTION_VERSION",
        "tabular-structure-projection/test-next",
    )
    changed_projection = tabular_structure._table_ref(source_sha256, 1, 1, "a" * 64)
    monkeypatch.setattr(
        tabular_structure,
        "PROJECTION_VERSION",
        "tabular-structure-projection/v4",
    )
    monkeypatch.setattr(
        tabular_structure,
        "STRUCTURE_PRODUCER_ALGORITHM_VERSION",
        "region-producer/test-next",
    )
    changed_algorithm = tabular_structure._table_ref(source_sha256, 1, 1, "a" * 64)
    monkeypatch.setattr(
        tabular_structure,
        "ENUMERATION_RULE_VERSION",
        "enumeration-rules/test-next",
        raising=False,
    )
    changed_enumeration_rules = tabular_structure._table_ref(source_sha256, 1, 1, "a" * 64)

    assert first.startswith("tbl_v2_" + "a" * 64 + "_")
    assert len(
        {
            first,
            changed_members,
            changed_schema,
            changed_projection,
            changed_algorithm,
            changed_enumeration_rules,
        }
    ) == 6


def test_projection_root_requires_the_current_enumeration_rule_version(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )

    assert projection["enumeration_rule_version"] == "enumeration-rules/v9"

    missing = dict(projection)
    missing.pop("enumeration_rule_version")
    with pytest.raises(ValueError, match="fixed top-level schema"):
        validate_tabular_structure_projection(missing)

    changed = dict(projection)
    changed["enumeration_rule_version"] = "enumeration-rules/unknown"
    with pytest.raises(ValueError, match="enumeration rule version"):
        validate_tabular_structure_projection(changed)


def test_table_manifest_requires_one_strict_enumeration_decision(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )
    table = projection["tables"][0]

    assert {
        "enumeration_status": table["enumeration_status"],
        "enumeration_reason": table["enumeration_reason"],
        "matched_rule": table["matched_rule"],
    } == {
        "enumeration_status": "supported_complete",
        "enumeration_reason": "record_axis_proven",
        "matched_rule": "L1-05",
    }

    for field in ("enumeration_status", "enumeration_reason", "matched_rule"):
        missing = json.loads(json.dumps(projection))
        missing["tables"][0].pop(field)
        with pytest.raises(ValueError, match="table manifest"):
            validate_tabular_structure_projection(missing)

    changed = json.loads(json.dumps(projection))
    changed["tables"][0]["enumeration_reason"] = "record_axis_not_proven"
    with pytest.raises(ValueError, match="enumeration decision"):
        validate_tabular_structure_projection(changed)

    invalid_rule = json.loads(json.dumps(projection))
    invalid_rule["tables"][0]["matched_rule"] = []
    with pytest.raises(ValueError, match="enumeration decision"):
        validate_tabular_structure_projection(invalid_rule)


def test_duplicate_source_member_ingestion_reaches_d2(table_parser, monkeypatch):
    original = tabular_structure._new_projected_item

    def duplicate_member_event(**kwargs):
        item = original(**kwargs)
        item["emitted_member_events"].append(item["emitted_member_events"][0])
        return item

    monkeypatch.setattr(tabular_structure, "_new_projected_item", duplicate_member_event)
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert projection["tables"][0]["matched_rule"] == "D2"
    assert projection["tables"][0]["enumeration_reason"] == "membership_not_closed"


def test_unassigned_source_member_ingestion_reaches_d2(table_parser, monkeypatch):
    original = tabular_structure._new_projected_item

    def drop_last_member_event(**kwargs):
        item = original(**kwargs)
        item["emitted_member_events"] = item["emitted_member_events"][:-1]
        return item

    monkeypatch.setattr(tabular_structure, "_new_projected_item", drop_last_member_event)
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )

    assert projection["tables"][0]["matched_rule"] == "D2"
    assert projection["tables"][0]["enumeration_reason"] == "membership_not_closed"


def test_proven_record_slot_count_mismatch_reaches_d3(table_parser, monkeypatch):
    original = tabular_structure._new_projected_item

    def add_unemitted_proven_slot(**kwargs):
        item = original(**kwargs)
        item["proven_record_slots"].append(99)
        return item

    monkeypatch.setattr(tabular_structure, "_new_projected_item", add_unemitted_proven_slot)
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert projection["tables"][0]["matched_rule"] == "D3"
    assert projection["tables"][0]["enumeration_reason"] == "record_count_mismatch"


def test_proven_record_slots_must_match_emitted_data_row_ordinals(table_parser, monkeypatch):
    original = tabular_structure._new_projected_item

    def replace_first_slot_with_header(**kwargs):
        item = original(**kwargs)
        item["proven_record_slots"][0] = min(row for row, _column in item["members"])
        return item

    monkeypatch.setattr(
        tabular_structure,
        "_new_projected_item",
        replace_first_slot_with_header,
    )
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert projection["tables"][0]["matched_rule"] == "D3"
    assert projection["tables"][0]["enumeration_reason"] == "record_count_mismatch"


def test_missing_unknown_projection_emits_d4_tombstone(table_parser, monkeypatch):
    workbook = Workbook()
    workbook.active["A1"] = "Standalone note"
    monkeypatch.setattr(tabular_structure, "_unknown_structure_region", lambda **_kwargs: None)

    projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"] == []
    assert projection["rows"] == []
    assert set(projection) == tabular_structure.PROJECTION_FIELDS
    assert audit["version"] == "tabular-structure-producer-audit/v1"
    assert audit["producer_generation_ref"] == projection["producer_generation_ref"]
    assert audit["enumeration_rule_version"] == projection["enumeration_rule_version"]
    assert audit["source_sha256"] == projection["source_sha256"]
    assert audit["defects"] == [
        {
            "row_kind": "defect_tombstone",
            "table_ref": None,
            "sheet_ordinal": 1,
            "source_region_ordinal": 1,
            "membership_sha256": tabular_structure._region_membership_sha256(1, {(1, 1)}),
            "enumeration_status": "defect",
            "enumeration_reason": "missing_projection",
            "matched_rule": "D4",
        }
    ]

    fabricated_identity = json.loads(json.dumps(audit))
    fabricated_identity["defects"][0]["table_ref"] = "tbl_v2_" + "0" * 128
    with pytest.raises(ValueError, match="cannot fabricate"):
        tabular_structure._validate_tabular_structure_producer_audit(fabricated_identity)

    wrong_reason = json.loads(json.dumps(audit))
    wrong_reason["defects"][0]["enumeration_reason"] = "membership_not_closed"
    with pytest.raises(ValueError, match="decision is invalid"):
        tabular_structure._validate_tabular_structure_producer_audit(wrong_reason)

    duplicate_region = json.loads(json.dumps(audit))
    duplicate_region["defects"].append(dict(duplicate_region["defects"][0]))
    with pytest.raises(ValueError, match="not unique deterministic"):
        tabular_structure._validate_tabular_structure_producer_audit(duplicate_region)


def test_producer_audit_exports_content_free_closed_object_evidence(table_parser):
    projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )

    assert set(projection) == tabular_structure.PROJECTION_FIELDS
    assert audit["defects"] == []
    assert len(audit["source_regions"]) == 1
    assert len(audit["output_objects"]) == 1
    source = audit["source_regions"][0]
    output = audit["output_objects"][0]
    assert source["assigned_object_ref"] == output["object_ref"] == projection["tables"][0]["table_ref"]
    assert source["assignment_count"] == 1
    assert output["row_kind"] == "object"
    assert output["identity_validation_status"] == "pending_independent_validation"
    assert output["enumeration_status"] == "supported_complete"
    tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_rejects_record_slot_outside_union_members(table_parser):
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )
    forged = json.loads(json.dumps(audit))
    output = forged["output_objects"][0]
    output["record_slot_coordinate_sets"][0] = ["1:999:999"]
    output["record_slot_sha256"] = tabular_structure._audit_digest(
        "adr039-record-slot/v1",
        ["|".join(slot) for slot in output["record_slot_coordinate_sets"]],
    )

    with pytest.raises(ValueError, match="record slot.*union"):
        tabular_structure._validate_tabular_structure_producer_audit(forged)


def test_producer_audit_rejects_duplicate_complete_record_slots(table_parser):
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )
    forged = json.loads(json.dumps(audit))
    output = forged["output_objects"][0]
    output["record_slot_coordinate_sets"][1] = list(
        output["record_slot_coordinate_sets"][0]
    )
    output["record_slot_sha256"] = tabular_structure._audit_digest(
        "adr039-record-slot/v1",
        ["|".join(slot) for slot in output["record_slot_coordinate_sets"]],
    )

    with pytest.raises(ValueError, match="record slots.*unique"):
        tabular_structure._validate_tabular_structure_producer_audit(forged)


def test_producer_audit_rejects_header_row_as_a_complete_record_slot(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code"])
    sheet.append(["S-1"])
    sheet.append(["S-2"])
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )
    forged = json.loads(json.dumps(audit))
    output = forged["output_objects"][0]
    assert output.get("emitted_data_row_ordinals") == [2, 3]
    output["record_slot_coordinate_sets"] = [["1:1:1"], ["1:2:1"]]
    output["record_slot_sha256"] = tabular_structure._audit_digest(
        "adr039-record-slot/v1",
        ["1:1:1", "1:2:1"],
    )

    with pytest.raises(ValueError, match="record slots.*data rows"):
        tabular_structure._validate_tabular_structure_producer_audit(forged)


@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("bbox", [1, 1, 999, 999], "source geometry"),
        ("row_count", 999, "source geometry"),
        ("column_count", 999, "source geometry"),
        ("worksheet_ordinal", 2, "source region reference"),
        ("source_region_ref", "1:99", "source region reference"),
    ],
)
def test_producer_audit_recomputes_source_region_geometry(
    table_parser,
    field,
    value,
    message,
):
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )
    audit["source_regions"][0][field] = value
    if field == "source_region_ref":
        audit["output_objects"][0]["component_region_refs"][0] = value

    with pytest.raises(ValueError, match=message):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_cannot_self_approve_identity_validation(table_parser):
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )
    audit["output_objects"][0]["identity_validation_status"] = "validated"

    with pytest.raises(ValueError, match="identity validation status"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_requires_each_d4_defect_to_match_its_tombstone(
    table_parser,
    monkeypatch,
):
    workbook = Workbook()
    workbook.active["A1"] = "Standalone note"
    monkeypatch.setattr(tabular_structure, "_unknown_structure_region", lambda **_kwargs: None)
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )
    audit["defects"] = []

    with pytest.raises(ValueError, match="D4 defect.*tombstone"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_rejects_overlapping_complete_components(table_parser):
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _vertical_complete_with_headerless_continuation_bytes(),
        parser=table_parser,
    )
    assert len(audit["source_regions"]) == 2
    first, second = audit["source_regions"]
    second["member_coordinate_set"].append(first["member_coordinate_set"][0])
    second["member_coordinate_set"].sort(key=tabular_structure._audit_coordinate_key)
    second["membership_sha256"] = hashlib.sha256(
        "\n".join(second["member_coordinate_set"]).encode("ascii")
    ).hexdigest()
    second["member_count"] = len(second["member_coordinate_set"])
    parsed_members = [
        tabular_structure._audit_coordinate_key(value)
        for value in second["member_coordinate_set"]
    ]
    member_rows = {row for _sheet, row, _column in parsed_members}
    member_columns = {column for _sheet, _row, column in parsed_members}
    second["bbox"] = [
        min(member_rows),
        min(member_columns),
        max(member_rows),
        max(member_columns),
    ]
    second["row_count"] = len(member_rows)
    second["column_count"] = len(member_columns)
    output = audit["output_objects"][0]
    output["component_membership_sha256_list"][1] = second["membership_sha256"]

    with pytest.raises(ValueError, match="component memberships.*disjoint"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_recomputes_table_identity(table_parser):
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )
    forged_ref = "tbl_v2_" + audit["output_objects"][0]["union_membership_sha256"] + "_" + "0" * 64
    output = audit["output_objects"][0]
    output["object_ref"] = forged_ref
    output["table_ref"] = forged_ref
    audit["source_regions"][0]["assigned_object_ref"] = forged_ref

    with pytest.raises(ValueError, match="table identity"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_binds_output_worksheet_to_components(table_parser):
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )
    audit["output_objects"][0]["worksheet_ordinal"] = 2

    with pytest.raises(ValueError, match="output worksheet"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_requires_numeric_component_order(table_parser):
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _vertical_complete_with_headerless_continuation_bytes(),
        parser=table_parser,
    )
    output = audit["output_objects"][0]
    output["component_region_refs"].reverse()
    output["component_membership_sha256_list"].reverse()

    with pytest.raises(ValueError, match="component references.*ordered"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_requires_multiple_assignments_to_reach_d2(table_parser):
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )
    first = audit["output_objects"][0]
    second = json.loads(json.dumps(first))
    second_ref = tabular_structure._table_ref(
        audit["source_sha256"],
        1,
        2,
        second["union_membership_sha256"],
    )
    second.update(
        {
            "object_ref": second_ref,
            "table_ref": second_ref,
            "matched_rule": "R8",
            "decision_chain_stop": "R8",
            "enumeration_status": "not_guaranteed_explained",
            "enumeration_reason": "record_axis_not_proven",
            "source_total_count": None,
        }
    )
    audit["output_objects"].append(second)
    audit["source_regions"][0]["assigned_object_ref"] = None
    audit["source_regions"][0]["assignment_count"] = 2

    with pytest.raises(ValueError, match="multiple assignments.*D2"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_requires_the_decision_stop_to_match_the_rule(table_parser):
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )
    audit["output_objects"][0]["decision_chain_stop"] = "R8"

    with pytest.raises(ValueError, match="decision chain stop"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_rejects_a_total_on_a_noncomplete_decision(table_parser):
    workbook = Workbook()
    workbook.active.append(["Code", "State"])
    workbook.active.append(["A-1", "Open"])
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )
    assert audit["output_objects"][0]["enumeration_status"] == "not_guaranteed_explained"
    audit["output_objects"][0]["source_total_count"] = 1

    with pytest.raises(ValueError, match="decision conflicts with source total"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_requires_unclosed_membership_to_be_d2(table_parser):
    workbook = Workbook()
    workbook.active.append(["Code", "State"])
    workbook.active.append(["A-1", "Open"])
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )
    output = audit["output_objects"][0]
    output["emitted_member_coordinate_multiset"] = output[
        "emitted_member_coordinate_multiset"
    ][:-1]
    emitted = output["emitted_member_coordinate_multiset"]
    output["emitted_cell_multiset_sha256"] = tabular_structure._audit_digest(
        "adr039-emitted-cell-multiset/v1",
        emitted,
    )
    output["emitted_member_occurrence_count"] = len(emitted)
    output["member_max_ingest_count"] = 1

    with pytest.raises(ValueError, match="unclosed membership.*D2"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_record_slots_cover_the_complete_source_rows(table_parser):
    workbook = Workbook()
    workbook.active.append(["Code", "State"])
    workbook.active.append(["A-1", "Open"])
    workbook.active.append(["A-2", "Closed"])
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )
    output = audit["output_objects"][0]
    output["record_slot_coordinate_sets"][0] = output[
        "record_slot_coordinate_sets"
    ][0][:-1]
    output["record_slot_sha256"] = tabular_structure._audit_digest(
        "adr039-record-slot/v1",
        ["|".join(slot) for slot in output["record_slot_coordinate_sets"]],
    )

    with pytest.raises(ValueError, match="record slot.*complete source row"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


@pytest.mark.parametrize("defect_rule", ["D2", "D3", "D4"])
def test_producer_audit_rejects_unsupported_defect_labels(table_parser, defect_rule):
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )
    output = audit["output_objects"][0]
    status, reason = tabular_structure.ENUMERATION_DECISIONS[defect_rule]
    output.update(
        {
            "matched_rule": defect_rule,
            "decision_chain_stop": defect_rule,
            "enumeration_status": status,
            "enumeration_reason": reason,
            "source_total_count": None,
        }
    )

    with pytest.raises(ValueError, match="defect decision.*evidence"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_cannot_hide_d3_evidence_as_l2(table_parser, monkeypatch):
    original = tabular_structure._new_projected_item

    def add_unemitted_proven_slot(**kwargs):
        item = original(**kwargs)
        item["proven_record_slots"].append(99)
        return item

    monkeypatch.setattr(tabular_structure, "_new_projected_item", add_unemitted_proven_slot)
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )
    output = audit["output_objects"][0]
    status, reason = tabular_structure.ENUMERATION_DECISIONS["R8"]
    output.update(
        {
            "matched_rule": "R8",
            "decision_chain_stop": "R8",
            "enumeration_status": status,
            "enumeration_reason": reason,
        }
    )

    with pytest.raises(ValueError, match="record count evidence.*D3"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)


def test_producer_audit_cannot_hide_d3_evidence_as_d1(table_parser, monkeypatch):
    original = tabular_structure._new_projected_item

    def add_unemitted_proven_slot(**kwargs):
        item = original(**kwargs)
        item["proven_record_slots"].append(99)
        return item

    monkeypatch.setattr(tabular_structure, "_new_projected_item", add_unemitted_proven_slot)
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )
    output = audit["output_objects"][0]
    status, reason = tabular_structure.ENUMERATION_DECISIONS["D1"]
    output.update(
        {
            "matched_rule": "D1",
            "decision_chain_stop": "D1",
            "enumeration_status": status,
            "enumeration_reason": reason,
        }
    )

    with pytest.raises(ValueError, match="record count evidence.*D3"):
        tabular_structure._validate_tabular_structure_producer_audit(audit)



def test_unproven_single_record_uses_the_default_l2_decision(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "State"])
    sheet.append(["A-1", "Open"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert {
        "enumeration_status": projection["tables"][0]["enumeration_status"],
        "enumeration_reason": projection["tables"][0]["enumeration_reason"],
        "matched_rule": projection["tables"][0]["matched_rule"],
    } == {
        "enumeration_status": "not_guaranteed_explained",
        "enumeration_reason": "record_axis_not_proven",
        "matched_rule": "R8",
    }


def _l2_reason_workbook_bytes(rule):
    workbook = Workbook()
    sheet = workbook.active
    if rule == "R1":
        sheet["A1"] = "Standalone note"
    elif rule == "R2":
        sheet.append(["Code", "Value"])
        sheet.append(["A-1", 1])
        sheet.append(["A-2", 2])
        sheet.append(["A-3", 3])
        sheet.row_dimensions[3].hidden = True
    elif rule == "R3":
        sheet.append([None, "Column A", "Column B"])
        sheet.append(["Row A", 1, 2])
        sheet.append(["Row B", 3, 4])
    elif rule == "R4":
        sheet.append(["Code", "Value"])
        sheet.append(["A-1", 1])
        sheet.append(["A-2", 2])
        sheet.append(["Aggregate", "=SUM(B2:B3)"])
    elif rule == "R5":
        sheet.append(["Code", "Value"])
        sheet.append(["A-1", 1])
        sheet.append(["A-2", 2])
        sheet.append(["Code", "Value"])
        sheet.append(["B-1", 3])
        sheet.append(["B-2", 4])
    elif rule == "R6":
        return _complete_table_with_partially_overlapping_unknown_bytes()
    elif rule == "R7":
        sheet.append(["Code", "Value"])
        fills = ("FFF2CC", "FFF2CC", "D9EAD3", "D9EAD3")
        for index, color in enumerate(fills, start=1):
            sheet.append([f"A-{index}", index])
            for cell in sheet[sheet.max_row]:
                cell.fill = PatternFill(fill_type="solid", fgColor=color)
    else:
        raise AssertionError(f"unhandled L2 rule: {rule}")
    return _save_workbook(workbook)


@pytest.mark.parametrize(
    ("rule", "reason"),
    [
        ("R1", "not_a_list"),
        ("R2", "total_unstable"),
        ("R3", "matrix_layout"),
        ("R4", "subtotal_rows_mixed"),
        ("R5", "multi_block_unseparated"),
        ("R6", "partial_overlap_continuation"),
        ("R7", "visual_only_boundary"),
    ],
)
def test_real_workbook_structures_reach_each_ordered_l2_reason(
    table_parser,
    rule,
    reason,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _l2_reason_workbook_bytes(rule),
        parser=table_parser,
    )

    assert projection["tables"]
    assert projection["rows"]
    assert all(table["source_total_count"] is None for table in projection["tables"])
    assert {table["matched_rule"] for table in projection["tables"]} == {rule}
    assert {table["enumeration_reason"] for table in projection["tables"]} == {reason}


def _collision_workbook_bytes(fixture_id):
    workbook = Workbook()
    sheet = workbook.active
    if fixture_id == "C12":
        sheet["A1"] = "=UNKNOWN_FUNCTION(1)"
        return _save_workbook(workbook)
    if fixture_id == "C16":
        sheet.cell(1, 1, "Earlier")
        sheet.cell(1, 2, "Detail")
        sheet.cell(4, 2, "Later")
        sheet.cell(4, 3, "Extra")
        return _save_workbook(workbook)
    if fixture_id == "C17":
        for column, color in enumerate(("FFF2CC", "FFF2CC", "D9EAD3", "D9EAD3"), start=1):
            sheet.cell(1, column, f"Text-{column}")
            sheet.cell(1, column).fill = PatternFill(fill_type="solid", fgColor=color)
        return _save_workbook(workbook)

    sheet.append(["Code", "State"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])

    if fixture_id in {"C23", "C34", "C37", "M02_R3_R4_R7"}:
        workbook = Workbook()
        sheet = workbook.active
        row_count = 5 if fixture_id in {"C37", "M02_R3_R4_R7"} else 4
        sheet.cell(1, 2, "Column A")
        sheet.cell(1, 3, "Column B")
        for row in range(2, row_count + 1):
            sheet.cell(row, 1, f"Row-{row}")
            sheet.cell(row, 2, row)
            sheet.cell(row, 3, row * 10)
        if fixture_id == "C23":
            sheet.row_dimensions[3].hidden = True
        if fixture_id in {"C34", "M02_R3_R4_R7"}:
            sheet.cell(row_count, 3, f"=SUM(C2:C{row_count - 1})")
        if fixture_id in {"C37", "M02_R3_R4_R7"}:
            for row, color in zip(range(2, row_count + 1), ("FFF2CC", "FFF2CC", "D9EAD3", "D9EAD3")):
                for column in range(1, 4):
                    sheet.cell(row, column).fill = PatternFill(fill_type="solid", fgColor=color)
        return _save_workbook(workbook)
    if fixture_id in {"C24", "C47", "M01_R2_R4_R7"}:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Code", "Value"])
        for row in range(2, 6):
            sheet.cell(row, 1, f"A-{row}")
            sheet.cell(row, 2, row)
        sheet.cell(6, 1, "Aggregate")
        sheet.cell(6, 2, "=SUM(B2:B5)")
        if fixture_id in {"C24", "M01_R2_R4_R7"}:
            sheet.row_dimensions[3].hidden = True
        if fixture_id in {"C47", "M01_R2_R4_R7"}:
            for row, color in zip(range(2, 7), ("FFF2CC", "FFF2CC", "D9EAD3", "D9EAD3", "D9EAD3")):
                for column in (1, 2):
                    sheet.cell(row, column).fill = PatternFill(fill_type="solid", fgColor=color)
        return _save_workbook(workbook)
    if fixture_id in {"C25", "C45", "C57"}:
        workbook = Workbook()
        sheet = workbook.active
        rows = (
            (1, ("Code", "Value")),
            (2, ("A-1", 1)),
            (3, ("A-2", 2)),
            (4, ("Code", "Value")),
            (5, ("B-1", 3)),
            (6, ("B-2", 4)),
            (7, ("B-3", 5)),
        )
        for row, values in rows:
            sheet.cell(row, 1, values[0])
            sheet.cell(row, 2, values[1])
        if fixture_id == "C25":
            sheet.row_dimensions[3].hidden = True
        if fixture_id == "C45":
            sheet.cell(7, 2, "=SUM(B5:B6)")
        if fixture_id == "C57":
            for row, color in zip(range(2, 8), ("FFF2CC", "FFF2CC", "FFF2CC", "FFF2CC", "D9EAD3", "D9EAD3")):
                for column in (1, 2):
                    sheet.cell(row, column).fill = PatternFill(fill_type="solid", fgColor=color)
        return _save_workbook(workbook)
    if fixture_id in {"C26", "C27"}:
        if fixture_id == "C26":
            sheet.cell(6, 2, "Later")
            sheet.cell(6, 3, "=UNKNOWN_FUNCTION(1)")
        else:
            for row, color in zip(range(2, 6), ("FFF2CC", "FFF2CC", "D9EAD3", "D9EAD3")):
                if row > 3:
                    sheet.cell(row, 1, f"A-{row}")
                    sheet.cell(row, 2, "Open")
                for column in (1, 2):
                    sheet.cell(row, column).fill = PatternFill(fill_type="solid", fgColor=color)
            sheet.row_dimensions[3].hidden = True
        return _save_workbook(workbook)
    if fixture_id == "C35":
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(1, 2, "Column A")
        sheet.cell(1, 3, "Column B")
        for row in range(2, 4):
            sheet.cell(row, 1, f"Row-{row}")
            sheet.cell(row, 2, row)
            sheet.cell(row, 3, row * 10)
        sheet.cell(4, 2, "Column A")
        sheet.cell(4, 3, "Column B")
        for row in range(5, 7):
            sheet.cell(row, 1, f"Row-{row}")
            sheet.cell(row, 2, row)
            sheet.cell(row, 3, row * 10)
        return _save_workbook(workbook)
    if fixture_id == "C36":
        sheet.cell(6, 2, None)
        sheet.cell(6, 3, "Column A")
        sheet.cell(6, 4, "Column B")
        sheet.cell(7, 2, "Row A")
        sheet.cell(7, 3, 1)
        sheet.cell(7, 4, 2)
        sheet.cell(8, 2, "Row B")
        sheet.cell(8, 3, 3)
        sheet.cell(8, 4, 4)
    elif fixture_id == "C46":
        sheet.cell(6, 2, "B-1")
        sheet.cell(6, 3, 1)
        sheet.cell(7, 2, "B-2")
        sheet.cell(7, 3, 2)
        sheet.cell(8, 2, "Aggregate")
        sheet.cell(8, 3, "=SUM(C6:C7)")
    elif fixture_id == "C56":
        for row, values in (
            (6, ("Code", "Value")),
            (7, ("B-1", 1)),
            (8, ("B-2", 2)),
            (9, ("Code", "Value")),
            (10, ("C-1", 3)),
            (11, ("C-2", 4)),
        ):
            sheet.cell(row, 2, values[0])
            sheet.cell(row, 3, values[1])
    elif fixture_id == "C67":
        fills = ("FFF2CC", "FFF2CC", "D9EAD3", "D9EAD3")
        sheet.cell(6, 2, "Code")
        sheet.cell(6, 3, "Value")
        for offset, color in enumerate(fills, start=7):
            sheet.cell(offset, 2, f"B-{offset}")
            sheet.cell(offset, 3, offset)
            for column in (2, 3):
                sheet.cell(offset, column).fill = PatternFill(fill_type="solid", fgColor=color)
    else:
        raise AssertionError(f"unhandled collision fixture: {fixture_id}")
    return _save_workbook(workbook)


@pytest.mark.parametrize(
    ("fixture_id", "expected_rules", "expected_stop"),
    [
        ("C12", {"R1", "R2"}, "R1"),
        ("C16", {"R1", "R6"}, "R1"),
        ("C17", {"R1", "R7"}, "R1"),
        ("C23", {"R2", "R3"}, "R2"),
        ("C24", {"R2", "R4"}, "R2"),
        ("C25", {"R2", "R5"}, "R2"),
        ("C26", {"R2", "R6"}, "R2"),
        ("C27", {"R2", "R7"}, "R2"),
        ("C34", {"R3", "R4"}, "R3"),
        ("C35", {"R3", "R5"}, "R3"),
        ("C36", {"R3", "R6"}, "R3"),
        ("C37", {"R3", "R7"}, "R3"),
        ("C45", {"R4", "R5"}, "R4"),
        ("C46", {"R4", "R6"}, "R4"),
        ("C47", {"R4", "R7"}, "R4"),
        ("C56", {"R5", "R6"}, "R5"),
        ("C57", {"R5", "R7"}, "R5"),
        ("C67", {"R6", "R7"}, "R6"),
        ("M01_R2_R4_R7", {"R2", "R4", "R7"}, "R2"),
        ("M02_R3_R4_R7", {"R3", "R4", "R7"}, "R3"),
    ],
)
def test_real_workbook_collision_vectors_use_production_predicates(
    table_parser,
    monkeypatch,
    fixture_id,
    expected_rules,
    expected_stop,
):
    captured = []
    ordered = tabular_structure._ordered_enumeration_rule

    def capture(predicates, l1_rule):
        captured.append(dict(predicates))
        return ordered(predicates, l1_rule)

    monkeypatch.setattr(tabular_structure, "_ordered_enumeration_rule", capture)
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _collision_workbook_bytes(fixture_id),
        parser=table_parser,
    )

    expected_vector = {rule: rule in expected_rules for rule in tabular_structure.NEGATIVE_ENUMERATION_RULES}
    assert expected_vector in captured
    assert expected_stop in {table["matched_rule"] for table in projection["tables"]}


def test_visual_only_boundary_uses_style_clusters_without_formula_or_merge(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value"])
    for row, color in ((2, "FFF2CC"), (3, "FFF2CC"), (4, "D9EAD3"), (5, "D9EAD3")):
        sheet.cell(row, 1, f"A-{row}")
        sheet.cell(row, 2, row)
        for column in (1, 2):
            sheet.cell(row, column).fill = PatternFill(fill_type="solid", fgColor=color)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["matched_rule"] == "R7"


def test_dense_matrix_with_occupied_corner_stops_at_matrix_rule(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Dimension", "Column A", "Column B"])
    sheet.append(["Row A", 1, 2])
    sheet.append(["Row B", 3, 4])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] is None
    assert projection["tables"][0]["matched_rule"] == "R3"


def test_dense_three_column_text_list_is_not_a_matrix(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Status", "Owner"])
    sheet.append(["A-1", "Open", "Team-1"])
    sheet.append(["A-2", "Closed", "Team-2"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] == 2
    assert projection["tables"][0]["matched_rule"] == "L1-07"


def test_dense_text_list_without_digit_markers_is_not_a_matrix(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "State", "Owner"])
    sheet.append(["Alpha", "Open", "East"])
    sheet.append(["Beta", "Closed", "West"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert projection["tables"][0]["matched_rule"] == "R8"


def test_dense_numeric_record_list_is_not_a_matrix(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Count", "Owner"])
    sheet.append(["A-1", 1, "East"])
    sheet.append(["A-2", 2, "West"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] == 2
    assert projection["tables"][0]["matched_rule"] == "L1-07"


def test_mixed_value_matrix_cannot_claim_a_complete_record_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Dimension", "Count", "Status"])
    sheet.append(["North", 1, "High"])
    sheet.append(["South", 2, "Low"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert projection["tables"][0]["matched_rule"] == "R3"


def test_sparse_matrix_cannot_claim_a_complete_record_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Dimension", "First", "Second"])
    sheet.append(["North", 1, None])
    sheet.append(["South", None, 2])
    sheet.append(["West", 3, 4])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert projection["tables"][0]["matched_rule"] == "R3"


def test_matrix_decision_considers_the_complete_candidate_region(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Count", "Score"])
    sheet.append(["Alice", 1, 10])
    sheet.append(["Bob", 2, 20])
    sheet.append(["Carol", 3, "Pending"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert projection["tables"][0]["matched_rule"] == "R8"


def test_single_style_cluster_does_not_trigger_visual_only_boundary(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value"])
    for row in range(2, 6):
        sheet.cell(row, 1, f"A-{row}")
        sheet.cell(row, 2, row)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["matched_rule"] == "L1-07"


def test_multicell_non_list_without_isomorphic_slots_reaches_r1(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Prepared by"
    sheet["B1"] = "Reviewed by"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["matched_rule"] == "R1"


def test_hidden_header_does_not_make_a_stable_record_total_unstable(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value"])
    sheet.append(["A-1", 1])
    sheet.append(["A-2", 2])
    sheet.row_dimensions[1].hidden = True

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] == 2
    assert projection["tables"][0]["matched_rule"] == "L1-07"


def test_uncached_nonaggregate_formula_reaches_total_unstable(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value"])
    sheet.append(["A-1", 1])
    sheet.append(["A-2", "=B2"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert projection["tables"][0]["matched_rule"] == "R2"


def test_table_ref_rejects_membership_identity_tampering(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )
    original = projection["tables"][0]["table_ref"]
    prefix, version, membership, identity = original.split("_")
    changed_membership = "0" * 64 if membership != "0" * 64 else "1" * 64
    tampered = f"{prefix}_{version}_{changed_membership}_{identity}"
    projection["tables"][0]["table_ref"] = tampered
    for row in projection["rows"]:
        row["table_ref_kwd"] = tampered
        row["row_ref_kwd"] = f"{tampered}:{row['row_ordinal_int']}"
        row["id"] = "tsr_v1_" + tabular_structure._versioned_digest(
            "tabular-row-record/v1",
            projection["producer_generation_ref"],
            row["row_ref_kwd"],
        )

    with pytest.raises(ValueError, match="table reference"):
        validate_tabular_structure_projection(projection)


def _workbook_bytes(
    *,
    headers=("Code", "Description", "Force"),
    include_unknown=False,
    include_merged_body=False,
    include_note=True,
    include_context=False,
    rows=3,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inspection"
    if include_context:
        sheet.merge_cells("A1:C1")
        sheet["A1"] = "\u9879\u76ee \u00b5\u03a9\u2103\u0085\u202e"
        sheet["A3"] = "Program"
        sheet["B3"] = "Platform-X"
        sheet["A4"], sheet["B4"], sheet["C4"] = headers
    else:
        sheet.append(list(headers))

    for index in range(1, rows + 1):
        sheet.append(["R-DUP" if index <= 2 else f"R-{index:04d}", "Repeated" if index <= 2 else "Item", 7.5])

        if include_merged_body and index == 1:
            body_row = sheet.max_row + 1
            sheet.merge_cells(start_row=body_row, start_column=1, end_row=body_row, end_column=3)
            sheet.cell(body_row, 1, "R-MERGED")

    if include_unknown:
        sheet.append(["Sparse body row", None, None])

    if include_note:
        note_row = sheet.max_row + 1
        sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)
        sheet.cell(note_row, 1, "For reference only")

    second = workbook.create_sheet("Secondary")
    second.append(["Part", "Status"])
    second.append(["P-1", "Approved"])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
def _generation_ref():
    return str(uuid.uuid4())


def _save_workbook(workbook):
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _vertical_multi_region_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Code", "Status"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    sheet.append([None, None])
    sheet.append([None, None])
    sheet.append(["Code", "Status"])
    sheet.append(["B-1", "Open"])
    sheet.append(["B-2", "Closed"])
    return _save_workbook(workbook)


def _horizontal_multi_region_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Left code", "Left status", None, None, "Right code", "Right status"])
    sheet.append(["L-1", "Open", None, None, "R-1", "Closed"])
    sheet.append(["L-2", "Closed", None, None, "R-2", "Open"])
    return _save_workbook(workbook)


def _horizontal_varying_merge_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous horizontal merges"
    sheet.append(["Field A", "Field B", "Field C", "Field D"])
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "A-1"
    sheet["C2"] = "C-1"
    sheet["D2"] = "D-1"
    sheet.merge_cells("B3:C3")
    sheet["A3"] = "A-2"
    sheet["B3"] = "B-2"
    sheet["D3"] = "D-2"
    sheet.merge_cells("C4:D4")
    sheet["A4"] = "A-3"
    sheet["B4"] = "B-3"
    sheet["C4"] = "C-3"
    return _save_workbook(workbook)


def _vertical_varying_merge_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous vertical merges"
    sheet.append(["Field A", "Field B", "Field C", "Field D"])
    sheet.merge_cells("A2:A3")
    sheet["A2"] = "A-group-1"
    sheet["B2"] = "B-1"
    sheet["C2"] = "C-1"
    sheet["D2"] = "D-1"
    sheet["B3"] = "B-2"
    sheet["C3"] = "C-2"
    sheet["D3"] = "D-2"
    sheet.merge_cells("B4:B5")
    sheet["A4"] = "A-3"
    sheet["B4"] = "B-group-2"
    sheet["C4"] = "C-3"
    sheet["D4"] = "D-3"
    sheet["A5"] = "A-4"
    sheet["C5"] = "C-4"
    sheet["D5"] = "D-4"
    return _save_workbook(workbook)


def _mixed_row_merge_supplier_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous mixed merges"
    sheet.append(["Group", "Material", "Supplier", "Evidence", "Note"])

    # Each record remains physically present, but each row has a different
    # combination of vertical and horizontal source merges.
    sheet.merge_cells("A2:A3")
    sheet["A2"] = "G-1"
    sheet.merge_cells("B2:C2")
    sheet["B2"] = "M-1"
    sheet["D2"] = "S-1"
    sheet["E2"] = "Verified"

    sheet.merge_cells("B3:D3")
    sheet["B3"] = "M-2 / S-2"
    sheet["E3"] = "Pending"

    sheet.merge_cells("A4:A5")
    sheet["A4"] = "G-2"
    sheet.merge_cells("C4:E4")
    sheet["C4"] = "S-3"
    sheet["B4"] = "M-3"

    sheet.merge_cells("B5:C5")
    sheet["B5"] = "M-4"
    sheet["D5"] = "S-4"
    sheet["E5"] = "Approved"

    # These columns are a disjoint link sidecar, not table columns.
    sheet["H1"] = "Sidecar label"
    sheet["I1"] = "https://example.invalid/material"
    sheet["I1"].hyperlink = "https://example.invalid/material"
    return _save_workbook(workbook)


def _vertical_complete_with_headerless_continuation_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Code", "Status"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    sheet.append([None, None])
    sheet.append([None, None])
    sheet.append(["B-1", "Open"])
    return _save_workbook(workbook)


def _sheet_level_repeated_form_segments_bytes(
    *,
    second_context_name="零件号",
    second_context_value="2906150-PH01",
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PFMEA"
    rows = (
        ("10", "外观检查", "缺陷", "5"),
        ("20", "尺寸检查", "超差", "7"),
        ("30", "装配", "漏装", "6"),
        ("40", "终检", "错装", "8"),
    )
    for segment_index, (segment_start, segment_rows) in enumerate(
        ((1, rows[:2]), (9, rows[2:])),
    ):
        sheet.cell(segment_start, 1, "潜在失效模式及后果分析（PFMEA）")
        sheet.cell(
            segment_start + 1,
            1,
            "零件号" if segment_index == 0 else second_context_name,
        )
        sheet.cell(
            segment_start + 1,
            2,
            "2906150-PH01" if segment_index == 0 else second_context_value,
        )
        for column, value in enumerate(
            ("过程编号", "过程名称", "潜在失效模式", "严重度"),
            start=1,
        ):
            sheet.cell(segment_start + 2, column, value)
        for row_offset, values in enumerate(segment_rows, start=3):
            for column, value in enumerate(values, start=1):
                sheet.cell(segment_start + row_offset, column, value)
    return _save_workbook(workbook)


def _four_page_repeated_form_with_context_value_changes_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PFMEA"
    pages = (
        ("2906150-PH01", (("10", "外观检查", "缺陷", "5"), ("20", "尺寸检查", "超差", "7"))),
        ("2906150-PH01", (("30", "装配", "漏装", "6"), ("40", "终检", "错装", "8"))),
        ("2906150-PH01", (("50", "包装", "破损", "4"), ("60", "发运", "错发", "3"))),
        ("2906151-SE01", (("10", "外观检查", "缺陷", "5"), ("20", "尺寸检查", "超差", "7"))),
    )
    for page_index, (part_number, records) in enumerate(pages):
        start = page_index * 8 + 1
        sheet.cell(start, 1, "潜在失效模式及后果分析（PFMEA）")
        sheet.cell(start + 1, 1, "零件号")
        sheet.cell(start + 1, 2, part_number)
        for column, value in enumerate(
            ("过程编号", "过程名称", "潜在失效模式", "严重度"),
            start=1,
        ):
            sheet.cell(start + 2, column, value)
        for row_offset, record in enumerate(records, start=3):
            for column, value in enumerate(record, start=1):
                sheet.cell(start + row_offset, column, value)
    return _save_workbook(workbook)


def _headerless_predecessor_with_named_following_page_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous paged register"
    records = (
        ("10", "Missing", "Delay", "4", "Cause A", "3", "Prevent A", "Detect A", "4", "48"),
        ("20", "Wrong", "Rework", "6", "Cause B", "2", "Prevent B", "Detect B", "3", "36"),
        ("30", "Damaged", "Reject", "5", "Cause C", "3", "Prevent C", "Detect C", "4", "60"),
        ("40", "Loose", "Noise", "7", "Cause D", "2", "Prevent D", "Detect D", "3", "42"),
    )

    def context(start):
        sheet.cell(start, 1, "Submission")
        sheet.cell(start + 2, 1, "Reference")
        sheet.cell(start + 2, 4, "Description")
        sheet.cell(start + 2, 9, "Program")
        sheet.cell(start + 2, 12, "Prepared")
        sheet.cell(start + 2, 13, "Person")
        sheet.cell(start + 2, 14, "Phone")
        sheet.merge_cells(
            start_row=start + 2,
            start_column=15,
            end_row=start + 3,
            end_column=18,
        )
        sheet.cell(start + 2, 15, "Contact")
        sheet.merge_cells(
            start_row=start + 3,
            start_column=1,
            end_row=start + 3,
            end_column=3,
        )
        sheet.cell(start + 3, 1, "R-001")
        sheet.merge_cells(
            start_row=start + 3,
            start_column=4,
            end_row=start + 3,
            end_column=8,
        )
        sheet.cell(start + 3, 4, "Assembly")
        sheet.merge_cells(
            start_row=start + 3,
            start_column=9,
            end_row=start + 3,
            end_column=11,
        )
        sheet.cell(start + 3, 9, "Program A")
        sheet.cell(start + 3, 12, "Approved")
        sheet.cell(start + 4, 1, "Core team")
        sheet.merge_cells(
            start_row=start + 4,
            start_column=2,
            end_row=start + 5,
            end_column=3,
        )
        sheet.cell(start + 4, 2, "Team")
        sheet.cell(start + 4, 4, "Modified")
        sheet.cell(start + 4, 12, "Approval")
        sheet.merge_cells(
            start_row=start + 4,
            start_column=15,
            end_row=start + 4,
            end_column=18,
        )
        sheet.merge_cells(
            start_row=start + 5,
            start_column=4,
            end_row=start + 5,
            end_column=7,
        )
        sheet.cell(start + 5, 12, "Approval date")
        sheet.merge_cells(
            start_row=start + 5,
            start_column=15,
            end_row=start + 5,
            end_column=18,
        )

    def record(row, values):
        for column, value in zip((1, 2, 3, 4, 6, 7, 8, 9, 10, 11), values):
            sheet.cell(row, column, value)

    context(1)
    for row, values in enumerate(records[:2], start=7):
        record(row, values)
    sheet.merge_cells("B10:N10")
    sheet["B10"] = "Anonymous paged register"

    context(15)
    for column, value in enumerate(
        (
            "Process",
            "Mode",
            "Effect",
            "Severity",
            "Class",
            "Cause",
            "Occurrence",
            "Prevention",
            "Detection",
            "Rating",
            "Priority",
            "Action",
            "Owner",
            "Result",
            "Final severity",
            "Final occurrence",
            "Final detection",
            "Final priority",
        ),
        start=1,
    ):
        sheet.cell(21, column, value)
    for row, values in enumerate(records[2:], start=22):
        record(row, values)
    return _save_workbook(workbook)


def _vertical_complete_with_subset_headerless_continuation_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Code", "Status", "Owner", "Revision", "Date"])
    sheet.append(["A-1", "Open", "Team-1", "R1", "2026-01-01"])
    sheet.append(["A-2", "Closed", "Team-2", "R2", "2026-01-02"])
    sheet.append([None, None, None, None, None])
    sheet.append([None, None, None, None, None])
    sheet.append(["B-1", "Open", "Team-3", "R3"])
    return _save_workbook(workbook)


def _vertical_complete_with_superset_headerless_continuation_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Code", "Status", "Owner", "Revision"])
    sheet.append(["A-1", "Open", "Team-1", "R1"])
    sheet.append(["A-2", "Closed", "Team-2", "R2"])
    sheet.append([None, None, None, None, None])
    sheet.append([None, None, None, None, None])
    sheet.append(["B-1", "Open", "Team-3", "R3", "2026-01-03"])
    return _save_workbook(workbook)


def _complete_table_with_partially_overlapping_unknown_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Code", "Status"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    sheet.append([None, None, None])
    sheet.append([None, None, None])
    sheet.append([None, "Unresolved", "Extra"])
    return _save_workbook(workbook)


def _horizontal_complete_with_headerless_sibling_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Left code", "Left status", None, None, "R-1", "Closed"])
    sheet.append(["L-1", "Open", None, None, "R-2", "Open"])
    sheet.append(["L-2", "Closed", None, None, "R-3", "Closed"])
    return _save_workbook(workbook)


def _complete_table_with_independent_repeated_header_table_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Left code", "Left status", None, None, "Right code", "Right status"])
    sheet.append(["L-1", "Open", None, None, "R-1", "Open"])
    sheet.append(["L-2", "Closed", None, None, "Right code", "Right status"])
    sheet.append([None, None, None, None, "R-2", "Closed"])
    return _save_workbook(workbook)


def _complete_table_with_isolated_annotation_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Code", "Status"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    sheet.cell(row=10, column=10, value="Annotation")
    return _save_workbook(workbook)


def _complete_table_with_axis_aligned_unknown_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Code", "Status"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    sheet.cell(row=10, column=1, value="Unresolved")
    return _save_workbook(workbook)


def _complete_table_with_context_only_unknown_overlap_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Context", "Code", "State"])
    sheet.append([None, "A-1", "Open"])
    sheet.append([None, "A-2", "Closed"])
    sheet.cell(row=10, column=1, value="Annotation")
    return _save_workbook(workbook)


def _complete_table_with_g_sensitive_sibling_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Code", "Status", None, None, "R1 code", "R1 status", None, "R2 code", "R2 status"])
    sheet.append(["A-1", "Open", None, None, "B-1", "Open", None, "C-1", "Closed"])
    sheet.append(["A-2", "Closed", None, None, "B-2", "Closed", None, "C-2", "Open"])
    return _save_workbook(workbook)


def _single_column_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Code"])
    sheet.append(["S-1"])
    sheet.append(["S-2"])
    return _save_workbook(workbook)


def _merged_header_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "Items"
    sheet["A2"] = "Code"
    sheet["B2"] = "Status"
    sheet.append(["M-1", "Open"])
    sheet.append(["M-2", "Closed"])
    return _save_workbook(workbook)


def _nested_sparse_header_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous nested header"
    sheet.merge_cells("A1:H2")
    sheet["A1"] = "Anonymous report"
    sheet.merge_cells("A3:B3")
    sheet["A3"] = "Identity"
    sheet.merge_cells("C3:H3")
    sheet["C3"] = "Details"
    sheet["A4"] = "No"
    sheet["B4"] = "Code"
    sheet.merge_cells("C4:D4")
    sheet["C4"] = "Material"
    sheet.merge_cells("E4:F4")
    sheet["E4"] = "Supplier"
    sheet.merge_cells("G4:H4")
    sheet["G4"] = "Process"
    for row, values in enumerate(
        (
            (1, "K-1", "Rubber", "Acme", "Extrude"),
            (2, "K-2", "Steel", "Beta", "Press"),
            (3, "K-3", "Wire", "Gamma", "Draw"),
        ),
        start=5,
    ):
        sheet.cell(row=row, column=1, value=values[0])
        sheet.cell(row=row, column=2, value=values[1])
        sheet.cell(row=row, column=3, value=values[2])
        sheet.cell(row=row, column=5, value=values[3])
        sheet.cell(row=row, column=7, value=values[4])
    return _save_workbook(workbook)


def _merged_header_continuation_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous header continuation"
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "Anonymous register"
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "Identity"
    sheet.merge_cells("C2:H2")
    sheet["C2"] = "Details"
    sheet.merge_cells("A3:B3")
    sheet["A3"] = "Code"
    sheet.merge_cells("C3:D3")
    sheet["C3"] = "Material"
    sheet.merge_cells("E3:F3")
    sheet["E3"] = "Supplier"
    sheet.merge_cells("G3:H3")
    sheet["G3"] = "Process"
    for column, value in enumerate(
        ("Code type", "Code value", "Code note", "Supplier name", "Supplier area"),
        start=2,
    ):
        sheet.cell(row=4, column=column, value=value)
    sheet["A4"] = "No"
    sheet["H4"] = "Process name"
    sheet.merge_cells("B5:C5")
    sheet["B5"] = "Code detail"
    sheet["D5"] = "Code note"
    sheet["E5"] = "Supplier name"
    sheet["F5"] = "Supplier area"
    sheet["G5"] = "Process type"
    sheet["H5"] = "Process name"
    sheet["A5"] = "No"
    sheet["H5"] = "Process name"
    for row, values in enumerate(
        (
            (None, None, "K-1", "Rubber", "Acme", "East", "Mix", "Extrude"),
            (None, None, "K-2", "Steel", "Beta", "West", "Mix", "Press"),
            (None, None, "K-3", "Wire", "Gamma", "North", "Mix", "Draw"),
        ),
        start=6,
    ):
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row, column=column, value=value)
    return _save_workbook(workbook)


def _optional_field_shape_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous optional field"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Anonymous register"
    for column, value in enumerate(("Code", "Description", "Status", "Optional"), start=1):
        sheet.cell(row=2, column=column, value=value)
    for row, values in enumerate(
        (
            ("A-1", "First", "Open", "Note"),
            ("A-2", "Second", "Open", None),
            ("A-3", "Third", "Closed", None),
        ),
        start=3,
    ):
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row, column=column, value=value)
    return _save_workbook(workbook)


def _multilevel_sparse_table_with_context_child_bytes(*, context_row_count=1):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.merge_cells("A2:H3")
    sheet["A2"] = "Anonymous catalogue"
    sheet.merge_cells("A4:B4")
    sheet["A4"] = "Identity"
    sheet.merge_cells("C4:D4")
    sheet["C4"] = "Classification"
    sheet.merge_cells("E4:G4")
    sheet["E4"] = "Attributes"
    for row in range(4, 4 + context_row_count):
        sheet.cell(row=row, column=10, value=f"Anonymous context {row}")
    sheet.merge_cells("A5:B5")
    sheet["A5"] = "Reference"
    sheet.merge_cells("C5:D5")
    sheet["C5"] = "Grouping"
    sheet.merge_cells("E5:G5")
    sheet["E5"] = "Details"
    sheet.merge_cells("A6:H6")
    sheet["A6"] = "Record fields"
    sheet.merge_cells("A7:A8")
    sheet["A7"] = "Number"
    sheet.merge_cells("B7:B8")
    sheet["B7"] = "Code"
    sheet.merge_cells("C7:D7")
    sheet["C7"] = "Group"
    sheet["C8"] = "Type"
    sheet["D8"] = "State"
    sheet.merge_cells("E7:E8")
    sheet["E7"] = "Owner"
    sheet.merge_cells("F7:G7")
    sheet["F7"] = "Limits"
    sheet["F8"] = "Lower"
    sheet["G8"] = "Upper"
    sheet.merge_cells("H7:H8")
    sheet["H7"] = "Optional"
    for index, row in enumerate(range(9, 19), start=1):
        values = [
            index,
            f"R-{index:02d}",
            "A",
            "Open",
            "Team",
            "0",
            "1",
            "present" if index == 8 else None,
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row, column=column, value=value)
    return _save_workbook(workbook)


def _competing_header_depth_candidate_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous candidate selection"
    sheet.merge_cells("A1:H2")
    sheet["A1"] = "Anonymous register"
    for row in (3, 4):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.cell(row=row, column=1, value=f"Context {row} A")
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        sheet.cell(row=row, column=3, value=f"Context {row} B")
        sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        sheet.cell(row=row, column=5, value=f"Context {row} C")
        sheet.cell(row=row, column=8, value=f"Context {row} D")
    sheet["J3"] = "Anonymous sidecar"
    sheet["J4"] = "Anonymous sidecar value"
    sheet.merge_cells("A5:H5")
    sheet["A5"] = "Record fields"
    sheet.merge_cells("A6:A7")
    sheet["A6"] = "Sequence"
    sheet.merge_cells("B6:B7")
    sheet["B6"] = "Code"
    sheet.merge_cells("C6:D6")
    sheet["C6"] = "Grouping"
    sheet["C7"] = "Group"
    sheet["D7"] = "State"
    sheet.merge_cells("E6:E7")
    sheet["E6"] = "Owner"
    sheet.merge_cells("F6:G7")
    sheet["F6"] = "Evidence"
    sheet.merge_cells("H6:H7")
    sheet["H6"] = "Optional"
    for index, row in enumerate(range(8, 18), start=1):
        values = [
            index,
            f"R-{index:02d}",
            "A",
            "Open",
            "Team",
            "0",
            "1",
            "present" if index == 8 else None,
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row, column=column, value=value)
        sheet.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
    return _save_workbook(workbook)


def _context_form_before_multilevel_header_workbook_bytes(*, record_count=2):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous performance report"
    sheet.merge_cells("A1:J1")
    sheet["A1"] = "Anonymous performance report"

    sheet.merge_cells("A2:E3")
    sheet["A2"] = "Unpaired context"
    sheet.merge_cells("F2:G2")
    sheet["F2"] = "Provider"
    sheet.merge_cells("H2:J2")
    sheet["H2"] = "Laboratory"
    sheet.merge_cells("F3:G3")
    sheet["F3"] = "Provider A"
    sheet.merge_cells("H3:J3")
    sheet["H3"] = "Laboratory A"
    sheet["K2"] = "Sidecar action"

    for row, values in (
        (4, ("Component", "Reference", "Revision")),
        (5, ("Component A", "Reference A", "Revision A")),
    ):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        sheet.cell(row, 1, values[0])
        sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        sheet.cell(row, 4, values[1])
        sheet.merge_cells(start_row=row, start_column=6, end_row=row, end_column=10)
        sheet.cell(row, 6, values[2])

    sheet.merge_cells("A6:B6")
    sheet["A6"] = "Fixture"
    sheet["C6"] = "Fixture A"
    sheet.merge_cells("D6:E6")
    sheet["D6"] = "Cavity count"
    sheet["F6"] = "Cavity count A"
    sheet["G6"] = "Cavity"
    sheet.merge_cells("H6:J6")
    sheet["H6"] = "Cavity A"

    for column, value in enumerate(
        (
            "Sequence",
            "Test",
            "Description",
            "Requirement",
            "Quantity",
            "Equipment",
            "Date",
        ),
        start=1,
    ):
        sheet.merge_cells(
            start_row=7,
            start_column=column,
            end_row=8,
            end_column=column,
        )
        sheet.cell(7, column, value)
    sheet.merge_cells("H7:J7")
    sheet["H7"] = "Measured"
    for column, value in enumerate(("Run 1", "Run 2", "Run 3"), start=8):
        sheet.cell(8, column, value)

    for index, row in enumerate(range(9, 9 + record_count), start=1):
        for column, value in enumerate(
            (
                index,
                f"Test {index}",
                f"Description {index}",
                f"Requirement {index}",
                "1",
                f"Equipment {index}",
                "2026-08-21",
                f"M-{index}-1",
                f"M-{index}-2",
                f"M-{index}-3",
            ),
            start=1,
        ):
            sheet.cell(row, column, value)
    return _save_workbook(workbook)


def _context_form_before_horizontally_merged_leaf_empty_axis_bytes(
    *,
    merge_record_fields=False,
    numeric_record_key=True,
    rectangular_leaf_merges=True,
    rectangular_header_continuation=False,
    record_count=0,
    single_semantic_path=False,
):
    workbook = load_workbook(
        BytesIO(
            _context_form_before_multilevel_header_workbook_bytes(
                record_count=record_count
            )
        )
    )
    sheet = workbook.active
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row >= 7:
            sheet.unmerge_cells(str(merged))
    for row in range(7, 9):
        for column in range(1, 11):
            sheet.cell(row, column).value = None

    if single_semantic_path:
        sheet.merge_cells("A7:J7")
        sheet["A7"] = "Record field"
        sheet.merge_cells("A8:J8")
        sheet["A8"] = "Record field"
        return _save_workbook(workbook)

    leaf_merges = (
        (
            ("A7:A8", "Sequence"),
            ("B7:C8", "Material"),
            ("D7:D8", "Requirement"),
            ("E7:F8", "Standard"),
            ("G7:G8", "Result"),
        )
        if rectangular_leaf_merges
        else (
            ("A7:A8", "Sequence"),
            ("B7:C7", "Material"),
            ("B8:C8", "Detail"),
            ("D7:D8", "Requirement"),
            ("E7:F7", "Standard"),
            ("E8:F8", "Clause"),
            ("G7:G8", "Result"),
        )
    )
    for cell_range, value in leaf_merges:
        sheet.merge_cells(cell_range)
        sheet[cell_range.split(":", 1)[0]] = value
    sheet.merge_cells("H7:J7")
    sheet["H7"] = "Outcome"
    for column, value in enumerate(("Measured", "Unit", "Conclusion"), start=8):
        sheet.cell(8, column).value = value
    if merge_record_fields:
        for row in range(9, 9 + record_count):
            for start_column, end_column in ((2, 3), (5, 6)):
                sheet.cell(row, end_column).value = None
                sheet.merge_cells(
                    start_row=row,
                    start_column=start_column,
                    end_row=row,
                    end_column=end_column,
                )
    if not numeric_record_key:
        for index, row in enumerate(range(9, 9 + record_count), start=1):
            sheet.cell(row, 1).value = f"Header level {index}"
    if rectangular_header_continuation:
        for row in range(9, 9 + record_count):
            for column in range(2, 4):
                sheet.cell(row, column).value = None
        sheet.merge_cells(
            start_row=9,
            start_column=2,
            end_row=8 + record_count,
            end_column=3,
        )
        sheet["B9"] = "Header continuation"
    return _save_workbook(workbook)


def _context_form_with_an_unused_trailing_table_column_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous characteristics"
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "Anonymous characteristics"
    for row, values in (
        (2, ("Provider", "Provider A", "Component", "Component A")),
        (3, ("Program", "Program A", "Reference", "Reference A")),
    ):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.cell(row, 1, values[0])
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        sheet.cell(row, 3, values[1])
        sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        sheet.cell(row, 5, values[2])
        sheet.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        sheet.cell(row, 8, values[3])
    sheet.merge_cells("A4:A5")
    sheet["A4"] = "Sequence"
    sheet.merge_cells("B4:B5")
    sheet["B4"] = "Operation"
    sheet.merge_cells("C4:D4")
    sheet["C4"] = "Characteristic"
    sheet["C5"] = "Product"
    sheet["D5"] = "Process"
    for column, value in enumerate(
        ("Class", "Requirement", "Method", "Note"),
        start=5,
    ):
        sheet.merge_cells(
            start_row=4,
            start_column=column,
            end_row=5,
            end_column=column,
        )
        sheet.cell(4, column, value)
    for row, values in (
        (6, (1, "Inspect", "Surface", "Fixture", "A", "Clean", "Visual")),
        (7, (2, "Measure", "Length", "Gauge", "B", "10", "Caliper")),
    ):
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column, value)
    return _save_workbook(workbook)


def _multilevel_repeated_form_with_optional_parent_bytes(
    *,
    second_process_header="Process",
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous paged register"
    for page_index, start in enumerate((1, 9)):
        sheet.merge_cells(
            start_row=start,
            start_column=1,
            end_row=start,
            end_column=4,
        )
        sheet.cell(start, 1, "Anonymous paged register")
        sheet.cell(start + 1, 1, "Reference")
        sheet.cell(start + 1, 2, f"R-{page_index + 1}")
        sheet.merge_cells(
            start_row=start + 2,
            start_column=1,
            end_row=start + 3,
            end_column=1,
        )
        sheet.cell(start + 2, 1, "Sequence")
        if page_index == 0:
            sheet.merge_cells(
                start_row=start + 2,
                start_column=2,
                end_row=start + 2,
                end_column=3,
            )
            sheet.cell(start + 2, 2, "Characteristic")
        sheet.cell(start + 3, 2, "Product")
        sheet.cell(
            start + 3,
            3,
            "Process" if page_index == 0 else second_process_header,
        )
        sheet.merge_cells(
            start_row=start + 2,
            start_column=4,
            end_row=start + 3,
            end_column=4,
        )
        sheet.cell(start + 2, 4, "Status")
        for row_offset, values in enumerate(
            (
                (page_index * 2 + 1, "A", "B", "Open"),
                (page_index * 2 + 2, "C", "D", "Closed"),
            ),
            start=4,
        ):
            for column, value in enumerate(values, start=1):
                sheet.cell(start + row_offset, column, value)
    return _save_workbook(workbook)


def _context_form_with_a_right_side_action_before_multilevel_header_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous characteristics"
    sheet.merge_cells("A2:H3")
    sheet["A2"] = "Anonymous characteristics"
    for row, values in (
        (4, ("Provider", "Provider A", "Component", "Component A")),
        (5, ("Program", "Program A", "Reference", "Reference A")),
    ):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.cell(row, 1, values[0])
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        sheet.cell(row, 3, values[1])
        sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        sheet.cell(row, 5, values[2])
        sheet.cell(row, 8, values[3])
    sheet["J4"] = "Sidecar action"
    sheet.merge_cells("A6:H6")
    sheet["A6"] = "Anonymous component class"
    for column, value in enumerate(
        ("Sequence", "Operation"),
        start=1,
    ):
        sheet.merge_cells(
            start_row=7,
            start_column=column,
            end_row=8,
            end_column=column,
        )
        sheet.cell(7, column, value)
    sheet.merge_cells("C7:D7")
    sheet["C7"] = "Characteristic"
    sheet["C8"] = "Product"
    sheet["D8"] = "Process"
    for column, value in enumerate(
        ("Class", "Requirement", "Method", "Note"),
        start=5,
    ):
        sheet.merge_cells(
            start_row=7,
            start_column=column,
            end_row=8,
            end_column=column,
        )
        sheet.cell(7, column, value)
    for row, values in (
        (9, (1, "Inspect", "Surface", "Fixture", "SC", "Clean", "Visual", "")),
        (10, (2, "Measure", "Length", "Gauge", "CC", "10", "Caliper", "")),
    ):
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column, value)
    return _save_workbook(workbook)


def _vertically_paired_context_repeated_form_bytes(
    *,
    first_context_values=True,
    second_context_name="Part",
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous paged form"
    for page_index, start in enumerate((1, 9)):
        sheet.merge_cells(
            start_row=start,
            start_column=1,
            end_row=start,
            end_column=4,
        )
        sheet.cell(start, 1, "Anonymous paged form")
        sheet.cell(start + 1, 1, "Part" if page_index == 0 else second_context_name)
        sheet.cell(start + 1, 3, "Supplier")
        sheet.merge_cells(
            start_row=start + 2,
            start_column=1,
            end_row=start + 2,
            end_column=2,
        )
        if page_index > 0 or first_context_values:
            sheet.cell(start + 2, 1, f"P-{page_index + 1}")
        sheet.merge_cells(
            start_row=start + 2,
            start_column=3,
            end_row=start + 2,
            end_column=4,
        )
        if page_index > 0 or first_context_values:
            sheet.cell(start + 2, 3, "Supplier A")
        for column, value in enumerate(
            ("Sequence", "Operation", "Characteristic", "Status"),
            start=1,
        ):
            sheet.cell(start + 3, column, value)
        for row_offset, values in enumerate(
            (
                (page_index * 2 + 1, "Inspect", "Surface", "Open"),
                (page_index * 2 + 2, "Measure", "Length", "Closed"),
            ),
            start=4,
        ):
            for column, value in enumerate(values, start=1):
                sheet.cell(start + row_offset, column, value)
    return _save_workbook(workbook)


def _vertically_merged_record_key_with_an_empty_display_row_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous sparse records"
    sheet.append(["Sequence", "Operation", "Characteristic"])
    sheet.merge_cells("A2:A5")
    sheet["A2"] = 10
    sheet["B2"] = "Form"
    sheet["C2"] = "Length"
    sheet["C3"] = "Width"
    sheet["C5"] = "Depth"
    sheet["A6"] = 20
    sheet["B6"] = "Inspect"
    sheet["C6"] = "Surface"
    return _save_workbook(workbook)


def _multilevel_table_after_context_gap_with_sidecar_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous supplier register"
    sheet.merge_cells("A1:R1")
    sheet["A1"] = "Anonymous supplier register"

    sheet.merge_cells("A3:D3")
    sheet["A3"] = "Program"
    sheet.merge_cells("E3:I3")
    sheet["E3"] = "Program A"
    sheet.merge_cells("J3:L3")
    sheet["J3"] = "Component"
    sheet.merge_cells("M3:R3")
    sheet["M3"] = "Component A"
    sheet["S3"] = "Sidecar action"

    sheet.merge_cells("A4:A5")
    sheet["A4"] = "Sequence"
    sheet.merge_cells("B4:G4")
    sheet["B4"] = "Level"
    for column, value in enumerate(range(1, 7), start=2):
        sheet.cell(5, column, value)
    for column, value in enumerate(
        (
            "Material",
            "Grade",
            "Standard",
            "Alternative",
            "Status A",
            "Status B",
            "Status C",
            "Status D",
            "Process",
            "Supplier",
            "Address",
        ),
        start=8,
    ):
        sheet.merge_cells(
            start_row=4,
            start_column=column,
            end_row=5,
            end_column=column,
        )
        sheet.cell(4, column, value)

    for index, row in enumerate(range(6, 11), start=1):
        sheet.cell(row, index + 1, "yes")
        for column, value in enumerate(
            (
                f"Material {index}",
                f"Grade {index}",
                "Standard A",
                "N",
                "Y",
                "N",
                "N",
                "N",
                f"Process {index}",
                f"Supplier {index}",
                f"Address {index}",
            ),
            start=8,
        ):
            sheet.cell(row, column, value)
    return _save_workbook(workbook)


def _competing_merge_free_tail_candidate_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous grouped register"
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "Anonymous grouped register"
    sheet.merge_cells("A2:A3")
    sheet["A2"] = "Sequence"
    sheet.merge_cells("B2:B3")
    sheet["B2"] = "Operation"
    sheet.merge_cells("C2:D2")
    sheet["C2"] = "Characteristic"
    sheet["C3"] = "Product"
    sheet["D3"] = "Process"
    sheet.merge_cells("E2:E3")
    sheet["E2"] = "Method"
    sheet.merge_cells("F2:F3")
    sheet["F2"] = "Response"

    sheet.merge_cells("A4:A6")
    sheet["A4"] = 10
    sheet.merge_cells("B4:B6")
    sheet["B4"] = "Grouped operation"
    for row in range(4, 7):
        sheet.cell(row, 3, f"Product {row}")
        sheet.cell(row, 4, f"Process {row}")
        sheet.cell(row, 5, f"Method {row}")
        sheet.cell(row, 6, f"Response {row}")

    for row, sequence in ((7, 20), (8, 30), (9, 40)):
        sheet.cell(row, 1, sequence)
        sheet.cell(row, 2, f"Operation {sequence}")
        sheet.cell(row, 3, f"Product {sequence}")
        sheet.cell(row, 4, f"Process {sequence}")
        sheet.cell(row, 5, f"Method {sequence}")
        sheet.cell(row, 6, f"Response {sequence}")
    return _save_workbook(workbook)


def _competing_multilevel_header_candidate_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous material register"
    sheet.merge_cells("A2:R3")
    sheet["A2"] = "Anonymous material register"
    sheet["S2"] = "Sidecar action"
    sheet.merge_cells("A4:D4")
    sheet["A4"] = "Reference"
    sheet.merge_cells("E4:I4")
    sheet["E4"] = "R-001"
    sheet.merge_cells("J4:L4")
    sheet["J4"] = "Description"
    sheet.merge_cells("M4:P4")
    sheet["M4"] = "Assembly"
    sheet.merge_cells("Q4:R4")
    sheet["Q4"] = "Source"

    sheet.merge_cells("A5:A6")
    sheet["A5"] = "Sequence"
    sheet.merge_cells("B5:G5")
    sheet["B5"] = "Level"
    for column, value in enumerate(range(1, 7), start=2):
        sheet.cell(6, column, value)
    sheet.merge_cells("H5:I5")
    sheet["H5"] = "Part number"
    sheet["H6"] = "Internal"
    sheet["I6"] = "Supplier"
    sheet.merge_cells("J5:J6")
    sheet["J5"] = "Material"
    sheet.merge_cells("K5:K6")
    sheet["K5"] = "Specification"
    sheet.merge_cells("L5:L6")
    sheet["L5"] = "Standard"
    sheet.merge_cells("M5:M6")
    sheet["M5"] = "Alternative"
    sheet.merge_cells("N5:R5")
    sheet["N5"] = "Status"
    for column, value in enumerate(
        ("Internal", "Domestic", "Imported", "Auxiliary", "Approved"),
        start=14,
    ):
        sheet.cell(6, column, value)

    for row, level_column in enumerate((4, 3, 3, 4, 3), start=7):
        sheet.cell(row, level_column, "yes")
        sheet.cell(row, 8, "/")
        sheet.cell(row, 9, "/")
        sheet.cell(row, 10, f"Material {row}")
        sheet.cell(row, 11, f"Grade {row}")
        sheet.cell(row, 12, "/")
        sheet.cell(row, 13, "Y")
        sheet.cell(row, 14 if row % 2 else 15, "yes")
        sheet.cell(row, 18, "approved")
    return _save_workbook(workbook)


def _single_level_header_after_context_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous risk register"
    sheet["A1"] = "Submission"
    sheet.merge_cells("O1:R1")
    sheet.merge_cells("O2:R2")
    sheet["A3"] = "Reference"
    sheet.merge_cells("D3:G3")
    sheet["D3"] = "Description"
    sheet["I3"] = "Program"
    sheet["L3"] = "Prepared"
    sheet["M3"] = "Person"
    sheet["N3"] = "Phone"
    sheet.merge_cells("O3:R4")
    sheet["O3"] = "Contact"
    sheet.merge_cells("A4:C4")
    sheet["A4"] = "Reference"
    sheet.merge_cells("D4:H4")
    sheet["D4"] = "Assembly"
    sheet.merge_cells("I4:K4")
    sheet["I4"] = "Program"
    sheet["L4"] = "Prepared"
    sheet["A5"] = "Core team"
    sheet.merge_cells("B5:C6")
    sheet["B5"] = "Team"
    sheet["D5"] = "Modified"
    sheet["L5"] = "Approval"
    sheet.merge_cells("O5:R5")
    sheet.merge_cells("D6:G6")
    sheet["L6"] = "Approval date"
    sheet.merge_cells("O6:R6")
    for column, value in enumerate(
        (
            "Process",
            "Mode",
            "Effect",
            "Severity",
            "Class",
            "Cause",
            "Occurrence",
            "Prevention",
            "Detection",
            "Rating",
            "Priority",
            "Action",
            "Owner",
            "Result",
            "Final severity",
            "Final occurrence",
            "Final detection",
            "Final priority",
        ),
        start=1,
    ):
        sheet.cell(7, column, value)
    sheet.merge_cells("A8:A12")
    sheet["A8"] = "10"
    sheet.merge_cells("A14:A16")
    sheet["A14"] = "20"
    sheet.merge_cells("A17:A18")
    sheet["A17"] = "30"
    sheet.merge_cells("A19:A20")
    sheet["A19"] = "40"
    sheet.merge_cells("H19:H20")
    sheet["H19"] = "Shared prevention"
    for row in range(8, 21):
        for column, value in {
            2: f"Mode {row}",
            3: f"Effect {row}",
            4: "5",
            6: f"Cause {row}",
            7: "3",
            8: f"Prevention {row}",
            9: f"Detection {row}",
            10: "4",
            11: "60",
        }.items():
            if (
                sheet.cell(row, column).__class__.__name__ != "MergedCell"
                and sheet.cell(row, column).value is None
            ):
                sheet.cell(row, column, value)
    return _save_workbook(workbook)


def _mixed_storage_numeric_key_candidate_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous mixed storage"
    sheet.merge_cells("A1:H2")
    sheet["A1"] = "Anonymous register"
    for row in (3, 4):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.cell(row=row, column=1, value=f"Context {row} A")
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        sheet.cell(row=row, column=3, value=f"Context {row} B")
        sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        sheet.cell(row=row, column=5, value=f"Context {row} C")
        sheet.cell(row=row, column=8, value=f"Context {row} D")
    sheet.merge_cells("A5:H5")
    sheet["A5"] = "Record fields"
    sheet.merge_cells("A6:A7")
    sheet["A6"] = "Sequence"
    sheet.merge_cells("B6:B7")
    sheet["B6"] = "Code"
    sheet.merge_cells("C6:D6")
    sheet["C6"] = "Grouping"
    sheet["C7"] = "Group"
    sheet["D7"] = "State"
    sheet.merge_cells("E6:E7")
    sheet["E6"] = "Owner"
    sheet.merge_cells("F6:G7")
    sheet["F6"] = "Evidence"
    sheet.merge_cells("H6:H7")
    sheet["H6"] = "Optional"
    for index, row in enumerate(range(8, 18), start=1):
        key = str(index) if index <= 2 else index
        values = [
            key,
            f"R-{index:02d}",
            "A",
            "Open",
            "Team",
            f"V-{index}",
            f"V-{index}",
            "present" if index == 8 else None,
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row, column=column, value=value)
        sheet.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
    return _save_workbook(workbook)


def _sparse_region_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Code", None, "Status"])
    sheet.append(["P-1", None, "Open"])
    sheet.append(["P-2", None, "Closed"])
    sheet.append([None, None, None])
    sheet.append(["P-3", None, "Open"])
    sheet.append(["P-4", None, "Closed"])
    return _save_workbook(workbook)


def _g_sensitive_horizontal_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.append(["Left code", "Left status", None, "Right code", "Right status"])
    sheet.append(["L-1", "Open", None, "R-1", "Closed"])
    sheet.append(["L-2", "Closed", None, "R-2", "Open"])
    return _save_workbook(workbook)


def _three_level_header_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "Items"
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "Identity"
    sheet["A3"] = "Code"
    sheet["B3"] = "State"
    sheet.append(["T-1", "Open"])
    sheet.append(["T-2", "Closed"])
    return _save_workbook(workbook)


def _four_level_header_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    for row_ordinal, label in enumerate(("Items", "Identity", "Detail"), start=1):
        sheet.merge_cells(
            start_row=row_ordinal,
            start_column=1,
            end_row=row_ordinal,
            end_column=2,
        )
        sheet.cell(row_ordinal, 1, label)
    sheet["A4"] = "Code"
    sheet["B4"] = "State"
    sheet.append(["Q-1", "Open"])
    sheet.append(["Q-2", "Closed"])
    return _save_workbook(workbook)


def _formula_only_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous regions"
    sheet["A1"] = '=CONCAT("F-", 1)'
    sheet["A2"] = '=CONCAT("F-", 2)'
    return _save_workbook(workbook)


def _complete_table_with_distant_multiline_signoff_bytes(record_count=3):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous report"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Anonymous register"
    sheet.append(["Sequence", "Item", "Measure", "Status"])
    for row_ordinal in range(3, 3 + record_count):
        sheet.append(
            [
                row_ordinal - 2,
                f"I-{row_ordinal}",
                row_ordinal * 10,
                "Open",
            ]
        )

    # A physically separated approval form reuses table columns but is not a
    # continuation of the already closed record axis.
    signoff_start = 3 + record_count + 2
    sheet.merge_cells(
        start_row=signoff_start,
        start_column=1,
        end_row=signoff_start,
        end_column=4,
    )
    sheet.cell(signoff_start, 1, "Anonymous approval")
    for row_ordinal in range(signoff_start + 1, signoff_start + 4):
        sheet.merge_cells(
            start_row=row_ordinal,
            start_column=1,
            end_row=row_ordinal,
            end_column=2,
        )
        sheet.cell(row_ordinal, 1, f"Role {row_ordinal}")
        sheet.merge_cells(
            start_row=row_ordinal,
            start_column=3,
            end_row=row_ordinal,
            end_column=4,
        )
        sheet.cell(row_ordinal, 3, f"Signature {row_ordinal}")
    return _save_workbook(workbook)


def _context_with_sparse_single_record_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous plan"
    sheet["G2"] = "Anonymous plan"
    sheet["K4"] = "Reference"
    sheet["N4"] = "Control"
    sheet["K5"] = "Reference value"
    sheet["N5"] = "Control value"
    for column, value in enumerate(
        (
            "Part",
            "P-001",
            None,
            "Name",
            "Component",
            "Mode",
            "Program",
            "Supplier",
            "Supplier A",
            None,
            "Owner",
            "Owner A",
            None,
            "Date",
        ),
        start=1,
    ):
        sheet.cell(row=6, column=column, value=value)
    sheet["A8"] = "Context"
    sheet.merge_cells("B8:C8")
    sheet["B8"] = "Context value"
    sheet["F8"] = "Approval"
    sheet.merge_cells("H8:J8")
    sheet["H8"] = "Approval value"
    sheet["N8"] = "Revision"
    sheet["N9"] = "Revision value"

    sheet["A11"] = "Sequence"
    sheet.merge_cells("B11:C11")
    sheet["B11"] = "Process"
    sheet.merge_cells("D11:E11")
    sheet["D11"] = "Characteristic"
    for column, value in enumerate(
        (
            "Method",
            "Frequency",
            "Target",
            "Class",
            "Requirement",
        ),
        start=6,
    ):
        sheet.cell(row=11, column=column, value=value)
    sheet["K11"] = "Auxiliary"
    sheet.merge_cells("L11:O11")
    sheet["L11"] = "Note"

    sheet["A12"] = 1
    sheet.merge_cells("B12:C12")
    sheet["B12"] = "Process A"
    sheet.merge_cells("D12:E12")
    sheet["D12"] = "Characteristic A"
    for column, value in enumerate(
        ("Method A", "Annual", "2026-01-01", "SC", "At least 1.33"),
        start=6,
    ):
        sheet.cell(row=12, column=column, value=value)
    sheet.merge_cells("L12:O12")
    return _save_workbook(workbook)


def _context_with_trailing_dense_empty_header_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous empty analysis"
    sheet.merge_cells("B2:N2")
    sheet["B2"] = "Anonymous analysis"
    sheet["L5"] = "Reference"
    sheet["N5"] = "Number"
    sheet["L6"] = "Reference value"
    sheet["N6"] = "Number value"
    sheet["A7"] = "Submission context"
    sheet.merge_cells("A9:C9")
    sheet["A9"] = "Part"
    sheet.merge_cells("D9:G9")
    sheet["D9"] = "Name"
    sheet["I9"] = "Program"
    sheet["L9"] = "Owner"
    sheet.merge_cells("O9:R9")
    sheet["O9"] = "Contact"
    sheet["L10"] = "Date"
    sheet["A11"] = "Team"
    sheet.merge_cells("D11:G11")
    sheet["D11"] = "Revision"
    sheet["L11"] = "Approval"
    sheet["L12"] = "Approved on"
    for column in range(1, 19):
        sheet.cell(row=14, column=column, value=f"Field {column}")
    return _save_workbook(workbook)


@pytest.fixture
def table_parser(monkeypatch):
    return _load_table_module(monkeypatch).Excel()


def test_same_bytes_keep_business_identity_but_get_a_new_generation(table_parser):
    source = _workbook_bytes()
    first = build_tabular_structure_projection(
        "anonymous.xlsx",
        source,
        parser=table_parser,
    )
    second = build_tabular_structure_projection(
        "anonymous.xlsx",
        source,
        parser=table_parser,
    )

    assert first["producer_generation_ref"] != second["producer_generation_ref"]
    assert [row["table_ref_kwd"] for row in first["rows"]] == [row["table_ref_kwd"] for row in second["rows"]]
    assert [row["row_ref_kwd"] for row in first["rows"]] == [row["row_ref_kwd"] for row in second["rows"]]
    assert [row["row_role_kwd"] for row in first["rows"]] == [row["row_role_kwd"] for row in second["rows"]]
    assert [row["data_row_index_int"] for row in first["rows"]] == [row["data_row_index_int"] for row in second["rows"]]
    assert [row["source_total_count_int"] for row in first["rows"]] == [row["source_total_count_int"] for row in second["rows"]]
    assert {row["producer_generation_ref_kwd"] for row in first["rows"]} == {first["producer_generation_ref"]}
    assert {row["id"] for row in first["rows"]}.isdisjoint({row["id"] for row in second["rows"]})


def test_vertical_tables_on_one_sheet_get_separate_stable_projections(table_parser):
    source = _vertical_multi_region_workbook_bytes()

    projections = [
        build_tabular_structure_projection(
            "anonymous.xlsx",
            source,
            producer_generation_ref=_generation_ref(),
            parser=table_parser,
        )
        for _ in range(3)
    ]

    expected_identity = [
        (table["table_ordinal"], table["table_ref"])
        for table in projections[0]["tables"]
    ]
    assert [table["table_ordinal"] for table in projections[0]["tables"]] == [1, 2]
    assert [table["source_total_count"] for table in projections[0]["tables"]] == [2, 2]
    assert all(
        [(table["table_ordinal"], table["table_ref"]) for table in projection["tables"]]
        == expected_identity
        for projection in projections[1:]
    )


def test_horizontal_tables_use_exact_columns_without_duplicate_ingestion(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _horizontal_multi_region_workbook_bytes(),
        parser=table_parser,
    )

    assert [table["table_ordinal"] for table in projection["tables"]] == [1, 2]
    rows_by_table = {
        table["table_ref"]: [
            json.loads(row["ordered_fields_list"])
            for row in projection["rows"]
            if row["table_ref_kwd"] == table["table_ref"] and row["row_role_kwd"] == "data"
        ]
        for table in projection["tables"]
    }
    assert rows_by_table[projection["tables"][0]["table_ref"]] == [
        [
            {
                "column_id": "col_v1:1:1",
                "column_ordinal": 1,
                "header_path": ["Left code"],
                "name": "Left code",
                "value": "L-1",
            },
            {
                "column_id": "col_v1:1:2",
                "column_ordinal": 2,
                "header_path": ["Left status"],
                "name": "Left status",
                "value": "Open",
            },
        ],
        [
            {
                "column_id": "col_v1:1:1",
                "column_ordinal": 1,
                "header_path": ["Left code"],
                "name": "Left code",
                "value": "L-2",
            },
            {
                "column_id": "col_v1:1:2",
                "column_ordinal": 2,
                "header_path": ["Left status"],
                "name": "Left status",
                "value": "Closed",
            },
        ],
    ]
    assert rows_by_table[projection["tables"][1]["table_ref"]] == [
        [
            {
                "column_id": "col_v1:1:5",
                "column_ordinal": 1,
                "header_path": ["Right code"],
                "name": "Right code",
                "value": "R-1",
            },
            {
                "column_id": "col_v1:1:6",
                "column_ordinal": 2,
                "header_path": ["Right status"],
                "name": "Right status",
                "value": "Closed",
            },
        ],
        [
            {
                "column_id": "col_v1:1:5",
                "column_ordinal": 1,
                "header_path": ["Right code"],
                "name": "Right code",
                "value": "R-2",
            },
            {
                "column_id": "col_v1:1:6",
                "column_ordinal": 2,
                "header_path": ["Right status"],
                "name": "Right status",
                "value": "Open",
            },
        ],
    ]


def test_vertical_headerless_continuation_downgrades_complete_sibling(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _vertical_complete_with_headerless_continuation_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] == 3
    assert projection["tables"][0]["matched_rule"] == "L1-02"


def test_repeated_form_segments_are_one_sheet_level_logical_table(table_parser):
    projection = build_tabular_structure_projection(
        "ppap-like.xls",
        _sheet_level_repeated_form_segments_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] == 4
    assert [row["row_ordinal_int"] for row in projection["rows"]] == [4, 5, 12, 13]
    assert all(row["row_role_kwd"] == "data" for row in projection["rows"])


def test_repeated_form_context_values_may_change_inside_one_sheet_object(table_parser):
    projection = build_tabular_structure_projection(
        "ppap-like.xls",
        _sheet_level_repeated_form_segments_bytes(
            second_context_value="2906151-SE01",
        ),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] == 4


def test_repeated_headers_with_a_conflicting_form_context_remain_separate(table_parser):
    projection = build_tabular_structure_projection(
        "ppap-like.xls",
        _sheet_level_repeated_form_segments_bytes(
            second_context_name="图号",
        ),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert [table["source_total_count"] for table in projection["tables"]] == [2, 2]


def test_four_repeated_pages_merge_in_physical_order_across_context_value_changes(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "ppap-like.xls",
        _four_page_repeated_form_with_context_value_changes_bytes(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert [table["source_total_count"] for table in complete] == [8]
    assert [
        [row["row_ordinal_int"] for row in projection["rows"] if row["table_ref_kwd"] == table["table_ref"]]
        for table in complete
    ] == [[4, 5, 12, 13, 20, 21, 28, 29]]


def test_headerless_predecessor_merges_into_the_context_bound_named_page(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _headerless_predecessor_with_named_following_page_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 4
    assert [column["header_path"] for column in complete[0]["ordered_columns"]] == [
        ["Process"],
        ["Mode"],
        ["Effect"],
        ["Severity"],
        ["Class"],
        ["Cause"],
        ["Occurrence"],
        ["Prevention"],
        ["Detection"],
        ["Rating"],
        ["Priority"],
    ]
    assert [
        row["row_ordinal_int"]
        for row in projection["rows"]
        if row["table_ref_kwd"] == complete[0]["table_ref"]
        and row["row_role_kwd"] == "data"
    ] == [7, 8, 22, 23]


def test_isolated_single_row_annotation_is_not_merged_as_a_continuation(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "State"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    sheet.append([None, None])
    sheet.append([None, None])
    sheet.append(["Prepared by", "Alice"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert projection["tables"][0]["source_total_count"] == 2
    assert projection["tables"][0]["matched_rule"] == "L1-07"
    assert projection["tables"][1]["source_total_count"] is None
    assert projection["tables"][1]["matched_rule"] == "R8"


def test_same_shape_single_row_annotation_is_not_merged_as_a_continuation(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Record Name", "Status", "Owner"])
    sheet.append(["North Unit", "Open", "Team-1"])
    sheet.append(["South Site", "Closed", "Team-2"])
    sheet.append([None, None, None])
    sheet.append([None, None, None])
    sheet.append(["Prepared by", "Alice", "Team-3"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert projection["tables"][0]["source_total_count"] == 2
    assert projection["tables"][0]["matched_rule"] == "L1-07"
    assert projection["tables"][1]["source_total_count"] is None
    assert projection["tables"][1]["matched_rule"] == "R8"


def test_continuation_union_preserves_custom_context_limits(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _vertical_complete_with_headerless_continuation_bytes(),
        table_context_entry_limit=1,
        table_context_value_bytes=5,
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    context = json.loads(projection["rows"][0]["table_context_list"])
    assert len(context) == 1
    assert len(context[0]["value"].encode("utf-8")) <= 5


def test_continuation_with_unknown_row_cannot_drop_it_and_claim_complete(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Status"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    sheet.append([None, None])
    sheet.append([None, None])
    sheet.append(["Code", "Status"])
    sheet.append(["B-1", "Open"])
    sheet.append(["B-2", "Closed"])
    sheet.append(["Code", "Status"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert all(table["source_total_count"] is None for table in projection["tables"])
    assert any(
        row["row_ordinal_int"] == 9 and row["row_role_kwd"] == "unknown"
        for row in projection["rows"]
    )


def test_multirow_named_continuation_without_a_proven_axis_cannot_claim_complete(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value"])
    sheet.append(["A-1", 1])
    sheet.append(["A-2", 2])
    sheet.append([None, None])
    sheet.append([None, None])
    sheet.append(["Code", "Value"])
    sheet.append(["B-1", 3])
    sheet.append(["Summary", "Current"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert all(table["source_total_count"] is None for table in projection["tables"])
    assert all(table["enumeration_status"] == "not_guaranteed_explained" for table in projection["tables"])


def test_three_equal_continuation_segments_form_one_complete_record_sequence(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Status"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    sheet.append([None, None])
    sheet.append([None, None])
    sheet.append(["B-1", "Open"])
    sheet.append([None, None])
    sheet.append([None, None])
    sheet.append(["C-1", "Closed"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] == 4
    assert projection["tables"][0]["matched_rule"] == "L1-02"


def test_multirow_named_superset_continuation_forms_one_complete_record_sequence(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Status"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    sheet.append([None, None, None])
    sheet.append([None, None, None])
    sheet.append(["Code", "Status", "Owner"])
    sheet.append(["B-1", "Open", "Team-1"])
    sheet.append(["B-2", "Closed", "Team-2"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] == 4
    assert projection["tables"][0]["matched_rule"] == "L1-02"


def test_named_continuation_left_expansion_rekeys_every_row_to_the_union_manifest(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=2, value="Code")
    sheet.cell(row=1, column=3, value="Status")
    sheet.cell(row=2, column=2, value="A-1")
    sheet.cell(row=2, column=3, value="Open")
    sheet.cell(row=3, column=2, value="A-2")
    sheet.cell(row=3, column=3, value="Closed")
    sheet.cell(row=6, column=1, value="Owner")
    sheet.cell(row=6, column=2, value="Code")
    sheet.cell(row=6, column=3, value="Status")
    sheet.cell(row=7, column=1, value="Team-1")
    sheet.cell(row=7, column=2, value="B-1")
    sheet.cell(row=7, column=3, value="Open")
    sheet.cell(row=8, column=1, value="Team-2")
    sheet.cell(row=8, column=2, value="B-2")
    sheet.cell(row=8, column=3, value="Closed")

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    table = projection["tables"][0]
    assert table["source_total_count"] == 4
    assert table["matched_rule"] == "L1-02"
    assert table["ordered_columns"] == [
        {
            "column_id": "col_v1:1:1",
            "column_ordinal": 1,
            "header_path": ["Owner"],
            "name": "Owner",
        },
        {
            "column_id": "col_v1:1:2",
            "column_ordinal": 2,
            "header_path": ["Code"],
            "name": "Code",
        },
        {
            "column_id": "col_v1:1:3",
            "column_ordinal": 3,
            "header_path": ["Status"],
            "name": "Status",
        },
    ]
    fields_by_row = {
        row["row_ordinal_int"]: json.loads(row["ordered_fields_list"])
        for row in projection["rows"]
    }
    assert [field["column_ordinal"] for field in fields_by_row[2]] == [2, 3]
    assert [field["column_ordinal"] for field in fields_by_row[3]] == [2, 3]
    assert [field["column_ordinal"] for field in fields_by_row[7]] == [1, 2, 3]
    assert [field["column_ordinal"] for field in fields_by_row[8]] == [1, 2, 3]
    validate_tabular_structure_projection(projection)


def test_supported_complete_table_requires_nonempty_ordered_columns(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Status"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )
    assert projection["tables"][0]["enumeration_status"] == "supported_complete"
    projection["tables"][0]["ordered_columns"] = []

    with pytest.raises(ValueError, match="supported complete table requires ordered columns"):
        validate_tabular_structure_projection(projection)


def test_manifest_downgrades_when_a_value_field_has_no_header_path():
    item = {
        "table": {
            "enumeration_status": "supported_complete",
            "enumeration_reason": "record_axis_proven",
            "matched_rule": "L1-07",
            "sheet_ordinal": 1,
            "source_total_count": 1,
            "table_label": "Anonymous register",
            "table_context": [],
        },
        "rows": [{
            "table_label_kwd": "Anonymous register",
            "table_context_list": "[]",
            "ordered_fields_list": json.dumps([
                {
                    "column_id": "col_v1:1:1",
                    "column_ordinal": 1,
                    "header_path": ["Code"],
                    "name": "Code",
                    "value": "A-1",
                },
                {
                    "column_id": "col_v1:1:2",
                    "column_ordinal": 2,
                    "header_path": [],
                    "name": "Optional",
                    "value": "present",
                },
            ]),
        }],
        "structure_evidence": {
            "headers_by_column": {1: "Code", 2: "Optional"},
            "header_paths_by_column": {1: ["Code"], 2: []},
        },
        "proven_record_slots": [1],
        "members": {(1, 1), (1, 2)},
        "worksheet_name": "Anonymous register",
    }

    tabular_structure._finalize_table_manifest_evidence(item)

    assert item["table"]["enumeration_status"] == "not_guaranteed_explained"
    assert item["table"]["enumeration_reason"] == "record_axis_not_proven"
    assert item["table"]["matched_rule"] == "R8"
    assert item["table"]["source_total_count"] is None
    assert item["table"]["ordered_columns"] == []


def test_manifest_ordinals_are_dense_over_the_selected_source_axis():
    item = {
        "table": {
            "enumeration_status": "supported_complete",
            "enumeration_reason": "record_axis_proven",
            "matched_rule": "L1-07",
            "sheet_ordinal": 1,
            "source_total_count": 1,
            "table_label": "Anonymous register",
            "table_context": [],
        },
        "rows": [{
            "table_label_kwd": "Anonymous register",
            "table_context_list": "[]",
            "ordered_fields_list": json.dumps([
                {
                    "column_id": "col_v1:1:3",
                    "column_ordinal": 1,
                    "header_path": ["Code"],
                    "name": "Code",
                    "value": "A-1",
                },
                {
                    "column_id": "col_v1:1:4",
                    "column_ordinal": 2,
                    "header_path": ["State"],
                    "name": "State",
                    "value": "Open",
                },
            ]),
        }],
        "structure_evidence": {
            "headers_by_column": {3: "Code", 4: "State"},
            "header_paths_by_column": {3: ["Code"], 4: ["State"]},
        },
        "proven_record_slots": [1],
        "members": {(1, 3), (1, 4)},
        "worksheet_name": "Anonymous register",
    }

    tabular_structure._finalize_table_manifest_evidence(item)

    assert [
        (column["column_id"], column["column_ordinal"])
        for column in item["table"]["ordered_columns"]
    ] == [("col_v1:1:3", 1), ("col_v1:1:4", 2)]


def test_hidden_headerless_continuation_row_cannot_claim_complete(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Status"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    sheet.append([None, None])
    sheet.append([None, None])
    sheet.append(["B-1", "Open"])
    sheet.row_dimensions[6].hidden = True

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] is None
    assert projection["tables"][0]["matched_rule"] == "R2"


def test_subset_headerless_continuation_downgrades_complete_sibling(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _vertical_complete_with_subset_headerless_continuation_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] == 3
    assert projection["tables"][0]["matched_rule"] == "L1-02"


def test_superset_headerless_continuation_downgrades_complete_sibling(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _vertical_complete_with_superset_headerless_continuation_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] is None
    assert projection["tables"][0]["matched_rule"] == "D1"
    assert all("Column_" not in row["ordered_fields_list"] for row in projection["rows"])


def test_unnamed_superset_still_runs_membership_closure(table_parser, monkeypatch):
    original = tabular_structure._merge_continuation_pair

    def duplicate_merged_member(**kwargs):
        merged = original(**kwargs)
        if merged is not None:
            merged["emitted_member_events"].append(merged["emitted_member_events"][0])
        return merged

    monkeypatch.setattr(tabular_structure, "_merge_continuation_pair", duplicate_merged_member)
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _vertical_complete_with_superset_headerless_continuation_bytes(),
        parser=table_parser,
    )

    assert projection["tables"][0]["matched_rule"] == "D2"
    assert projection["tables"][0]["enumeration_reason"] == "membership_not_closed"


def test_partially_overlapping_unknown_downgrades_complete_sibling(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _complete_table_with_partially_overlapping_unknown_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert all(table["source_total_count"] is None for table in projection["tables"])
    assert {table["matched_rule"] for table in projection["tables"]} == {"R6"}


def test_partial_overlap_does_not_downgrade_a_disjoint_table_on_the_same_sheet(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Status", None, None, None, "Other", "Value"])
    sheet.append(["A-1", "Open", None, None, None, "X-1", 1])
    sheet.append(["A-2", "Closed", None, None, None, "X-2", 2])
    sheet.cell(row=6, column=2, value="Later")
    sheet.cell(row=6, column=3, value="Extra")

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 3
    assert [table["source_total_count"] for table in projection["tables"]] == [None, 2, None]
    assert [table["matched_rule"] for table in projection["tables"]] == ["R6", "L1-03", "R6"]


def test_partial_overlap_before_a_complete_table_is_not_a_continuation(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=2, value="Earlier")
    sheet.cell(row=1, column=3, value="Detail")
    sheet.append([None, None, None])
    sheet.append([None, None, None])
    sheet.append(["Code", "Status"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert [table["source_total_count"] for table in projection["tables"]] == [None, 2]
    assert [table["matched_rule"] for table in projection["tables"]] == ["R8", "L1-07"]


def test_horizontal_headerless_record_axis_does_not_downgrade_disjoint_sibling(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _horizontal_complete_with_headerless_sibling_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert [table["source_total_count"] for table in projection["tables"]] == [2, None]
    assert [table["matched_rule"] for table in projection["tables"]] == ["L1-07", "R8"]


def test_projected_table_with_unknown_row_does_not_downgrade_complete_sibling(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _complete_table_with_independent_repeated_header_table_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert [table["source_total_count"] for table in projection["tables"]] == [2, None]


def test_axis_aligned_partial_context_does_not_downgrade_complete_sibling(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _complete_table_with_axis_aligned_unknown_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert [table["source_total_count"] for table in projection["tables"]] == [2, None]


def test_context_only_unknown_overlap_does_not_downgrade_complete_sibling(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _complete_table_with_context_only_unknown_overlap_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert projection["tables"][0]["source_total_count"] == 2
    assert projection["tables"][1]["source_total_count"] is None


def test_isolated_annotation_does_not_downgrade_complete_sibling(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _complete_table_with_isolated_annotation_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert projection["tables"][0]["source_total_count"] == 2
    assert projection["tables"][1]["source_total_count"] is None


def test_g_sensitive_unknown_does_not_downgrade_disjoint_sibling(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _complete_table_with_g_sensitive_sibling_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert [table["source_total_count"] for table in projection["tables"]] == [2, None]
    assert [table["matched_rule"] for table in projection["tables"]] == ["L1-07", "R8"]
    assert all(
        row["row_role_kwd"] == "unknown"
        for row in projection["rows"]
        if row["table_ref_kwd"] == projection["tables"][1]["table_ref"]
    )


def test_g_sensitive_horizontal_boundary_cannot_produce_complete(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _g_sensitive_horizontal_workbook_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] is None
    assert all(row["row_role_kwd"] == "unknown" for row in projection["rows"])


def test_single_column_record_axis_gets_a_complete_projection(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _single_column_workbook_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] == 2
    assert [
        json.loads(row["ordered_fields_list"])
        for row in projection["rows"]
        if row["row_role_kwd"] == "data"
    ] == [
        [
            {
                "column_id": "col_v1:1:1",
                "column_ordinal": 1,
                "header_path": ["Code"],
                "name": "Code",
                "value": "S-1",
            }
        ],
        [
            {
                "column_id": "col_v1:1:1",
                "column_ordinal": 1,
                "header_path": ["Code"],
                "name": "Code",
                "value": "S-2",
            }
        ],
    ]
    expected_membership = tabular_structure._region_membership_sha256(
        1,
        {(1, 1), (2, 1), (3, 1)},
    )
    assert projection["tables"][0]["table_ref"].startswith(
        f"tbl_v2_{expected_membership}_"
    )


def test_single_column_free_text_block_cannot_prove_a_complete_record_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Instruction"])
    sheet.append(["Review the source before use"])
    sheet.append(["Confirm the revision before release"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None


def test_single_record_slot_cannot_prove_a_complete_record_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Status"])
    sheet.append(["S-1", "Open"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["data_row_count"] == 1
    assert projection["tables"][0]["source_total_count"] is None

    projection["tables"][0]["source_total_count"] = 1
    projection["rows"][0]["source_total_count_int"] = 1
    with pytest.raises(ValueError, match="source total"):
        validate_tabular_structure_projection(projection)


def test_structurally_bounded_single_record_proves_a_complete_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:E1")
    sheet["A1"] = "Anonymous register"
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "Context"
    sheet.merge_cells("C2:E2")
    sheet["C2"] = "Value"
    sheet["A3"] = "Sequence"
    sheet.merge_cells("B3:C3")
    sheet["B3"] = "Item"
    sheet["D3"] = "Measure"
    sheet["E3"] = "Status"
    sheet["A4"] = 1
    sheet.merge_cells("B4:C4")
    sheet["B4"] = "I-1"
    sheet["D4"] = 2
    sheet["E4"] = "Open"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 1
    rows = [
        row
        for row in projection["rows"]
        if row["table_ref_kwd"] == complete[0]["table_ref"]
    ]
    assert [(row["row_ordinal_int"], row["row_role_kwd"]) for row in rows] == [
        (4, "data"),
    ]


def test_single_free_text_slot_remains_fail_closed(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Anonymous note"
    sheet.merge_cells("A2:D2")
    sheet["A2"] = "Review this statement before release"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert all(table["source_total_count"] is None for table in projection["tables"])


def test_separated_footer_does_not_pollute_a_proven_record_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Anonymous register"
    sheet.append(["Sequence", "Item", "Measure", "Status"])
    sheet.append([1, "I-1", 2, "Open"])
    sheet.append([2, "I-2", 3, "Closed"])
    sheet.append([3, "I-3", 4, "Open"])
    sheet.append([None, None, None, None])
    sheet.append(["Prepared by", "Person", "Approved by", "Person"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 3
    rows = [
        row
        for row in projection["rows"]
        if row["table_ref_kwd"] == complete[0]["table_ref"]
    ]
    assert [row["row_ordinal_int"] for row in rows if row["row_role_kwd"] == "data"] == [
        3,
        4,
        5,
    ]
    footer = next(row for row in rows if row["row_ordinal_int"] == 7)
    assert footer["row_role_kwd"] == "note"
    assert footer["data_row_index_int"] is None
    assert all(row["source_total_count_int"] == 3 for row in rows)


def test_unseparated_footer_still_invalidates_completeness(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Sequence", "Item", "Measure", "Status"])
    sheet.append([1, "I-1", 2, "Open"])
    sheet.append([2, "I-2", 3, "Closed"])
    sheet.append(["Prepared by", "Person", "Approved by", "Person"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert all(table["source_total_count"] is None for table in projection["tables"])


def test_candidate_selection_prefers_the_closed_record_axis_over_wider_context(
    table_parser,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:G1")
    sheet["A1"] = "Anonymous register"
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "Context A"
    sheet.merge_cells("C2:G2")
    sheet["C2"] = "Value A"
    sheet.merge_cells("A3:C3")
    sheet["A3"] = "Context B"
    sheet.merge_cells("D3:G3")
    sheet["D3"] = "Value B"
    sheet.merge_cells("A4:B4")
    sheet["A4"] = "Sequence"
    sheet.merge_cells("C4:D4")
    sheet["C4"] = "Item"
    sheet.merge_cells("E4:G4")
    sheet["E4"] = "Measure"
    for row_ordinal, values in enumerate(
        ((1, "I-1", 10), (2, "I-2", 20), (3, "I-3", 30)),
        start=5,
    ):
        sheet.cell(row=row_ordinal, column=1, value=values[0])
        sheet.merge_cells(
            start_row=row_ordinal,
            start_column=1,
            end_row=row_ordinal,
            end_column=2,
        )
        sheet.cell(row=row_ordinal, column=3, value=values[1])
        sheet.merge_cells(
            start_row=row_ordinal,
            start_column=3,
            end_row=row_ordinal,
            end_column=4,
        )
        sheet.cell(row=row_ordinal, column=5, value=values[2])
        sheet.merge_cells(
            start_row=row_ordinal,
            start_column=5,
            end_row=row_ordinal,
            end_column=7,
        )

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 3
    assert [
        row["row_ordinal_int"]
        for row in projection["rows"]
        if row["table_ref_kwd"] == complete[0]["table_ref"]
        and row["row_role_kwd"] == "data"
    ] == [5, 6, 7]


def test_numeric_record_key_keeps_value_shape_changes_on_one_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Anonymous register"
    sheet.append(["Sequence", "Item", "Measure", "Status"])
    sheet.append([1, "I-1", 10, "Open"])
    sheet.append([2, "I-2", "pending", "Open"])
    sheet.append([3, "I-3", "pending", "Closed"])
    sheet.append([4, "I-4", 40, "Closed"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] == 4
    assert [row["row_role_kwd"] for row in projection["rows"]] == [
        "data",
        "data",
        "data",
        "data",
    ]


def test_record_key_axis_accepts_lossless_mixed_numeric_storage():
    rows = [
        (1, ["01", "A"], False),
        (2, ["2.0", "B"], False),
        (3, [3, "C"], False),
        (4, [4.0, "D"], False),
    ]
    original_keys = [row[1][0] for row in rows]

    assert tabular_structure._record_key_axis_proven(rows, {0}) is True
    assert [row[1][0] for row in rows] == original_keys


@pytest.mark.parametrize(
    "values",
    [
        ["1", "one", 3],
        ["1", "1.0", 2],
        ["1", "3", 2],
        ["1", "NaN", 3],
        ["１", "2", 3],
        ["1", "2", "3"],
        [True, 2, 3],
    ],
)
def test_record_key_axis_rejects_non_numeric_duplicate_or_non_monotonic_values(values):
    rows = [
        (row_ordinal, [value, f"R-{row_ordinal}"], False)
        for row_ordinal, value in enumerate(values, start=1)
    ]

    assert tabular_structure._record_key_axis_proven(rows, {0}) is False


def test_repeated_context_block_separated_from_the_table_does_not_break_g1(
    table_parser,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "Anonymous context"
    sheet["A2"] = "Context key"
    sheet["A3"] = "Context value"
    sheet.merge_cells("A5:F5")
    sheet["A5"] = "Anonymous register"
    sheet.append(["Sequence", "Item", "Measure", "Status", "Owner", "Note"])
    sheet.append([1, "I-1", 10, "Open", "A", None])
    sheet.append([2, "I-2", 20, "Open", "B", None])
    sheet.append([3, "I-3", 30, "Closed", "C", None])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 3


def test_separated_note_section_does_not_downgrade_a_complete_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Anonymous register"
    sheet.append(["Sequence", "Item", "Measure", "Status"])
    for row_ordinal in range(3, 7):
        sheet.append([row_ordinal - 2, f"I-{row_ordinal}", row_ordinal * 10, "Open"])
    sheet.merge_cells("A8:D9")
    sheet["A8"] = "Anonymous note section"
    sheet.merge_cells("A10:B10")
    sheet["A10"] = "Prepared"
    sheet.merge_cells("C10:D10")
    sheet["C10"] = "Approved"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 4


def test_distant_multiline_signoff_does_not_downgrade_a_closed_record_axis(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _complete_table_with_distant_multiline_signoff_bytes(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 3
    complete_rows = [
        row
        for row in projection["rows"]
        if row["table_ref_kwd"] == complete[0]["table_ref"]
    ]
    assert [row["row_ordinal_int"] for row in complete_rows] == [3, 4, 5]
    assert all(
        row["row_role_kwd"] == "unknown"
        for row in projection["rows"]
        if row["row_ordinal_int"] >= 8
    )


@pytest.mark.parametrize("record_count", (3, 8))
def test_short_signoff_gap_does_not_depend_on_record_count(table_parser, record_count):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _complete_table_with_distant_multiline_signoff_bytes(record_count),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == record_count


def test_context_and_g1_splits_preserve_a_sparse_single_record_axis(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _context_with_sparse_single_record_bytes(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 1
    complete_rows = [
        row
        for row in projection["rows"]
        if row["table_ref_kwd"] == complete[0]["table_ref"]
    ]
    assert [row["row_ordinal_int"] for row in complete_rows] == [12]
    assert complete[0]["ordered_columns"]
    assert all(column["header_path"] for column in complete[0]["ordered_columns"])


def test_context_does_not_hide_a_trailing_dense_empty_record_axis(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _context_with_trailing_dense_empty_header_bytes(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 0
    assert complete[0]["matched_rule"] == "L1-08"
    assert len(complete[0]["ordered_columns"]) == 18

    # The trailing empty axis is the only independently enumerable structure
    # in this worksheet. Earlier signoff/context regions must not become
    # same-label sibling candidates and poison downstream completeness.
    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["matched_rule"] == "L1-08"
    assert projection["tables"][0]["source_total_count"] == 0


def test_trailing_empty_axis_does_not_absorb_an_independent_complete_table(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "State", "Owner"])
    sheet.append([1, "Open", "A"])
    sheet.append([2, "Closed", "B"])
    sheet.merge_cells("A6:F6")
    sheet["A6"] = "Anonymous empty register"
    for column in range(1, 7):
        sheet.cell(row=7, column=column, value=f"Empty field {column}")

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert [table["source_total_count"] for table in projection["tables"]] == [2, 0]
    assert [table["matched_rule"] for table in projection["tables"]] == [
        "L1-03",
        "L1-08",
    ]


def test_two_empty_axes_on_one_sheet_remain_separate(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    for title_row, header_row in ((1, 2), (6, 7)):
        sheet.merge_cells(
            start_row=title_row,
            start_column=1,
            end_row=title_row,
            end_column=4,
        )
        sheet.cell(title_row, 1, f"Anonymous register {title_row}")
        for column in range(1, 5):
            sheet.cell(header_row, column, f"Field {title_row}-{column}")

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert all(
        table["matched_rule"] == "L1-08"
        and table["source_total_count"] == 0
        for table in projection["tables"]
    )


def test_signoff_after_an_empty_axis_remains_a_separate_unknown(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Anonymous register"
    for column in range(1, 5):
        sheet.cell(row=2, column=column, value=f"Field {column}")
    sheet["A6"] = "Prepared by"
    sheet["B6"] = "Person"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert projection["tables"][0]["matched_rule"] == "L1-08"
    assert projection["tables"][0]["source_total_count"] == 0
    assert projection["tables"][1]["source_total_count"] is None


def test_unseparated_multiline_form_still_cannot_claim_a_complete_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Anonymous form"
    for row_ordinal in range(2, 5):
        sheet.merge_cells(
            start_row=row_ordinal,
            start_column=1,
            end_row=row_ordinal,
            end_column=2,
        )
        sheet.cell(row_ordinal, 1, f"Role {row_ordinal}")
        sheet.merge_cells(
            start_row=row_ordinal,
            start_column=3,
            end_row=row_ordinal,
            end_column=4,
        )
        sheet.cell(row_ordinal, 3, f"Signature {row_ordinal}")

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert all(
        table["enumeration_status"] != "supported_complete"
        for table in projection["tables"]
    )


def test_single_record_after_a_separated_structural_header_is_complete(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Anonymous context"
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "Context key"
    sheet.merge_cells("C2:D2")
    sheet["C2"] = "Context value"
    sheet["A4"] = "Sequence"
    sheet.merge_cells("B4:C4")
    sheet["B4"] = "Item"
    sheet["D4"] = "Status"
    sheet["A5"] = 1
    sheet.merge_cells("B5:C5")
    sheet["B5"] = "I-1"
    sheet["D5"] = "Open"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 1
    assert [row["row_ordinal_int"] for row in projection["rows"]] == [5]


def test_intermediate_title_belongs_to_the_following_structured_table(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Sequence", "Item", "Status"])
    sheet.append([1, "I-1", "Open"])
    sheet.append([2, "I-2", "Closed"])
    sheet["A8"] = "Anonymous following section"
    sheet.merge_cells("A10:C10")
    sheet["A10"] = "Anonymous following register"
    sheet.append(["Code", "Owner", "State"])
    sheet.append(["F-1", "Team", "Open"])
    sheet.append(["F-2", "Team", "Closed"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert [
        table["source_total_count"]
        for table in projection["tables"]
        if table["source_total_count"] is not None
    ] == [2, 2]


def test_numeric_key_only_slots_are_preserved_but_not_counted_as_business_records(
    table_parser,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "Anonymous register"
    for column, header in enumerate(
        ("Sequence", "Item", "Measure A", "Measure B", "Status", "Note"),
        start=1,
    ):
        sheet.cell(row=2, column=column, value=header)
    sheet.append([1, "I-1", 10, None, "Open", None])
    sheet.append([2, None, None, None, None, None])
    sheet.append([3, None, None, None, None, None])
    sheet.append([4, None, None, None, None, None])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 1
    rows = [
        row
        for row in projection["rows"]
        if row["table_ref_kwd"] == complete[0]["table_ref"]
    ]
    assert [row["row_ordinal_int"] for row in rows] == [3, 4, 5, 6]
    assert [row["row_role_kwd"] for row in rows] == [
        "data",
        "note",
        "note",
        "note",
    ]
    assert [row["data_row_index_int"] for row in rows] == [1, None, None, None]
    assert all(row["source_total_count_int"] == 1 for row in rows)


def test_numeric_key_only_slot_between_filled_rows_is_not_a_business_record(
    table_parser,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Sequence", "Item", "Status"])
    sheet.append([1, "I-1", "Open"])
    sheet.append([2, None, None])
    sheet.append([3, "I-3", "Closed"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] == 2
    assert [row["row_role_kwd"] for row in projection["rows"]] == [
        "data",
        "note",
        "data",
    ]
    assert [row["data_row_index_int"] for row in projection["rows"]] == [
        1,
        None,
        2,
    ]


def test_single_numeric_key_only_slot_is_not_a_business_record(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Sequence", "Item", "Status"])
    sheet.append([1, "I-1", "Open"])
    sheet.append([2, "I-2", "Closed"])
    sheet.append([3, None, None])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] == 2
    assert [row["row_role_kwd"] for row in projection["rows"]] == [
        "data",
        "data",
        "note",
    ]


def test_sparse_row_with_any_non_key_field_remains_a_business_record(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Sequence", "Item", "Status"])
    sheet.append([1, "I-1", "Open"])
    sheet.append([2, "I-2", None])
    sheet.append([3, "I-3", "Closed"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] == 3
    assert all(row["row_role_kwd"] == "data" for row in projection["rows"])


def test_single_column_numeric_records_are_not_treated_as_key_only_slots(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Reading"])
    sheet.append([10])
    sheet.append([20])
    sheet.append([30])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] == 3
    assert all(row["row_role_kwd"] == "data" for row in projection["rows"])


def test_header_only_axis_ignores_a_nonadjacent_sidecar_column(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:G1")
    sheet["A1"] = "Anonymous register"
    for column, header in enumerate(
        (
            "Sequence",
            "Reference",
            "Received",
            "Issuer",
            "Issued",
            "Reason",
            "Category",
        ),
        start=1,
    ):
        sheet.cell(row=2, column=column, value=header)
    sheet["I2"] = "https://example.invalid/evidence"

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 0
    assert complete[0]["matched_rule"] == "L1-08"


def test_headerless_record_slots_preserve_every_row_and_cannot_claim_complete(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    sheet.append(["A-3", "Open"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert [row["row_ordinal_int"] for row in projection["rows"]] == [1, 2, 3]
    assert all(row["row_role_kwd"] == "unknown" for row in projection["rows"])


def test_unrelated_uncached_formula_does_not_downgrade_complete_sibling(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Status"])
    sheet.append(["S-1", "Open"])
    sheet.append(["S-2", "Closed"])
    sheet.cell(row=20_001, column=20, value="=SUM(1, 1)")

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 2
    assert [table["source_total_count"] for table in projection["tables"]] == [2, None]
    assert [table["matched_rule"] for table in projection["tables"]] == ["L1-07", "R1"]


def test_multilevel_merged_header_stays_with_its_record_axis(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _merged_header_workbook_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] == 2
    assert [row["row_ordinal_int"] for row in projection["rows"]] == [3, 4]


def test_nested_sparse_header_is_not_emitted_as_a_data_row(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _nested_sparse_header_workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    table = projection["tables"][0]
    assert table["source_total_count"] == 3
    assert table["matched_rule"] == "L1-04"
    assert [row["row_ordinal_int"] for row in projection["rows"]] == [5, 6, 7]
    assert [row["row_role_kwd"] for row in projection["rows"]] == [
        "data",
        "data",
        "data",
    ]
    assert [
        column["column_id"] for column in table["ordered_columns"]
    ] == [f"col_v1:1:{ordinal}" for ordinal in range(1, 8)]
    assert [
        column["column_ordinal"] for column in table["ordered_columns"]
    ] == list(range(1, 8))


def test_merged_header_continuation_is_not_emitted_as_a_data_row(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _merged_header_continuation_workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    table = projection["tables"][0]
    assert table["source_total_count"] == 3
    assert [row["row_ordinal_int"] for row in projection["rows"]] == [6, 7, 8]
    assert all(row["row_role_kwd"] == "data" for row in projection["rows"])
    assert [
        column["column_id"] for column in table["ordered_columns"]
    ] == [f"col_v1:1:{ordinal}" for ordinal in range(1, 9)]


def test_optional_field_shape_change_does_not_break_the_record_axis(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _optional_field_shape_workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    table = projection["tables"][0]
    assert table["source_total_count"] == 3
    assert [row["row_ordinal_int"] for row in projection["rows"]] == [3, 4, 5]
    assert all(row["row_role_kwd"] == "data" for row in projection["rows"])


def test_continuous_record_axis_ignores_required_field_shape_changes(
    table_parser,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Shape-changing records"
    sheet.append(
        ["Process", "Operation", "Characteristic", "Requirement", "Method"]
    )
    sheet.append(["80", "Press", "Seat", "No damage", "Visual"])
    sheet.append(["90", "Rivet", "Torque", "< 8N.m", "Tester"])
    sheet.append(["90", None, "Angle", "25 deg min", "Tester"])
    sheet.append(["100", "Oil", "Grease", "0.7-1.0 g", "Valve"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    table = projection["tables"][0]
    assert table["source_total_count"] == 4
    assert [row["row_ordinal_int"] for row in projection["rows"]] == [2, 3, 4, 5]
    assert [row["row_role_kwd"] for row in projection["rows"]] == [
        "data",
        "data",
        "data",
        "data",
    ]


def test_multilevel_sparse_table_ignores_context_only_g1_disagreement(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _multilevel_sparse_table_with_context_child_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    table = projection["tables"][0]
    assert table["matched_rule"] == "L1-04"
    assert table["source_total_count"] == 10
    assert [row["row_ordinal_int"] for row in projection["rows"]] == list(
        range(9, 19)
    )
    assert all(row["row_role_kwd"] == "data" for row in projection["rows"])
    assert [
        column["column_id"] for column in table["ordered_columns"]
    ] == [f"col_v1:1:{ordinal}" for ordinal in range(1, 9)]


def test_nested_record_axis_g1_children_are_covered_by_one_dominant_axis():
    workbook = Workbook()
    worksheet = workbook.active
    dominant = {
        (3, 1),
        (3, 2),
        (3, 3),
        (4, 1),
        (4, 2),
        (5, 1),
        (5, 2),
    }
    nested = {(4, 3), (5, 3)}
    region = {
        "members": dominant | nested,
        "g1_children": [dominant, nested],
    }

    assert tabular_structure._g1_disagreement_is_outside_record_axis(
        worksheet,
        region,
        {3, 4, 5},
    )


def test_side_by_side_record_axis_g1_children_are_not_treated_as_nested():
    workbook = Workbook()
    worksheet = workbook.active
    left = {(3, 1), (4, 1), (5, 1)}
    right = {(3, 3), (4, 3), (5, 3)}
    region = {
        "members": left | right,
        "g1_children": [left, right],
    }

    assert not tabular_structure._g1_disagreement_is_outside_record_axis(
        worksheet,
        region,
        {3, 4, 5},
    )


def test_repeated_axis_g1_child_before_records_is_not_swallowed_as_context(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _multilevel_sparse_table_with_context_child_bytes(context_row_count=2),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] is None
    assert projection["tables"][0]["matched_rule"] == "R8"


def test_g_sensitive_sparse_table_remains_one_unknown_projection(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _sparse_region_workbook_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] is None
    assert {row["row_ordinal_int"] for row in projection["rows"]} >= {1, 2, 3, 5, 6}
    assert all(row["row_role_kwd"] == "unknown" for row in projection["rows"])


def test_two_non_isomorphic_rows_cannot_prove_a_record_axis(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Status", "Owner"])
    sheet.append(["A-1", "Open", None])
    sheet.append(["A-2", None, "North"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None


def test_three_level_merged_header_does_not_become_a_data_row(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _three_level_header_workbook_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] == 2
    assert [row["row_ordinal_int"] for row in projection["rows"]] == [4, 5]


def test_four_level_merged_header_has_no_fixed_depth_limit(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _four_level_header_workbook_bytes(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] == 2
    assert [row["row_ordinal_int"] for row in projection["rows"]] == [5, 6]


def test_formula_only_region_is_preserved_as_unknown(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _formula_only_workbook_bytes(),
        parser=table_parser,
    )

    assert projection["tables"]
    assert projection["tables"][0]["source_total_count"] is None
    assert len(projection["rows"]) >= 1
    assert all(row["row_role_kwd"] == "unknown" for row in projection["rows"])


def test_duplicate_values_remain_distinct_and_rows_use_only_fixed_fields(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(include_context=False),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )
    target = [row for row in projection["rows"] if row["table_label_kwd"] == "Inspection"]
    data_rows = [row for row in target if row["row_role_kwd"] == "data"]
    uncertain_rows = [row for row in target if row["row_role_kwd"] == "unknown"]

    assert len(data_rows) == 3
    assert len({row["row_ref_kwd"] for row in data_rows}) == 3
    assert [row["data_row_index_int"] for row in data_rows] == [1, 2, 3]
    assert {row["source_total_count_int"] for row in target} == {None}
    assert len(uncertain_rows) == 1
    assert uncertain_rows[0]["data_row_index_int"] is None
    assert all(set(row) == PROJECTION_ROW_FIELDS for row in projection["rows"])
    assert json.loads(data_rows[0]["ordered_fields_list"]) == [
        {
            "column_id": "col_v1:1:1",
            "column_ordinal": 1,
            "header_path": ["Code"],
            "name": "Code",
            "value": "R-DUP",
        },
        {
            "column_id": "col_v1:1:2",
            "column_ordinal": 2,
            "header_path": ["Description"],
            "name": "Description",
            "value": "Repeated",
        },
        {
            "column_id": "col_v1:1:3",
            "column_ordinal": 3,
            "header_path": ["Force"],
            "name": "Force",
            "value": "7.5",
        },
    ]
    assert all("Code" not in set(row) for row in projection["rows"])

    alternate = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(headers=("Component", "Supplier", "Unit")),
        parser=table_parser,
    )
    assert {frozenset(row) for row in alternate["rows"]} == {PROJECTION_ROW_FIELDS}


def test_ordered_fields_remove_only_insignificant_float_zeroes():
    fields = _ordered_fields(
        ["Integral", "Decimal", "Text"],
        [1.0, 1.5, "01-A"],
        note=False,
    )

    assert fields == [
        {"name": "Integral", "value": "1"},
        {"name": "Decimal", "value": "1.5"},
        {"name": "Text", "value": "01-A"},
    ]


def test_stable_plain_text_rows_are_data_and_prove_the_denominator(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PlainText"
    sheet.append(["Characteristic", "Type", "Requirement"])
    sheet.append(["Rolling resistance", "SC", "<= 7.5 N/KN"])
    sheet.append(["Uniformity", "SC", "RFV <= 170 N"])
    sheet.append(["Stiffness", "SC", "Radial >= 320 N/mm"])
    output = BytesIO()
    workbook.save(output)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        output.getvalue(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )
    target = [row for row in projection["rows"] if row["table_label_kwd"] == "PlainText"]

    assert [row["row_role_kwd"] for row in target] == ["data", "data", "data"]
    assert [row["data_row_index_int"] for row in target] == [1, 2, 3]
    assert {row["source_total_count_int"] for row in target} == {3}


def test_repeated_partial_merge_rows_are_data_when_the_shape_is_stable(table_parser, monkeypatch):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MergedText"
    sheet.append(["Characteristic", "Type", "Requirement"])
    for index, requirement in enumerate(("<= 7.5 N/KN", "RFV <= 170 N", ">= 320 N/mm"), start=2):
        sheet.merge_cells(start_row=index, start_column=1, end_row=index, end_column=2)
        sheet.cell(index, 1, f"Characteristic {index}")
        sheet.cell(index, 3, requirement)
    monkeypatch.setattr(
        table_parser,
        "_parse_sheet_structure",
        lambda _worksheet, _rows: (["Characteristic", "Type", "Requirement"], 0, 1),
    )
    output = BytesIO()
    workbook.save(output)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        output.getvalue(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )
    target = [row for row in projection["rows"] if row["table_label_kwd"] == "MergedText"]

    assert [row["row_role_kwd"] for row in target] == ["data", "data", "data"]
    assert {row["source_total_count_int"] for row in target} == {3}


@pytest.mark.parametrize(
    ("workbook_factory", "sheet_name", "column_count", "expected_row_count"),
    [
        (_horizontal_varying_merge_workbook_bytes, "Anonymous horizontal merges", 4, 3),
        (_vertical_varying_merge_workbook_bytes, "Anonymous vertical merges", 4, 4),
        (_mixed_row_merge_supplier_workbook_bytes, "Anonymous mixed merges", 5, 4),
    ],
)
def test_varying_record_merge_shapes_preserve_the_record_axis(
    table_parser,
    monkeypatch,
    workbook_factory,
    sheet_name,
    column_count,
    expected_row_count,
):
    monkeypatch.setattr(
        table_parser,
        "_parse_sheet_structure",
        lambda _worksheet, _rows: (
            [f"Field {chr(65 + index)}" for index in range(column_count)],
            0,
            1,
        ),
    )
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        workbook_factory(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    table = next(
        table
        for table in projection["tables"]
        if table["table_label"] == sheet_name
        and table["row_count"] == expected_row_count
    )
    rows = [
        row
        for row in projection["rows"]
        if row["table_ref_kwd"] == table["table_ref"]
    ]

    assert [row["row_ordinal_int"] for row in rows] == list(
        range(2, expected_row_count + 2)
    )
    assert all(
        "Sidecar label" not in row["ordered_fields_list"]
        and "example.invalid" not in row["ordered_fields_list"]
        for row in rows
    )
    assert [row["row_role_kwd"] for row in rows] == ["data"] * expected_row_count
    assert [row["data_row_index_int"] for row in rows] == list(
        range(1, expected_row_count + 1)
    )
    assert table["source_total_count"] == expected_row_count


def test_varying_record_merge_shapes_are_proven_by_the_real_candidate_stage(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _mixed_row_merge_supplier_workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    table = next(
        table
        for table in projection["tables"]
        if table["table_label"] == "Anonymous mixed merges"
    )
    rows = [
        row
        for row in projection["rows"]
        if row["table_ref_kwd"] == table["table_ref"]
    ]

    assert [row["row_ordinal_int"] for row in rows] == [2, 3, 4, 5]
    assert [row["row_role_kwd"] for row in rows] == ["data"] * 4
    assert [row["data_row_index_int"] for row in rows] == [1, 2, 3, 4]
    assert table["source_total_count"] == 4


def test_header_candidate_selection_prefers_the_largest_proven_record_axis(
    table_parser,
):
    workbook = table_parser._load_excel_to_workbook(
        BytesIO(_competing_header_depth_candidate_workbook_bytes())
    )
    worksheet = workbook.active
    rows, _populated_rows, _unresolved_rows = (
        tabular_structure._complete_worksheet_rows(worksheet)
    )

    headers, header_start, data_start = tabular_structure._parse_region_structure(
        table_parser,
        worksheet,
        rows,
    )

    assert data_start == 7, (header_start, data_start, headers)
    assert len(rows[data_start:]) == 10


def test_context_form_does_not_become_part_of_the_multilevel_header(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _context_form_before_multilevel_header_workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    table = next(
        table
        for table in projection["tables"]
        if table["table_label"] == "Anonymous performance report"
        and table["source_total_count"] == 2
    )
    assert [column["header_path"] for column in table["ordered_columns"]] == [
        ["Sequence"],
        ["Test"],
        ["Description"],
        ["Requirement"],
        ["Quantity"],
        ["Equipment"],
        ["Date"],
        ["Measured", "Run 1"],
        ["Measured", "Run 2"],
        ["Measured", "Run 3"],
    ]


def test_context_form_pairs_metadata_by_source_merge_geometry(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _context_form_before_multilevel_header_workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    table = next(
        table
        for table in projection["tables"]
        if table["table_label"] == "Anonymous performance report"
        and table["source_total_count"] == 2
    )
    assert table["table_context"] == [
        {"name": "context", "value": "Anonymous performance report"},
        {"name": "context", "value": "Unpaired context"},
        {"name": "Provider", "value": "Provider A"},
        {"name": "Laboratory", "value": "Laboratory A"},
        {"name": "Component", "value": "Component A"},
        {"name": "Reference", "value": "Reference A"},
        {"name": "Revision", "value": "Revision A"},
        {"name": "Fixture", "value": "Fixture A"},
    ]


def test_unused_trailing_table_column_does_not_hide_right_side_context(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _context_form_with_an_unused_trailing_table_column_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    table = next(
        table
        for table in projection["tables"]
        if table["source_total_count"] == 2
    )
    assert [column["header_path"] for column in table["ordered_columns"]] == [
        ["Sequence"],
        ["Operation"],
        ["Characteristic", "Product"],
        ["Characteristic", "Process"],
        ["Class"],
        ["Requirement"],
        ["Method"],
    ]
    assert table["table_context"][:5] == [
        {"name": "context", "value": "Anonymous characteristics"},
        {"name": "Provider", "value": "Provider A"},
        {"name": "Component", "value": "Component A"},
        {"name": "Program", "value": "Program A"},
        {"name": "Reference", "value": "Reference A"},
    ]


def test_right_side_action_does_not_expand_the_header_candidate_context_width(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _context_form_with_a_right_side_action_before_multilevel_header_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    table = next(
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
        and table["source_total_count"] == 2
    )
    assert [column["header_path"] for column in table["ordered_columns"]] == [
        ["Sequence"],
        ["Operation"],
        ["Characteristic", "Product"],
        ["Characteristic", "Process"],
        ["Class"],
        ["Requirement"],
        ["Method"],
    ]
    assert table["table_context"][:5] == [
        {"name": "context", "value": "Anonymous characteristics"},
        {"name": "Provider", "value": "Provider A"},
        {"name": "Component", "value": "Component A"},
        {"name": "Program", "value": "Program A"},
        {"name": "Reference", "value": "Reference A"},
    ]


def test_repeated_pages_allow_an_optional_parent_header_path(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _multilevel_repeated_form_with_optional_parent_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 4


def test_repeated_pages_with_a_conflicting_leaf_header_remain_separate(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _multilevel_repeated_form_with_optional_parent_bytes(
            second_process_header="Method",
        ),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 2
    assert [table["source_total_count"] for table in complete] == [2, 2]


def test_vertically_paired_context_values_do_not_split_repeated_pages(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _vertically_paired_context_repeated_form_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 4
    assert complete[0]["table_context"][:3] == [
        {"name": "context", "value": "Anonymous paged form"},
        {"name": "Part", "value": "P-1"},
        {"name": "Supplier", "value": "Supplier A"},
    ]


def test_conflicting_vertically_paired_context_keys_keep_pages_separate(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _vertically_paired_context_repeated_form_bytes(
            second_context_name="Drawing",
        ),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert [table["source_total_count"] for table in complete] == [2, 2]


def test_empty_context_values_keep_source_labels_as_repeated_form_identity(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _vertically_paired_context_repeated_form_bytes(
            first_context_values=False,
        ),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 4


def test_vertical_record_merge_proves_an_axis_across_an_empty_display_row(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _vertically_merged_record_key_with_an_empty_display_row_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 4
    assert [
        row["row_ordinal_int"]
        for row in projection["rows"]
        if row["row_role_kwd"] == "data"
    ] == [2, 3, 5, 6]


def test_context_sidecar_cannot_turn_a_proven_multilevel_table_into_an_empty_axis(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _multilevel_table_after_context_gap_with_sidecar_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 5
    assert [
        row["row_ordinal_int"]
        for row in projection["rows"]
        if row["table_ref_kwd"] == complete[0]["table_ref"]
        and row["row_role_kwd"] == "data"
    ] == [6, 7, 8, 9, 10]


def test_context_form_followed_by_a_multilevel_header_proves_an_empty_axis(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _context_form_before_multilevel_header_workbook_bytes(record_count=0),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 0
    assert complete[0]["matched_rule"] == "L1-08"
    assert [column["header_path"] for column in complete[0]["ordered_columns"]] == [
        ["Sequence"],
        ["Test"],
        ["Description"],
        ["Requirement"],
        ["Quantity"],
        ["Equipment"],
        ["Date"],
        ["Measured", "Run 1"],
        ["Measured", "Run 2"],
        ["Measured", "Run 3"],
    ]
    assert projection["rows"] == []


def test_context_form_with_horizontally_merged_leaf_headers_proves_an_empty_axis(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _context_form_before_horizontally_merged_leaf_empty_axis_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 0
    assert complete[0]["matched_rule"] == "L1-08"
    assert [column["header_path"] for column in complete[0]["ordered_columns"]] == [
        ["Sequence"],
        ["Material"],
        ["Material"],
        ["Requirement"],
        ["Standard"],
        ["Standard"],
        ["Result"],
        ["Outcome", "Measured"],
        ["Outcome", "Unit"],
        ["Outcome", "Conclusion"],
    ]
    assert [column["column_ordinal"] for column in complete[0]["ordered_columns"]] == list(
        range(1, 11)
    )
    assert projection["rows"] == []


def test_proven_records_take_precedence_over_a_horizontally_merged_empty_axis_candidate(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _context_form_before_horizontally_merged_leaf_empty_axis_bytes(
            merge_record_fields=True,
            record_count=2,
        ),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 2
    assert complete[0]["matched_rule"] == "L1-04"
    assert [
        row["row_ordinal_int"]
        for row in projection["rows"]
        if row["table_ref_kwd"] == complete[0]["table_ref"]
        and row["row_role_kwd"] == "data"
    ] == [9, 10]


def test_header_levels_without_a_proven_record_key_remain_an_empty_axis(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _context_form_before_horizontally_merged_leaf_empty_axis_bytes(
            numeric_record_key=False,
            rectangular_header_continuation=True,
            record_count=2,
        ),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 0
    assert complete[0]["matched_rule"] == "L1-08"
    assert projection["rows"] == []


def test_source_backed_records_after_a_rectangular_header_boundary_take_precedence(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _context_form_before_horizontally_merged_leaf_empty_axis_bytes(
            merge_record_fields=True,
            numeric_record_key=False,
            record_count=2,
        ),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    complete = [
        table
        for table in projection["tables"]
        if table["enumeration_status"] == "supported_complete"
    ]
    assert len(complete) == 1
    assert complete[0]["source_total_count"] == 2
    assert complete[0]["matched_rule"] == "L1-04"
    assert [
        row["row_ordinal_int"]
        for row in projection["rows"]
        if row["table_ref_kwd"] == complete[0]["table_ref"]
        and row["row_role_kwd"] == "data"
    ] == [9, 10]


def test_duplicate_context_empty_axis_preserves_an_existing_trailing_proof():
    context = (
        ["Sequence", "Material", "Material"],
        [["Sequence"], ["Material"], ["Material"]],
        4,
        13,
    )
    trailing = (
        ["Sequence", "Material", "Requirement"],
        [["Sequence"], ["Material"], ["Requirement"]],
        2,
        13,
    )

    assert tabular_structure._preferred_empty_record_axis_structure(
        context,
        trailing,
    ) is trailing


def test_unique_context_empty_axis_keeps_its_existing_precedence():
    context = (
        ["Sequence", "Material", "Requirement"],
        [["Sequence"], ["Material"], ["Requirement"]],
        4,
        13,
    )
    trailing = (
        ["Sequence", "Material", "Requirement"],
        [["Sequence"], ["Material"], ["Requirement"]],
        2,
        13,
    )

    assert tabular_structure._preferred_empty_record_axis_structure(
        context,
        trailing,
    ) is context


def test_one_dimensional_duplicate_header_paths_do_not_prove_an_empty_axis(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _context_form_before_horizontally_merged_leaf_empty_axis_bytes(
            rectangular_leaf_merges=False,
        ),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    assert not any(
        table["enumeration_status"] == "supported_complete"
        and table["source_total_count"] == 0
        for table in projection["tables"]
    )


def test_context_form_with_only_one_merged_semantic_path_is_not_an_empty_axis(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _context_form_before_horizontally_merged_leaf_empty_axis_bytes(
            single_semantic_path=True
        ),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    assert not any(
        table["enumeration_status"] == "supported_complete"
        and table["source_total_count"] == 0
        for table in projection["tables"]
    )


def test_context_preceded_empty_axis_discovery_has_linear_membership_checks(
    table_parser,
):
    class MembershipCountingRows(list):
        def __init__(self, values):
            super().__init__(values)
            self.contains_calls = 0

        def __contains__(self, value):
            self.contains_calls += 1
            return super().__contains__(value)

    workbook = Workbook()
    sheet = workbook.active
    for row_ordinal in range(1, 201):
        sheet.cell(row_ordinal, 1, f"Row {row_ordinal}")
    rows = list(sheet.iter_rows())
    populated_rows = MembershipCountingRows(range(1, 201))

    assert tabular_structure._context_preceded_multilevel_empty_axis(
        table_parser,
        sheet,
        rows,
        populated_rows,
        [],
    ) is None
    assert populated_rows.contains_calls <= len(populated_rows) * 2


def test_header_candidate_does_not_fold_a_merged_record_group_into_the_header(
    table_parser,
):
    workbook = table_parser._load_excel_to_workbook(
        BytesIO(_competing_merge_free_tail_candidate_workbook_bytes())
    )
    worksheet = workbook.active
    rows, _populated_rows, _unresolved_rows = (
        tabular_structure._complete_worksheet_rows(worksheet)
    )

    headers, header_start, data_start = tabular_structure._parse_region_structure(
        table_parser,
        worksheet,
        rows,
    )

    assert data_start == 3, (header_start, data_start, headers)
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _competing_merge_free_tail_candidate_workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )
    assert projection["tables"][0]["source_total_count"] == 6
    assert [
        row["row_ordinal_int"]
        for row in projection["rows"]
        if row["row_role_kwd"] == "data"
    ] == [4, 5, 6, 7, 8, 9]


def test_header_candidate_excludes_all_rows_of_a_multilevel_header(table_parser):
    workbook = table_parser._load_excel_to_workbook(
        BytesIO(_competing_multilevel_header_candidate_workbook_bytes())
    )
    worksheet = workbook.active
    rows, _populated_rows, _unresolved_rows = (
        tabular_structure._complete_worksheet_rows(worksheet)
    )

    headers, header_start, data_start = tabular_structure._parse_region_structure(
        table_parser,
        worksheet,
        rows,
    )

    assert data_start == 6, (header_start, data_start, headers)
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _competing_multilevel_header_candidate_workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )
    assert projection["tables"][0]["source_total_count"] == 5
    assert [
        row["row_ordinal_int"]
        for row in projection["rows"]
        if row["row_role_kwd"] == "data"
    ] == [7, 8, 9, 10, 11]


def test_header_candidate_excludes_a_single_level_header_after_context(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _single_level_header_after_context_workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] == 13
    assert [
        row["row_ordinal_int"]
        for row in projection["rows"]
        if row["row_role_kwd"] == "data"
    ] == list(range(8, 21))


def test_header_candidate_accepts_lossless_mixed_storage_numeric_keys(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _mixed_storage_numeric_key_candidate_workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    table = next(
        table
        for table in projection["tables"]
        if table["table_label"] == "Anonymous mixed storage"
    )
    rows = [
        row
        for row in projection["rows"]
        if row["table_ref_kwd"] == table["table_ref"]
    ]

    assert table["source_total_count"] == 10
    assert [row["row_ordinal_int"] for row in rows] == list(range(8, 18))
    assert [row["row_role_kwd"] for row in rows] == ["data"] * 10
    assert [row["data_row_index_int"] for row in rows] == list(range(1, 11))
    assert [
        json.loads(row["ordered_fields_list"])[0]["value"] for row in rows
    ] == [str(value) for value in range(1, 11)]


def test_vertical_merge_parent_values_are_inherited_and_emitted_once(
    table_parser,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _vertical_varying_merge_workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )

    table = next(
        table
        for table in projection["tables"]
        if table["table_label"] == "Anonymous vertical merges"
    )
    rows = [
        row
        for row in projection["rows"]
        if row["table_ref_kwd"] == table["table_ref"]
    ]
    expected_parent_values = {
        2: "A-group-1",
        3: "A-group-1",
        4: "A-3",
        5: "A-4",
    }

    assert [row["row_role_kwd"] for row in rows] == ["data"] * 4
    for row in rows:
        fields = json.loads(row["ordered_fields_list"])
        parent_fields = [field for field in fields if field["column_ordinal"] == 1]
        assert len(parent_fields) == 1
        assert parent_fields[0]["value"] == expected_parent_values[row["row_ordinal_int"]]
    assert table["source_total_count"] == 4


def test_repeated_header_inside_a_table_still_invalidates_completeness(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RepeatedHeader"
    sheet.append(["Code", "Description"])
    sheet.append(["A-1", "First"])
    sheet.append(["Code", "Description"])
    sheet.append(["A-2", "Second"])
    output = BytesIO()
    workbook.save(output)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        output.getvalue(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )
    target = [row for row in projection["rows"] if row["table_label_kwd"] == "RepeatedHeader"]
    repeated_header = next(row for row in target if row["row_ordinal_int"] == 3)

    assert repeated_header["row_role_kwd"] == "unknown"
    assert repeated_header["data_row_index_int"] is None
    assert {row["source_total_count_int"] for row in target} == {None}


def test_table_parser_exposes_the_projection_producer_without_using_chunk_output(monkeypatch):
    table = _load_table_module(monkeypatch)

    projection = table.build_structure_projection("anonymous.xlsx", _workbook_bytes())

    assert projection["version"] == "tabular-structure-projection/v6"
    assert projection["rows"]
    assert all("content_with_weight" not in row for row in projection["rows"])


def test_unknown_body_row_invalidates_the_table_denominator_without_becoming_a_note(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(include_unknown=True),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )
    target = [row for row in projection["rows"] if row["table_label_kwd"] == "Inspection"]
    sparse = next(row for row in target if "Sparse body row" in row["ordered_fields_list"])

    assert sparse["row_role_kwd"] == "unknown"
    assert sparse["data_row_index_int"] is None
    assert {row["source_total_count_int"] for row in target} == {None}


def test_full_width_merged_row_inside_the_body_fails_closed_instead_of_becoming_a_note(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(include_merged_body=True),
        parser=table_parser,
    )
    target = [row for row in projection["rows"] if row["table_label_kwd"] == "Inspection"]
    merged_body = next(row for row in target if "R-MERGED" in row["ordered_fields_list"])

    assert merged_body["row_role_kwd"] == "unknown"
    assert merged_body["data_row_index_int"] is None
    assert {row["source_total_count_int"] for row in target} == {None}


def test_lone_full_width_merged_body_row_cannot_be_silently_excluded_as_a_note(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Description"])
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "R-MERGED"
    output = BytesIO()
    workbook.save(output)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        output.getvalue(),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert projection["rows"][0]["row_role_kwd"] == "unknown"
    assert projection["rows"][0]["data_row_index_int"] is None


def test_final_full_width_merged_row_after_data_cannot_be_assumed_to_be_a_note(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Description"])
    sheet.append(["A-1", "Ordinary"])
    sheet.merge_cells("A3:B3")
    sheet["A3"] = "A-2"
    output = BytesIO()
    workbook.save(output)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        output.getvalue(),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert projection["rows"][-1]["row_role_kwd"] == "unknown"


def test_repeated_full_width_body_rows_need_positive_note_evidence(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Status"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    for row_ordinal, value in ((4, "A-3"), (5, "A-4")):
        sheet.merge_cells(
            start_row=row_ordinal,
            start_column=1,
            end_row=row_ordinal,
            end_column=2,
        )
        sheet.cell(row_ordinal, 1, value)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert [row["row_ordinal_int"] for row in projection["rows"]] == [2, 3, 4, 5]
    assert [row["row_role_kwd"] for row in projection["rows"][-2:]] == ["unknown", "unknown"]


def test_mixed_generation_and_inconsistent_totals_fail_validation(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )
    projection["rows"][0]["producer_generation_ref_kwd"] = _generation_ref()

    with pytest.raises(ValueError, match="mixed producer generations"):
        validate_tabular_structure_projection(projection)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )
    projection["rows"][0]["source_total_count_int"] = 999

    with pytest.raises(ValueError, match="source total"):
        validate_tabular_structure_projection(projection)


def test_manifest_and_record_identity_tampering_fail_validation(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(),
        parser=table_parser,
    )
    projection["tables"][0]["row_count"] += 1

    with pytest.raises(ValueError, match="table manifest"):
        validate_tabular_structure_projection(projection)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(),
        parser=table_parser,
    )
    projection["rows"][0]["id"] = "tsr_v1_tampered"

    with pytest.raises(ValueError, match="record identity"):
        validate_tabular_structure_projection(projection)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(),
        parser=table_parser,
    )
    projection["tables"][0]["table_ref"] = "tbl_v1_" + "0" * 64

    with pytest.raises(ValueError, match="table reference"):
        validate_tabular_structure_projection(projection)


def test_multi_region_manifest_requires_contiguous_physical_order(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _vertical_multi_region_workbook_bytes(),
        parser=table_parser,
    )
    projection["tables"].reverse()

    with pytest.raises(ValueError, match="physical order"):
        validate_tabular_structure_projection(projection)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _vertical_multi_region_workbook_bytes(),
        parser=table_parser,
    )
    second_table_ref = projection["tables"][1]["table_ref"]
    projection["tables"] = [projection["tables"][1]]
    projection["rows"] = [
        row for row in projection["rows"] if row["table_ref_kwd"] == second_table_ref
    ]

    with pytest.raises(ValueError, match="contiguous"):
        validate_tabular_structure_projection(projection)


def test_malformed_json_evidence_fields_fail_validation(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(),
        parser=table_parser,
    )
    projection["rows"][0]["ordered_fields_list"] = "not-json"

    with pytest.raises(ValueError, match="ordered fields"):
        validate_tabular_structure_projection(projection)


@pytest.mark.parametrize("payload", [5, None, True, {"unexpected": "value"}])
def test_fixed_field_schema_rejects_non_list_json_payloads(table_parser, payload):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(),
        parser=table_parser,
    )
    projection["rows"][0]["ordered_fields_list"] = json.dumps(payload)

    with pytest.raises(ValueError, match="ordered fields must use the fixed field schema"):
        validate_tabular_structure_projection(projection)


@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("tabular_structure_version_kwd", "tabular-row/v0", "row version"),
        ("producer_schema_version_kwd", "table-producer/v0", "producer schema"),
        ("structure_kind_kwd", "ordinary_chunk", "structure kind"),
        ("row_ordinal_int", "5", "row ordinal"),
        ("data_row_index_int", "1", "data row index"),
        ("source_total_count_int", "3", "source total"),
    ],
)
def test_fixed_row_schema_rejects_wrong_versions_kinds_and_integer_types(
    table_parser,
    field,
    value,
    message,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(),
        parser=table_parser,
    )
    projection["rows"][0][field] = value

    with pytest.raises(ValueError, match=message):
        validate_tabular_structure_projection(projection)


def test_fixed_row_schema_rejects_non_string_table_label(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(),
        parser=table_parser,
    )
    projection["rows"][0]["table_label_kwd"] = 123

    with pytest.raises(ValueError, match="table label"):
        validate_tabular_structure_projection(projection)


@pytest.mark.parametrize("field", ["sheet_ordinal", "table_ordinal", "row_count", "data_row_count", "source_total_count"])
def test_manifest_rejects_boolean_integer_fields(table_parser, field):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(),
        parser=table_parser,
    )
    projection["tables"][0][field] = True

    with pytest.raises(ValueError, match="table manifest"):
        validate_tabular_structure_projection(projection)


def test_header_only_sheet_produces_no_false_table_manifest(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Description"])
    output = BytesIO()
    workbook.save(output)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        output.getvalue(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] is None
    assert all(row["row_role_kwd"] == "unknown" for row in projection["rows"])


def test_second_table_after_an_internal_separator_fails_closed(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value"])
    sheet.append(["A-1", 1])
    sheet.append(["A-2", 2])
    sheet.append([None, None])
    sheet.append(["Supplier", "Status"])
    sheet.append(["North", "Approved"])
    output = BytesIO()
    workbook.save(output)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        output.getvalue(),
        parser=table_parser,
    )

    assert len(projection["tables"]) == 1
    assert projection["tables"][0]["source_total_count"] is None
    assert all(row["row_role_kwd"] == "unknown" for row in projection["rows"])


def test_adjacent_second_table_with_a_new_row_shape_fails_closed(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value"])
    sheet.append(["A-1", 1])
    sheet.append(["A-2", 2])
    sheet.append(["Supplier", "Status"])
    sheet.append(["North", "Approved"])
    output = BytesIO()
    workbook.save(output)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        output.getvalue(),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert any(row["row_role_kwd"] == "unknown" for row in projection["rows"])


def test_adjacent_second_table_with_the_same_text_shape_fails_closed(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value"])
    sheet.append(["A-1", "One"])
    sheet.append(["A-2", "Two"])
    sheet.append(["Supplier", "Status"])
    sheet.append(["North", "Approved"])
    output = BytesIO()
    workbook.save(output)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        output.getvalue(),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert any(row["row_role_kwd"] == "unknown" for row in projection["rows"])


def test_complete_projection_does_not_drop_data_after_a_large_blank_region(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value"])
    sheet.append(["A-1", 1])
    sheet.cell(row=20_001, column=1, value="A-2")
    sheet.cell(row=20_001, column=2, value=2)
    output = BytesIO()
    workbook.save(output)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        output.getvalue(),
        parser=table_parser,
    )

    assert any(row["row_ordinal_int"] == 20_001 for row in projection["rows"])
    assert projection["tables"][0]["source_total_count"] is None


def test_large_gap_before_the_first_body_row_also_fails_closed(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value"])
    sheet.cell(row=20_001, column=1, value="A-1")
    sheet.cell(row=20_001, column=2, value=1)
    output = BytesIO()
    workbook.save(output)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        output.getvalue(),
        parser=table_parser,
    )

    assert projection["tables"][0]["source_total_count"] is None
    assert projection["rows"][0]["row_role_kwd"] == "unknown"


def test_complete_projection_reads_to_the_last_source_row_without_a_fixed_probe(
    monkeypatch,
    table_parser,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value"])
    sheet.append(["A-1", 1])
    sheet.cell(row=20_001, column=1, value="A-2")
    sheet.cell(row=20_001, column=2, value=2)
    original_iter_rows = sheet.iter_rows

    observed_max_rows = []

    def full_source_iter_rows(*args, **kwargs):
        observed_max_rows.append(kwargs.get("max_row", 0))
        return original_iter_rows(*args, **kwargs)

    monkeypatch.setattr(sheet, "iter_rows", full_source_iter_rows)
    monkeypatch.setattr(table_parser, "_load_excel_to_workbook", lambda _source: workbook)

    tabular_structure._complete_worksheet_rows(sheet)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        b"complete-source-bytes",
        parser=table_parser,
    )

    assert any(row["row_ordinal_int"] == 20_001 for row in projection["rows"])
    assert max(observed_max_rows) == 20_001


def test_uncached_formula_only_row_is_preserved_as_unknown(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value"])
    sheet.append(["A-1", 1])
    sheet.append(["=CONCAT(\"A-\", 2)", "=SUM(1, 1)"])
    output = BytesIO()
    workbook.save(output)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        output.getvalue(),
        parser=table_parser,
    )

    formula_row = next(row for row in projection["rows"] if row["row_ordinal_int"] == 3)
    assert formula_row["row_role_kwd"] == "unknown"
    assert projection["tables"][0]["source_total_count"] is None


def test_row_with_value_and_uncached_formula_is_emitted_once_as_unknown(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value"])
    sheet.append(["A-1", 1])
    sheet.append(["A-2", "=SUM(1, 1)"])
    output = BytesIO()
    workbook.save(output)

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        output.getvalue(),
        parser=table_parser,
    )

    target_rows = [row for row in projection["rows"] if row["row_ordinal_int"] == 3]
    assert len(target_rows) == 1
    assert target_rows[0]["row_role_kwd"] == "unknown"
    assert projection["tables"][0]["source_total_count"] is None


def test_all_uncached_formula_rows_are_preserved_as_unknown(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Status"])
    sheet.append(["A-1", "Open"])
    sheet.cell(row=3, column=1, value='=CONCAT("A-", 2)')
    sheet.cell(row=3, column=2, value='=CONCAT("Closed")')
    sheet.cell(row=4, column=1, value='=CONCAT("A-", 3)')
    sheet.cell(row=4, column=2, value='=CONCAT("Open")')

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )

    assert {row["row_ordinal_int"] for row in projection["rows"]} >= {2, 3, 4}
    assert {
        row["row_ordinal_int"]: row["row_role_kwd"]
        for row in projection["rows"]
    } == {2: "data", 3: "unknown", 4: "unknown"}
    assert all(table["source_total_count"] is None for table in projection["tables"])


def test_biff_formula_inventory_uses_only_sheet_and_cell_coordinates():
    global_bof = struct.pack("<HHHH", 0x0809, 4, 0x0600, 0x0005)
    worksheet_bof = struct.pack("<HHHH", 0x0809, 4, 0x0600, 0x0010)
    formula = struct.pack("<HHHH", 0x0006, 4, 2, 3)
    eof = struct.pack("<HH", 0x000A, 0)
    sheet_offset = len(global_bof) + 4 + 6 + len(eof)
    boundsheet_payload = struct.pack("<IBB", sheet_offset, 0, 0)
    boundsheet = struct.pack("<HH", 0x0085, len(boundsheet_payload)) + boundsheet_payload
    stream = global_bof + boundsheet + eof + worksheet_bof + formula + eof

    assert _formula_coordinates_from_biff_stream(stream) == [{(3, 4)}]


def test_biff_formula_inventory_preserves_nonworksheet_sheet_ordinals():
    global_bof = struct.pack("<HHHH", 0x0809, 4, 0x0600, 0x0005)
    macro_bof = struct.pack("<HHHH", 0x0809, 4, 0x0600, 0x0040)
    worksheet_bof = struct.pack("<HHHH", 0x0809, 4, 0x0600, 0x0010)
    formula = struct.pack("<HHHH", 0x0006, 4, 2, 3)
    eof = struct.pack("<HH", 0x000A, 0)
    first_boundsheet_length = 4 + 8
    second_boundsheet_length = 4 + 8
    macro_offset = len(global_bof) + first_boundsheet_length + second_boundsheet_length + len(eof)
    worksheet_offset = macro_offset + len(macro_bof) + len(eof)
    macro_payload = struct.pack("<IBBBB", macro_offset, 2, 1, 0, 0)
    worksheet_payload = struct.pack("<IBBBB", worksheet_offset, 0, 0, 0, 0)
    macro_boundsheet = struct.pack("<HH", 0x0085, len(macro_payload)) + macro_payload
    worksheet_boundsheet = (
        struct.pack("<HH", 0x0085, len(worksheet_payload)) + worksheet_payload
    )
    stream = (
        global_bof
        + macro_boundsheet
        + worksheet_boundsheet
        + eof
        + macro_bof
        + eof
        + worksheet_bof
        + formula
        + eof
    )

    assert _formula_coordinates_from_biff_stream(stream) == [set(), {(3, 4)}]


def test_biff_formula_cached_result_inventory_preserves_result_kind_and_ordinals():
    global_bof = struct.pack("<HHHH", 0x0809, 4, 0x0600, 0x0005)
    macro_bof = struct.pack("<HHHH", 0x0809, 4, 0x0600, 0x0040)
    worksheet_bof = struct.pack("<HHHH", 0x0809, 4, 0x0600, 0x0010)
    numeric_formula_payload = (
        struct.pack("<HHH", 2, 3, 0)
        + struct.pack("<d", 60.0)
        + b"\x00" * 6
    )
    empty_formula_payload = (
        struct.pack("<HHH", 4, 5, 0)
        + b"\x03\x00\x00\x00\x00\x00\xff\xff"
        + b"\x00" * 6
    )
    numeric_formula = struct.pack(
        "<HH", 0x0006, len(numeric_formula_payload)
    ) + numeric_formula_payload
    empty_formula = struct.pack(
        "<HH", 0x0006, len(empty_formula_payload)
    ) + empty_formula_payload
    eof = struct.pack("<HH", 0x000A, 0)
    first_boundsheet_length = 4 + 8
    second_boundsheet_length = 4 + 8
    macro_offset = len(global_bof) + first_boundsheet_length + second_boundsheet_length + len(eof)
    worksheet_offset = macro_offset + len(macro_bof) + len(eof)
    macro_payload = struct.pack("<IBBBB", macro_offset, 2, 1, 0, 0)
    worksheet_payload = struct.pack("<IBBBB", worksheet_offset, 0, 0, 0, 0)
    stream = (
        global_bof
        + struct.pack("<HH", 0x0085, len(macro_payload))
        + macro_payload
        + struct.pack("<HH", 0x0085, len(worksheet_payload))
        + worksheet_payload
        + eof
        + macro_bof
        + eof
        + worksheet_bof
        + numeric_formula
        + empty_formula
        + eof
    )

    assert _formula_cached_result_kinds_from_biff_stream(stream) == [
        {},
        {(3, 4): "numeric", (5, 6): "empty"},
    ]


def test_cached_empty_formula_coordinate_remains_in_source_region_membership(
    table_parser,
    monkeypatch,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Value", "Calculated"])
    sheet.append(["A-1", 1, None])
    binary = _save_workbook(workbook)
    monkeypatch.setattr(
        tabular_structure,
        "_formula_coordinates_by_sheet",
        lambda _binary: ([{(2, 3)}], True),
    )
    monkeypatch.setattr(
        tabular_structure,
        "_formula_cached_result_kinds_by_sheet",
        lambda _binary: [{(2, 3): "empty"}],
    )
    captured = []
    original = tabular_structure._worksheet_structure_regions

    def capture(
        parser,
        worksheet,
        sheet_ordinal,
        unresolved_formula_coordinates=None,
        formula_coordinates=None,
    ):
        captured.append(
            (
                set(unresolved_formula_coordinates or ()),
                set(formula_coordinates or ()),
            )
        )
        return original(
            parser,
            worksheet,
            sheet_ordinal,
            unresolved_formula_coordinates,
            formula_coordinates,
        )

    monkeypatch.setattr(tabular_structure, "_worksheet_structure_regions", capture)

    build_tabular_structure_projection("anonymous.xls", binary, parser=table_parser)

    assert captured == [(set(), {(2, 3)})]


def test_generation_reference_rejects_customer_text(table_parser):
    with pytest.raises(ValueError, match="UUID/ULID"):
        build_tabular_structure_projection(
            "anonymous.xlsx",
            _workbook_bytes(),
            producer_generation_ref="supplier-run-2026",
            parser=table_parser,
        )


def test_context_is_bounded_and_removes_controls_without_a_language_allowlist(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inspection"
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "\u9879\u76ee \u00b5\u03a9\u2103\u0085\u202e"
    sheet.append(["Code", "Description", "Force"])
    sheet.append(["A-1", "Open", 1])
    sheet.append(["A-2", "Closed", 2])
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        producer_generation_ref=_generation_ref(),
        table_context_entry_limit=1,
        table_context_value_bytes=32,
        parser=table_parser,
    )
    target = next(row for row in projection["rows"] if row["table_label_kwd"] == "Inspection")
    context = json.loads(target["table_context_list"])

    assert len(context) == 1
    assert all(len(item["value"].encode("utf-8")) <= 32 for item in context)
    assert "\u0085" not in json.dumps(context, ensure_ascii=False)
    assert "\u202e" not in json.dumps(context, ensure_ascii=False)
    assert "\u9879\u76ee \u00b5\u03a9\u2103" in json.dumps(context, ensure_ascii=False)


def test_global_indices_and_totals_survive_projection_part_boundaries(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(rows=3002, include_note=False, include_context=False),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )
    parts = partition_tabular_structure_projection(projection, rows_per_part=3000)
    target = [row for row in projection["rows"] if row["table_label_kwd"] == "Inspection"]
    data_rows = [row for row in target if row["row_role_kwd"] == "data"]

    assert len(parts) == 2
    assert [row["data_row_index_int"] for row in data_rows] == list(range(1, 3003))
    assert {row["source_total_count_int"] for row in target} == {3002}
    assert len({row["row_ref_kwd"] for row in target}) == len(target)


def test_projection_build_does_not_change_ordinary_table_chunks(monkeypatch, table_parser):
    table = _load_table_module(monkeypatch)
    source = _workbook_bytes()
    before = table.chunk(
        "anonymous.xlsx",
        binary=source,
        callback=lambda *_args: None,
        kb_id="ordinary-kb",
    )

    build_tabular_structure_projection(
        "anonymous.xlsx",
        source,
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )
    after = table.chunk(
        "anonymous.xlsx",
        binary=source,
        callback=lambda *_args: None,
        kb_id="ordinary-kb",
    )

    assert json.dumps(after, ensure_ascii=False, sort_keys=True) == json.dumps(before, ensure_ascii=False, sort_keys=True)


class _VerifiedStorage:
    def __init__(self, fail_on=None):
        self.objects = {}
        self.fail_on = fail_on

    def put(self, bucket, name, binary, tenant_id=None):
        if self.fail_on and self.fail_on(name):
            return None
        self.objects[(bucket, name)] = bytes(binary)

    def obj_exist(self, bucket, name, tenant_id=None):
        return (bucket, name) in self.objects

    def get(self, bucket, name, tenant_id=None):
        return self.objects.get((bucket, name))


class _StorageWithoutTenantReadArguments(_VerifiedStorage):
    def obj_exist(self, bucket, name):
        return (bucket, name) in self.objects

    def get(self, bucket, name):
        return self.objects.get((bucket, name))


def test_sidecar_storage_writes_verified_parts_before_the_immutable_manifest(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )
    storage = _VerifiedStorage()

    receipt = store_tabular_structure_projection(
        storage,
        bucket="dataset-id",
        document_id="document-id",
        projection=projection,
        rows_per_part=2,
        tenant_id="tenant-id",
    )

    names = [name for bucket, name in storage.objects if bucket == "dataset-id"]
    assert "/manifest-" in names[-1] and names[-1].endswith(".json")
    assert all("/part-" not in name or len(name.rsplit("-", 1)[-1].removesuffix(".json")) == 64 for name in names)
    assert all("anonymous" not in name and "Inspection" not in name for name in names)
    manifest = json.loads(storage.objects[("dataset-id", receipt["manifest_object_name"])])
    assert manifest["producer_generation_ref"] == projection["producer_generation_ref"]
    for part in manifest["parts"]:
        payload = storage.objects[("dataset-id", part["object_name"])]
        assert hashlib.sha256(payload).hexdigest() == part["sha256"]


def test_sidecar_storage_never_publishes_a_manifest_after_a_part_write_failure(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(),
        producer_generation_ref=_generation_ref(),
        parser=table_parser,
    )
    storage = _VerifiedStorage(fail_on=lambda name: "/part-000002-" in name)

    with pytest.raises(IOError, match="verify structure projection object"):
        store_tabular_structure_projection(
            storage,
            bucket="dataset-id",
            document_id="document-id",
            projection=projection,
            rows_per_part=2,
        )

    assert not any("/manifest-" in name for _bucket, name in storage.objects)


def test_sidecar_storage_supports_backends_without_tenant_read_arguments(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _workbook_bytes(),
        parser=table_parser,
    )
    storage = _StorageWithoutTenantReadArguments()

    receipt = store_tabular_structure_projection(
        storage,
        bucket="dataset-id",
        document_id="document-id",
        projection=projection,
        tenant_id="tenant-id",
    )

    assert "/manifest-" in receipt["manifest_object_name"]
