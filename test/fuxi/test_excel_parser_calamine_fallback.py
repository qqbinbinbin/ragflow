import importlib.util
import sys
import types
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook


def _load_excel_parser_class():
    rag_nlp = types.ModuleType("rag.nlp")
    rag_nlp.find_codec = lambda *_args, **_kwargs: "utf-8"
    lazy_image = types.ModuleType("rag.utils.lazy_image")
    lazy_image.LazyImage = object
    sys.modules.setdefault("rag.nlp", rag_nlp)
    sys.modules.setdefault("rag.utils.lazy_image", lazy_image)
    path = Path("deepdoc/parser/excel_parser.py")
    spec = importlib.util.spec_from_file_location("fuxi_excel_parser", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def test_calamine_fallback_preserves_raw_rows_merges_and_empty_cells(monkeypatch):
    module = _load_excel_parser_class()
    RAGFlowExcelParser = module.RAGFlowExcelParser

    class FakeSheet:
        start = (1, 0)
        merged_cell_ranges = [((1, 0), (1, 2)), ((3, 0), (4, 0))]

        @staticmethod
        def iter_rows():
            return iter([
                ["", "", ""],
                ["Supplier report", "", ""],
                ["", "", ""],
                ["Part", "Supplier", "Status"],
                ["", "Name", "Quality"],
                ["P-1", "North", "Approved"],
            ])

    class FakeEmptySheet:
        start = None
        merged_cell_ranges = []

        @staticmethod
        def iter_rows():
            return iter([])

    class FakeCalamineWorkbook:
        sheet_names = ["Cover", "Supplier matrix"]

        @staticmethod
        def from_filelike(_file):
            return FakeCalamineWorkbook()

        @staticmethod
        def get_sheet_by_name(name):
            return FakeEmptySheet() if name == "Cover" else FakeSheet()

    def fake_load_workbook(*_args, **_kwargs):
        raise ValueError("legacy OLE workbook")

    monkeypatch.setattr(module, "load_workbook", fake_load_workbook)
    monkeypatch.setitem(
        sys.modules,
        "python_calamine",
        types.SimpleNamespace(CalamineWorkbook=FakeCalamineWorkbook),
    )

    workbook = RAGFlowExcelParser._load_excel_to_workbook(
        BytesIO(b"\xd0\xcf\x11\xe0fixture")
    )

    assert workbook.sheetnames == ["Cover", "Supplier matrix"]
    assert workbook["Cover"].max_row == 1
    sheet = workbook["Supplier matrix"]
    assert sheet["A1"].value is None
    assert sheet["A2"].value == "Supplier report"
    assert sheet["A3"].value is None
    assert sheet["A4"].value == "Part"
    assert sheet["A6"].value == "P-1"
    assert sheet["B6"].value == "North"
    assert str(sheet["B3"].value).lower() != "nan"
    assert {str(cell_range) for cell_range in sheet.merged_cells.ranges} == {
        "A2:C2",
        "A4:A5",
    }


def test_dataframe_workbook_conversion_keeps_missing_values_empty():
    module = _load_excel_parser_class()
    workbook = module.RAGFlowExcelParser._dataframe_to_workbook(
        pd.DataFrame({"Part": ["P-1"], "Optional": [float("nan")]})
    )

    assert workbook.active["B2"].value is None
