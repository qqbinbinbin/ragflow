import json
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

import rag.app.tabular_structure as tabular_structure
from rag.app.tabular_structure import build_tabular_structure_projection
from rag.app.tabular_structure import (
    StructureSnapshotChanged,
    load_tabular_structure_projection,
    store_tabular_structure_projection,
)
from test.fuxi.test_adr039_lovable_review_export import (
    _build_anonymous_workbook,
    _load_baseline,
)
from test.fuxi.test_table_semantic_rows import _load_table_module


def _save_workbook(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _multilevel_header_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous structure"
    sheet.merge_cells("A1:G1")
    sheet["A1"] = "Shared context"
    sheet.merge_cells("A2:A3")
    sheet["A2"] = "Code"
    sheet.merge_cells("B2:C2")
    sheet["B2"] = "Product"
    sheet.merge_cells("D2:E2")
    sheet["D2"] = "Process"
    sheet.merge_cells("F2:G2")
    sheet["F2"] = "Evidence"
    for column, value in enumerate(
        ["Status", "Measure", "Status", "Measure", "Status", "Measure"],
        start=2,
    ):
        sheet.cell(row=3, column=column, value=value)
    for index in range(1, 11):
        sheet.append(
            [
                f"A-{index:02d}",
                "Open" if index % 2 else "Closed",
                index,
                "Closed" if index % 2 else "Open",
                index * 2,
                "Ready",
                index * 3,
            ]
        )
    return _save_workbook(workbook)


def _repeated_text_anchor_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous structure"
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "Same text"
    sheet.merge_cells("A2:B2")
    sheet["A2"] = "Same text"
    sheet["A3"] = "Code / type: A|B\nC"
    sheet["B3"] = "State"
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    return _save_workbook(workbook)


def _unnamed_superset_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous structure"
    sheet.append(["Code", "State"])
    sheet.append(["A-1", "Open"])
    sheet.append(["A-2", "Closed"])
    sheet.append([None, None, None])
    sheet.append([None, None, None])
    sheet.append(["B-1", "Open", "Unproven"])
    return _save_workbook(workbook)


def _horizontally_merged_data_anchor_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous structure"
    sheet.append(["Code", "Criteria", None, "Note"])
    sheet.merge_cells("B1:C1")
    for row_ordinal, code, criteria, note in (
        (2, "A-1", "Within limit", "Reviewed"),
        (3, "A-2", "At limit", "Pending"),
        (4, "A-3", "Above limit", "Escalated"),
    ):
        sheet.cell(row=row_ordinal, column=1, value=code)
        sheet.cell(row=row_ordinal, column=2, value=criteria)
        sheet.merge_cells(
            start_row=row_ordinal,
            start_column=2,
            end_row=row_ordinal,
            end_column=3,
        )
        sheet.cell(row=row_ordinal, column=4, value=note)
    return _save_workbook(workbook)


@pytest.fixture
def table_parser(monkeypatch):
    return _load_table_module(monkeypatch).Excel()


def _data_fields(projection):
    data_rows = [
        row for row in projection["rows"] if row["row_role_kwd"] == "data"
    ]
    assert data_rows
    return [json.loads(row["ordered_fields_list"]) for row in data_rows]


class _MemoryStorage:
    def __init__(self):
        self.objects = {}

    def obj_exist(self, bucket, object_name, tenant_id=None):
        return (bucket, object_name) in self.objects

    def put(self, bucket, object_name, payload, tenant_id=None):
        self.objects[(bucket, object_name)] = payload

    def get(self, bucket, object_name, tenant_id=None):
        return self.objects.get((bucket, object_name))


def test_header_structure_contract_uses_the_reviewed_strict_versions():
    assert tabular_structure.TABULAR_STRUCTURE_VERSION == "tabular-row/v2"
    assert tabular_structure.PRODUCER_SCHEMA_VERSION == "table-producer/v6"
    assert tabular_structure.PROJECTION_VERSION == "tabular-structure-projection/v6"
    assert tabular_structure.PROJECTION_PART_VERSION == "tabular-structure-part/v3"
    assert tabular_structure.STRUCTURE_PRODUCER_ALGORITHM_VERSION == "region-producer/v20"
    assert tabular_structure.ENUMERATION_RULE_VERSION == "enumeration-rules/v9"


def test_table_identity_remains_bound_to_every_reviewed_source_version(monkeypatch):
    source_sha256 = "a" * 64
    membership_sha256 = "b" * 64
    baseline = tabular_structure._table_ref(source_sha256, 1, 1, membership_sha256)
    changed = set()
    for name, value in [
        ("PRODUCER_SCHEMA_VERSION", "table-producer/test-next"),
        ("PROJECTION_VERSION", "tabular-structure-projection/test-next"),
        ("STRUCTURE_PRODUCER_ALGORITHM_VERSION", "region-producer/test-next"),
        ("ENUMERATION_RULE_VERSION", "enumeration-rules/test-next"),
    ]:
        with monkeypatch.context() as scoped:
            scoped.setattr(tabular_structure, name, value)
            changed.add(
                tabular_structure._table_ref(
                    source_sha256,
                    1,
                    1,
                    membership_sha256,
                )
            )

    assert baseline not in changed
    assert len(changed) == 4


def test_projection_part_version_is_strictly_enforced_inside_ragflow_storage(
    table_parser,
    monkeypatch,
):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _multilevel_header_workbook(),
        parser=table_parser,
    )
    storage = _MemoryStorage()
    monkeypatch.setattr(
        tabular_structure,
        "PROJECTION_PART_VERSION",
        "tabular-structure-part/v2",
    )
    receipt = store_tabular_structure_projection(
        storage,
        bucket="anonymous-bucket",
        document_id="anonymous-document",
        projection=projection,
    )

    monkeypatch.setattr(
        tabular_structure,
        "PROJECTION_PART_VERSION",
        "tabular-structure-part/v3",
    )
    with pytest.raises(StructureSnapshotChanged, match="part generation changed"):
        load_tabular_structure_projection(
            storage,
            bucket="anonymous-bucket",
            document_id="anonymous-document",
            producer_generation_ref=receipt["producer_generation_ref"],
            manifest_object_name=receipt["manifest_object_name"],
            manifest_sha256=receipt["manifest_sha256"],
        )


def test_multilevel_header_projection_exposes_source_structural_column_evidence(
    table_parser,
):
    source = _multilevel_header_workbook()
    workbook = load_workbook(BytesIO(source), data_only=False)
    headers, header_start, data_start = tabular_structure._parse_region_structure(
        table_parser,
        workbook.active,
        list(workbook.active.iter_rows()),
    )

    assert header_start == 0
    assert data_start == 3
    assert headers == [
        "Shared context-Code",
        "Shared context-Product-Status",
        "Shared context-Product-Measure",
        "Shared context-Process-Status",
        "Shared context-Process-Measure",
        "Shared context-Evidence-Status",
        "Shared context-Evidence-Measure",
    ]

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        source,
        parser=table_parser,
    )

    assert projection["tables"][0]["enumeration_status"] == "supported_complete"
    assert projection["tables"][0]["source_total_count"] == 10
    fields_by_row = _data_fields(projection)
    for fields in fields_by_row:
        for field in fields:
            assert "column_id" in field
            assert "column_ordinal" in field
            assert "header_path" in field
    assert all(
        [field["column_id"] for field in fields]
        == [f"col_v1:1:{column}" for column in range(1, 8)]
        for fields in fields_by_row
    )
    assert all(
        [field["column_ordinal"] for field in fields] == list(range(1, 8))
        for fields in fields_by_row
    )
    assert [field["header_path"] for field in fields_by_row[0]] == [
        ["Shared context", "Code"],
        ["Shared context", "Product", "Status"],
        ["Shared context", "Product", "Measure"],
        ["Shared context", "Process", "Status"],
        ["Shared context", "Process", "Measure"],
        ["Shared context", "Evidence", "Status"],
        ["Shared context", "Evidence", "Measure"],
    ]
    assert all(
        set(field) == {
            "column_id",
            "column_ordinal",
            "header_path",
            "name",
            "value",
        }
        for fields in fields_by_row
        for field in fields
    )


def test_header_path_deduplicates_merge_identity_not_equal_text(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _repeated_text_anchor_workbook(),
        parser=table_parser,
    )

    fields = _data_fields(projection)[0]
    assert "header_path" in fields[0]
    assert fields[0]["header_path"] == [
        "Same text",
        "Same text",
        "Code / type: A|B\nC",
    ]


def test_horizontal_merged_data_anchor_is_emitted_once_per_record(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _horizontally_merged_data_anchor_workbook(),
        parser=table_parser,
    )

    assert projection["tables"][0]["enumeration_status"] == "supported_complete"
    assert projection["tables"][0]["source_total_count"] == 3
    for fields in _data_fields(projection):
        assert [field["column_id"] for field in fields] == [
            "col_v1:1:1",
            "col_v1:1:2",
            "col_v1:1:4",
        ]
        assert [field["column_ordinal"] for field in fields] == [1, 2, 4]
        assert [field["header_path"] for field in fields] == [
            ["Code"],
            ["Criteria"],
            ["Note"],
        ]
        assert sum(
            field["value"] in {"Within limit", "At limit", "Above limit"}
            for field in fields
        ) == 1


def test_single_level_header_remains_one_atomic_path_segment(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous structure"
    header = "Code / type: A|B\nC"
    sheet.append([header, "State"])
    sheet.append(["A-1", "Open"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )
    fields = _data_fields(projection)[0]

    assert "header_path" in fields[0]
    assert fields[0]["header_path"] == [header]


@pytest.mark.parametrize(
    "sample_code",
    [
        "L1-CONTINUATION-EQUAL",
        "L1-CONTINUATION-SUBSET",
        "L1-CONTINUATION-SUPERSET-NAMED",
    ],
)
def test_continuation_keeps_structural_column_identity_and_real_paths(
    sample_code,
    table_parser,
):
    sample = next(
        item for item in _load_baseline()["samples"]
        if item["sample_code"] == sample_code
    )
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _build_anonymous_workbook(sample),
        parser=table_parser,
    )
    fields_by_row = _data_fields(projection)

    for fields in fields_by_row:
        for field in fields:
            assert "column_id" in field
            assert "column_ordinal" in field
            assert "header_path" in field
            assert field["header_path"]
    identity_by_ordinal = {
        field["column_ordinal"]: field["column_id"]
        for fields in fields_by_row
        for field in fields
        if field["value"]
    }
    for fields in fields_by_row:
        assert all(
            identity_by_ordinal[field["column_ordinal"]] == field["column_id"]
            for field in fields
        )


def test_continuation_sparse_values_keep_full_table_column_ordinals(table_parser):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous structure"
    sheet.append(["Code", "State", "Owner"])
    sheet.append(["A-1", "Open", "Team-1"])
    sheet.append(["A-2", "Closed", "Team-2"])
    sheet.append([None, None, None])
    sheet.append([None, None, None])
    sheet.append(["B-1", None, "Team-3"])

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _save_workbook(workbook),
        parser=table_parser,
    )
    continuation_row = next(
        fields
        for fields in _data_fields(projection)
        if any(field["value"] == "B-1" for field in fields)
    )

    assert [field["column_ordinal"] for field in continuation_row] == [1, 3]
    assert [field["column_id"] for field in continuation_row] == [
        "col_v1:1:1",
        "col_v1:1:3",
    ]
    assert [field["header_path"] for field in continuation_row] == [
        ["Code"],
        ["Owner"],
    ]


def test_unnamed_superset_remains_d1_without_synthetic_header_paths(table_parser):
    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        _unnamed_superset_workbook(),
        parser=table_parser,
    )

    assert projection["tables"][0]["enumeration_status"] == "defect"
    assert projection["tables"][0]["matched_rule"] == "D1"
    assert projection["tables"][0]["enumeration_reason"] == "unnamed_required_field"
    assert all("Column_" not in row["ordered_fields_list"] for row in projection["rows"])


@pytest.mark.parametrize(
    "sample",
    _load_baseline()["samples"],
    ids=lambda sample: sample["sample_code"],
)
def test_new_header_contract_cannot_flip_an_anonymous_l1_baseline(sample, table_parser):
    binary = _build_anonymous_workbook(sample)
    kwargs = {}
    if sample["sample_code"] == "L1-ADR044-XLS-EQUIVALENT":
        import hashlib

        kwargs["adr044_conversion_receipt"] = {
            "original_source_sha256": "1" * 64,
            "converted_source_sha256": hashlib.sha256(binary).hexdigest(),
            "converter_version": "anonymous-converter/v1",
        }

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        binary,
        parser=table_parser,
        **kwargs,
    )
    tables = projection["tables"]

    assert [table["enumeration_status"] for table in tables] == sample[
        "expected_statuses"
    ]
    assert [table["enumeration_reason"] for table in tables] == [
        "record_axis_proven"
    ] * sample["expected_object_count"]
    assert [table["source_total_count"] for table in tables] == sample[
        "expected_source_totals"
    ]
    for fields in _data_fields(projection):
        for field in fields:
            assert "header_path" in field
            assert field["header_path"]
            assert not field["name"].startswith("Column_")
