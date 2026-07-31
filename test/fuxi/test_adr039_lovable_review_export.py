import hashlib
import importlib.util
import json
import subprocess
import sys
import unicodedata
from collections import Counter
from datetime import date
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from test.fuxi.test_table_semantic_rows import _load_table_module

import rag.app.tabular_structure as tabular_structure

from rag.app.tabular_structure import build_tabular_structure_projection


BASELINE_PATH = Path(__file__).parent / "fixtures" / "adr039-l1-regression-baseline.json"
EXPORTER_PATH = Path(__file__).parents[2] / "tools" / "export_adr039_lovable_review.py"

FROZEN_L1_EXPECTATIONS = {
    "L1-FIRST-OR-DELAYED-HEADER": (1, (3,), ("supported_complete",), ("L1-07",)),
    "L1-MERGED-MULTILEVEL-10": (1, (10,), ("supported_complete",), ("L1-04",)),
    "L1-SEPARATED-MULTI-LIST": (
        2,
        (2, 2),
        ("supported_complete", "supported_complete"),
        ("L1-03", "L1-03"),
    ),
    "L1-CONTINUATION-EQUAL": (1, (3,), ("supported_complete",), ("L1-02",)),
    "L1-CONTINUATION-SUBSET": (1, (3,), ("supported_complete",), ("L1-02",)),
    "L1-CONTINUATION-SUPERSET-NAMED": (1, (3,), ("supported_complete",), ("L1-02",)),
    "L1-SINGLE-COLUMN-CATALOGUE": (1, (2,), ("supported_complete",), ("L1-05",)),
    "L1-SPARSE-STABLE-SLOTS": (1, (3,), ("supported_complete",), ("L1-06",)),
    "L1-ADR044-XLS-EQUIVALENT": (1, (10,), ("supported_complete",), ("L1-01",)),
}

NEGATIVE_RULES = ("R1", "R2", "R3", "R4", "R5", "R6", "R7")


def _truth_vector(*rules):
    return {rule: rule in rules for rule in NEGATIVE_RULES}


ORDERED_COLLISION_FIXTURES = {
    "C12": (_truth_vector("R1", "R2"), "R1"),
    "C16": (_truth_vector("R1", "R6"), "R1"),
    "C17": (_truth_vector("R1", "R7"), "R1"),
    "C23": (_truth_vector("R2", "R3"), "R2"),
    "C24": (_truth_vector("R2", "R4"), "R2"),
    "C25": (_truth_vector("R2", "R5"), "R2"),
    "C26": (_truth_vector("R2", "R6"), "R2"),
    "C27": (_truth_vector("R2", "R7"), "R2"),
    "C34": (_truth_vector("R3", "R4"), "R3"),
    "C35": (_truth_vector("R3", "R5"), "R3"),
    "C36": (_truth_vector("R3", "R6"), "R3"),
    "C37": (_truth_vector("R3", "R7"), "R3"),
    "C45": (_truth_vector("R4", "R5"), "R4"),
    "C46": (_truth_vector("R4", "R6"), "R4"),
    "C47": (_truth_vector("R4", "R7"), "R4"),
    "C56": (_truth_vector("R5", "R6"), "R5"),
    "C57": (_truth_vector("R5", "R7"), "R5"),
    "C67": (_truth_vector("R6", "R7"), "R6"),
    "M01_R2_R4_R7": (_truth_vector("R2", "R4", "R7"), "R2"),
    "M02_R3_R4_R7": (_truth_vector("R3", "R4", "R7"), "R3"),
}


def _load_baseline():
    return json.loads(BASELINE_PATH.read_text(encoding="utf-8"))


def _load_exporter_module():
    spec = importlib.util.spec_from_file_location("adr039_lovable_review_exporter", EXPORTER_PATH)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _coordinate_digest(coordinates):
    payload = "\n".join(f"1:{row}:{column}" for row, column in sorted(coordinates)).encode("ascii")
    return hashlib.sha256(payload).hexdigest()


def _multiset_digest(coordinates):
    lines = ["adr039-emitted-cell-multiset/v1"]
    lines.extend(f"1:{row}:{column}" for row, column in sorted(coordinates))
    return hashlib.sha256("\n".join(lines).encode("ascii")).hexdigest()


def _typed_value_map_digest(cell_values):
    lines = ["adr039-cell-value-map/v1"]
    for row, column, value_type, value in sorted(cell_values):
        normalized = unicodedata.normalize("NFC", value) if value_type == "text" else value
        lines.append(
            json.dumps(
                [1, row, column, value_type, normalized],
                ensure_ascii=True,
                separators=(",", ":"),
            )
        )
    return hashlib.sha256("\n".join(lines).encode("ascii")).hexdigest()


def _merge_map_digest(merged_ranges):
    lines = ["adr039-merged-ranges/v1"]
    lines.extend(
        json.dumps([1, *merged_range], separators=(",", ":"))
        for merged_range in sorted(merged_ranges)
    )
    return hashlib.sha256("\n".join(lines).encode("ascii")).hexdigest()


def _rows_by_columns(rows, columns):
    return {(row, column) for row in rows for column in columns}


def _sample_objects(sample):
    shape = sample["shape"]
    case = sample["builder_case"]
    if case == "delayed_header_two_columns_three_records":
        return [_rows_by_columns(shape["header_rows"] + shape["record_rows"], shape["occupied_columns"])]
    if case in {
        "three_merged_header_levels_two_columns_ten_records",
        "adr044_biff8_conversion_of_three_merged_header_levels_two_columns_ten_records",
    }:
        assert shape["merged_non_anchor_membership"] == "included_when_anchor_nonblank"
        members = {
            (row, column)
            for min_row, min_column, max_row, max_column in shape["merged_ranges"]
            for row in range(min_row, max_row + 1)
            for column in range(min_column, max_column + 1)
        }
        members |= _rows_by_columns([shape["leaf_header_row"]], shape["occupied_columns"])
        members |= _rows_by_columns(shape["record_rows"], shape["occupied_columns"])
        return [members]
    if case == "two_vertical_lists_two_blank_rows_two_records_each":
        return [
            _rows_by_columns([header] + rows, shape["occupied_columns"])
            for header, rows in zip(shape["object_header_rows"], shape["object_record_rows"])
        ]
    if case in {
        "equal_columns_headerless_single_row_continuation",
        "subset_columns_headerless_single_row_continuation",
        "superset_columns_named_header_and_single_record",
    }:
        main = _rows_by_columns(
            [shape["main_header_row"]] + shape["main_record_rows"],
            shape["main_columns"],
        )
        continuation_rows = shape["continuation_record_rows"][:]
        if "continuation_header_row" in shape:
            continuation_rows.insert(0, shape["continuation_header_row"])
        continuation = _rows_by_columns(continuation_rows, shape["continuation_columns"])
        expected_components = sample["component_membership_sha256"]
        assert [_coordinate_digest(main), _coordinate_digest(continuation)] == expected_components
        assert main.isdisjoint(continuation)
        return [main | continuation]
    if case == "single_header_one_column_two_records":
        return [_rows_by_columns(shape["header_rows"] + shape["record_rows"], shape["occupied_columns"])]
    if case == "three_columns_anchor_first_three_sparse_records":
        members = _rows_by_columns(shape["header_rows"] + shape["record_rows"], shape["occupied_columns"])
        return [members - {tuple(cell) for cell in shape["empty_cells"]}]
    raise AssertionError(f"unhandled builder case: {case}")


def _save_workbook(workbook):
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _set_row(sheet, row, columns, values):
    assert len(columns) == len(values)
    for column, value in zip(columns, values):
        sheet.cell(row=row, column=column, value=value)


def _build_anonymous_workbook(sample):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous structure"
    shape = sample["shape"]
    case = sample["builder_case"]

    if case == "delayed_header_two_columns_three_records":
        _set_row(sheet, shape["header_rows"][0], shape["occupied_columns"], ["Code", "State"])
        for index, row in enumerate(shape["record_rows"], start=1):
            _set_row(sheet, row, shape["occupied_columns"], [f"R-{index:02d}", f"S-{index:02d}"])
    elif case in {
        "three_merged_header_levels_two_columns_ten_records",
        "adr044_biff8_conversion_of_three_merged_header_levels_two_columns_ten_records",
    }:
        labels = ["Anonymous level one", "Anonymous level two", "Anonymous level three"]
        for merged_range, label in zip(shape["merged_ranges"], labels):
            min_row, min_column, max_row, max_column = merged_range
            sheet.merge_cells(
                start_row=min_row,
                start_column=min_column,
                end_row=max_row,
                end_column=max_column,
            )
            sheet.cell(row=min_row, column=min_column, value=label)
        _set_row(sheet, shape["leaf_header_row"], shape["occupied_columns"], ["Code", "Date"])
        for index, row in enumerate(shape["record_rows"], start=1):
            _set_row(
                sheet,
                row,
                shape["occupied_columns"],
                [f"R-{index:02d}", date(2026, 1, index)],
            )
    elif case == "two_vertical_lists_two_blank_rows_two_records_each":
        for object_index, (header_row, record_rows) in enumerate(
            zip(shape["object_header_rows"], shape["object_record_rows"]),
            start=1,
        ):
            _set_row(sheet, header_row, shape["occupied_columns"], ["Code", "State"])
            for record_index, row in enumerate(record_rows, start=1):
                _set_row(
                    sheet,
                    row,
                    shape["occupied_columns"],
                    [f"R-{object_index}-{record_index}", f"S-{record_index}"],
                )
    elif case in {
        "equal_columns_headerless_single_row_continuation",
        "subset_columns_headerless_single_row_continuation",
        "superset_columns_named_header_and_single_record",
    }:
        main_columns = shape["main_columns"]
        field_names = ["Code", "State", "Owner", "Revision", "Date"]
        _set_row(
            sheet,
            shape["main_header_row"],
            main_columns,
            field_names[: len(main_columns)],
        )
        for record_index, row in enumerate(shape["main_record_rows"], start=1):
            record_values = [
                f"A-{record_index}",
                "Open" if record_index % 2 else "Closed",
                f"Team-{record_index}",
                f"R{record_index}",
                f"2026-01-0{record_index}",
            ]
            _set_row(
                sheet,
                row,
                main_columns,
                record_values[: len(main_columns)],
            )
        continuation_columns = shape["continuation_columns"]
        if "continuation_header_row" in shape:
            _set_row(
                sheet,
                shape["continuation_header_row"],
                continuation_columns,
                field_names[: len(continuation_columns)],
            )
        for record_index, row in enumerate(shape["continuation_record_rows"], start=3):
            record_values = [
                f"B-{record_index}",
                "Open",
                f"Team-{record_index}",
                f"R{record_index}",
                f"2026-01-0{record_index}",
            ]
            _set_row(
                sheet,
                row,
                continuation_columns,
                record_values[: len(continuation_columns)],
            )
    elif case == "single_header_one_column_two_records":
        _set_row(sheet, shape["header_rows"][0], shape["occupied_columns"], ["Code"])
        for index, row in enumerate(shape["record_rows"], start=1):
            _set_row(sheet, row, shape["occupied_columns"], [f"R-{index:02d}"])
    elif case == "three_columns_anchor_first_three_sparse_records":
        _set_row(sheet, shape["header_rows"][0], shape["occupied_columns"], ["Code", "State", "Note"])
        empty_cells = {tuple(cell) for cell in shape["empty_cells"]}
        for index, row in enumerate(shape["record_rows"], start=1):
            for column in shape["occupied_columns"]:
                if (row, column) not in empty_cells:
                    sheet.cell(row=row, column=column, value=index * 100 + column)
    else:
        raise AssertionError(f"unhandled builder case: {case}")

    return _save_workbook(workbook)


def _logical_workbook_members(binary):
    workbook = load_workbook(BytesIO(binary), data_only=False, read_only=False)
    sheet = workbook.active
    members = {
        (cell.row, cell.column)
        for cell in sheet._cells.values()
        if cell.value is not None and str(cell.value).strip()
    }
    for merged in sheet.merged_cells.ranges:
        anchor = sheet.cell(merged.min_row, merged.min_col).value
        if anchor is None or not str(anchor).strip():
            continue
        members.update(
            (row, column)
            for row in range(merged.min_row, merged.max_row + 1)
            for column in range(merged.min_col, merged.max_col + 1)
        )
    return members


def _table_membership_sha256(table):
    return table["table_ref"].split("_")[2]


@pytest.fixture
def table_parser(monkeypatch):
    return _load_table_module(monkeypatch).Excel()


def test_l1_baseline_freezes_reviewed_object_counts_totals_statuses_and_rules():
    baseline = _load_baseline()

    assert baseline["schema"] == "adr039-l1-regression-baseline/v2"
    assert baseline["runtime_input"] is False
    assert baseline["builder_contract"]["production_import_forbidden"] is True
    actual = {
        sample["sample_code"]: (
            sample["expected_object_count"],
            tuple(sample["expected_source_totals"]),
            tuple(sample["expected_statuses"]),
            tuple(sample["expected_rules"]),
        )
        for sample in baseline["samples"]
    }
    assert actual == FROZEN_L1_EXPECTATIONS


def test_l1_baseline_recomputes_all_membership_and_ingest_evidence():
    baseline = _load_baseline()

    assert baseline["builder_contract"]["membership_hash_input"].endswith("no trailing newline")
    assert "preserving duplicates" in baseline["builder_contract"]["emitted_cell_multiset_hash_input"]
    for sample in baseline["samples"]:
        objects = _sample_objects(sample)
        assert len(objects) == sample["expected_object_count"]
        assert [_coordinate_digest(members) for members in objects] == sample["expected_membership_sha256"]
        assert [
            _multiset_digest(members) for members in objects
        ] == sample["expected_emitted_cell_multiset_sha256"]
        assert sample["expected_emitted_cell_occurrence_count"] == [len(members) for members in objects]
        assert sample["expected_member_max_ingest_count"] == [1] * len(objects)


def test_l1_baseline_binds_duplicate_ingest_and_adr044_content_equivalence():
    baseline = _load_baseline()
    duplicate = baseline["defect_samples"][0]
    emitted = [tuple(cell) for cell in duplicate["emitted_member_coordinate_multiset"]]
    counts = Counter(emitted)

    assert duplicate["expected_check"] == "D2"
    assert duplicate["expected_status"] == "defect"
    assert duplicate["expected_reason"] == "membership_not_closed"
    assert max(counts.values()) == duplicate["expected_member_max_ingest_count"] == 2
    assert _coordinate_digest(set(emitted)) == duplicate["expected_membership_sha256"]
    assert len(emitted) == duplicate["expected_emitted_cell_occurrence_count"]
    assert _multiset_digest(emitted) == duplicate["expected_emitted_cell_multiset_sha256"]
    assert duplicate["expected_emitted_cell_multiset_sha256"] != duplicate["expected_membership_sha256"]

    adr044 = next(sample for sample in baseline["samples"] if sample["sample_code"] == "L1-ADR044-XLS-EQUIVALENT")
    equivalence = adr044["conversion_value_equivalence"]
    assert adr044["converter_version_contract"] == "immutable_converter_build_or_service_version_not_protocol_version"
    original_values = equivalence["original_typed_cell_values"]
    converted_values = equivalence["converted_typed_cell_values"]
    assert len(original_values) == len(converted_values) == equivalence["expected_value_entry_count"]
    assert _typed_value_map_digest(original_values) == equivalence["expected_original_value_map_sha256"]
    assert _typed_value_map_digest(converted_values) == equivalence["expected_converted_value_map_sha256"]
    assert original_values == converted_values

    original_merges = equivalence["original_merged_ranges"]
    converted_merges = equivalence["converted_merged_ranges"]
    assert _merge_map_digest(original_merges) == equivalence["expected_original_merge_map_sha256"]
    assert _merge_map_digest(converted_merges) == equivalence["expected_converted_merge_map_sha256"]
    assert original_merges == converted_merges == adr044["shape"]["merged_ranges"]


def test_anonymous_builder_recreates_every_reviewed_member_coordinate():
    baseline = _load_baseline()

    for sample in baseline["samples"]:
        binary = _build_anonymous_workbook(sample)
        expected_members = set().union(*_sample_objects(sample))
        assert _logical_workbook_members(binary) == expected_members, sample["sample_code"]


@pytest.mark.parametrize("sample", _load_baseline()["samples"], ids=lambda sample: sample["sample_code"])
def test_anonymous_l1_baseline_runs_through_the_real_producer(sample, table_parser):
    binary = _build_anonymous_workbook(sample)
    kwargs = {}
    if sample["sample_code"] == "L1-ADR044-XLS-EQUIVALENT":
        kwargs["adr044_conversion_receipt"] = {
            "original_source_sha256": "1" * 64,
            "converted_source_sha256": hashlib.sha256(binary).hexdigest(),
            "converter_version": "anonymous-converter/v1",
        }

    projection = build_tabular_structure_projection(
        "anonymous.xlsx",
        binary,
        parser=table_parser,
        **kwargs,
    )
    tables = projection["tables"]

    assert len(tables) == sample["expected_object_count"]
    assert [table["source_total_count"] for table in tables] == sample["expected_source_totals"]
    assert [table["enumeration_status"] for table in tables] == sample["expected_statuses"]
    assert [table["matched_rule"] for table in tables] == sample["expected_rules"]
    assert [_table_membership_sha256(table) for table in tables] == sample["expected_membership_sha256"]

    rows_by_table = {
        table["table_ref"]: [
            row for row in projection["rows"] if row["table_ref_kwd"] == table["table_ref"]
        ]
        for table in tables
    }
    for table in tables:
        data_rows = [row for row in rows_by_table[table["table_ref"]] if row["row_role_kwd"] == "data"]
        assert [row["data_row_index_int"] for row in data_rows] == list(
            range(1, table["source_total_count"] + 1)
        )
        assert all(row["source_total_count_int"] == table["source_total_count"] for row in data_rows)


def test_review_b_export_uses_exact_approved_columns_and_filters_private_audit_fields(
    table_parser,
):
    exporter = _load_exporter_module()
    workbook = Workbook()
    workbook.active.append(["Code"])
    workbook.active.append(["DO_NOT_EXPORT"])
    workbook.active.append(["ANOTHER_PRIVATE_VALUE"])
    binary = _save_workbook(workbook)
    generation_ref = exporter.review_generation_ref_from_bytes("anonymous-document", binary)
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "customer-secret.xls",
        binary,
        producer_generation_ref=generation_ref,
        parser=table_parser,
    )

    source_tsv, output_tsv = exporter.render_review_b_tsv(
        audit,
        expected_generation_ref=generation_ref,
    )

    assert source_tsv.splitlines()[0].split("\t") == list(exporter.SOURCE_REGION_FIELDS)
    assert output_tsv.splitlines()[0].split("\t") == list(exporter.OUTPUT_OBJECT_FIELDS)
    assert "emitted_data_row_ordinals" not in output_tsv
    assert "DO_NOT_EXPORT" not in source_tsv + output_tsv
    assert "ANOTHER_PRIVATE_VALUE" not in source_tsv + output_tsv
    assert "customer-secret.xls" not in source_tsv + output_tsv
    assert "\tvalidated\t" in output_tsv


def test_review_b_export_is_deterministic_and_serializes_nested_values_as_compact_json(
    table_parser,
):
    exporter = _load_exporter_module()
    binary = _build_anonymous_workbook(_load_baseline()["samples"][0])
    generation_ref = exporter.review_generation_ref_from_bytes("anonymous-document", binary)
    _projection, audit = tabular_structure._build_tabular_structure_projection_with_audit(
        "anonymous.xlsx",
        binary,
        producer_generation_ref=generation_ref,
        parser=table_parser,
    )

    first = exporter.render_review_b_tsv(
        audit,
        expected_generation_ref=generation_ref,
    )
    second = exporter.render_review_b_tsv(
        json.loads(json.dumps(audit)),
        expected_generation_ref=generation_ref,
    )

    assert first == second
    assert "NULL" in first[1] or "[" in first[0]
    assert all("\r" not in payload for payload in first)


def test_source_bound_review_b_export_stops_before_parsing_on_sha_mismatch(
    tmp_path,
    monkeypatch,
):
    exporter = _load_exporter_module()
    source = tmp_path / "source.bin"
    source.write_bytes(b"not a workbook")
    called = False

    def fail_if_parsed(*_args, **_kwargs):
        nonlocal called
        called = True
        raise AssertionError("Producer must not run before the SHA gate")

    monkeypatch.setattr(exporter, "_build_audit", fail_if_parsed)

    with pytest.raises(ValueError, match="source SHA-256 mismatch"):
        exporter.export_source_bound_review_b(
            source,
            "0" * 64,
            tmp_path / "out",
            "anonymous-document",
        )
    assert called is False


def test_review_b_export_cli_resolves_the_repository_import_root(tmp_path):
    workbook = Workbook()
    workbook.active.append(["Code"])
    workbook.active.append(["A-1"])
    workbook.active.append(["A-2"])
    source = tmp_path / "source.xlsx"
    source.write_bytes(_save_workbook(workbook))
    output = tmp_path / "output"

    result = subprocess.run(
        [
            sys.executable,
            str(EXPORTER_PATH),
            "--source",
            str(source),
            "--expected-source-sha256",
            hashlib.sha256(source.read_bytes()).hexdigest(),
            "--output-directory",
            str(output),
            "--document-id",
            "anonymous-document",
        ],
        cwd=tmp_path,
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0, result.stderr
    assert (output / "source_regions.tsv").is_file()
    assert (output / "output_objects.tsv").is_file()


def test_adr044_rule_requires_a_receipt_bound_to_the_converted_bytes(table_parser):
    sample = next(
        sample
        for sample in _load_baseline()["samples"]
        if sample["sample_code"] == "L1-ADR044-XLS-EQUIVALENT"
    )
    binary = _build_anonymous_workbook(sample)

    with pytest.raises(ValueError, match="converted source SHA-256"):
        build_tabular_structure_projection(
            "anonymous.xlsx",
            binary,
            parser=table_parser,
            adr044_conversion_receipt={
                "original_source_sha256": "1" * 64,
                "converted_source_sha256": "2" * 64,
                "converter_version": "anonymous-converter/v1",
            },
        )


@pytest.mark.parametrize(
    ("fixture_id", "truth_vector", "expected_rule"),
    [
        (fixture_id, truth_vector, expected_rule)
        for fixture_id, (truth_vector, expected_rule) in ORDERED_COLLISION_FIXTURES.items()
    ],
    ids=list(ORDERED_COLLISION_FIXTURES),
)
def test_ordered_collision_truth_vectors_stop_at_the_first_matching_rule(
    fixture_id,
    truth_vector,
    expected_rule,
):
    assert set(truth_vector) == set(NEGATIVE_RULES), fixture_id
    assert tabular_structure._ordered_enumeration_rule(truth_vector, None) == expected_rule
