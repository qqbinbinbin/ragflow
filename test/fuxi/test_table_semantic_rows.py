import importlib.util
import sys
import types
from io import BytesIO
from importlib.machinery import SourceFileLoader
from pathlib import Path

from openpyxl import Workbook


def _load_excel_parser_module(monkeypatch):
    rag_nlp = types.ModuleType("rag.nlp")
    rag_nlp.find_codec = lambda *_args, **_kwargs: "utf-8"
    lazy_image = types.ModuleType("rag.utils.lazy_image")
    lazy_image.LazyImage = object
    monkeypatch.setitem(sys.modules, "rag.nlp", rag_nlp)
    monkeypatch.setitem(sys.modules, "rag.utils.lazy_image", lazy_image)
    path = Path("deepdoc/parser/excel_parser.py")
    spec = importlib.util.spec_from_file_location("semantic_excel_parser", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def _load_table_module(monkeypatch):
    excel_parser = _load_excel_parser_module(monkeypatch)

    class KnowledgebaseService:
        @classmethod
        def delete_field_map(cls, _kb_id):
            return None

        @classmethod
        def update_parser_config(cls, _kb_id, _config):
            return None

    knowledgebase_service = types.ModuleType("api.db.services.knowledgebase_service")
    knowledgebase_service.KnowledgebaseService = KnowledgebaseService
    figure_parser = types.ModuleType("deepdoc.parser.figure_parser")
    figure_parser.vision_figure_parser_figure_xlsx_wrapper = lambda **_kwargs: []
    parser_utils = types.ModuleType("deepdoc.parser.utils")
    parser_utils.get_text = lambda *_args, **_kwargs: ""
    deepdoc_parser = types.ModuleType("deepdoc.parser")
    deepdoc_parser.ExcelParser = excel_parser.RAGFlowExcelParser
    rag_nlp = types.ModuleType("rag.nlp")
    rag_nlp.rag_tokenizer = types.SimpleNamespace(tokenize=lambda text: text.split())
    rag_nlp.tokenize = lambda document, text, _english, **_kwargs: document.update(
        {"content_with_weight": text, "content_ltks": text.split()}
    )
    rag_nlp.tokenize_table = lambda *_args, **_kwargs: []
    common = types.ModuleType("common")
    common.settings = types.SimpleNamespace(
        DOC_ENGINE_INFINITY=False,
        DOC_ENGINE_OCEANBASE=False,
    )
    common_constants = types.ModuleType("common.constants")
    common_constants.MAXIMUM_TASK_PAGE_NUMBER = 100000
    xpinyin = types.ModuleType("xpinyin")
    xpinyin.Pinyin = type(
        "Pinyin",
        (),
        {"get_pinyins": lambda _self, value, _separator: [str(value)]},
    )

    for name, module in {
        "api.db.services.knowledgebase_service": knowledgebase_service,
        "deepdoc.parser.figure_parser": figure_parser,
        "deepdoc.parser.utils": parser_utils,
        "deepdoc.parser": deepdoc_parser,
        "rag.nlp": rag_nlp,
        "common": common,
        "common.constants": common_constants,
        "xpinyin": xpinyin,
    }.items():
        monkeypatch.setitem(sys.modules, name, module)

    loader = SourceFileLoader("semantic_table_parser", "rag/app/table.py")
    spec = importlib.util.spec_from_loader("semantic_table_parser", loader)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def _semantic_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Supplier matrix"
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "Approved material sources"
    sheet["A3"] = "Program"
    sheet["B3"] = "Platform-X"
    sheet["C3"] = "Assembly"
    sheet["D3"] = "Wheel"
    sheet["E3"] = "Primary supplier"
    sheet["F3"] = "Acme"
    sheet.merge_cells("A4:A5")
    sheet["A4"] = "Part"
    sheet.merge_cells("B4:B5")
    sheet["B4"] = "Material"
    sheet.merge_cells("C4:C5")
    sheet["C4"] = "Specification"
    sheet.merge_cells("D4:F4")
    sheet["D4"] = "Supplier information"
    sheet["D5"] = "Name"
    sheet["E5"] = "Location"
    sheet["F5"] = 2026
    sheet.append(["P-1", "Cord", "1440", "North", "Chengdu", "IATF"])
    sheet.append(["P-2", "Wire", "1.3HT", "South", "Wuhan", "IATF"])

    second = workbook.create_sheet("Simple inventory")
    second.append(["Code", "Description"])
    second.append(["I-1", "Bearing"])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_excel_parser_detects_delayed_multilevel_header_and_preserves_context(monkeypatch):
    table = _load_table_module(monkeypatch)

    frames, _tables = table.Excel()(
        "supplier.xlsx",
        binary=_semantic_workbook_bytes(),
        callback=lambda *_args: None,
    )

    assert len(frames) == 2
    matrix = frames[0]
    assert list(matrix.columns) == [
        "Part",
        "Material",
        "Specification",
        "Supplier information-Name",
        "Supplier information-Location",
        "Supplier information-2026",
    ]
    assert matrix.iloc[0].to_dict() == {
        "Part": "P-1",
        "Material": "Cord",
        "Specification": "1440",
        "Supplier information-Name": "North",
        "Supplier information-Location": "Chengdu",
        "Supplier information-2026": "IATF",
    }
    assert matrix.attrs["sheet_name"] == "Supplier matrix"
    assert "Program: Platform-X" in matrix.attrs["sheet_context"]
    assert "Primary supplier: Acme" in matrix.attrs["sheet_context"]
    assert "Cord" not in " ".join(matrix.columns)

    inventory = frames[1]
    assert list(inventory.columns) == ["Code", "Description"]
    assert inventory.iloc[0].to_dict() == {"Code": "I-1", "Description": "Bearing"}
    assert inventory.attrs["sheet_name"] == "Simple inventory"


def test_table_chunks_include_sheet_context_headers_and_complete_row(monkeypatch):
    table = _load_table_module(monkeypatch)

    chunks = table.chunk(
        "supplier.xlsx",
        binary=_semantic_workbook_bytes(),
        callback=lambda *_args: None,
        kb_id="kb-es",
    )

    target = next(chunk for chunk in chunks if "Cord" in chunk["content_with_weight"])
    content = target["content_with_weight"]
    assert "- Sheet: Supplier matrix" in content
    assert "- Sheet context:" in content
    assert "Program: Platform-X" in content
    assert "- Material: Cord" in content
    assert "- Specification: 1440" in content
    assert "- Supplier information-Name: North" in content
    assert "- Supplier information-2026: IATF" in content
    assert "Cord-Cord" not in content
    assert "nan" not in content.lower()


def test_simple_table_with_multiple_rows_keeps_first_row_as_header(monkeypatch):
    table = _load_table_module(monkeypatch)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventory"
    sheet.append(["Code", "Description", "Quantity"])
    sheet.append(["I-1", "Bearing", 2])
    sheet.append(["I-2", "Bolt", 40])
    sheet.append(["I-3", "Seal", 6])
    output = BytesIO()
    workbook.save(output)

    frames, _tables = table.Excel()(
        "inventory.xlsx",
        binary=output.getvalue(),
        callback=lambda *_args: None,
    )

    assert list(frames[0].columns) == ["Code", "Description", "Quantity"]
    assert frames[0].to_dict(orient="records") == [
        {"Code": "I-1", "Description": "Bearing", "Quantity": 2},
        {"Code": "I-2", "Description": "Bolt", "Quantity": 40},
        {"Code": "I-3", "Description": "Seal", "Quantity": 6},
    ]


def test_delayed_header_without_merges_prefers_earliest_consistent_row(monkeypatch):
    table = _load_table_module(monkeypatch)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Supplier register"
    sheet.append(["Supplier register", None, None])
    sheet.append([None, None, None])
    sheet.append(["Part", "Supplier", "Status"])
    sheet.append(["P-1", "North", "Approved"])
    sheet.append(["P-2", "South", "Approved"])
    output = BytesIO()
    workbook.save(output)

    frames, _tables = table.Excel()(
        "register.xlsx",
        binary=output.getvalue(),
        callback=lambda *_args: None,
    )

    assert list(frames[0].columns) == ["Part", "Supplier", "Status"]
    assert frames[0].to_dict(orient="records") == [
        {"Part": "P-1", "Supplier": "North", "Status": "Approved"},
        {"Part": "P-2", "Supplier": "South", "Status": "Approved"},
    ]
    assert frames[0].attrs["sheet_context"] == "Supplier register"
