#
#  Copyright 2025 The InfiniFlow Authors. All Rights Reserved.
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

"""Immutable, no-vector row projections for tabular completeness checks."""

from __future__ import annotations

import hashlib
import inspect
import json
import re
import struct
import uuid
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any


TABULAR_STRUCTURE_VERSION = "tabular-row/v2"
PRODUCER_SCHEMA_VERSION = "table-producer/v6"
PROJECTION_VERSION = "tabular-structure-projection/v6"
PROJECTION_PART_VERSION = "tabular-structure-part/v3"
STRUCTURE_PRODUCER_ALGORITHM_VERSION = "region-producer/v21"
ENUMERATION_RULE_VERSION = "enumeration-rules/v9"
ROW_PAGE_TRANSPORT_VERSION = "tabular-row-page-compact/v1"
_CURRENT_PROJECTION_CONTRACT = (
    PRODUCER_SCHEMA_VERSION,
    PROJECTION_VERSION,
    STRUCTURE_PRODUCER_ALGORITHM_VERSION,
    ENUMERATION_RULE_VERSION,
)
_KNOWN_BACKFILL_PROJECTION_CONTRACTS = frozenset(
    {
        ("table-producer/v6", "tabular-structure-projection/v6", "region-producer/v10", "enumeration-rules/v3"),
        ("table-producer/v6", "tabular-structure-projection/v6", "region-producer/v11", "enumeration-rules/v3"),
        ("table-producer/v6", "tabular-structure-projection/v6", "region-producer/v12", "enumeration-rules/v4"),
        ("table-producer/v6", "tabular-structure-projection/v6", "region-producer/v13", "enumeration-rules/v5"),
        ("table-producer/v6", "tabular-structure-projection/v6", "region-producer/v14", "enumeration-rules/v6"),
        ("table-producer/v6", "tabular-structure-projection/v6", "region-producer/v15", "enumeration-rules/v7"),
        ("table-producer/v6", "tabular-structure-projection/v6", "region-producer/v16", "enumeration-rules/v8"),
        ("table-producer/v6", "tabular-structure-projection/v6", "region-producer/v17", "enumeration-rules/v9"),
        ("table-producer/v6", "tabular-structure-projection/v6", "region-producer/v18", "enumeration-rules/v9"),
        ("table-producer/v6", "tabular-structure-projection/v6", "region-producer/v19", "enumeration-rules/v9"),
        ("table-producer/v6", "tabular-structure-projection/v6", "region-producer/v20", "enumeration-rules/v9"),
        _CURRENT_PROJECTION_CONTRACT,
    }
)
PROJECTION_FIELDS = frozenset(
    {
        "version",
        "producer_schema_version",
        "producer_generation_ref",
        "structure_algorithm_version",
        "enumeration_rule_version",
        "source_sha256",
        "tables",
        "rows",
    }
)

DEFAULT_CONTEXT_ENTRY_LIMIT = 8
DEFAULT_CONTEXT_VALUE_BYTES = 128
DEFAULT_TABLE_LABEL_BYTES = 128
DEFAULT_ROWS_PER_PART = 3000

ENUMERATION_DECISIONS = {
    "L1-01": ("supported_complete", "record_axis_proven"),
    "L1-02": ("supported_complete", "record_axis_proven"),
    "L1-03": ("supported_complete", "record_axis_proven"),
    "L1-04": ("supported_complete", "record_axis_proven"),
    "L1-05": ("supported_complete", "record_axis_proven"),
    "L1-06": ("supported_complete", "record_axis_proven"),
    "L1-07": ("supported_complete", "record_axis_proven"),
    "L1-08": ("supported_complete", "empty_record_axis_proven"),
    "R1": ("not_guaranteed_explained", "not_a_list"),
    "R2": ("not_guaranteed_explained", "total_unstable"),
    "R3": ("not_guaranteed_explained", "matrix_layout"),
    "R4": ("not_guaranteed_explained", "subtotal_rows_mixed"),
    "R5": ("not_guaranteed_explained", "multi_block_unseparated"),
    "R6": ("not_guaranteed_explained", "partial_overlap_continuation"),
    "R7": ("not_guaranteed_explained", "visual_only_boundary"),
    "R8": ("not_guaranteed_explained", "record_axis_not_proven"),
    "D1": ("defect", "unnamed_required_field"),
    "D2": ("defect", "membership_not_closed"),
    "D3": ("defect", "record_count_mismatch"),
    "D4": ("defect", "missing_projection"),
}
NEGATIVE_ENUMERATION_RULES = ("R1", "R2", "R3", "R4", "R5", "R6", "R7")

PROJECTION_ROW_FIELDS = frozenset(
    {
        "id",
        "tabular_structure_version_kwd",
        "structure_kind_kwd",
        "producer_schema_version_kwd",
        "producer_generation_ref_kwd",
        "table_ref_kwd",
        "table_label_kwd",
        "table_context_list",
        "row_ref_kwd",
        "row_ordinal_int",
        "data_row_index_int",
        "row_role_kwd",
        "source_total_count_int",
        "ordered_fields_list",
    }
)

_ULID_RE = re.compile(r"^[0-7][0-9A-HJKMNP-TV-Z]{25}$")
_UNTRUSTED_CONTROL_RE = re.compile("[\x00-\x1f\x7f-\x9f\u202a-\u202e\u2066-\u2069]")
_ASCII_DECIMAL_SCALAR_RE = re.compile(
    r"^[+-]?(?:[0-9]+(?:\.[0-9]*)?|\.[0-9]+)(?:[eE][+-]?[0-9]+)?$"
)


class StructureGenerationConflict(RuntimeError):
    """The persistent generation state violates its single-active invariant."""


class StructureSnapshotMissing(LookupError):
    """The requested immutable structure generation is not readable."""


class StructureSnapshotChanged(RuntimeError):
    """The requested immutable structure generation no longer matches its digest."""

    def __init__(self, message: str, active_generation_ref: str | None = None):
        super().__init__(message)
        self.active_generation_ref = active_generation_ref


def _canonical_json(value: Any) -> bytes:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")


def _versioned_digest(kind: str, *parts: object) -> str:
    payload = "\x00".join([kind, *(str(part) for part in parts)]).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def tabular_structure_projection_prefix(
    document_id: str,
    producer_generation_ref: str,
) -> str:
    if not isinstance(document_id, str) or not document_id:
        raise ValueError("document_id is required")
    _validate_generation_ref(producer_generation_ref)
    document_ref = _versioned_digest("tabular-structure-document/v1", document_id)
    return (
        f"_fuxi/tabular-structure/v1/{document_ref}/"
        f"{producer_generation_ref}/"
    )


def _table_ref(
    source_sha256: str,
    sheet_ordinal: int,
    table_ordinal: int,
    membership_sha256: str,
) -> str:
    return _table_ref_for_contract(
        source_sha256,
        sheet_ordinal,
        table_ordinal,
        membership_sha256,
        producer_schema_version=PRODUCER_SCHEMA_VERSION,
        projection_version=PROJECTION_VERSION,
        structure_algorithm_version=STRUCTURE_PRODUCER_ALGORITHM_VERSION,
        enumeration_rule_version=ENUMERATION_RULE_VERSION,
    )


def _table_ref_for_contract(
    source_sha256: str,
    sheet_ordinal: int,
    table_ordinal: int,
    membership_sha256: str,
    *,
    producer_schema_version: str,
    projection_version: str,
    structure_algorithm_version: str,
    enumeration_rule_version: str,
) -> str:
    if not re.fullmatch(r"[0-9a-f]{64}", membership_sha256):
        raise ValueError("table membership SHA-256 is invalid")
    identity = _versioned_digest(
        "tabular-table/v2",
        producer_schema_version,
        projection_version,
        structure_algorithm_version,
        enumeration_rule_version,
        source_sha256,
        sheet_ordinal,
        table_ordinal,
        membership_sha256,
    )
    return f"tbl_v2_{membership_sha256}_{identity}"


def _validate_generation_ref(producer_generation_ref: str) -> None:
    if not isinstance(producer_generation_ref, str):
        raise ValueError("producer_generation_ref must be an opaque UUID/ULID-compatible value")
    try:
        parsed_uuid = uuid.UUID(producer_generation_ref)
        is_canonical_uuid = str(parsed_uuid) == producer_generation_ref.lower()
    except ValueError:
        is_canonical_uuid = False
    if not is_canonical_uuid and not _ULID_RE.fullmatch(producer_generation_ref):
        raise ValueError("producer_generation_ref must be an opaque UUID/ULID-compatible value")


def _apply_enumeration_decision(table: dict[str, Any], matched_rule: str) -> None:
    status, reason = ENUMERATION_DECISIONS[matched_rule]
    table.update(
        {
            "enumeration_status": status,
            "enumeration_reason": reason,
            "matched_rule": matched_rule,
        }
    )


def _ordered_enumeration_rule(
    negative_predicates: dict[str, bool],
    l1_rule: str | None,
) -> str:
    if (
        not isinstance(negative_predicates, dict)
        or set(negative_predicates) != set(NEGATIVE_ENUMERATION_RULES)
        or any(not isinstance(value, bool) for value in negative_predicates.values())
    ):
        raise ValueError("enumeration predicate vector must contain seven booleans")
    if l1_rule is not None and l1_rule not in ENUMERATION_DECISIONS:
        raise ValueError("enumeration L1 rule is invalid")
    for rule in NEGATIVE_ENUMERATION_RULES:
        if negative_predicates[rule]:
            return rule
    return l1_rule or "R8"


def _clear_complete_decision(
    table: dict[str, Any],
    rows: list[dict[str, Any]],
    matched_rule: str = "R8",
) -> None:
    table["source_total_count"] = None
    _apply_enumeration_decision(table, matched_rule)
    for row in rows:
        row["source_total_count_int"] = None


def _validate_adr044_conversion_receipt(
    receipt: dict[str, str] | None,
    converted_source_sha256: str,
) -> bool:
    if receipt is None:
        return False
    required = {"original_source_sha256", "converted_source_sha256", "converter_version"}
    if not isinstance(receipt, dict) or set(receipt) != required:
        raise ValueError("ADR-044 conversion receipt does not match the fixed schema")
    for field in ("original_source_sha256", "converted_source_sha256"):
        if not isinstance(receipt[field], str) or not re.fullmatch(r"[0-9a-f]{64}", receipt[field]):
            raise ValueError(f"ADR-044 {field.replace('_', ' ')} is invalid")
    if receipt["converted_source_sha256"] != converted_source_sha256:
        raise ValueError("ADR-044 converted source SHA-256 does not match the workbook bytes")
    converter_version = receipt["converter_version"]
    if (
        not isinstance(converter_version, str)
        or not converter_version.strip()
        or _sanitize_untrusted_text(converter_version) != converter_version
    ):
        raise ValueError("ADR-044 converter version is invalid")
    return True


def _sanitize_untrusted_text(value: object) -> str:
    return _UNTRUSTED_CONTROL_RE.sub("", "" if value is None else str(value)).strip()


def _truncate_utf8(value: str, byte_limit: int) -> str:
    if byte_limit < 1:
        raise ValueError("UTF-8 byte limit must be positive")
    encoded = value.encode("utf-8")
    bounded = (
        value
        if len(encoded) <= byte_limit
        else encoded[:byte_limit].decode("utf-8", errors="ignore")
    )
    return bounded.strip()


def _cell_value(parser, worksheet, row_ordinal: int, column_ordinal: int, merged_ranges):
    value = worksheet.cell(row=row_ordinal, column=column_ordinal).value
    if value is not None:
        return value
    return parser._get_merged_cell_value(worksheet, row_ordinal, column_ordinal, merged_ranges)


def _source_cell_anchor(row_ordinal: int, column_ordinal: int, merged_ranges):
    for merged in merged_ranges:
        if (
            merged.min_row <= row_ordinal <= merged.max_row
            and merged.min_col <= column_ordinal <= merged.max_col
        ):
            return (
                "merge",
                merged.min_row,
                merged.min_col,
                merged.max_row,
                merged.max_col,
            )
    return ("cell", row_ordinal, column_ordinal)


def _complete_worksheet_rows(worksheet):
    """Return all physical worksheet rows and their source row ordinals."""

    if not worksheet.max_row:
        return [], [], []
    cells = getattr(worksheet, "_cells", None)
    if not isinstance(cells, dict):
        raise ValueError("worksheet backend cannot prove complete sparse row coverage")
    # Snapshot sparse cells before iter_rows materializes empty cells. The
    # worksheet dimension can include formatting-only rows, so derive the
    # physical bound from source cells, merges, and unresolved coordinates.
    physical_cells = list(cells.values())
    populated_rows = sorted(
        {
            cell.row
            for cell in physical_cells
            if cell.value is not None and str(cell.value).strip() != ""
        }
    )
    unresolved_rows = sorted(
        {
            row_ordinal
            for row_ordinal, _column_ordinal in getattr(
                worksheet,
                "_fuxi_unresolved_coordinates",
                set(),
            )
        }
    )
    physical_row_bound = max(
        (
            [cell.row for cell in physical_cells]
            + [merged.max_row for merged in worksheet.merged_cells.ranges]
            + unresolved_rows
        ),
        default=0,
    )
    header_rows = list(
        worksheet.iter_rows(min_row=1, max_row=physical_row_bound)
    )
    return header_rows, populated_rows, unresolved_rows


def _ordered_fields(
    headers: list[str],
    values: list[object],
    *,
    note: bool,
    sheet_ordinal: int | None = None,
    header_paths: list[list[str]] | None = None,
    column_ordinals: list[int] | None = None,
    absolute_column_ordinals: list[int] | None = None,
    source_anchors: list[tuple] | None = None,
) -> list[dict[str, Any]]:
    fields = []
    emitted_anchors = set()
    for index, (name, value) in enumerate(zip(headers, values)):
        if value is None or (isinstance(value, str) and not value.strip()):
            continue
        rendered = (
            str(int(value))
            if isinstance(value, float) and value.is_integer()
            else str(value).strip()
        )
        if not rendered:
            continue
        source_anchor = source_anchors[index] if source_anchors is not None else None
        if source_anchor is not None:
            if source_anchor in emitted_anchors:
                continue
            emitted_anchors.add(source_anchor)
        field = {"name": str(name), "value": rendered}
        if sheet_ordinal is not None:
            column_ordinal = (
                column_ordinals[index]
                if column_ordinals is not None
                else index + 1
            )
            absolute_column_ordinal = (
                absolute_column_ordinals[index]
                if absolute_column_ordinals is not None
                else column_ordinal
            )
            field = {
                "column_id": f"col_v1:{sheet_ordinal}:{absolute_column_ordinal}",
                "column_ordinal": column_ordinal,
                "header_path": list(header_paths[index]) if header_paths is not None else [],
                **field,
            }
        fields.append(field)
        if note:
            break
    return fields


def _is_full_width_merge(row_ordinal: int, width: int, merged_ranges) -> bool:
    return any(
        merged.min_row <= row_ordinal <= merged.max_row
        and merged.min_col == 1
        and merged.max_col >= width
        for merged in merged_ranges
    )


def _has_partial_row_merge(row_ordinal: int, width: int, merged_ranges) -> bool:
    return any(
        merged.min_row <= row_ordinal <= merged.max_row
        and (merged.min_col != 1 or merged.max_col < width)
        for merged in merged_ranges
    )


def _row_merge_signature(row_ordinal: int, merged_ranges) -> tuple[tuple[int, int], ...]:
    return tuple(
        sorted(
            (merged.min_col, merged.max_col)
            for merged in merged_ranges
            if merged.min_row <= row_ordinal <= merged.max_row
        )
    )


def _classify_body_row(
    row_ordinal: int,
    values: list[object],
    merged_ranges,
    *,
    record_axis_evidence: dict[str, Any] | None = None,
) -> str:
    width = len(values)
    if record_axis_evidence is not None:
        if row_ordinal in record_axis_evidence.get("note_row_ordinals", ()):
            return "note"
        if row_ordinal in record_axis_evidence.get("unknown_row_ordinals", ()):
            return "unknown"
        if row_ordinal not in record_axis_evidence.get("record_row_ordinals", ()):
            return "unknown"
    if _is_full_width_merge(row_ordinal, width, merged_ranges):
        return "unknown"
    if record_axis_evidence is not None:
        if _has_partial_row_merge(row_ordinal, width, merged_ranges) and not _row_merge_signature(
            row_ordinal, merged_ranges
        ):
            return "unknown"
        row_offsets = set(
            _record_field_offsets(
                values,
                row_ordinal=row_ordinal,
                merged_ranges=merged_ranges,
            )
        )
        required_offsets = set(record_axis_evidence["required_offsets"])
        if required_offsets and required_offsets.issubset(row_offsets) and (
            len(values) == 1
            or len(row_offsets) >= 2
            or record_axis_evidence.get("record_key_axis_proven") is True
            or record_axis_evidence.get("single_record_axis_proven") is True
        ):
            return "data"
        return "unknown"
    populated = sum(value is not None and str(value).strip() != "" for value in values)
    if populated >= min(2, width):
        return "data"
    return "unknown"


def _row_shape(values: list[object], *, distinguish_text_digits: bool = False) -> tuple[str, ...]:
    shape = []
    for value in values:
        if value is None or (isinstance(value, str) and not value.strip()):
            shape.append("empty")
        elif isinstance(value, bool):
            shape.append("boolean")
        elif isinstance(value, (int, float)):
            shape.append("number")
        else:
            rendered = str(value).strip()
            if distinguish_text_digits and any(char.isdigit() for char in rendered):
                shape.append("text_with_digit")
            else:
                shape.append("text")
    return tuple(shape)


def _record_axis_value_shape(value: object) -> tuple[str, ...]:
    if value is None or (isinstance(value, str) and not value.strip()):
        return ("empty",)
    if isinstance(value, bool):
        return ("boolean",)
    if isinstance(value, (int, float)):
        return ("number",)
    return ("text", *_text_structure(value))


def _record_key_numeric_value(value: object) -> Decimal | None:
    """Normalize finite numeric cell scalars without changing source values."""

    if isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        normalized = value
    elif isinstance(value, (int, float)):
        try:
            normalized = Decimal(str(value))
        except InvalidOperation:
            return None
    elif isinstance(value, str):
        rendered = value.strip()
        if not _ASCII_DECIMAL_SCALAR_RE.fullmatch(rendered):
            return None
        try:
            normalized = Decimal(rendered)
        except InvalidOperation:
            return None
    else:
        return None
    return normalized if normalized.is_finite() else None


def _record_key_axis_proven(
    rows: list[tuple[int, list[object], bool]],
    required_offsets: set[int],
) -> bool:
    """Prove a source-backed numeric record key on the leftmost required field."""

    if len(rows) < 2 or not required_offsets:
        return False
    key_offset = min(required_offsets)
    source_values = [row[1][key_offset] for row in rows]
    native_numeric = [
        value
        for value in source_values
        if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)
    ]
    text_numeric = [value for value in source_values if isinstance(value, str)]
    if text_numeric and not native_numeric:
        return False
    values = [_record_key_numeric_value(value) for value in source_values]
    if any(value is None for value in values):
        return False
    if len(set(values)) != len(values):
        return False
    deltas = [right - left for left, right in zip(values, values[1:])]
    return bool(deltas) and (all(delta > 0 for delta in deltas) or all(delta < 0 for delta in deltas))


def _record_key_only_slots(
    rows: list[tuple[int, list[object], bool]],
    row_offsets: list[tuple[int, ...]],
    required_offsets: set[int],
) -> tuple[int, ...]:
    """Identify multi-field template slots that contain only a numeric key."""

    if len(required_offsets) != 1 or not _record_key_axis_proven(rows, required_offsets):
        return ()
    key_offset = next(iter(required_offsets))
    if not any(any(offset != key_offset for offset in offsets) for offsets in row_offsets):
        return ()
    return tuple(
        row[0]
        for row, offsets in zip(rows, row_offsets)
        if set(offsets) == {key_offset}
    )


def _is_repeated_header_row(headers: list[str], values: list[object]) -> bool:
    """Treat an exact repeated header as a structural boundary."""

    if len(headers) != len(values):
        return False
    normalized_values = ["" if value is None else str(value).strip() for value in values]
    if not all(normalized_values):
        return False
    return normalized_values == [str(header).strip() for header in headers]


def _source_anchors_by_row(
    worksheet,
    rows,
    *,
    row_limit: int,
    width: int,
    merged_ranges,
) -> dict[int, list[dict[str, Any]]]:
    anchors_by_row: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row_ordinal, row in enumerate(rows[:row_limit], start=1):
        for column_ordinal in range(1, min(width, len(row)) + 1):
            source = _source_cell_anchor(
                row_ordinal,
                column_ordinal,
                merged_ranges,
            )
            if source[0] == "merge":
                _kind, min_row, min_column, max_row, max_column = source
                if (row_ordinal, column_ordinal) != (min_row, min_column):
                    continue
            else:
                _kind, min_row, min_column = source
                max_row, max_column = min_row, min_column
            value = _sanitize_untrusted_text(
                worksheet.cell(min_row, min_column).value
            )
            if not value:
                continue
            anchors_by_row[row_ordinal].append(
                {
                    "row": min_row,
                    "min_column": min_column,
                    "max_row": max_row,
                    "max_column": min(max_column, width),
                    "value": value,
                }
            )
    return anchors_by_row


def _table_context(
    parser,
    worksheet,
    rows,
    header_start: int,
    *,
    width: int,
    source_geometry: bool = False,
    entry_limit: int,
    value_bytes: int,
) -> list[dict[str, str]]:
    if entry_limit < 1:
        raise ValueError("table context entry limit must be positive")
    if width < 1:
        raise ValueError("table context width must be positive")
    merged_ranges = list(worksheet.merged_cells.ranges)
    if not source_geometry:
        entries = []
        for row_ordinal, row in enumerate(rows[:header_start], start=1):
            values = []
            for column_ordinal in range(1, len(row) + 1):
                value = _cell_value(
                    parser,
                    worksheet,
                    row_ordinal,
                    column_ordinal,
                    merged_ranges,
                )
                value = _sanitize_untrusted_text(value)
                if value and (not values or values[-1] != value):
                    values.append(value)
            if not values:
                continue
            pairs = []
            if len(values) == 1:
                pairs.append(("context", values[0]))
            else:
                pairs.extend(zip(values[::2], values[1::2]))
                if len(values) % 2:
                    pairs.append(("context", values[-1]))
            for name, value in pairs:
                entries.append(
                    {
                        "name": _truncate_utf8(
                            _sanitize_untrusted_text(name),
                            value_bytes,
                        ),
                        "value": _truncate_utf8(
                            _sanitize_untrusted_text(value),
                            value_bytes,
                        ),
                    }
                )
                if len(entries) == entry_limit:
                    return entries
        return entries

    anchors_by_row = _source_anchors_by_row(
        worksheet,
        rows,
        row_limit=header_start,
        width=width,
        merged_ranges=merged_ranges,
    )

    entries = []
    consumed: set[tuple[int, int]] = set()

    def append_entry(name: str, value: str) -> bool:
        entries.append(
            {
                "name": _truncate_utf8(_sanitize_untrusted_text(name), value_bytes),
                "value": _truncate_utf8(_sanitize_untrusted_text(value), value_bytes),
            }
        )
        return len(entries) == entry_limit

    for row_ordinal in range(1, header_start + 1):
        anchors = [
            anchor
            for anchor in anchors_by_row.get(row_ordinal, [])
            if (anchor["row"], anchor["min_column"]) not in consumed
        ]
        if not anchors:
            continue
        next_row_anchors = [
            anchor
            for anchor in anchors_by_row.get(row_ordinal + 1, [])
            if (anchor["row"], anchor["min_column"]) not in consumed
        ]
        next_anchors = {
            (anchor["min_column"], anchor["max_column"]): anchor
            for anchor in next_row_anchors
        }
        next_anchors_by_start = {
            anchor["min_column"]: anchor
            for anchor in next_row_anchors
            if sum(
                candidate["min_column"] == anchor["min_column"]
                for candidate in next_row_anchors
            )
            == 1
        }
        inline_segments = _inline_context_segments(
            worksheet,
            row_ordinal=row_ordinal,
            width=width,
            merged_ranges=merged_ranges,
        )
        if inline_segments is not None:
            for name_segment, value_segment in zip(
                inline_segments[::2],
                inline_segments[1::2],
            ):
                name = name_segment["value"]
                value = value_segment["value"]
                if append_entry(name if value else "context", value or name):
                    return entries
            continue
        vertical_pairs = []
        if len(anchors) >= 2:
            for anchor_index, anchor in enumerate(anchors):
                vertical_value = next_anchors.get(
                    (anchor["min_column"], anchor["max_column"])
                )
                if vertical_value is None:
                    aligned_value = next_anchors_by_start.get(anchor["min_column"])
                    next_label_start = (
                        anchors[anchor_index + 1]["min_column"]
                        if anchor_index + 1 < len(anchors)
                        else width + 1
                    )
                    if (
                        aligned_value is not None
                        and aligned_value["max_column"] >= anchor["max_column"]
                        and aligned_value["max_column"] < next_label_start
                    ):
                        vertical_value = aligned_value
                if anchor["max_row"] == row_ordinal and vertical_value is not None:
                    vertical_pairs.append((anchor, vertical_value))
        if len(vertical_pairs) >= 2:
            pair_by_anchor = {
                (anchor["row"], anchor["min_column"]): vertical_value
                for anchor, vertical_value in vertical_pairs
            }
            for anchor in anchors:
                vertical_value = pair_by_anchor.get(
                    (anchor["row"], anchor["min_column"])
                )
                if vertical_value is None:
                    if append_entry("context", anchor["value"]):
                        return entries
                    continue
                if append_entry(anchor["value"], vertical_value["value"]):
                    return entries
                consumed.add(
                    (vertical_value["row"], vertical_value["min_column"])
                )
            continue
        horizontal_run = []

        def flush_horizontal_run() -> bool:
            if len(horizontal_run) == 2:
                if append_entry(
                    horizontal_run[0]["value"],
                    horizontal_run[1]["value"],
                ):
                    return True
            else:
                for anchor in horizontal_run:
                    if append_entry("context", anchor["value"]):
                        return True
            horizontal_run.clear()
            return False

        for anchor in anchors:
            horizontal_run.append(anchor)
        if flush_horizontal_run():
            return entries
    return entries


def _region_membership_sha256(sheet_ordinal: int, members: set[tuple[int, int]]) -> str:
    payload = "\n".join(
        f"{sheet_ordinal}:{row_ordinal}:{column_ordinal}"
        for row_ordinal, column_ordinal in sorted(members)
    ).encode("ascii")
    return hashlib.sha256(payload).hexdigest()


def _logical_occupied_cells(
    parser,
    worksheet,
    unresolved_formula_coordinates: set[tuple[int, int]] | None = None,
    formula_coordinates: set[tuple[int, int]] | None = None,
) -> set[tuple[int, int]]:
    cells = getattr(worksheet, "_cells", None)
    if not isinstance(cells, dict):
        raise ValueError("worksheet backend cannot prove exact cell membership")
    occupied = {
        (cell.row, cell.column)
        for cell in cells.values()
        if cell.value is not None and str(cell.value).strip() != ""
    }
    for merged in worksheet.merged_cells.ranges:
        anchor = worksheet.cell(merged.min_row, merged.min_col).value
        if anchor is None or str(anchor).strip() == "":
            continue
        occupied.update(
            (row_ordinal, column_ordinal)
            for row_ordinal in range(merged.min_row, merged.max_row + 1)
            for column_ordinal in range(merged.min_col, merged.max_col + 1)
        )
    occupied.update(unresolved_formula_coordinates or set())
    # A cached-empty BIFF formula is still a source cell.  Keep its coordinate
    # in the immutable region membership while keeping calculation uncertainty
    # in the separate unresolved set above.
    occupied.update(formula_coordinates or set())
    return occupied


def _connected_cell_regions(
    occupied: set[tuple[int, int]],
    *,
    tolerance: int = 2,
) -> list[set[tuple[int, int]]]:
    remaining = set(occupied)
    regions = []
    while remaining:
        start = min(remaining)
        remaining.remove(start)
        members = {start}
        pending = [start]
        while pending:
            row_ordinal, column_ordinal = pending.pop()
            for row_delta in range(-tolerance, tolerance + 1):
                for column_delta in range(-tolerance, tolerance + 1):
                    if row_delta == 0 and column_delta == 0:
                        continue
                    neighbor = (row_ordinal + row_delta, column_ordinal + column_delta)
                    if neighbor not in remaining:
                        continue
                    remaining.remove(neighbor)
                    members.add(neighbor)
                    pending.append(neighbor)
        regions.append(members)
    return regions


def _region_bbox(members: set[tuple[int, int]]) -> tuple[int, int, int, int]:
    rows = [row for row, _column in members]
    columns = [column for _row, column in members]
    return min(rows), min(columns), max(rows), max(columns)


def _cell_distance_to_members(
    coordinate: tuple[int, int],
    members: set[tuple[int, int]],
) -> int:
    row_ordinal, column_ordinal = coordinate
    return min(
        max(abs(row_ordinal - member_row), abs(column_ordinal - member_column))
        for member_row, member_column in members
    )


def _split_closed_empty_axis_region(
    parser,
    worksheet,
    members: set[tuple[int, int]],
) -> list[set[tuple[int, int]]] | None:
    candidates = []
    merged_ranges = list(worksheet.merged_cells.ranges)
    for title_merge in merged_ranges:
        if title_merge.max_col <= title_merge.min_col:
            continue
        title = worksheet.cell(title_merge.min_row, title_merge.min_col).value
        if title is None or not str(title).strip():
            continue
        header_row = title_merge.max_row + 1
        table_columns = range(title_merge.min_col, title_merge.max_col + 1)
        if any(
            (header_row, column) not in members
            or worksheet.cell(header_row, column).value is None
            or not str(worksheet.cell(header_row, column).value).strip()
            or _source_cell_anchor(header_row, column, merged_ranges)[0] != "cell"
            for column in table_columns
        ):
            continue
        headers = [
            _sanitize_untrusted_text(worksheet.cell(header_row, column).value).strip()
            for column in table_columns
        ]
        if (
            any(not header or header.startswith("Column_") for header in headers)
            or len(set(headers)) != len(headers)
        ):
            continue
        if any(
            row > header_row
            and title_merge.min_col <= column <= title_merge.max_col
            for row, column in members
        ):
            continue
        table_members = {
            coordinate
            for coordinate in members
            if title_merge.min_row <= coordinate[0] <= header_row
            and title_merge.min_col <= coordinate[1] <= title_merge.max_col
        }
        outside_members = members - table_members
        if not outside_members or any(
            title_merge.min_col <= column <= title_merge.max_col
            for _row, column in outside_members
        ):
            continue
        candidates.append((table_members, outside_members))
    if len(candidates) != 1:
        return None
    table_members, outside_members = candidates[0]
    return [
        table_members,
        *_connected_cell_regions(outside_members, tolerance=2),
    ]


def _worksheet_structure_regions(
    parser,
    worksheet,
    sheet_ordinal: int,
    unresolved_formula_coordinates: set[tuple[int, int]] | None = None,
    formula_coordinates: set[tuple[int, int]] | None = None,
) -> list[dict[str, Any]]:
    occupied = _logical_occupied_cells(
        parser,
        worksheet,
        unresolved_formula_coordinates,
        formula_coordinates,
    )
    if not occupied:
        return []
    g1_regions = _connected_cell_regions(occupied, tolerance=1)
    g2_regions = _connected_cell_regions(occupied, tolerance=2)
    split_regions = []
    for members in g2_regions:
        split = _split_closed_empty_axis_region(parser, worksheet, members)
        if split is None:
            split_regions.append(members)
            continue
        for component in split:
            split_regions.append(component)
    g2_regions = split_regions
    bboxes = [_region_bbox(members) for members in g2_regions]
    unresolved = set(unresolved_formula_coordinates or set())
    unresolved_by_region = [set() for _region in g2_regions]
    has_unbound_unresolved = False
    for coordinate in unresolved:
        owners = [
            index
            for index, members in enumerate(g2_regions)
            if _cell_distance_to_members(coordinate, members) <= 2
        ]
        if len(owners) == 1:
            unresolved_by_region[owners[0]].add(coordinate)
        else:
            has_unbound_unresolved = True

    result = []
    for members, unresolved_members, bbox in zip(g2_regions, unresolved_by_region, bboxes):
        g1_children = [child for child in g1_regions if child.issubset(members)]
        result.append(
            {
                "members": members,
                "unresolved_members": unresolved_members,
                "bbox": bbox,
                "membership_sha256": _region_membership_sha256(sheet_ordinal, members),
                "has_unbound_unresolved": has_unbound_unresolved,
                "g1_children": sorted(
                    g1_children,
                    key=lambda child: (*_region_bbox(child), _region_membership_sha256(sheet_ordinal, child)),
                ),
            }
        )
    result.sort(key=lambda region: (*region["bbox"], region["membership_sha256"]))
    return result


def _formula_coordinates_from_biff_stream(stream: bytes) -> list[set[tuple[int, int]]]:
    boundsheets = []
    cursor = 0
    while cursor + 4 <= len(stream):
        record_id, record_length = struct.unpack_from("<HH", stream, cursor)
        payload_start = cursor + 4
        payload_end = payload_start + record_length
        if payload_end > len(stream):
            raise ValueError("BIFF record exceeds workbook stream")
        payload = stream[payload_start:payload_end]
        if record_id == 0x0085 and record_length >= 6:
            boundsheets.append((struct.unpack_from("<I", payload, 0)[0], payload[5]))
        cursor = payload_end

    result = []
    for sheet_offset, sheet_type in boundsheets:
        if sheet_type != 0:
            result.append(set())
            continue
        formulas = set()
        cursor = sheet_offset
        saw_worksheet_bof = False
        saw_eof = False
        while cursor + 4 <= len(stream):
            record_id, record_length = struct.unpack_from("<HH", stream, cursor)
            payload_start = cursor + 4
            payload_end = payload_start + record_length
            if payload_end > len(stream):
                raise ValueError("BIFF sheet record exceeds workbook stream")
            payload = stream[payload_start:payload_end]
            if record_id == 0x0809 and record_length >= 4:
                saw_worksheet_bof = struct.unpack_from("<H", payload, 2)[0] == 0x0010
            elif record_id == 0x0006 and record_length >= 4:
                row_ordinal, column_ordinal = struct.unpack_from("<HH", payload, 0)
                formulas.add((row_ordinal + 1, column_ordinal + 1))
            elif record_id == 0x000A:
                saw_eof = True
                break
            cursor = payload_end
        if not saw_worksheet_bof or not saw_eof:
            raise ValueError("BIFF worksheet substream is incomplete")
        result.append(formulas)
    if not result:
        raise ValueError("BIFF workbook has no worksheet boundsheets")
    return result


def _formula_cached_result_kinds_from_biff_stream(
    stream: bytes,
) -> list[dict[tuple[int, int], str]]:
    """Read BIFF formula result kinds while preserving worksheet ordinals.

    BIFF stores the cached formula result in the Formula record, but not an
    OOXML-style expression string that can be fed to the existing reference
    parser. The cached kind is still authoritative for whether the source
    loader has a result for that cell; row completeness remains fail-closed
    when a formula is the only source value in its row.
    """

    boundsheets = []
    cursor = 0
    while cursor + 4 <= len(stream):
        record_id, record_length = struct.unpack_from("<HH", stream, cursor)
        payload_start = cursor + 4
        payload_end = payload_start + record_length
        if payload_end > len(stream):
            raise ValueError("BIFF record exceeds workbook stream")
        payload = stream[payload_start:payload_end]
        if record_id == 0x0085 and record_length >= 6:
            boundsheets.append((struct.unpack_from("<I", payload, 0)[0], payload[5]))
        cursor = payload_end

    result = []
    for sheet_offset, sheet_type in boundsheets:
        if sheet_type != 0:
            result.append({})
            continue
        cached = {}
        cursor = sheet_offset
        saw_worksheet_bof = False
        saw_eof = False
        while cursor + 4 <= len(stream):
            record_id, record_length = struct.unpack_from("<HH", stream, cursor)
            payload_start = cursor + 4
            payload_end = payload_start + record_length
            if payload_end > len(stream):
                raise ValueError("BIFF sheet record exceeds workbook stream")
            payload = stream[payload_start:payload_end]
            if record_id == 0x0809 and record_length >= 4:
                saw_worksheet_bof = struct.unpack_from("<H", payload, 2)[0] == 0x0010
            elif record_id == 0x0006 and record_length >= 14:
                row_ordinal, column_ordinal = struct.unpack_from("<HH", payload, 0)
                result_bytes = payload[6:14]
                if result_bytes[6:8] == b"\xff\xff":
                    kind = {
                        0: "string",
                        1: "boolean",
                        2: "error",
                        3: "empty",
                    }.get(result_bytes[0], "unknown")
                else:
                    kind = "numeric"
                cached[(row_ordinal + 1, column_ordinal + 1)] = kind
            elif record_id == 0x000A:
                saw_eof = True
                break
            cursor = payload_end
        if not saw_worksheet_bof or not saw_eof:
            raise ValueError("BIFF worksheet substream is incomplete")
        result.append(cached)
    if not result:
        raise ValueError("BIFF workbook has no worksheet boundsheets")
    return result


def _formula_coordinates_by_sheet(binary: bytes) -> tuple[list[set[tuple[int, int]]], bool]:
    """Recover formula locations without exposing formula text to projections."""

    if binary.startswith(b"\xd0\xcf\x11\xe0"):
        try:
            import olefile

            with olefile.OleFileIO(BytesIO(binary)) as ole:
                stream_name = next(
                    name
                    for name in ("Workbook", "Book")
                    if ole.exists(name)
                )
                stream = ole.openstream(stream_name).read()
            return _formula_coordinates_from_biff_stream(stream), True
        except Exception:
            return [], False
    if not binary.startswith(b"PK\x03\x04"):
        return [], True
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(binary), data_only=False, read_only=False)
    except Exception:
        return [], False
    result = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        result.append(
            {
                (cell.row, cell.column)
                for cell in getattr(worksheet, "_cells", {}).values()
                if cell.__class__.__name__ != "MergedCell" and cell.data_type == "f"
            }
        )
    return result, True


def _formula_cached_result_kinds_by_sheet(binary: bytes) -> list[dict[tuple[int, int], str]]:
    if not binary.startswith(b"\xd0\xcf\x11\xe0"):
        return []
    try:
        import olefile

        with olefile.OleFileIO(BytesIO(binary)) as ole:
            stream_name = next(
                name
                for name in ("Workbook", "Book")
                if ole.exists(name)
            )
            stream = ole.openstream(stream_name).read()
        return _formula_cached_result_kinds_from_biff_stream(stream)
    except Exception:
        return []


def _formula_values_by_sheet(binary: bytes) -> list[dict[tuple[int, int], str]]:
    """Return OOXML formula expressions for structural dependency checks."""

    if not binary.startswith(b"PK\x03\x04"):
        return []
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(binary), data_only=False, read_only=False)
    except Exception:
        return []
    result = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        result.append(
            {
                (cell.row, cell.column): cell.value
                for cell in getattr(worksheet, "_cells", {}).values()
                if cell.__class__.__name__ != "MergedCell"
                and cell.data_type == "f"
                and isinstance(cell.value, str)
            }
        )
    return result


def _formula_reference_ranges(
    formula: str,
    sheet_name: str,
) -> tuple[list[tuple[int, int, int, int]], bool]:
    """Parse local A1 references without exposing formula text downstream."""

    from openpyxl.formula.tokenizer import Tokenizer
    from openpyxl.utils.cell import range_boundaries

    ranges = []
    unresolved = False
    try:
        tokens = Tokenizer(formula).items
    except Exception:
        return [], True
    for token in tokens:
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue
        reference = token.value
        if "!" in reference:
            reference_sheet, reference = reference.rsplit("!", 1)
            reference_sheet = reference_sheet.strip("'").replace("''", "'")
            if reference_sheet != sheet_name:
                unresolved = True
                continue
        reference = reference.replace("$", "")
        try:
            min_column, min_row, max_column, max_row = range_boundaries(reference)
        except ValueError:
            unresolved = True
            continue
        if None in (min_column, min_row, max_column, max_row):
            unresolved = True
            continue
        ranges.append((min_row, min_column, max_row, max_column))
    return ranges, unresolved


def _copy_structure_region(parser, worksheet, region: dict[str, Any]):
    from openpyxl import Workbook

    region_workbook = Workbook()
    target = region_workbook.active
    min_row, min_column, _max_row, _max_column = region["bbox"]
    members = region["members"]
    merged_ranges = list(worksheet.merged_cells.ranges)
    merged_coordinates = {
        (row_ordinal, column_ordinal)
        for merged in merged_ranges
        for row_ordinal in range(merged.min_row, merged.max_row + 1)
        for column_ordinal in range(merged.min_col, merged.max_col + 1)
    }

    for row_ordinal, column_ordinal in sorted(members):
        if (row_ordinal, column_ordinal) in merged_coordinates:
            continue
        value = worksheet.cell(row_ordinal, column_ordinal).value
        if value is not None and str(value).strip() != "":
            target.cell(
                row=row_ordinal - min_row + 1,
                column=column_ordinal - min_column + 1,
                value=value,
            )
        elif (row_ordinal, column_ordinal) in members:
            # Materialize formula coordinates whose cached value is empty so
            # the copied worksheet retains source-cell geometry.
            target.cell(
                row=row_ordinal - min_row + 1,
                column=column_ordinal - min_column + 1,
            )

    for merged in merged_ranges:
        coordinates = {
            (row_ordinal, column_ordinal)
            for row_ordinal in range(merged.min_row, merged.max_row + 1)
            for column_ordinal in range(merged.min_col, merged.max_col + 1)
        }
        if not coordinates.issubset(members):
            continue
        value = worksheet.cell(merged.min_row, merged.min_col).value
        target.cell(
            row=merged.min_row - min_row + 1,
            column=merged.min_col - min_column + 1,
            value=value,
        )
        target.merge_cells(
            start_row=merged.min_row - min_row + 1,
            start_column=merged.min_col - min_column + 1,
            end_row=merged.max_row - min_row + 1,
            end_column=merged.max_col - min_column + 1,
        )

    for row_ordinal, column_ordinal in region["unresolved_members"]:
        target.cell(
            row=row_ordinal - min_row + 1,
            column=column_ordinal - min_column + 1,
        )
    target._fuxi_unresolved_coordinates = {
        (row_ordinal - min_row + 1, column_ordinal - min_column + 1)
        for row_ordinal, column_ordinal in region["unresolved_members"]
    }
    target._fuxi_source_column_offset = min_column - 1
    return target, min_row - 1


def _parse_region_structure(parser, worksheet, rows):
    fallback = parser._parse_sheet_structure(worksheet, rows)
    max_scan_rows = len(rows)
    candidates = []
    merged_ranges = list(worksheet.merged_cells.ranges)
    for start in range(max_scan_rows):
        if parser._is_empty_row([cell.value for cell in rows[start]]):
            continue
        if (
            start > 0
            and not merged_ranges
            and not parser._is_empty_row([cell.value for cell in rows[start - 1]])
        ):
            # A contiguous text block has no structural evidence of a new
            # table boundary. Keep the earlier candidate or parser fallback.
            continue
        for depth in range(1, max_scan_rows - start + 1):
            end = start + depth
            if parser._is_empty_row([cell.value for cell in rows[end - 1]]):
                continue
            parent_rows = range(start + 1, end)
            if not all(
                any(
                    merged.min_row <= row_ordinal <= merged.max_row
                    and (merged.max_col > merged.min_col or merged.max_row > merged.min_row)
                    for merged in merged_ranges
                )
                for row_ordinal in parent_rows
            ):
                continue

            headers = parser._build_headers_for_region(worksheet, rows, start, end)
            if not headers:
                continue
            if start > 0 and any(
                merged.min_row <= start + 1 <= merged.max_row
                and merged.min_col == 1
                and merged.max_col >= len(headers)
                for merged in merged_ranges
            ):
                previous_nonempty_row = next(
                    (
                        row_index
                        for row_index in range(start, 0, -1)
                        if not parser._is_empty_row(
                            [cell.value for cell in rows[row_index - 1]]
                        )
                    ),
                    None,
                )
                if previous_nonempty_row is not None and not any(
                    merged.min_row <= previous_nonempty_row <= merged.max_row
                    and merged.min_col == 1
                    and merged.max_col >= len(headers)
                    for merged in merged_ranges
                ):
                    continue
            if any(
                merged.min_row <= end
                and merged.max_row >= end + 1
                and merged.max_row > merged.min_row
                for merged in merged_ranges
            ):
                # A candidate boundary may not cut through a vertically
                # spanning source merge. The following row is still part of
                # the structural header in that case.
                continue
            structural_width = max(
                (
                    merged.max_col
                    for merged in merged_ranges
                    if merged.min_row <= end
                    and merged.max_row >= start + 1
                    and worksheet.cell(merged.min_row, merged.min_col).value is not None
                    and str(worksheet.cell(merged.min_row, merged.min_col).value).strip()
                ),
                default=0,
            )
            contiguous_header_width = 0
            for column_index, header in enumerate(headers, start=1):
                if not header or header.startswith("Column_"):
                    break
                contiguous_header_width = column_index
            if structural_width:
                structural_width = min(
                    max(structural_width, contiguous_header_width),
                    len(headers),
                )
            else:
                structural_width = len(headers)
            if structural_width < 1:
                continue
            candidate_headers = headers[:structural_width]
            if any(
                merged.min_row <= end < merged.max_row
                and merged.min_col <= structural_width
                and merged.max_col >= 1
                for merged in merged_ranges
            ):
                continue
            if _is_full_width_merge(end, structural_width, merged_ranges) and any(
                merged.min_row > end
                and merged.min_col <= structural_width
                and merged.max_col >= 1
                for merged in merged_ranges
            ):
                continue
            if depth > 1 and not _row_merge_signature(end, merged_ranges):
                if any(
                    merged.min_row <= end - 1 <= merged.max_row
                    and merged.max_row > merged.min_row
                    for merged in merged_ranges
                ):
                    continue
            first_following_row = next(
                (
                    row_index
                    for row_index, row in enumerate(rows[end:], start=end + 1)
                    if not parser._is_empty_row([cell.value for cell in row])
                ),
                None,
            )
            if first_following_row is not None:
                first_values = [
                    _cell_value(parser, worksheet, first_following_row, column_index, merged_ranges)
                    for column_index in range(1, structural_width + 1)
                ]
                inherited_merge_columns = {
                    column_index
                    for column_index in range(1, structural_width + 1)
                    if any(
                        merged.min_row <= end
                        and merged.max_row >= first_following_row
                        and merged.min_col <= column_index <= merged.max_col
                        for merged in merged_ranges
                    )
                }
                first_row_spanning_columns = {
                    column_index
                    for column_index in range(1, structural_width + 1)
                    if any(
                        merged.min_row == first_following_row
                        and merged.max_row > first_following_row
                        and merged.min_col <= column_index <= merged.max_col
                        for merged in merged_ranges
                    )
                }
                if (
                    all(value is not None and str(value).strip() for value in first_values)
                    and (
                        inherited_merge_columns == set(range(1, structural_width + 1))
                        or first_row_spanning_columns == set(range(1, structural_width + 1))
                    )
                ):
                    # Every value is a continuation of a merge that began in
                    # the candidate header. The boundary is still inside a
                    # multi-row header, not at the record axis.
                    continue
            following_rows = []
            full_width_before_axis = False
            for row_index, row in enumerate(rows[end:], start=end + 1):
                if parser._is_empty_row([cell.value for cell in row]):
                    continue
                if _is_full_width_merge(row_index, structural_width, merged_ranges):
                    if not following_rows:
                        full_width_before_axis = True
                    else:
                        values = [
                            _cell_value(parser, worksheet, row_index, column_index, merged_ranges)
                            for column_index in range(1, structural_width + 1)
                        ]
                        following_rows.append((row_index, values, False))
                    continue
                values = [
                    _cell_value(parser, worksheet, row_index, column_index, merged_ranges)
                    for column_index in range(1, structural_width + 1)
                ]
                following_rows.append((row_index, values, False))
            if full_width_before_axis:
                continue
            if following_rows and following_rows[0][0] != end + 1:
                continue
            evidence = _record_axis_evidence(candidate_headers, following_rows, merged_ranges)
            if evidence is None:
                continue
            if evidence["single_record_axis_proven"] and not _single_record_header_boundary_proven(
                worksheet,
                header_start=start,
                data_start=end,
                width=structural_width,
                merged_ranges=merged_ranges,
                record_axis_evidence=evidence,
            ):
                continue
            if (
                len(following_rows) >= 2
                and _row_merge_signature(following_rows[0][0], merged_ranges)
                and not _row_merge_signature(following_rows[1][0], merged_ranges)
                and all(
                    not _row_merge_signature(row_ordinal, merged_ranges)
                    for row_ordinal, _values, _follows_body_gap in following_rows[1:]
                )
            ):
                # A merged row immediately followed by an unmerged stable axis
                # is still part of the header band, not the first record.
                continue
            record_offsets = evidence["occupied_offsets"]
            record_count = len(evidence["record_row_ordinals"])
            if any(
                offset >= len(candidate_headers) or candidate_headers[offset].startswith("Column_")
                for offset in record_offsets
            ):
                continue
            distinct_headers = {candidate_headers[offset] for offset in record_offsets}
            if len(distinct_headers) < min(2, len(record_offsets)):
                continue
            if not (
                len(candidate_headers) == 1
                and _single_column_axis_proven(candidate_headers, following_rows)
            ) and not _header_boundary_proven(
                parser,
                worksheet,
                header_start=start,
                data_start=end,
                headers=headers,
                body_rows=following_rows,
                merged_ranges=merged_ranges,
                record_axis_evidence=evidence,
            ):
                continue
            candidates.append(
                (
                    evidence["record_key_axis_proven"]
                    or evidence["single_record_axis_proven"],
                    evidence["record_key_axis_proven"],
                    evidence["single_record_axis_proven"],
                    not evidence["unknown_row_ordinals"],
                    evidence["record_axis_contiguous"],
                    len(record_offsets),
                    depth,
                    record_count,
                    -start,
                    headers,
                    start,
                    end,
                    not any(
                        _row_merge_signature(row_ordinal, merged_ranges)
                        for row_ordinal, _values, _follows_body_gap in following_rows
                    ),
                    tuple(evidence["record_row_ordinals"]),
                    len(distinct_headers) == len(record_offsets),
                    max(record_offsets) + 1,
                )
            )
    if candidates:
        def absorbs_proven_record_rows(candidate, witness) -> bool:
            if candidate is witness or candidate[10] != witness[10]:
                return False
            candidate_records = set(candidate[13])
            witness_records = set(witness[13])
            if (
                not candidate_records
                or not candidate_records < witness_records
                or candidate[11] not in witness_records
                or witness[11] >= candidate[11]
                or candidate[5] != witness[5]
                or not witness[3]
                or not witness[4]
            ):
                return False
            return all(
                candidate_header == witness_header
                or candidate_header.startswith(f"{witness_header}-")
                for candidate_header, witness_header in zip(
                    candidate[9],
                    witness[9],
                )
            )

        selection_candidates = [
            candidate
            for candidate in candidates
            if not any(
                absorbs_proven_record_rows(candidate, witness)
                for witness in candidates
            )
        ]
        strongest_candidate = max(
            selection_candidates,
            key=lambda candidate: (
                candidate[1],
                candidate[3],
                candidate[4],
                candidate[14],
                candidate[7],
                candidate[0],
                candidate[2],
                candidate[5],
                candidate[6],
                candidate[8],
            ),
        )

        def record_axis_rank(candidate):
            return (
                candidate[1],
                candidate[3],
                candidate[4],
                candidate[14],
                candidate[7],
                candidate[0],
                candidate[2],
                candidate[5],
            )

        fallback_candidate = next(
            (
                candidate
                for candidate in selection_candidates
                if candidate[9] == fallback[0]
                and candidate[10] == fallback[1]
                and candidate[11] == fallback[2]
            ),
            None,
        )
        fallback_context_anchors = (
            _source_anchors_by_row(
                worksheet,
                rows,
                row_limit=fallback_candidate[10],
                width=len(fallback_candidate[9]),
                merged_ranges=merged_ranges,
            )
            if fallback_candidate is not None
            else {}
        )
        fallback_excludes_proven_context = (
            fallback_candidate is not None
            and strongest_candidate[10] < fallback_candidate[10]
            and any(
                _inline_context_row_proven(
                    worksheet,
                    rows,
                    row_ordinal=row_ordinal,
                    width=context_width,
                    merged_ranges=merged_ranges,
                )
                for row_ordinal in range(
                    strongest_candidate[10] + 1,
                    fallback_candidate[10] + 1,
                )
                for context_width in {
                    anchor["max_column"]
                    for anchor in fallback_context_anchors.get(row_ordinal, [])
                    if fallback_candidate[15]
                    <= anchor["max_column"]
                    <= len(fallback_candidate[9])
                }
            )
        )
        selected_candidate = (
            fallback_candidate
            if fallback_candidate is not None
            and record_axis_rank(fallback_candidate)
            == record_axis_rank(strongest_candidate)
            and fallback_excludes_proven_context
            else strongest_candidate
        )
        (
            _closed_axis_proven,
            _record_key_axis_proven,
            _single_record_axis_proven,
            _has_no_unknown_rows,
            _record_axis_contiguous,
            _field_count,
            _depth,
            _record_count,
            _negative_start,
            headers,
            header_start,
            data_start,
            _body_merge_free,
            _record_row_ordinals,
            _record_field_identities_closed,
            _record_axis_width,
        ) = selected_candidate
        return headers, header_start, data_start
    return fallback


def _header_paths_for_region(
    parser,
    worksheet,
    rows,
    header_start: int,
    data_start: int,
    *,
    expected_width: int | None = None,
):
    paths = parser._build_header_paths_for_region(
        worksheet,
        rows,
        header_start,
        data_start,
    )
    normalized = [
        [str(segment).strip() for segment in path if str(segment).strip()]
        for path in paths
    ]
    if (
        expected_width is not None
        and len(normalized) > expected_width
        and all(not path for path in normalized[expected_width:])
    ):
        return normalized[:expected_width]
    return normalized


def _physical_region_rows(worksheet, header_start: int, data_start: int):
    """Read only a proven physical region while retaining its source offset."""

    if header_start < 0 or data_start <= header_start:
        raise ValueError("invalid physical structure region")
    return list(
        worksheet.iter_rows(
            min_row=header_start + 1,
            max_row=data_start,
        )
    )


def _header_structure_for_physical_region(
    parser,
    worksheet,
    header_start: int,
    data_start: int,
):
    rows = _physical_region_rows(worksheet, header_start, data_start)
    local_end = len(rows)
    headers = parser._build_headers_for_region(
        worksheet,
        rows,
        0,
        local_end,
        row_offset=header_start,
    )
    header_paths = parser._build_header_paths_for_region(
        worksheet,
        rows,
        0,
        local_end,
        row_offset=header_start,
    )
    return headers, [
        [str(segment).strip() for segment in path if str(segment).strip()]
        for path in header_paths
    ]


def _inline_context_segments(
    worksheet,
    *,
    row_ordinal: int,
    width: int,
    merged_ranges,
) -> list[dict[str, Any]] | None:
    segments = []
    column = 1
    while column <= width:
        merged = next(
            (
                candidate
                for candidate in merged_ranges
                if candidate.min_row <= row_ordinal <= candidate.max_row
                and candidate.min_col <= column <= candidate.max_col
            ),
            None,
        )
        if merged is not None:
            if merged.min_row != row_ordinal or merged.max_row != row_ordinal:
                return None
            min_column = max(1, merged.min_col)
            max_column = min(width, merged.max_col)
            value = worksheet.cell(merged.min_row, merged.min_col).value
        else:
            min_column = max_column = column
            value = worksheet.cell(row_ordinal, column).value
        segments.append(
            {
                "min_column": min_column,
                "max_column": max_column,
                "value": _sanitize_untrusted_text(value),
            }
        )
        column = max_column + 1
    coalesced_segments = []
    for segment in segments:
        if (
            coalesced_segments
            and not segment["value"]
            and not coalesced_segments[-1]["value"]
        ):
            coalesced_segments[-1]["max_column"] = segment["max_column"]
        else:
            coalesced_segments.append(segment)
    segments = coalesced_segments
    if len(segments) < 4 or len(segments) % 2:
        return None
    if segments[0]["min_column"] != 1 or segments[-1]["max_column"] != width:
        return None
    if not all(segment["value"] for segment in segments[::2]):
        return None
    if not any(
        segment["max_column"] > segment["min_column"]
        for segment in segments
    ) or not all(
        left["max_column"] + 1 == right["min_column"]
        for left, right in zip(segments, segments[1:])
    ):
        return None
    return segments


def _inline_context_row_proven(
    worksheet,
    rows,
    *,
    row_ordinal: int,
    width: int,
    merged_ranges,
) -> bool:
    return _inline_context_segments(
        worksheet,
        row_ordinal=row_ordinal,
        width=width,
        merged_ranges=merged_ranges,
    ) is not None


def _context_preceded_multilevel_empty_axis(
    parser,
    worksheet,
    rows,
    populated_rows,
    merged_ranges,
    *,
    proof: dict[str, bool] | None = None,
):
    populated_set = set(populated_rows)
    last_populated = max(populated_set)
    merge_covered_rows = set()
    candidate_starts = set()
    for merged in merged_ranges:
        if merged.max_row > last_populated:
            continue
        merge_covered_rows.update(range(merged.min_row, merged.max_row + 1))
        if 2 <= merged.min_row < last_populated:
            candidate_starts.add(merged.min_row)

    trailing_populated_start = last_populated
    while trailing_populated_start - 1 in populated_set:
        trailing_populated_start -= 1
    trailing_merged_start = last_populated
    while trailing_merged_start in merge_covered_rows:
        trailing_merged_start -= 1
    trailing_merged_start += 1
    minimum_candidate_start = max(
        trailing_populated_start,
        trailing_merged_start,
    )
    candidates = []
    for first_header_row in sorted(candidate_starts):
        if first_header_row < minimum_candidate_start:
            continue
        header_merges = [
            merged
            for merged in merged_ranges
            if merged.min_row >= first_header_row
            and merged.max_row <= last_populated
        ]
        if not header_merges:
            continue
        if any(
            merged.min_row < first_header_row <= merged.max_row
            for merged in merged_ranges
        ):
            continue

        header_start = first_header_row - 1
        headers, header_paths = _header_structure_for_physical_region(
            parser,
            worksheet,
            header_start,
            last_populated,
        )
        while (
            headers
            and header_paths
            and headers[-1].startswith("Column_")
            and not header_paths[-1]
        ):
            headers.pop()
            header_paths.pop()
        width = len(headers)
        distinct_header_path_count = len({tuple(path) for path in header_paths})
        duplicate_paths_have_rectangular_merge = any(
            merged.max_row > merged.min_row
            and merged.max_col > merged.min_col
            for merged in header_merges
        )
        title_backed_context = any(
            merged.min_col == 1
            and merged.max_col >= width
            and merged.max_row < first_header_row
            and worksheet.cell(merged.min_row, merged.min_col).value is not None
            and str(worksheet.cell(merged.min_row, merged.min_col).value).strip()
            for merged in merged_ranges
        )
        header_has_vertical_merge = any(
            merged.max_row > merged.min_row
            for merged in header_merges
        )
        separated_title_backed_context = (
            title_backed_context
            and header_has_vertical_merge
            and first_header_row - 1 not in populated_set
        )
        if (
            width < 2
            or len(header_paths) != width
            or any(
                not header
                or header.startswith("Column_")
                or not path
                for header, path in zip(headers, header_paths)
            )
            or distinct_header_path_count < 2
            or (
                distinct_header_path_count != width
                and not duplicate_paths_have_rectangular_merge
            )
            or not (
                separated_title_backed_context
                or _inline_context_row_proven(
                    worksheet,
                    rows,
                    row_ordinal=first_header_row - 1,
                    width=width,
                    merged_ranges=merged_ranges,
                )
            )
        ):
            continue
        candidates.append(
            (
                headers,
                header_paths,
                header_start,
                last_populated,
                separated_title_backed_context,
            )
        )
    if len(candidates) != 1:
        inline_candidates = [candidate for candidate in candidates if not candidate[4]]
        if inline_candidates:
            candidates = inline_candidates
        if len(candidates) != 1:
            return None
    headers, header_paths, header_start, data_start, title_backed = candidates[0]
    if proof is not None:
        proof["title_backed_multilevel"] = title_backed
    return headers, header_paths, header_start, data_start


def _empty_record_axis_structure(parser, worksheet, rows, populated_rows, unresolved_rows):
    """Prove a titled, source-backed header whose record axis is exactly empty."""

    if unresolved_rows or len(populated_rows) < 2:
        return None
    merged_ranges = list(worksheet.merged_cells.ranges)
    occupied = _logical_occupied_cells(parser, worksheet)
    context_preceded_multilevel = _context_preceded_multilevel_empty_axis(
        parser,
        worksheet,
        rows,
        populated_rows,
        merged_ranges,
    )
    if context_preceded_multilevel is not None:
        return context_preceded_multilevel
    title_spans = [
        merged
        for merged in merged_ranges
        if merged.min_col == 1
        and merged.max_col > 1
        and worksheet.cell(merged.min_row, merged.min_col).value is not None
        and str(worksheet.cell(merged.min_row, merged.min_col).value).strip()
    ]
    last_populated = max(populated_rows)
    trailing_start = last_populated
    while trailing_start - 1 in populated_rows:
        trailing_start -= 1
    if trailing_start > min(populated_rows) and trailing_start - 1 not in populated_rows:
        header_start = trailing_start - 1
        data_start = last_populated
        headers, header_paths = _header_structure_for_physical_region(
            parser,
            worksheet,
            header_start,
            data_start,
        )
        header_merges = [
            merged
            for merged in merged_ranges
            if merged.min_row >= trailing_start
            and merged.max_row <= data_start
        ]
        preceding_width = max(
            (
                len(
                    {
                        column
                        for row, column in occupied
                        if row == row_ordinal
                    }
                )
                for row_ordinal in populated_rows
                if row_ordinal < trailing_start
            ),
            default=0,
        )
        single_row_dense_header = (
            trailing_start == data_start
            and len(headers) >= 2
            and preceding_width < len(headers)
            and all(
                _source_cell_anchor(data_start, column, merged_ranges)[0] == "cell"
                for column in range(1, len(headers) + 1)
            )
        )
        multilevel_structural_header = (
            trailing_start < data_start
            and any(
                merged.max_col > merged.min_col or merged.max_row > merged.min_row
                for merged in header_merges
            )
        )
        if (
            len(headers) >= 2
            and len(header_paths) == len(headers)
            and all(
                header
                and not header.startswith("Column_")
                and path
                for header, path in zip(headers, header_paths)
            )
            and len({tuple(path) for path in header_paths}) >= 2
            and (single_row_dense_header or multilevel_structural_header)
        ):
            return headers, header_paths, header_start, data_start

    candidates = []
    for title in title_spans:
        header_row = title.max_row + 1
        header_columns = []
        for column in range(title.min_col, title.max_col + 1):
            value = _cell_value(parser, worksheet, header_row, column, merged_ranges)
            if (
                value is None
                or not str(value).strip()
                or _source_cell_anchor(header_row, column, merged_ranges)[0] != "cell"
            ):
                break
            header_columns.append(column)
        if len(header_columns) < 2:
            continue
        if any(
            _cell_value(parser, worksheet, header_row, column, merged_ranges) is not None
            and str(_cell_value(parser, worksheet, header_row, column, merged_ranges)).strip()
            for column in range(header_columns[-1] + 1, title.max_col + 1)
        ):
            continue
        if any(
            row > header_row and column <= header_columns[-1]
            for row, column in occupied
        ):
            continue
        headers = [
            _sanitize_untrusted_text(
                _cell_value(parser, worksheet, header_row, column, merged_ranges)
            ).strip()
            for column in header_columns
        ]
        if (
            any(not header or header.startswith("Column_") for header in headers)
            or len(set(headers)) != len(headers)
        ):
            continue
        candidates.append((headers, [[header] for header in headers], header_row - 1, header_row))

    for header_row in populated_rows:
        if not any(title.max_row < header_row for title in title_spans):
            continue
        anchors = [
            column
            for column in range(1, worksheet.max_column + 1)
            if (header_row, column) in occupied
            and _source_cell_anchor(header_row, column, merged_ranges)[0] == "cell"
        ]
        if len(anchors) < 2 or anchors != list(range(anchors[0], anchors[-1] + 1)):
            continue
        if (
            header_row > min(populated_rows) + 1
            and header_row - 1 in populated_rows
            and not any(
                merged.max_row == header_row - 1
                and merged.min_row < merged.max_row
                for merged in merged_ranges
            )
        ):
            continue
        if any(
            row > header_row and anchors[0] <= column <= anchors[-1]
            for row, column in occupied
        ):
            continue
        headers = [
            _sanitize_untrusted_text(
                _cell_value(parser, worksheet, header_row, column, merged_ranges)
            ).strip()
            for column in anchors
        ]
        if (
            any(not header or header.startswith("Column_") for header in headers)
            or len(set(headers)) != len(headers)
        ):
            continue
        earlier_members = {
            (row, column)
            for row, column in occupied
            if row < header_row and anchors[0] <= column <= anchors[-1]
        }
        if not earlier_members or not any(
            title.min_col <= anchors[0]
            and title.max_col >= anchors[-1]
            and title.max_row < header_row
            for title in title_spans
        ):
            continue
        candidates.append((headers, [[header] for header in headers], header_row - 1, header_row))
    unique_candidates = {}
    for candidate in candidates:
        key = (
            tuple(candidate[0]),
            tuple(tuple(path) for path in candidate[1]),
            candidate[2],
            candidate[3],
        )
        unique_candidates[key] = candidate
    if len(unique_candidates) != 1:
        return None
    return next(iter(unique_candidates.values()))


def _trailing_empty_record_axis_structure(
    parser,
    worksheet,
    rows,
    populated_rows,
    unresolved_rows,
):
    """Prove a final dense header band before earlier context is parsed as data."""

    if unresolved_rows or len(populated_rows) < 2:
        return None
    last_populated = max(populated_rows)
    trailing_start = last_populated
    while trailing_start - 1 in populated_rows:
        trailing_start -= 1
    if (
        trailing_start == min(populated_rows)
        or trailing_start - 1 in populated_rows
    ):
        return None

    header_start = trailing_start - 1
    data_start = last_populated
    headers, header_paths = _header_structure_for_physical_region(
        parser,
        worksheet,
        header_start,
        data_start,
    )
    merged_ranges = list(worksheet.merged_cells.ranges)
    occupied = _logical_occupied_cells(parser, worksheet)
    preceding_width = max(
        (
            len({column for row, column in occupied if row == row_ordinal})
            for row_ordinal in populated_rows
            if row_ordinal < trailing_start
        ),
        default=0,
    )
    header_merges = [
        merged
        for merged in merged_ranges
        if merged.min_row >= trailing_start
        and merged.max_row <= data_start
    ]
    single_row_dense_header = (
        trailing_start == data_start
        and preceding_width < len(headers)
        and all(
            _source_cell_anchor(data_start, column, merged_ranges)[0] == "cell"
            for column in range(1, len(headers) + 1)
        )
    )
    multilevel_structural_header = (
        trailing_start < data_start
        and any(
            merged.max_col > merged.min_col or merged.max_row > merged.min_row
            for merged in header_merges
        )
    )
    if (
        len(headers) >= 2
        and len(header_paths) == len(headers)
        and all(
            header
            and not header.startswith("Column_")
            and path
            for header, path in zip(headers, header_paths)
        )
        and len({tuple(path) for path in header_paths}) >= 2
        and (single_row_dense_header or multilevel_structural_header)
    ):
        return headers, header_paths, header_start, data_start
    return None


def _nonempty_record_axis_structure(
    parser,
    worksheet,
    rows,
    populated_rows,
    *,
    require_record_key_axis: bool = False,
):
    """Return parsed structure only when its body independently proves records."""

    headers, header_start, data_start = _parse_region_structure(parser, worksheet, rows)
    if not headers:
        return None
    fallback_headers, fallback_header_start, fallback_data_start = (
        parser._parse_sheet_structure(worksheet, rows)
    )
    parser_boundary_agrees = (
        headers == fallback_headers
        and header_start == fallback_header_start
        and data_start == fallback_data_start
    )
    merged_ranges = list(worksheet.merged_cells.ranges)
    body_rows = [
        (
            row_ordinal,
            [
                _cell_value(parser, worksheet, row_ordinal, column_ordinal, merged_ranges)
                for column_ordinal in range(1, len(headers) + 1)
            ],
            False,
        )
        for row_ordinal in populated_rows
        if row_ordinal > data_start
    ]
    evidence = _record_axis_evidence(headers, body_rows, merged_ranges)
    if evidence is None:
        return None
    occupied_offsets = tuple(evidence["occupied_offsets"])
    if parser_boundary_agrees and occupied_offsets:
        record_width = max(occupied_offsets) + 1
        trailing_headers = headers[record_width:]
        if trailing_headers and all(
            header.startswith("Column_") for header in trailing_headers
        ):
            headers = headers[:record_width]
            body_rows = [
                (row_ordinal, values[:record_width], follows_body_gap)
                for row_ordinal, values, follows_body_gap in body_rows
            ]
            evidence = _record_axis_evidence(headers, body_rows, merged_ranges)
            if evidence is None:
                return None
    if evidence["single_record_axis_proven"] and not _single_record_header_boundary_proven(
        worksheet,
        header_start=header_start,
        data_start=data_start,
        width=len(headers),
        merged_ranges=merged_ranges,
        record_axis_evidence=evidence,
    ):
        return None
    if not _header_boundary_proven(
        parser,
        worksheet,
        header_start=header_start,
        data_start=data_start,
        headers=headers,
        body_rows=body_rows,
        merged_ranges=merged_ranges,
        record_axis_evidence=evidence,
    ):
        return None
    if require_record_key_axis and not evidence["record_key_axis_proven"]:
        return None
    return headers, header_start, data_start


def _primary_record_axis_structures(
    parser,
    worksheet,
    rows,
    populated_rows,
    *,
    allow_context_preceded_empty_axis: bool,
    context_preceded_empty_axis_proof: dict[str, bool] | None = None,
):
    """Resolve overlapping nonempty and context-preceded empty-axis proofs."""

    context_preceded_empty_structure = (
        _context_preceded_multilevel_empty_axis(
            parser,
            worksheet,
            rows,
            populated_rows,
            list(worksheet.merged_cells.ranges),
            proof=context_preceded_empty_axis_proof,
        )
        if allow_context_preceded_empty_axis
        else None
    )
    nonempty_structure = _nonempty_record_axis_structure(
        parser,
        worksheet,
        rows,
        populated_rows,
    )
    if context_preceded_empty_structure is not None and nonempty_structure is not None:
        _empty_headers, empty_header_paths, empty_header_start, empty_data_start = (
            context_preceded_empty_structure
        )
        _record_headers, _record_header_start, record_data_start = nonempty_structure
        duplicate_header_paths = len(
            {tuple(path) for path in empty_header_paths}
        ) != len(empty_header_paths)
        rectangular_header_merge_extends_into_record_axis = any(
            merged.min_row >= empty_header_start + 1
            and merged.max_row <= empty_data_start
            and merged.max_row > merged.min_row
            and merged.max_col > merged.min_col
            and merged.max_row > record_data_start
            for merged in worksheet.merged_cells.ranges
        )
        record_key_axis_proven = _nonempty_record_axis_structure(
            parser,
            worksheet,
            rows,
            populated_rows,
            require_record_key_axis=True,
        ) is not None
        if record_key_axis_proven or (
            duplicate_header_paths
            and not rectangular_header_merge_extends_into_record_axis
        ):
            context_preceded_empty_structure = None
        else:
            nonempty_structure = None
    return nonempty_structure, context_preceded_empty_structure


def _preferred_empty_record_axis_structure(
    context_preceded_empty_structure,
    trailing_empty_structure,
):
    if context_preceded_empty_structure is None:
        return trailing_empty_structure
    if trailing_empty_structure is None:
        return context_preceded_empty_structure
    _headers, header_paths, _header_start, _data_start = (
        context_preceded_empty_structure
    )
    if len({tuple(path) for path in header_paths}) != len(header_paths):
        return trailing_empty_structure
    return context_preceded_empty_structure


def _record_field_offsets(
    values: list[object],
    *,
    row_ordinal: int | None = None,
    merged_ranges=None,
) -> tuple[int, ...]:
    occupied = {
        index
        for index, value in enumerate(values)
        if value is not None and str(value).strip() != ""
    }
    if row_ordinal is None or merged_ranges is None:
        return tuple(sorted(occupied))
    width = len(values)
    for merged in merged_ranges:
        if not (merged.min_row <= row_ordinal <= merged.max_row):
            continue
        if merged.min_col > width or merged.max_col < 1:
            continue
        if any(
            values[index] is not None and str(values[index]).strip() != ""
            for index in range(max(0, merged.min_col - 1), min(width, merged.max_col))
        ):
            occupied.update(
                range(max(0, merged.min_col - 1), min(width, merged.max_col))
            )
    return tuple(sorted(occupied))


def _record_rows_are_semantically_adjacent(
    left_ordinal: int,
    right_ordinal: int,
    width: int,
    merged_ranges,
) -> bool:
    if right_ordinal == left_ordinal + 1:
        return True
    return any(
        merged.max_row > merged.min_row
        and merged.min_row <= left_ordinal
        and merged.max_row >= right_ordinal
        and merged.min_col <= width
        and merged.max_col >= 1
        for merged in merged_ranges
    )


def _record_axis_body_rows(
    parser,
    worksheet,
    headers: list[str],
    data_start: int,
    body_ordinals,
    merged_ranges,
) -> list[tuple[int, list[object], bool]]:
    """Build the shared source-backed body-row view for record-axis decisions."""

    body_rows = []
    previous_body_ordinal = data_start
    for row_ordinal in sorted(body_ordinals):
        follows_body_gap = (
            row_ordinal > previous_body_ordinal + 1
            and not _record_rows_are_semantically_adjacent(
                previous_body_ordinal,
                row_ordinal,
                len(headers),
                merged_ranges,
            )
        )
        body_rows.append(
            (
                row_ordinal,
                [
                    _cell_value(
                        parser,
                        worksheet,
                        row_ordinal,
                        column_ordinal,
                        merged_ranges,
                    )
                    for column_ordinal in range(1, len(headers) + 1)
                ],
                follows_body_gap,
            )
        )
        previous_body_ordinal = row_ordinal
    return body_rows


def _record_axis_evidence(
    headers: list[str],
    body_rows: list[tuple[int, list[object], bool]],
    merged_ranges,
) -> dict[str, Any] | None:
    """Prove one physical record axis from geometry and source-backed values."""

    if not body_rows:
        return None

    def evaluate(rows, note_rows, unknown_rows=()):
        row_ordinals = [row[0] for row in rows]
        record_axis_contiguous = not any(
            not _record_rows_are_semantically_adjacent(
                left,
                right,
                len(headers),
                merged_ranges,
            )
            for left, right in zip(row_ordinals, row_ordinals[1:])
        )
        if any(_is_full_width_merge(row[0], len(headers), merged_ranges) for row in rows):
            return None
        if any(_is_repeated_header_row(headers, row[1]) for row in rows):
            return None
        row_offsets = [
            _record_field_offsets(
                values,
                row_ordinal=row_ordinal,
                merged_ranges=merged_ranges,
            )
            for row_ordinal, values, _gap in rows
        ]
        if any(not offsets for offsets in row_offsets):
            return None
        if (
            len(headers) > 1
            and not merged_ranges
            and len(rows) == 2
            and len(set(row_offsets)) > 1
        ):
            return None
        common_offsets = set(row_offsets[0]).intersection(
            *(set(offsets) for offsets in row_offsets[1:])
        )
        occupied_offsets = set().union(*(set(offsets) for offsets in row_offsets))
        if not common_offsets or any(
            offset >= len(headers) or headers[offset].startswith("Column_")
            for offset in occupied_offsets
        ):
            return None
        if len({headers[offset] for offset in common_offsets}) < min(2, len(common_offsets)):
            return None

        record_key_axis_proven = _record_key_axis_proven(rows, common_offsets)
        key_only_slots = (
            _record_key_only_slots(rows, row_offsets, common_offsets)
            if len(headers) > 1
            else ()
        )
        key_only_slot_set = set(key_only_slots)
        record_row_ordinals = tuple(
            row_ordinal
            for row_ordinal in row_ordinals
            if row_ordinal not in key_only_slot_set
        )
        if not record_row_ordinals:
            return None
        single_axis = (
            len(rows) == 1
            and len(headers) >= 2
            and len(occupied_offsets) >= 2
            and _has_partial_row_merge(row_ordinals[0], len(headers), merged_ranges)
        )
        if len(rows) < 2 and not single_axis:
            return None
        return {
            "record_row_ordinals": record_row_ordinals,
            "row_ordinals": tuple(row[0] for row in body_rows),
            "note_row_ordinals": tuple(
                sorted({row[0] for row in note_rows} | key_only_slot_set)
            ),
            "unknown_row_ordinals": tuple(row[0] for row in unknown_rows),
            "record_axis_contiguous": record_axis_contiguous,
            "row_offsets": tuple(row_offsets),
            "required_offsets": tuple(sorted(common_offsets)),
            "optional_offsets": tuple(sorted(occupied_offsets - common_offsets)),
            "occupied_offsets": tuple(sorted(occupied_offsets)),
            "record_key_axis_proven": record_key_axis_proven,
            "single_record_axis_proven": single_axis,
        }

    segments = []
    current = []
    for row in body_rows:
        if current and not _record_rows_are_semantically_adjacent(
            current[-1][0],
            row[0],
            len(headers),
            merged_ranges,
        ):
            segments.append(current)
            current = []
        current.append(row)
    if current:
        segments.append(current)

    record_rows = list(segments[0])
    note_rows = []
    unknown_rows = []
    for segment in segments[1:]:
        target = (
            note_rows
            if len(segment) == 1
            or (
                segment
                and _is_full_width_merge(segment[0][0], len(headers), merged_ranges)
            )
            else unknown_rows
        )
        target.extend(segment)
    full_width_rows = [
        row
        for row in record_rows
        if _is_full_width_merge(row[0], len(headers), merged_ranges)
    ]
    full_width_indexes = [
        index
        for index, row in enumerate(record_rows)
        if _is_full_width_merge(row[0], len(headers), merged_ranges)
    ]
    if full_width_indexes:
        candidate_records = [row for row in record_rows if row not in full_width_rows]
        if len(candidate_records) < 2:
            return None
        candidate_notes = []
        unknown_rows = [*unknown_rows, *full_width_rows]
        only_trailing_full_width = full_width_indexes == list(
            range(full_width_indexes[0], len(record_rows))
        )
        key_shapes = {
            _record_axis_value_shape(row[1][0])
            for row in candidate_records
            if row[1] and row[1][0] is not None and str(row[1][0]).strip()
        }
        if only_trailing_full_width and len(full_width_rows) == 1:
            note_shape = _record_axis_value_shape(full_width_rows[0][1][0])
            if len(key_shapes) == 1 and note_shape not in key_shapes:
                candidate_notes = list(full_width_rows)
                unknown_rows = [
                    row for row in unknown_rows if row not in full_width_rows
                ]
        record_rows = candidate_records
        note_rows = [*candidate_notes, *note_rows]
        return evaluate(record_rows, note_rows, unknown_rows)

    return evaluate(record_rows, note_rows, unknown_rows)


def _header_boundary_proven(
    parser,
    worksheet,
    *,
    header_start: int,
    data_start: int,
    headers: list[str],
    body_rows: list[tuple[int, list[object], bool]],
    merged_ranges,
    record_axis_evidence: dict[str, Any] | None = None,
) -> bool:
    if any(
        merged.max_row == data_start
        and merged.min_row >= header_start + 1
        and merged.max_row > merged.min_row
        for merged in merged_ranges
    ):
        return True

    header_values = [
        _cell_value(parser, worksheet, data_start, column_ordinal, merged_ranges)
        for column_ordinal in range(1, len(headers) + 1)
    ]
    body_values = [
        values
        for row_ordinal, values, _follows_body_gap in body_rows
        if _classify_body_row(
            row_ordinal,
            values,
            merged_ranges,
            record_axis_evidence=record_axis_evidence,
        ) == "data"
        and not _is_repeated_header_row(headers, values)
    ]
    if len(body_values) < 2:
        return True
    boundary_offsets = (
        record_axis_evidence["required_offsets"]
        if record_axis_evidence is not None
        else range(len(header_values))
    )
    for column_index in boundary_offsets:
        header_value = header_values[column_index]
        header_kind = _row_shape([header_value], distinguish_text_digits=True)[0]
        body_kinds = {
            _row_shape([values[column_index]], distinguish_text_digits=True)[0]
            for values in body_values
            if column_index < len(values)
            and values[column_index] is not None
            and str(values[column_index]).strip() != ""
        }
        if body_kinds and header_kind not in body_kinds:
            return True
    return False


def _single_record_header_boundary_proven(
    worksheet,
    *,
    header_start: int,
    data_start: int,
    width: int,
    merged_ranges,
    record_axis_evidence: dict[str, Any] | None = None,
) -> bool:
    header_merges = [
        merged
        for merged in merged_ranges
        if merged.min_row >= header_start + 1
        and merged.max_row <= data_start
        and merged.max_col >= 1
        and merged.min_col <= width
    ]
    leaf_sources = {
        _source_cell_anchor(data_start, column, merged_ranges)
        for column in range(1, width + 1)
    }
    contextual_boundary = (
        any(merged.max_col > merged.min_col for merged in header_merges)
        and any(source[0] == "cell" for source in leaf_sources)
        and (
            data_start - header_start >= 2
            or (
                header_start > 0
                and any(
                    merged.max_row < header_start + 1
                    for merged in merged_ranges
                )
                or (
                    header_start > 1
                    and all(
                        worksheet.cell(header_start, column).value is None
                        for column in range(1, width + 1)
                    )
                    and any(
                        worksheet.cell(row, column).value is not None
                        for row in range(1, header_start)
                        for column in range(1, width + 1)
                    )
                )
            )
        )
    )
    if contextual_boundary:
        return True

    record_rows = tuple(
        (record_axis_evidence or {}).get("record_row_ordinals", ())
    )
    required_offsets = tuple(
        (record_axis_evidence or {}).get("required_offsets", ())
    )
    if len(record_rows) != 1 or len(required_offsets) < 2:
        return False
    record_row = record_rows[0]

    def horizontal_spans(row_ordinal: int) -> set[tuple[int, int]]:
        return {
            (max(1, merged.min_col), min(width, merged.max_col))
            for merged in merged_ranges
            if merged.min_row <= row_ordinal <= merged.max_row
            and merged.max_col > merged.min_col
            and merged.min_col <= width
            and merged.max_col >= 1
        }

    header_spans = horizontal_spans(data_start)
    record_spans = horizontal_spans(record_row)
    key_value = worksheet.cell(record_row, min(required_offsets) + 1).value
    return (
        bool(header_spans)
        and header_spans == record_spans
        and isinstance(key_value, (int, float))
        and not isinstance(key_value, bool)
        and any(
            _source_cell_anchor(record_row, column, merged_ranges)[0] == "cell"
            for column in range(1, width + 1)
        )
    )


def _text_structure(value: object) -> tuple[str, ...]:
    rendered = str(value).strip()
    result = []
    for char in rendered:
        if char.isdigit():
            kind = "digit"
        elif char.isalpha():
            kind = "alpha"
        elif char.isspace():
            kind = "space"
        else:
            kind = "symbol"
        if not result or result[-1] != kind:
            result.append(kind)
    return tuple(result)


def _single_column_axis_proven(
    headers: list[str],
    body_rows: list[tuple[int, list[object], bool]],
) -> bool:
    if len(headers) != 1:
        return True
    values = [
        values[0]
        for _row_ordinal, values, _follows_body_gap in body_rows
        if values and values[0] is not None and str(values[0]).strip() != ""
    ]
    if len(values) < 2:
        return False
    if all(isinstance(value, (bool, int, float)) for value in values):
        return True
    shapes = {_text_structure(value) for value in values}
    header_shape = _text_structure(headers[0])
    return (
        len(shapes) == 1
        and next(iter(shapes)) != header_shape
        and any(kind in {"digit", "symbol"} for kind in next(iter(shapes)))
    )


def _sparse_record_axis_proven(
    headers: list[str],
    body_rows: list[tuple[int, list[object], bool]],
    data_field_offsets: list[tuple[int, ...]],
) -> bool:
    if len(data_field_offsets) < 2 or len(data_field_offsets) != len(body_rows):
        return False
    if any(follows_body_gap for _row, _values, follows_body_gap in body_rows):
        return False
    offset_sets = [set(offsets) for offsets in data_field_offsets]
    common_offsets = set.intersection(*offset_sets)
    occupied_offsets = set.union(*offset_sets)
    optional_offsets = occupied_offsets - common_offsets
    if not optional_offsets or len(set(data_field_offsets)) < 2:
        return False
    if any(
        offset >= len(headers) or headers[offset].startswith("Column_")
        for offset in occupied_offsets
    ):
        return False
    return True


def _g1_disagreement_is_outside_record_axis(
    worksheet,
    region: dict[str, Any],
    record_row_ordinals: set[int],
) -> bool:
    if not record_row_ordinals:
        return False
    record_members = {
        coordinate
        for coordinate in region["members"]
        if coordinate[0] in record_row_ordinals
    }
    record_children = [
        child for child in region["g1_children"] if child & record_members
    ]
    non_record_children = [
        child for child in region["g1_children"] if not child & record_members
    ]
    record_columns = {column for _row, column in record_members}
    record_child_columns = (
        {column for _row, column in record_children[0]}
        if len(record_children) == 1
        else set()
    )
    merged_ranges = list(worksheet.merged_cells.ranges)

    dominant_record_children = []
    for candidate in record_children:
        candidate_rows = {row for row, _column in candidate}
        candidate_columns = {column for _row, column in candidate}
        candidate_bbox = _region_bbox(candidate)
        if not record_row_ordinals.issubset(candidate_rows):
            continue
        if not record_columns.issubset(candidate_columns):
            continue
        if all(
            candidate_bbox[0] <= child_bbox[0]
            and candidate_bbox[1] <= child_bbox[1]
            and candidate_bbox[2] >= child_bbox[2]
            and candidate_bbox[3] >= child_bbox[3]
            for child_bbox in (_region_bbox(child) for child in record_children)
        ):
            dominant_record_children.append(candidate)

    record_axis_closed = (
        len(record_children) == 1
        and record_members.issubset(record_children[0])
    ) or len(dominant_record_children) == 1

    def child_is_outside_context(child: set[tuple[int, int]]) -> bool:
        child_min_row, _min_column, child_max_row, _max_column = _region_bbox(child)
        if child_max_row < min(record_row_ordinals):
            child_columns = {column for _row, column in child}
            return (
                child_columns.issubset(record_child_columns)
                or not _members_prove_repeated_axis(child)
                or any(
                merged.min_row >= child_min_row
                and merged.max_row <= child_max_row
                and merged.min_col <= min(record_columns)
                and merged.max_col >= max(record_columns)
                for merged in merged_ranges
            )
            )
        if len({row for row, _column in child}) == 1:
            return True
        child_columns = {column for _row, column in child}
        if not child_columns.issubset(record_columns):
            return not _members_prove_repeated_axis(child)
        return any(
            merged.min_row >= child_min_row
            and merged.max_row <= child_max_row
            and merged.min_col <= min(record_columns)
            and merged.max_col >= max(record_columns)
            for merged in merged_ranges
        )

    return (
        bool(record_members)
        and record_axis_closed
        and all(child_is_outside_context(child) for child in non_record_children)
        and all(
            max(row for row, _column in child) < min(record_row_ordinals)
            or min(row for row, _column in child) > max(record_row_ordinals)
            for child in non_record_children
        )
    )


def _g1_child_region(region: dict[str, Any], members: set[tuple[int, int]]) -> dict[str, Any]:
    return {
        "members": members,
        "unresolved_members": region["unresolved_members"] & members,
        "bbox": _region_bbox(members),
        "g1_children": [members],
    }


def _context_with_headers(
    context: list[dict[str, str]],
    headers: list[str],
    *,
    entry_limit: int,
    value_bytes: int,
) -> list[dict[str, str]]:
    result = list(context[:entry_limit])
    if len(result) >= entry_limit:
        return result
    for header in headers:
        value = _truncate_utf8(_sanitize_untrusted_text(header), value_bytes)
        if not value or any(item["name"] == "field" and item["value"] == value for item in result):
            continue
        result.append({"name": "field", "value": value})
        if len(result) == entry_limit:
            break
    return result


def _project_structure_region(
    *,
    parser,
    worksheet,
    sheet_name: str,
    sheet_ordinal: int,
    table_ordinal: int,
    membership_sha256: str,
    source_sha256: str,
    producer_generation_ref: str,
    row_offset: int,
    table_context_entry_limit: int,
    table_context_value_bytes: int,
    force_unknown_total: bool = False,
) -> tuple[dict[str, Any], list[dict[str, Any]]] | None:
    header_rows, populated_row_ordinals, unresolved_row_ordinals = _complete_worksheet_rows(worksheet)
    if not header_rows or not populated_row_ordinals:
        return None
    nonempty_structure, context_preceded_empty_structure = (
        _primary_record_axis_structures(
            parser,
            worksheet,
            header_rows,
            populated_row_ordinals,
            allow_context_preceded_empty_axis=(
                not force_unknown_total
                and not unresolved_row_ordinals
                and len(populated_row_ordinals) >= 2
            ),
        )
    )
    trailing_empty_candidate = (
        None
        if force_unknown_total
        else _trailing_empty_record_axis_structure(
            parser,
            worksheet,
            header_rows,
            populated_row_ordinals,
            unresolved_row_ordinals,
        )
    )
    nonempty_has_unnamed_headers = (
        nonempty_structure is not None
        and any(
            not header or header.startswith("Column_")
            for header in nonempty_structure[0]
        )
    )
    trailing_empty_structure = (
        trailing_empty_candidate
        if trailing_empty_candidate is not None
        and (nonempty_structure is None or nonempty_has_unnamed_headers)
        else None
    )
    if trailing_empty_structure is not None:
        nonempty_structure = None
    preferred_empty_structure = _preferred_empty_record_axis_structure(
        context_preceded_empty_structure,
        trailing_empty_structure,
    )
    empty_structure = preferred_empty_structure or (
        None
        if force_unknown_total or nonempty_structure is not None
        else _empty_record_axis_structure(
            parser,
            worksheet,
            header_rows,
            populated_row_ordinals,
            unresolved_row_ordinals,
        )
    )
    if empty_structure is not None:
        headers, header_paths, header_start, data_start = empty_structure
    else:
        headers, header_start, data_start = nonempty_structure or _parse_region_structure(
            parser,
            worksheet,
            header_rows,
        )
        header_paths = _header_paths_for_region(
            parser,
            worksheet,
            header_rows,
            header_start,
            data_start,
            expected_width=len(headers),
        )
    if not headers:
        return None
    if len(header_paths) != len(headers):
        return None
    source_geometry_context = header_start > 0
    source_column_offset = getattr(worksheet, "_fuxi_source_column_offset", 0)

    body_ordinals = {ordinal for ordinal in populated_row_ordinals if ordinal > data_start}
    unresolved_body_ordinals = {ordinal for ordinal in unresolved_row_ordinals if ordinal > data_start}
    body_ordinals.update(unresolved_body_ordinals)
    if not body_ordinals:
        if empty_structure is None:
            return None
        table_ref = _table_ref(source_sha256, sheet_ordinal, table_ordinal, membership_sha256)
        table = {
            "table_ref": table_ref,
            "sheet_ordinal": sheet_ordinal,
            "table_ordinal": table_ordinal,
            "row_count": 0,
            "data_row_count": 0,
            "source_total_count": 0,
            "table_label": _sanitize_untrusted_text(sheet_name),
            "table_context": _context_with_headers(
                _table_context(
                    parser,
                    worksheet,
                    header_rows,
                    header_start,
                    width=max(
                        (
                            index
                            for index, header in enumerate(headers, start=1)
                            if header and not header.startswith("Column_")
                        ),
                        default=len(headers),
                    ),
                    source_geometry=source_geometry_context,
                    entry_limit=table_context_entry_limit,
                    value_bytes=table_context_value_bytes,
                ),
                headers,
                entry_limit=table_context_entry_limit,
                value_bytes=table_context_value_bytes,
            ),
        }
        _apply_enumeration_decision(table, "L1-08")
        return table, []

    table_ref = _table_ref(source_sha256, sheet_ordinal, table_ordinal, membership_sha256)
    context = _context_with_headers(
        _table_context(
            parser,
            worksheet,
            header_rows,
            header_start,
            width=max(
                (
                    index
                    for index, header in enumerate(headers, start=1)
                    if header and not header.startswith("Column_")
                ),
                default=len(headers),
            ),
            source_geometry=source_geometry_context,
            entry_limit=table_context_entry_limit,
            value_bytes=table_context_value_bytes,
        ),
        headers,
        entry_limit=table_context_entry_limit,
        value_bytes=table_context_value_bytes,
    )
    merged_ranges = list(worksheet.merged_cells.ranges)
    body_rows = _record_axis_body_rows(
        parser,
        worksheet,
        headers,
        data_start,
        body_ordinals,
        merged_ranges,
    )

    record_axis_evidence = _record_axis_evidence(headers, body_rows, merged_ranges)
    if (
        record_axis_evidence is not None
        and record_axis_evidence["single_record_axis_proven"]
        and not _single_record_header_boundary_proven(
            worksheet,
            header_start=header_start,
            data_start=data_start,
            width=len(headers),
            merged_ranges=merged_ranges,
            record_axis_evidence=record_axis_evidence,
        )
    ):
        record_axis_evidence = None
    if not _header_boundary_proven(
        parser,
        worksheet,
        header_start=header_start,
        data_start=data_start,
        headers=headers,
        body_rows=body_rows,
        merged_ranges=merged_ranges,
        record_axis_evidence=record_axis_evidence,
    ):
        return None
    record_body_rows = (
        [
            row
            for row in body_rows
            if row[0] in record_axis_evidence["record_row_ordinals"]
        ]
        if record_axis_evidence is not None
        else []
    )
    sparse_record_axis_evidence = (
        record_axis_evidence is not None
        and _sparse_record_axis_proven(
            headers,
            record_body_rows,
            list(record_axis_evidence["row_offsets"]),
        )
    )
    distinguish_text_digits = not any(
        isinstance(value, (int, float)) and not isinstance(value, bool)
        for _row_ordinal, values, _follows_body_gap in body_rows
        for value in values
    )
    pending_rows = []
    data_row_index = 0
    has_unknown = False
    established_shape = None
    for body_index, (local_row_ordinal, values, follows_body_gap) in enumerate(body_rows):
        row_role = "unknown"
        row_has_unresolved = any(
            row_ordinal == local_row_ordinal
            for row_ordinal, _column_ordinal in getattr(
                worksheet,
                "_fuxi_unresolved_coordinates",
                set(),
            )
        )
        if row_has_unresolved:
            row_role = "unknown"
        else:
            row_role = _classify_body_row(
                local_row_ordinal,
                values,
                merged_ranges,
                record_axis_evidence=record_axis_evidence,
            )
            current_shape = _row_shape(values, distinguish_text_digits=distinguish_text_digits)
            next_shape = (
                _row_shape(
                    body_rows[body_index + 1][1],
                    distinguish_text_digits=distinguish_text_digits,
                )
                if body_index + 1 < len(body_rows)
                else None
            )
            if _is_repeated_header_row(headers, values):
                row_role = "unknown"
            if record_axis_evidence is not None:
                required_offsets = record_axis_evidence["required_offsets"]
                required_shape = tuple(
                    current_shape[offset]
                    for offset in required_offsets
                    if offset < len(current_shape)
                )
                next_required_shape = (
                    tuple(
                        next_shape[offset]
                        for offset in required_offsets
                        if offset < len(next_shape)
                    )
                    if next_shape is not None
                    else None
                )
                established_required_shape = (
                    tuple(
                        established_shape[offset]
                        for offset in required_offsets
                        if offset < len(established_shape)
                    )
                    if established_shape is not None
                    else None
                )
                if (
                    row_role == "data"
                    and not record_axis_evidence.get("record_key_axis_proven")
                    and not sparse_record_axis_evidence
                    and established_required_shape is not None
                    and required_shape != established_required_shape
                    and required_shape == next_required_shape
                ):
                    row_role = "unknown"
        if row_role == "data":
            data_row_index += 1
            current_data_index = data_row_index
            established_shape = established_shape or _row_shape(
                values,
                distinguish_text_digits=distinguish_text_digits,
            )
        else:
            current_data_index = None
            has_unknown = has_unknown or row_role == "unknown"
        row_ordinal = local_row_ordinal + row_offset
        row_ref = f"{table_ref}:{row_ordinal}"
        pending_rows.append(
            {
                "id": "tsr_v1_" + _versioned_digest(
                    "tabular-row-record/v1",
                    producer_generation_ref,
                    row_ref,
                ),
                "tabular_structure_version_kwd": TABULAR_STRUCTURE_VERSION,
                "structure_kind_kwd": "table_row",
                "producer_schema_version_kwd": PRODUCER_SCHEMA_VERSION,
                "producer_generation_ref_kwd": producer_generation_ref,
                "table_ref_kwd": table_ref,
                "table_label_kwd": _sanitize_untrusted_text(sheet_name),
                "table_context_list": json.dumps(context, ensure_ascii=False, separators=(",", ":")),
                "row_ref_kwd": row_ref,
                "row_ordinal_int": row_ordinal,
                "data_row_index_int": current_data_index,
                "row_role_kwd": row_role,
                "source_total_count_int": None,
                "ordered_fields_list": json.dumps(
                    _ordered_fields(
                        headers,
                        values,
                        note=row_role == "note",
                        sheet_ordinal=sheet_ordinal,
                        header_paths=header_paths,
                        column_ordinals=list(range(1, len(headers) + 1)),
                        absolute_column_ordinals=list(
                            range(
                                source_column_offset + 1,
                                source_column_offset + len(headers) + 1,
                            )
                        ),
                        source_anchors=[
                            _source_cell_anchor(
                                local_row_ordinal,
                                column_ordinal,
                                merged_ranges,
                            )
                            for column_ordinal in range(1, len(headers) + 1)
                        ],
                    ),
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
            }
        )

    data_field_offsets = [
        _record_field_offsets(
            values,
            row_ordinal=_row_ordinal,
            merged_ranges=merged_ranges,
        )
        for _row_ordinal, values, _follows_body_gap in body_rows
        if _classify_body_row(
            _row_ordinal,
            values,
            merged_ranges,
            record_axis_evidence=record_axis_evidence,
        ) == "data"
        and not _is_repeated_header_row(headers, values)
    ]
    data_value_shapes = [
        tuple(
            (
                "boolean"
                if isinstance(value, bool)
                else "number"
                if isinstance(value, (int, float))
                else ("text", *_text_structure(value))
                if len(headers) == 1
                else "text"
            )
            for value in values
            if value is not None and str(value).strip() != ""
        )
        for _row_ordinal, values, _follows_body_gap in body_rows
        if _classify_body_row(
            _row_ordinal,
            values,
            merged_ranges,
            record_axis_evidence=record_axis_evidence,
        ) == "data"
        and not _is_repeated_header_row(headers, values)
    ]
    regular_record_axis_proven = (
        record_axis_evidence is not None
        and record_axis_evidence["record_axis_contiguous"]
        and not record_axis_evidence["unknown_row_ordinals"]
        and data_row_index == len(record_axis_evidence["record_row_ordinals"])
        and data_row_index == len(data_field_offsets)
        and bool(data_field_offsets)
        and (
            record_axis_evidence.get("record_key_axis_proven") is True
            or (
                len(set(data_field_offsets)) == 1
                and len(set(data_value_shapes)) == 1
            )
        )
        and _single_column_axis_proven(headers, body_rows)
    )
    sparse_record_axis_proven = (
        data_row_index == len(data_field_offsets)
        and record_axis_evidence is not None
        and _sparse_record_axis_proven(
            headers,
            [
                row
                for row in body_rows
                if row[0] in record_axis_evidence["record_row_ordinals"]
            ],
            data_field_offsets,
        )
    )
    record_axis_proven = regular_record_axis_proven or sparse_record_axis_proven
    source_total_count = (
        data_row_index
        if record_axis_proven and not has_unknown and not force_unknown_total
        else None
    )
    for record in pending_rows:
        record["source_total_count_int"] = source_total_count
    table = {
        "table_ref": table_ref,
        "sheet_ordinal": sheet_ordinal,
        "table_ordinal": table_ordinal,
        "row_count": len(pending_rows),
        "data_row_count": data_row_index,
        "source_total_count": source_total_count,
    }
    matched_rule = (
        "L1-05"
        if source_total_count is not None and len(headers) == 1
        else "L1-04"
        if source_total_count is not None and data_start - header_start > 1
        else "L1-06"
        if source_total_count is not None and sparse_record_axis_proven
        else "L1-07"
        if source_total_count is not None
        else "R8"
    )
    _apply_enumeration_decision(
        table,
        matched_rule,
    )
    return table, pending_rows


def _members_prove_repeated_axis(members: set[tuple[int, int]]) -> bool:
    rows: dict[int, set[int]] = {}
    for row_ordinal, column_ordinal in members:
        rows.setdefault(row_ordinal, set()).add(column_ordinal)
    signatures = [
        tuple(column - min(columns) for column in sorted(columns))
        for columns in rows.values()
        if columns
    ]
    return len(signatures) >= 2 and len(set(signatures)) == 1


def _column_sets_intersect(left: set[int], right: set[int]) -> bool:
    return bool(left) and bool(right) and not left.isdisjoint(right)


def _column_sets_are_nested(left: set[int], right: set[int]) -> bool:
    return bool(left) and bool(right) and (
        left.issubset(right) or right.issubset(left)
    )


def _unknown_region_can_extend_proven_axis(
    *,
    parser,
    worksheet,
    proven: dict[str, Any],
    unknown: dict[str, Any],
) -> bool:
    if (
        proven["bbox"][2] >= unknown["bbox"][0]
        or not _column_sets_intersect(
            proven["member_columns"],
            unknown["member_columns"],
        )
    ):
        return False
    proven_evidence = proven.get("structure_evidence")
    proven_rows = [
        row for row in proven["rows"] if row["row_role_kwd"] == "data"
    ]
    if not proven_evidence or not proven_rows:
        return False
    continuation_rows = _continuation_record_rows(unknown)
    return bool(continuation_rows) and _continuation_rows_match_proven_axis(
        parser=parser,
        worksheet=worksheet,
        main=proven,
        continuation=unknown,
        continuation_rows=continuation_rows,
    )


def _record_axis_details(item: dict[str, Any]) -> tuple[dict[str, Any], int, int] | None:
    evidence = item.get("structure_evidence")
    if not evidence:
        return None
    axis = evidence.get("record_axis_evidence")
    source_column = evidence.get("record_axis_source_column")
    width = evidence.get("record_axis_width")
    if (
        not isinstance(axis, dict)
        or not isinstance(source_column, int)
        or not isinstance(width, int)
        or width < 1
        or not axis.get("record_row_ordinals")
        or not axis.get("required_offsets")
    ):
        return None
    return axis, source_column, width


def _candidate_record_axis_source_column(item: dict[str, Any]) -> int | None:
    """Return a candidate's source-column origin without requiring a table parse."""

    evidence = item.get("structure_evidence")
    source_column = (
        evidence.get("record_axis_source_column")
        if isinstance(evidence, dict)
        else None
    )
    if isinstance(source_column, int):
        return source_column
    bbox = item.get("bbox")
    if (
        isinstance(bbox, tuple)
        and len(bbox) == 4
        and isinstance(bbox[1], int)
    ):
        return bbox[1]
    return None


def _source_row_offsets(
    item: dict[str, Any],
    row_ordinal: int,
    *,
    source_column: int,
) -> set[int]:
    """Rebase an item's source columns onto one earlier record-axis origin."""

    return {
        column_ordinal - source_column
        for column_ordinal in _member_rows(item["members"]).get(row_ordinal, set())
    }


def _source_row_merge_signature(
    worksheet,
    row_ordinal: int,
    *,
    source_column: int,
    width: int,
) -> tuple[tuple[int, int], ...]:
    axis_end = source_column + width - 1
    return tuple(
        sorted(
            (
                max(source_column, merged.min_col) - source_column,
                min(axis_end, merged.max_col) - source_column,
            )
            for merged in worksheet.merged_cells.ranges
            if merged.min_row <= row_ordinal <= merged.max_row
            and merged.max_col >= source_column
            and merged.min_col <= axis_end
        )
    )


def _candidate_record_axis_is_empty(
    parser,
    worksheet,
    row_ordinal: int,
    *,
    source_column: int,
    width: int,
    merged_ranges,
) -> bool:
    """Prove a candidate row has no source-backed value on the record axis."""

    return all(
        (
            (value := _cell_value(parser, worksheet, row_ordinal, column, merged_ranges))
        ) is None
        or not str(value).strip()
        for column in range(source_column, source_column + width)
    )


def _candidate_row_is_record_under(
    *,
    worksheet,
    candidate: dict[str, Any],
    row_ordinal: int,
    record_axis_evidence: dict[str, Any],
    source_column: int,
    candidate_source_column: int,
    width: int,
    proven_merge_signatures: set[tuple[tuple[int, int], ...]],
) -> bool:
    """Apply the established axis' row predicate in its source-column basis."""

    merged_ranges = list(worksheet.merged_cells.ranges)
    axis_end = source_column + width - 1
    row_offsets = _source_row_offsets(
        candidate,
        row_ordinal,
        # Compare candidate occupancy in the established axis' physical
        # column basis; a shifted region must not be re-numbered as a record.
        source_column=source_column,
    )
    full_width_merge = any(
        merged.min_row <= row_ordinal <= merged.max_row
        and merged.min_col <= source_column
        and merged.max_col >= axis_end
        for merged in merged_ranges
    )
    if full_width_merge:
        return False
    partial_merge = any(
        merged.min_row <= row_ordinal <= merged.max_row
        and merged.max_col >= source_column
        and merged.min_col <= axis_end
        and not (
            merged.min_col <= source_column and merged.max_col >= axis_end
        )
        for merged in merged_ranges
    )
    merge_signature = _source_row_merge_signature(
        worksheet,
        row_ordinal,
        source_column=source_column,
        width=width,
    )
    if partial_merge and not merge_signature:
        return False
    if merge_signature not in proven_merge_signatures:
        return False
    required_offsets = set(record_axis_evidence["required_offsets"])
    if not required_offsets:
        return False
    record_shape = (
        len(row_offsets) >= 2
        or record_axis_evidence.get("record_key_axis_proven") is True
        or record_axis_evidence.get("single_record_axis_proven") is True
    )
    if not record_shape:
        return False
    if required_offsets.issubset(row_offsets):
        return True
    # A missing key is still a continuation risk when the remaining required
    # fields and the established merge signature match a proven record row.
    if record_axis_evidence.get("record_key_axis_proven") is True:
        key_offset = min(required_offsets)
        return (required_offsets - {key_offset}).issubset(row_offsets)
    return False


def _anchor_shape(value: object) -> tuple[str, ...]:
    if value is None or (isinstance(value, str) and not value.strip()):
        return ()
    return _text_structure(value)


def _axis_anchor_is_not_extendable(
    *,
    parser,
    worksheet,
    earlier: dict[str, Any],
    later: dict[str, Any],
    record_axis_evidence: dict[str, Any],
    source_column: int,
    candidate_source_column: int,
    width: int,
) -> bool:
    """Prove a stable non-numeric anchor cannot be continued by later rows."""

    required_offsets = record_axis_evidence.get("required_offsets", ())
    if not required_offsets:
        return False
    anchor_column = source_column + min(required_offsets)
    candidate_anchor_column = candidate_source_column + min(required_offsets)
    merged_ranges = list(worksheet.merged_cells.ranges)
    later_rows = later.get("rows", ())
    if not later_rows:
        return False
    if record_axis_evidence.get("record_key_axis_proven") is True:
        if all(
            _candidate_record_axis_is_empty(
                parser,
                worksheet,
                row["row_ordinal_int"],
                source_column=candidate_source_column,
                width=width,
                merged_ranges=merged_ranges,
            )
            for row in later_rows
        ):
            return True
        key_offset = min(required_offsets)
        for row in later_rows:
            candidate_row_offsets = _source_row_offsets(
                later,
                row["row_ordinal_int"],
                source_column=candidate_source_column,
            )
            if candidate_row_offsets != {key_offset}:
                continue
            candidate_key = _record_key_numeric_value(
                _cell_value(
                    parser,
                    worksheet,
                    row["row_ordinal_int"],
                    candidate_anchor_column,
                    merged_ranges,
                )
            )
            if candidate_key is not None:
                # A numeric key-only slot is ambiguous with a continuation;
                # preserve the existing fail-closed behavior. Other footer
                # values are handled by the record-shape predicate below.
                return False
        return True
    main_anchor_shapes = {
        _anchor_shape(
            _cell_value(parser, worksheet, row["row_ordinal_int"], anchor_column, merged_ranges)
        )
        for row in earlier["rows"]
        if row["row_role_kwd"] == "data"
    }
    if (
        not main_anchor_shapes
        or () in main_anchor_shapes
        or not any(
            kind in {"digit", "symbol"}
            for shape in main_anchor_shapes
            for kind in shape
        )
    ):
        return False
    candidate_anchor_shapes = [
        _anchor_shape(
            _cell_value(
                parser,
                worksheet,
                row["row_ordinal_int"],
                candidate_anchor_column,
                merged_ranges,
            )
        )
        for row in later_rows
    ]
    return all(shape and shape not in main_anchor_shapes for shape in candidate_anchor_shapes)


def _axis_closure_proven(
    *,
    parser,
    worksheet,
    earlier: dict[str, Any],
    later: dict[str, Any],
) -> bool:
    """Prove a later unknown region cannot be an unaccounted record continuation.

    The proof is deliberately independent of row distance and footer wording.
    An established record axis plus non-record candidate rows are both required;
    missing page-footer values are not business records and do not invalidate
    the proof by themselves.
    """

    details = _record_axis_details(earlier)
    later_rows = later.get("rows", ())
    if details is None or not later_rows:
        return False
    record_axis_evidence, source_column, width = details
    candidate_source_column = _candidate_record_axis_source_column(later)
    if not isinstance(candidate_source_column, int):
        return False
    proven_merge_signatures = {
        _source_row_merge_signature(
            worksheet,
            row["row_ordinal_int"],
            source_column=source_column,
            width=width,
        )
        for row in earlier["rows"]
        if row["row_role_kwd"] == "data"
    }
    if not proven_merge_signatures:
        return False
    axis_not_extendable = _axis_anchor_is_not_extendable(
        parser=parser,
        worksheet=worksheet,
        earlier=earlier,
        later=later,
        record_axis_evidence=record_axis_evidence,
        source_column=source_column,
        candidate_source_column=candidate_source_column,
        width=width,
    )
    candidate_rows_are_not_records = all(
        not _candidate_row_is_record_under(
            worksheet=worksheet,
            candidate=later,
            row_ordinal=row["row_ordinal_int"],
            record_axis_evidence=record_axis_evidence,
            source_column=source_column,
            candidate_source_column=candidate_source_column,
            width=width,
            proven_merge_signatures=proven_merge_signatures,
        )
        for row in later_rows
    )
    return axis_not_extendable and candidate_rows_are_not_records


def _empty_axis_header_row(item: dict[str, Any]) -> int | None:
    evidence = item.get("structure_evidence")
    if not evidence or item["table"].get("matched_rule") != "L1-08":
        return None
    header_columns = set(evidence.get("headers_by_column", {}))
    if not header_columns:
        return None
    rows = _member_rows(item["members"])
    header_rows = [
        row_ordinal
        for row_ordinal, columns in rows.items()
        if header_columns.issubset(columns)
    ]
    return max(header_rows, default=None)


def _is_safe_empty_axis_context_component(
    candidate: dict[str, Any],
    *,
    projected: list[dict[str, Any]],
    empty_axis: dict[str, Any],
    empty_axis_header_row: int,
    source_regions: dict[tuple[int, int], dict[str, Any]],
) -> bool:
    """Recognize metadata-only regions before one proven trailing empty axis.

    A region is eligible only when it has no proven total or data row, its
    structural evidence is incomplete, and it is physically before the empty
    table header. This keeps uncertain standalone tables fail-closed.
    """

    candidate_is_before_axis = candidate["bbox"][2] < empty_axis_header_row
    candidate_is_after_axis = candidate["bbox"][0] > empty_axis_header_row
    if (
        candidate.get("positive_rule") is not None
        or candidate["table"].get("source_total_count") is not None
        or not candidate["members"]
        or not (candidate_is_before_axis or candidate_is_after_axis)
        or any(
            source_regions[source_region_key].get("unresolved_members")
            for source_region_key in candidate["source_components"]
        )
    ):
        return False

    evidence = candidate.get("structure_evidence")
    if evidence is None:
        has_after_context_anchor = any(
            other is not empty_axis
            and other["bbox"][0] > empty_axis_header_row
            and other.get("structure_evidence") is not None
            for other in projected
        )
        title_backed_empty_axis = (
            empty_axis.get("structure_evidence", {}).get(
                "record_axis_context_kind"
            )
            == "title_backed_multilevel"
        )
        return (
            candidate_is_before_axis
            or (
                candidate_is_after_axis
                and (
                    has_after_context_anchor
                    or title_backed_empty_axis
                )
            )
        ) and len({row for row, _column in candidate["members"]}) <= 1

    body_rows = set(evidence.get("body_row_ordinals", ()))
    headers = evidence.get("headers_by_column", {})
    header_paths = evidence.get("header_paths_by_column", {})
    has_complete_header = bool(headers) and all(
        header
        and not header.startswith("Column_")
        and header_paths.get(column)
        for column, header in headers.items()
    )
    if body_rows and has_complete_header:
        candidate_rows = {row for row, _column in candidate["members"]}
        interleaved_context = any(
            other is not candidate
            and other is not empty_axis
            and candidate_rows
            & {row for row, _column in other["members"]}
            and len(other["member_columns"]) > len(candidate["member_columns"])
            and other["bbox"][2] < empty_axis_header_row
            for other in projected
        )
        if not interleaved_context:
            return False
    # A complete header followed by multiple populated rows remains separate
    # unless row-interleaving evidence proves it is metadata for this object.
    return True


def _merge_unique_empty_axis_context(
    projected: list[dict[str, Any]],
    source_regions: dict[tuple[int, int], dict[str, Any]],
) -> list[dict[str, Any]]:
    """Attach pre-axis metadata to one uniquely proven empty table.

    The source regions remain individually represented in ``source_components``
    and the emitted event stream, so audit membership closure is preserved.
    Context-only regions are not emitted as rows because a verified empty table
    must retain zero projected records.
    """

    empty_axes = [
        item
        for item in projected
        if item["table"].get("matched_rule") == "L1-08"
        and item["table"].get("source_total_count") == 0
    ]
    if len(empty_axes) != 1:
        return projected
    empty_axis = empty_axes[0]
    header_row = _empty_axis_header_row(empty_axis)
    evidence = empty_axis.get("structure_evidence")
    if header_row is None or not evidence:
        return projected
    contexts = [
        candidate
        for candidate in projected
        if candidate is not empty_axis
        and _is_safe_empty_axis_context_component(
            candidate,
            projected=projected,
            empty_axis=empty_axis,
            empty_axis_header_row=header_row,
            source_regions=source_regions,
        )
    ]
    if not contexts:
        return projected

    for context in contexts:
        # Keep each source component's original set immutable. The union is
        # the output object's membership, not a rewrite of its source digest.
        empty_axis["members"] = empty_axis["members"] | context["members"]
        empty_axis["member_columns"] = (
            empty_axis["member_columns"] | context["member_columns"]
        )
        empty_axis["source_components"].update(context["source_components"])
        empty_axis["emitted_member_events"].extend(
            context["emitted_member_events"]
        )
        empty_axis["bbox"] = (
            min(empty_axis["bbox"][0], context["bbox"][0]),
            min(empty_axis["bbox"][1], context["bbox"][1]),
            max(empty_axis["bbox"][2], context["bbox"][2]),
            max(empty_axis["bbox"][3], context["bbox"][3]),
        )

    context_ids = {id(context) for context in contexts}
    return [item for item in projected if id(item) not in context_ids]


def _region_structure_evidence(parser, worksheet, region: dict[str, Any]) -> dict[str, Any] | None:
    region_worksheet, row_offset = _copy_structure_region(parser, worksheet, region)
    header_rows, populated_rows, _unresolved_rows = _complete_worksheet_rows(region_worksheet)
    if not header_rows or not populated_rows:
        return None
    context_preceded_empty_axis_proof = {}
    nonempty_structure, context_preceded_empty_structure = (
        _primary_record_axis_structures(
            parser,
            region_worksheet,
            header_rows,
            populated_rows,
            allow_context_preceded_empty_axis=(
                not _unresolved_rows and len(populated_rows) >= 2
            ),
            context_preceded_empty_axis_proof=context_preceded_empty_axis_proof,
        )
    )
    trailing_empty_candidate = _trailing_empty_record_axis_structure(
        parser,
        region_worksheet,
        header_rows,
        populated_rows,
        _unresolved_rows,
    )
    trailing_empty_structure = (
        trailing_empty_candidate
        if trailing_empty_candidate is not None
        and (
            nonempty_structure is None
            or any(
                not header or header.startswith("Column_")
                for header in nonempty_structure[0]
            )
        )
        else None
    )
    if trailing_empty_structure is not None:
        nonempty_structure = None
    preferred_empty_structure = _preferred_empty_record_axis_structure(
        context_preceded_empty_structure,
        trailing_empty_structure,
    )
    empty_structure = preferred_empty_structure or (
        None
        if nonempty_structure is not None
        else _empty_record_axis_structure(
            parser,
            region_worksheet,
            header_rows,
            populated_rows,
            _unresolved_rows,
        )
    )
    if empty_structure is not None:
        headers, header_paths, header_start, data_start = empty_structure
    else:
        headers, header_start, data_start = nonempty_structure or _parse_region_structure(
            parser,
            region_worksheet,
            header_rows,
        )
        header_paths = _header_paths_for_region(
            parser,
            region_worksheet,
            header_rows,
            header_start,
            data_start,
            expected_width=len(headers),
        )
    body_rows = [row for row in populated_rows if row > data_start]
    if not headers or (not body_rows and empty_structure is None):
        return None
    min_column = region["bbox"][1]
    if len(header_paths) != len(headers):
        return None

    merged_ranges = list(region_worksheet.merged_cells.ranges)
    record_axis_body_rows = _record_axis_body_rows(
        parser,
        region_worksheet,
        headers,
        data_start,
        body_rows,
        merged_ranges,
    )
    record_axis_evidence = _record_axis_evidence(
        headers,
        record_axis_body_rows,
        merged_ranges,
    )
    record_axis_context_kind = (
        "title_backed_multilevel"
        if context_preceded_empty_structure is not None
        and empty_structure == context_preceded_empty_structure
        and context_preceded_empty_axis_proof.get("title_backed_multilevel")
        else None
    )
    optional_parent_prefix_lengths = {}
    for column_offset, header_path in enumerate(header_paths):
        column_ordinal = column_offset + 1
        path_sources = []
        seen_sources = set()
        for row_index in range(header_start, data_start):
            row_ordinal = row_index + 1
            merged = next(
                (
                    candidate
                    for candidate in merged_ranges
                    if candidate.min_row <= row_ordinal <= candidate.max_row
                    and candidate.min_col <= column_ordinal <= candidate.max_col
                ),
                None,
            )
            if merged is not None:
                source = (
                    "merge",
                    merged.min_row,
                    merged.min_col,
                    merged.max_row,
                    merged.max_col,
                )
                value = region_worksheet.cell(merged.min_row, merged.min_col).value
                is_horizontal_parent = merged.max_col > merged.min_col
            else:
                source = ("cell", row_ordinal, column_ordinal)
                value = region_worksheet.cell(row_ordinal, column_ordinal).value
                is_horizontal_parent = False
            if source in seen_sources or parser._is_empty_value(value):
                continue
            rendered = str(value).strip()
            is_child_header = row_index > header_start and bool(path_sources)
            if parser._is_valid_header_part(rendered) or is_child_header:
                path_sources.append((rendered, is_horizontal_parent))
                seen_sources.add(source)

        prefix_length = 0
        if [segment for segment, _is_parent in path_sources] == header_path:
            for _segment, is_parent in path_sources:
                if not is_parent:
                    break
                prefix_length += 1
        optional_parent_prefix_lengths[min_column + column_offset] = prefix_length

    return {
        "headers_by_column": {
            min_column + offset: header for offset, header in enumerate(headers)
        },
        "header_paths_by_column": {
            min_column + offset: path for offset, path in enumerate(header_paths)
        },
        "body_row_ordinals": [row + row_offset for row in body_rows],
        "header_depth": data_start - header_start,
        "optional_parent_prefix_lengths_by_column": optional_parent_prefix_lengths,
        # Internal-only closure evidence; it is not part of the projection schema.
        "record_axis_evidence": record_axis_evidence,
        "record_axis_context_kind": record_axis_context_kind,
        "record_axis_source_column": min_column,
        "record_axis_width": len(headers),
    }


def _new_projected_item(
    *,
    parser,
    worksheet,
    region: dict[str, Any],
    source_region_key: tuple[int, int],
    table: dict[str, Any],
    rows: list[dict[str, Any]],
    positive_rule: str | None,
    structure_region: dict[str, Any] | None = None,
) -> dict[str, Any]:
    members = set(region["members"])
    structure_evidence = _region_structure_evidence(
        parser,
        worksheet,
        structure_region or region,
    )
    emitted_row_ordinals = {row["row_ordinal_int"] for row in rows}
    body_row_ordinals = (
        set(structure_evidence["body_row_ordinals"])
        if structure_evidence is not None
        else set()
    )
    context_row_ordinals = (
        {row for row, _column in members} - body_row_ordinals
        if structure_evidence is not None
        else set()
    )
    emitted_member_events = [
        coordinate
        for coordinate in members
        if coordinate[0] in emitted_row_ordinals
        or coordinate[0] in context_row_ordinals
    ]
    return {
        "table": table,
        "rows": rows,
        "worksheet_name": worksheet.title,
        "bbox": region["bbox"],
        "members": members,
        "member_columns": {column for _row, column in members},
        "structure_evidence": structure_evidence,
        "positive_rule": positive_rule,
        "source_components": {source_region_key: members},
        # This is an internal event stream. It must not be reconstructed from a
        # deduplicated set when continuation components are combined.
        "emitted_member_events": emitted_member_events,
        "proven_record_slots": [
            row["row_ordinal_int"]
            for row in rows
            if positive_rule is not None and row["row_role_kwd"] == "data"
        ],
    }


def _member_rows(members: set[tuple[int, int]]) -> dict[int, set[int]]:
    rows: dict[int, set[int]] = {}
    for row_ordinal, column_ordinal in members:
        rows.setdefault(row_ordinal, set()).add(column_ordinal)
    return rows


def _formula_is_unstable(
    *,
    worksheet,
    sheet_name: str,
    members: set[tuple[int, int]],
    formula_coordinates: set[tuple[int, int]],
    formula_values: dict[tuple[int, int], str],
    formula_cached_result_kinds: dict[tuple[int, int], str],
    formula_inventory_proven: bool,
) -> bool:
    if not formula_inventory_proven:
        return True
    local_formula_coordinates = members & formula_coordinates
    missing_formula_values = local_formula_coordinates - set(formula_values)
    for coordinate in missing_formula_values:
        if coordinate not in formula_cached_result_kinds:
            return True
        row_ordinal = coordinate[0]
        if not any(
            (row_ordinal, column_ordinal) in members
            and (row_ordinal, column_ordinal) not in local_formula_coordinates
            and worksheet.cell(row_ordinal, column_ordinal).value is not None
            and str(worksheet.cell(row_ordinal, column_ordinal).value).strip()
            for column_ordinal in sorted({column for row, column in members if row == row_ordinal})
        ):
            return True
    for coordinate in local_formula_coordinates:
        if coordinate not in formula_values:
            continue
        ranges, unresolved = _formula_reference_ranges(
            formula_values[coordinate],
            sheet_name,
        )
        if unresolved or not ranges:
            return True
        for min_row, min_column, max_row, max_column in ranges:
            references = {
                (row_ordinal, column_ordinal)
                for row_ordinal in range(min_row, max_row + 1)
                for column_ordinal in range(min_column, max_column + 1)
            }
            if not references.issubset(members):
                return True
    return bool(local_formula_coordinates & set(formula_values))


def _has_hidden_record_member(
    worksheet,
    members: set[tuple[int, int]],
    structure_evidence: dict[str, Any] | None,
) -> bool:
    from openpyxl.utils import get_column_letter

    member_columns = {column_ordinal for _row_ordinal, column_ordinal in members}
    candidate_rows = (
        set(structure_evidence["body_row_ordinals"])
        if structure_evidence
        else {row_ordinal for row_ordinal, _column_ordinal in members}
    )
    return any(worksheet.row_dimensions[row_ordinal].hidden for row_ordinal in candidate_rows) or any(
        worksheet.column_dimensions[get_column_letter(column_ordinal)].hidden
        for column_ordinal in member_columns
    )


def _is_matrix_layout(worksheet, members: set[tuple[int, int]]) -> bool:
    min_row, min_column, max_row, max_column = _region_bbox(members)
    if max_row - min_row < 2 or max_column - min_column < 2:
        return False
    body_rows = list(range(min_row + 1, max_row + 1))
    measure_columns = list(range(min_column + 1, max_column + 1))
    column_axis = {(min_row, column) for column in measure_columns}
    if not column_axis.issubset(members):
        return False
    column_axis_values = tuple(
        worksheet.cell(min_row, column).value for column in measure_columns
    )
    repeated_column_axis_rows = {
        row
        for row in body_rows
        if (row, min_column) not in members
        and all((row, column) in members for column in measure_columns)
        and tuple(worksheet.cell(row, column).value for column in measure_columns)
        == column_axis_values
    }
    record_rows = [row for row in body_rows if row not in repeated_column_axis_rows]
    row_axis = {(row, min_column) for row in record_rows}
    if len(record_rows) < 2 or not row_axis.issubset(members):
        return False
    if any(
        not any((row, column) in members for column in measure_columns)
        for row in record_rows
    ):
        return False

    row_axis_shapes = [
        _row_shape(
            [worksheet.cell(row, min_column).value],
            distinguish_text_digits=True,
        )[0]
        for row in record_rows
    ]
    row_axis_values = [
        str(worksheet.cell(row, min_column).value).strip()
        for row in record_rows
    ]
    corner_is_empty = (min_row, min_column) not in members
    if (
        not corner_is_empty
        and set(row_axis_shapes) != {"text"}
    ) or len(set(row_axis_values)) != len(row_axis_values):
        return False

    column_body_shapes = []
    for column in measure_columns:
        occupied_cells = [
            worksheet.cell(row, column)
            for row in record_rows
            if (row, column) in members
        ]
        if len(occupied_cells) < 2:
            return False
        non_formula_shapes = {
            _row_shape([cell.value], distinguish_text_digits=True)[0]
            for cell in occupied_cells
            if cell.data_type != "f" and cell.value is not None
        }
        if len(non_formula_shapes) > 1:
            return False
        column_body_shapes.append(non_formula_shapes)

    return bool(column_body_shapes) and any(
        isinstance(worksheet.cell(row, column).value, (bool, int, float))
        for row in record_rows
        for column in measure_columns
        if (row, column) in members
    )


def _has_mixed_aggregate_rows(
    *,
    sheet_name: str,
    members: set[tuple[int, int]],
    formula_values: dict[tuple[int, int], str],
) -> bool:
    member_rows = _member_rows(members)
    for (formula_row, formula_column), formula in formula_values.items():
        if (formula_row, formula_column) not in members:
            continue
        ranges, unresolved = _formula_reference_ranges(formula, sheet_name)
        if unresolved:
            continue
        for min_row, min_column, max_row, max_column in ranges:
            referenced_rows = list(range(min_row, max_row + 1))
            if (
                len(referenced_rows) < 2
                or max_row >= formula_row
                or formula_row != max_row + 1
                or formula_column < min_column
                or formula_column > max_column
            ):
                continue
            signatures = [member_rows.get(row, set()) for row in referenced_rows]
            if signatures and signatures[0] and all(signature == signatures[0] for signature in signatures):
                return True
    return False


def _has_unseparated_multiple_blocks(
    worksheet,
    members: set[tuple[int, int]],
) -> bool:
    member_rows = _member_rows(members)
    row_ordinals = sorted(member_rows)
    if len(row_ordinals) < 6:
        return False
    for repeated_index in range(3, len(row_ordinals) - 2):
        first_row = row_ordinals[0]
        repeated_row = row_ordinals[repeated_index]
        if repeated_row != row_ordinals[repeated_index - 1] + 1:
            continue
        if member_rows[first_row] != member_rows[repeated_row]:
            continue
        first_values = tuple(
            worksheet.cell(first_row, column).value for column in sorted(member_rows[first_row])
        )
        repeated_values = tuple(
            worksheet.cell(repeated_row, column).value
            for column in sorted(member_rows[repeated_row])
        )
        if first_values != repeated_values:
            continue
        before = [member_rows[row] for row in row_ordinals[1:repeated_index]]
        after = [member_rows[row] for row in row_ordinals[repeated_index + 1 :]]
        if len(before) >= 2 and len(after) >= 2 and len(set(map(frozenset, before + after))) == 1:
            return True
    return False


def _has_visual_only_boundary(
    worksheet,
    members: set[tuple[int, int]],
    structure_evidence: dict[str, Any] | None,
) -> bool:
    merged_coordinates = {
        (row_ordinal, column_ordinal)
        for merged in worksheet.merged_cells.ranges
        for row_ordinal in range(merged.min_row, merged.max_row + 1)
        for column_ordinal in range(merged.min_col, merged.max_col + 1)
    }
    if merged_coordinates & members:
        return False
    member_rows = _member_rows(members)

    def axis_implies_boundary(
        ordinals: list[int],
        slot_coordinates: list[list[tuple[int, int]]],
    ) -> bool:
        if len(ordinals) < 4 or any(
            right != left + 1 for left, right in zip(ordinals, ordinals[1:])
        ):
            return False
        relative_geometry = [
            tuple(
                (row - min(row for row, _column in coordinates), column - min(column for _row, column in coordinates))
                for row, column in coordinates
            )
            for coordinates in slot_coordinates
        ]
        if len(set(relative_geometry)) != 1:
            return False
        style_signatures = [
            tuple(worksheet.cell(row, column).style_id for row, column in coordinates)
            for coordinates in slot_coordinates
        ]
        if len(set(style_signatures)) < 2 or not any(
            style_id != 0 for signature in style_signatures for style_id in signature
        ):
            return False
        style_runs = []
        for signature in style_signatures:
            if not style_runs or style_runs[-1][0] != signature:
                style_runs.append([signature, 1])
            else:
                style_runs[-1][1] += 1
        if len(style_runs) < 2 or any(length < 2 for _signature, length in style_runs):
            return False
        style_boundaries = {
            index
            for index in range(1, len(style_signatures))
            if style_signatures[index - 1] != style_signatures[index]
        }
        value_shapes = [
            _row_shape(
                [worksheet.cell(row, column).value for row, column in coordinates],
                distinguish_text_digits=True,
            )
            for coordinates in slot_coordinates
        ]
        value_boundaries = {
            index
            for index in range(1, len(value_shapes))
            if value_shapes[index - 1] != value_shapes[index]
        }
        return style_boundaries.isdisjoint(value_boundaries)

    body_rows = (
        structure_evidence["body_row_ordinals"]
        if structure_evidence
        else sorted(member_rows)
    )
    if structure_evidence and len(body_rows) < 4:
        body_rows = sorted(member_rows)[1:]
    row_slots = [
        [(row, column) for column in sorted(member_rows[row])]
        for row in body_rows
        if row in member_rows
    ]
    if axis_implies_boundary(body_rows, row_slots):
        return True

    body_row_set = set(body_rows)
    columns = sorted({column for row, column in members if row in body_row_set})
    column_slots = [
        [(row, column) for row in body_rows if (row, column) in members]
        for column in columns
    ]
    return axis_implies_boundary(columns, column_slots)


def _region_negative_predicates(
    *,
    worksheet,
    sheet_name: str,
    item: dict[str, Any],
    siblings: list[dict[str, Any]],
    formula_coordinates: set[tuple[int, int]],
    formula_values: dict[tuple[int, int], str],
    formula_cached_result_kinds: dict[tuple[int, int], str],
    formula_inventory_proven: bool,
    partial_overlap: bool,
) -> dict[str, bool]:
    members = item["members"]
    member_rows = _member_rows(members)
    row_signatures = [frozenset(columns) for columns in member_rows.values() if columns]
    outside_record_axes = not any(
        sibling.get("positive_rule") is not None
        and _column_sets_intersect(item["member_columns"], sibling["member_columns"])
        for sibling in siblings
        if sibling is not item
    )
    no_repeated_slots = len(row_signatures) < 2 or len(set(row_signatures)) == len(row_signatures)
    aggregate_rows = _has_mixed_aggregate_rows(
        sheet_name=sheet_name,
        members=members,
        formula_values=formula_values,
    )
    return {
        "R1": outside_record_axes
        and no_repeated_slots
        and item.get("structure_evidence") is None,
        "R2": _has_hidden_record_member(
            worksheet,
            members,
            item.get("structure_evidence"),
        )
        or (
            not aggregate_rows
            and _formula_is_unstable(
                worksheet=worksheet,
                sheet_name=sheet_name,
                members=members,
                formula_coordinates=formula_coordinates,
                formula_values=formula_values,
                formula_cached_result_kinds=formula_cached_result_kinds,
                formula_inventory_proven=formula_inventory_proven,
            )
        ),
        "R3": _is_matrix_layout(worksheet, members),
        "R4": aggregate_rows,
        "R5": _has_unseparated_multiple_blocks(worksheet, members),
        "R6": partial_overlap,
        "R7": _has_visual_only_boundary(
            worksheet,
            members,
            item.get("structure_evidence"),
        ),
    }


def _canonical_union_context(
    main_rows: list[dict[str, Any]],
    headers_by_column: dict[int, str],
    *,
    entry_limit: int,
    value_bytes: int,
) -> str:
    existing = json.loads(main_rows[0]["table_context_list"]) if main_rows else []
    context_only = [item for item in existing if item["name"] != "field"]
    context = _context_with_headers(
        context_only,
        [headers_by_column[column] for column in sorted(headers_by_column)],
        entry_limit=entry_limit,
        value_bytes=value_bytes,
    )
    return json.dumps(context, ensure_ascii=False, separators=(",", ":"))


def _rekey_projected_item(
    item: dict[str, Any],
    *,
    table_ordinal: int,
    source_sha256: str,
    producer_generation_ref: str,
) -> None:
    table = item["table"]
    membership_sha256 = _region_membership_sha256(table["sheet_ordinal"], item["members"])
    table_ref = _table_ref(
        source_sha256,
        table["sheet_ordinal"],
        table_ordinal,
        membership_sha256,
    )
    table["table_ref"] = table_ref
    table["table_ordinal"] = table_ordinal
    for row in item["rows"]:
        row_ref = f"{table_ref}:{row['row_ordinal_int']}"
        row["table_ref_kwd"] = table_ref
        row["row_ref_kwd"] = row_ref
        row["id"] = "tsr_v1_" + _versioned_digest(
            "tabular-row-record/v1",
            producer_generation_ref,
            row_ref,
        )


def _finalize_table_manifest_evidence(item: dict[str, Any]) -> None:
    table = item["table"]
    rows = item["rows"]
    if rows:
        table["table_label"] = rows[0]["table_label_kwd"]
        table["table_context"] = json.loads(rows[0]["table_context_list"])
    else:
        table.setdefault("table_label", _sanitize_untrusted_text(item["worksheet_name"]))
        table.setdefault("table_context", [])

    evidence = item.get("structure_evidence")
    if evidence is None:
        table["ordered_columns"] = []
        return
    row_fields = [json.loads(row["ordered_fields_list"]) for row in rows]
    record_axis_columns = {
        column
        for row, column in item["members"]
        if row in set(item["proven_record_slots"])
    }
    evidence_columns = sorted(evidence["headers_by_column"])
    if rows and record_axis_columns:
        # Preserve leading structural columns that have no value in a sparse
        # record row, while stopping before disjoint trailing sidecars.
        evidence_columns = [
            absolute_column
            for absolute_column in evidence_columns
            if absolute_column <= max(record_axis_columns)
        ]
    if evidence_columns:
        # Manifest ordinals are dense over the proven table axis, while field
        # ordinals retain their source-column coordinates. Fill only structural
        # columns inside that axis; disjoint sidecars remain excluded.
        axis_start = min(evidence_columns)
        axis_end = max(evidence_columns)
        evidence_columns = [
            absolute_column
            for absolute_column in range(axis_start, axis_end + 1)
            if absolute_column in evidence["headers_by_column"]
        ]
    column_ordinals = {
        absolute_column: column_ordinal
        for column_ordinal, absolute_column in enumerate(evidence_columns, start=1)
    }
    ordered_columns = []
    for absolute_column in evidence_columns:
        ordered_columns.append(
            {
                "column_id": f"col_v1:{table['sheet_ordinal']}:{absolute_column}",
                "column_ordinal": column_ordinals[absolute_column],
                "header_path": list(evidence["header_paths_by_column"][absolute_column]),
                "name": evidence["headers_by_column"][absolute_column],
            }
        )
    columns_by_id = {column["column_id"]: column for column in ordered_columns}
    row_evidence_matches = all(
        all(
            field.get("header_path")
            and (column := columns_by_id.get(field.get("column_id"))) is not None
            and all(
                field.get(key) == column[key]
                for key in ("column_id", "column_ordinal", "header_path", "name")
            )
            for field in fields
        )
        for fields in row_fields
    )
    has_complete_paths = bool(ordered_columns) and all(
        column["header_path"] for column in ordered_columns
    )
    if not has_complete_paths or not row_evidence_matches:
        table["ordered_columns"] = []
        if table.get("enumeration_status") == "supported_complete":
            _clear_complete_decision(table, rows, "R8")
        return
    table["ordered_columns"] = ordered_columns


def _continuation_record_rows(item: dict[str, Any]) -> list[dict[str, Any]]:
    data_rows = [row for row in item["rows"] if row["row_role_kwd"] == "data"]
    if data_rows:
        if any(row["row_role_kwd"] == "unknown" for row in item["rows"]):
            return []
        return data_rows
    if item["table"]["data_row_count"] == 0:
        return list(item["rows"])
    return []


def _continuation_rows_match_proven_axis(
    *,
    parser,
    worksheet,
    main: dict[str, Any],
    continuation: dict[str, Any],
    continuation_rows: list[dict[str, Any]],
) -> bool:
    main_evidence = main.get("structure_evidence")
    if not main_evidence:
        return False
    shared_columns = sorted(main["member_columns"] & continuation["member_columns"])
    if not shared_columns:
        return False
    merged_ranges = list(worksheet.merged_cells.ranges)

    def shape(row_ordinal: int) -> tuple[tuple[str, ...], ...]:
        return tuple(
            _record_axis_value_shape(
                _cell_value(parser, worksheet, row_ordinal, column, merged_ranges)
            )
            for column in shared_columns
        )

    proven_shapes = {
        shape(row["row_ordinal_int"])
        for row in main["rows"]
        if row["row_role_kwd"] == "data"
        and row["row_ordinal_int"] in main_evidence["body_row_ordinals"]
    }
    shapes_match = bool(proven_shapes) and all(
        shape(row["row_ordinal_int"]) in proven_shapes
        for row in continuation_rows
    )
    if not shapes_match or len(continuation_rows) != 1:
        return shapes_match

    anchor_column = shared_columns[0]
    main_anchor_shapes = {
        _text_structure(
            _cell_value(
                parser,
                worksheet,
                row["row_ordinal_int"],
                anchor_column,
                merged_ranges,
            )
        )
        for row in main["rows"]
        if row["row_role_kwd"] == "data"
        and row["row_ordinal_int"] in main_evidence["body_row_ordinals"]
    }
    continuation_anchor_shape = _text_structure(
        _cell_value(
            parser,
            worksheet,
            continuation_rows[0]["row_ordinal_int"],
            anchor_column,
            merged_ranges,
        )
    )
    return (
        main_anchor_shapes == {continuation_anchor_shape}
        and any(kind in {"digit", "symbol"} for kind in continuation_anchor_shape)
    )


def _repeated_form_header_union(
    main_evidence: dict[str, Any],
    continuation_evidence: dict[str, Any],
) -> dict[str, dict[int, Any]] | None:
    main_paths = main_evidence["header_paths_by_column"]
    continuation_paths = continuation_evidence["header_paths_by_column"]
    if set(main_paths) != set(continuation_paths):
        return None

    main_optional = main_evidence.get(
        "optional_parent_prefix_lengths_by_column",
        {},
    )
    continuation_optional = continuation_evidence.get(
        "optional_parent_prefix_lengths_by_column",
        {},
    )
    headers_by_column = {}
    header_paths_by_column = {}
    optional_parent_prefix_lengths_by_column = {}
    largest_omitted_prefix = 0
    for column in sorted(main_paths):
        main_path = list(main_paths[column])
        continuation_path = list(continuation_paths[column])
        if not main_path or not continuation_path or main_path[-1] != continuation_path[-1]:
            return None

        if main_path == continuation_path:
            use_main = main_optional.get(column, 0) >= continuation_optional.get(column, 0)
        elif (
            len(main_path) > len(continuation_path)
            and main_path[-len(continuation_path) :] == continuation_path
            and len(main_path) - len(continuation_path) <= main_optional.get(column, 0)
        ):
            use_main = True
            largest_omitted_prefix = max(
                largest_omitted_prefix,
                len(main_path) - len(continuation_path),
            )
        elif (
            len(continuation_path) > len(main_path)
            and continuation_path[-len(main_path) :] == main_path
            and len(continuation_path) - len(main_path)
            <= continuation_optional.get(column, 0)
        ):
            use_main = False
            largest_omitted_prefix = max(
                largest_omitted_prefix,
                len(continuation_path) - len(main_path),
            )
        else:
            return None

        selected_evidence = main_evidence if use_main else continuation_evidence
        selected_optional = main_optional if use_main else continuation_optional
        headers_by_column[column] = selected_evidence["headers_by_column"][column]
        header_paths_by_column[column] = list(
            selected_evidence["header_paths_by_column"][column]
        )
        optional_parent_prefix_lengths_by_column[column] = selected_optional.get(column, 0)

    if (
        main_evidence["header_depth"] != continuation_evidence["header_depth"]
        and (
            largest_omitted_prefix == 0
            or abs(
                main_evidence["header_depth"]
                - continuation_evidence["header_depth"]
            )
            > largest_omitted_prefix
        )
    ):
        return None
    return {
        "headers_by_column": headers_by_column,
        "header_paths_by_column": header_paths_by_column,
        "optional_parent_prefix_lengths_by_column": (
            optional_parent_prefix_lengths_by_column
        ),
    }


def _repeated_form_identity_matches(
    main: dict[str, Any],
    continuation: dict[str, Any],
) -> bool:
    """Prove that two complete regions are repeated pages of one form.

    A repeated column header alone is not enough: two independent vertical
    tables can legitimately reuse the same columns. A non-field context is
    the source-backed form identity that distinguishes a repeated page from
    that layout, while the normalized header paths bind the record schema.
    """

    if main["table"].get("table_label") != continuation["table"].get("table_label"):
        return False
    main_evidence = main.get("structure_evidence")
    continuation_evidence = continuation.get("structure_evidence")
    if not main_evidence or not continuation_evidence:
        return False
    if _repeated_form_header_union(main_evidence, continuation_evidence) is None:
        return False

    def context_identity(
        item: dict[str, Any],
    ) -> tuple[frozenset[str], frozenset[str]]:
        rows = item.get("rows") or []
        if not rows:
            return frozenset(), frozenset()
        try:
            context = json.loads(rows[0]["table_context_list"])
        except (KeyError, TypeError, json.JSONDecodeError):
            return frozenset(), frozenset()
        table_label = item["table"].get("table_label") or item.get("worksheet_name")
        paired_names = frozenset(
            entry["name"]
            for entry in context
            if isinstance(entry, dict)
            and isinstance(entry.get("name"), str)
            and isinstance(entry.get("value"), str)
            and entry["name"] not in {"context", "field"}
        )
        labels = paired_names | frozenset(
            entry["value"]
            for entry in context
            if isinstance(entry, dict)
            and entry.get("name") == "context"
            and isinstance(entry.get("value"), str)
            and entry["value"] != table_label
        )
        return paired_names, labels

    main_paired, main_labels = context_identity(main)
    continuation_paired, continuation_labels = context_identity(continuation)
    paired_identity_matches = bool(main_paired and continuation_paired) and (
        main_paired.issubset(continuation_paired)
        or continuation_paired.issubset(main_paired)
    )
    label_identity_matches = len(main_labels & continuation_labels) >= 2
    return paired_identity_matches or label_identity_matches


def _source_row_signature(
    parser,
    worksheet,
    row_ordinal: int,
    columns: list[int],
    merged_ranges,
) -> tuple[tuple[int, str], ...]:
    return tuple(
        (column, _sanitize_untrusted_text(value))
        for column in columns
        if (
            value := _cell_value(
                parser,
                worksheet,
                row_ordinal,
                column,
                merged_ranges,
            )
        )
        is not None
        and str(value).strip()
    )


def _source_row_occupied_columns(
    parser,
    worksheet,
    row_ordinal: int,
    columns: list[int],
    merged_ranges,
) -> frozenset[int]:
    return frozenset(
        column
        for column in columns
        if (
            value := _cell_value(
                parser,
                worksheet,
                row_ordinal,
                column,
                merged_ranges,
            )
        )
        is not None
        and str(value).strip()
    )


def _is_wide_structural_note_row(
    parser,
    worksheet,
    row_ordinal: int,
    columns: list[int],
    merged_ranges,
) -> bool:
    signature = _source_row_signature(
        parser,
        worksheet,
        row_ordinal,
        columns,
        merged_ranges,
    )
    if len({value for _column, value in signature}) != 1:
        return False
    column_set = set(columns)
    return any(
        merged.min_row <= row_ordinal <= merged.max_row
        and len(
            column_set
            & set(range(merged.min_col, merged.max_col + 1))
        )
        >= max(2, len(columns) // 2)
        for merged in merged_ranges
    )


def _preceding_headerless_record_rows(
    *,
    parser,
    worksheet,
    main: dict[str, Any],
    predecessor: dict[str, Any],
) -> list[dict[str, Any]]:
    if predecessor["bbox"][2] >= main["bbox"][0]:
        return []
    main_evidence = main.get("structure_evidence")
    main_data_rows = [
        row
        for row in main["rows"]
        if row["row_role_kwd"] == "data"
    ]
    if not main_evidence or len(main_data_rows) < 2:
        return []
    if predecessor["member_columns"] != main["member_columns"]:
        return []
    predecessor_evidence = predecessor.get("structure_evidence")
    if (
        predecessor["table"].get("source_total_count") is not None
        and predecessor_evidence is not None
        and _repeated_form_header_union(predecessor_evidence, main_evidence)
        is not None
    ):
        return []

    columns = sorted(main["member_columns"])
    merged_ranges = list(worksheet.merged_cells.ranges)
    proven_patterns = {
        _source_row_occupied_columns(
            parser,
            worksheet,
            row["row_ordinal_int"],
            columns,
            merged_ranges,
        )
        for row in main_data_rows
    }
    predecessor_row_ordinals = sorted(
        {row for row, _column in predecessor["members"]}
    )
    candidate_ordinals = [
        row_ordinal
        for row_ordinal in predecessor_row_ordinals
        if _source_row_occupied_columns(
            parser,
            worksheet,
            row_ordinal,
            columns,
            merged_ranges,
        )
        in proven_patterns
    ]
    if len(candidate_ordinals) < 2:
        return []

    first_candidate = candidate_ordinals[0]
    last_candidate = candidate_ordinals[-1]
    main_first_record = min(row["row_ordinal_int"] for row in main_data_rows)
    main_context_signatures = {
        signature
        for row_ordinal in sorted({row for row, _column in main["members"]})
        if row_ordinal < main_first_record
        and (
            signature := _source_row_signature(
                parser,
                worksheet,
                row_ordinal,
                columns,
                merged_ranges,
            )
        )
    }
    predecessor_context_rows = {
        row_ordinal
        for row_ordinal in predecessor_row_ordinals
        if row_ordinal < first_candidate
        and _source_row_signature(
            parser,
            worksheet,
            row_ordinal,
            columns,
            merged_ranges,
        )
        in main_context_signatures
    }
    matched_context_signatures = {
        _source_row_signature(
            parser,
            worksheet,
            row_ordinal,
            columns,
            merged_ranges,
        )
        for row_ordinal in predecessor_context_rows
    }
    if len(matched_context_signatures) < 2:
        return []
    if any(
        row_ordinal not in predecessor_context_rows
        for row_ordinal in predecessor_row_ordinals
        if row_ordinal < first_candidate
    ):
        return []
    if any(
        row_ordinal not in candidate_ordinals
        for row_ordinal in predecessor_row_ordinals
        if first_candidate <= row_ordinal <= last_candidate
    ):
        return []
    if any(
        not _is_wide_structural_note_row(
            parser,
            worksheet,
            row_ordinal,
            columns,
            merged_ranges,
        )
        for row_ordinal in predecessor_row_ordinals
        if row_ordinal > last_candidate
    ):
        return []

    rows_by_ordinal = {
        row["row_ordinal_int"]: row
        for row in predecessor["rows"]
    }
    if any(row_ordinal not in rows_by_ordinal for row_ordinal in candidate_ordinals):
        return []
    return [rows_by_ordinal[row_ordinal] for row_ordinal in candidate_ordinals]


def _merge_continuation_pair(
    *,
    parser,
    worksheet,
    main: dict[str, Any],
    continuation: dict[str, Any],
    source_sha256: str,
    producer_generation_ref: str,
    table_context_entry_limit: int,
    table_context_value_bytes: int,
) -> dict[str, Any] | None:
    main_total = main["table"]["source_total_count"]
    continuation_precedes = continuation["bbox"][2] < main["bbox"][0]
    if main_total is None or (
        not continuation_precedes
        and main["bbox"][2] >= continuation["bbox"][0]
    ):
        return None
    main_columns = main["member_columns"]
    continuation_columns = continuation["member_columns"]
    if not _column_sets_are_nested(main_columns, continuation_columns):
        return None
    repeated_form = (
        not continuation_precedes
        and
        continuation["table"]["source_total_count"] is not None
        and main_columns == continuation_columns
        and _repeated_form_identity_matches(main, continuation)
    )
    if (
        not continuation_precedes
        and
        continuation["table"]["source_total_count"] is not None
        and main_columns == continuation_columns
        and not repeated_form
    ):
        return None
    continuation_rows = (
        _preceding_headerless_record_rows(
            parser=parser,
            worksheet=worksheet,
            main=main,
            predecessor=continuation,
        )
        if continuation_precedes
        else _continuation_record_rows(continuation)
    )
    if not continuation_rows or len(continuation_columns) < 2:
        return None

    main_evidence = main.get("structure_evidence")
    if not main_evidence:
        return None
    headers_by_column = dict(main_evidence["headers_by_column"])
    header_paths_by_column = dict(main_evidence["header_paths_by_column"])
    optional_parent_prefix_lengths_by_column = dict(
        main_evidence.get("optional_parent_prefix_lengths_by_column", {})
    )
    continuation_evidence = continuation.get("structure_evidence")
    if (
        not continuation_precedes
        and continuation["table"]["data_row_count"] == 0
        and continuation_evidence is not None
    ):
        return None
    named_continuation = (
        not continuation_precedes
        and
        continuation["table"]["data_row_count"] == len(continuation_rows)
        and continuation_evidence is not None
        and set(continuation_evidence["body_row_ordinals"])
        == {
            row["row_ordinal_int"]
            for row in continuation_rows
        }
        | {
            row["row_ordinal_int"]
            for row in continuation["rows"]
            if row["row_role_kwd"] == "note"
        }
    )
    if (
        named_continuation
        and len(continuation_rows) > 1
        and continuation.get("positive_rule") is None
    ):
        return None
    if named_continuation:
        continuation_headers = continuation_evidence["headers_by_column"]
        continuation_header_paths = continuation_evidence["header_paths_by_column"]
        if repeated_form:
            header_union = _repeated_form_header_union(
                main_evidence,
                continuation_evidence,
            )
            if header_union is None:
                return None
            headers_by_column = header_union["headers_by_column"]
            header_paths_by_column = header_union["header_paths_by_column"]
            optional_parent_prefix_lengths_by_column = header_union[
                "optional_parent_prefix_lengths_by_column"
            ]
        else:
            if any(
                column in headers_by_column
                and (
                    headers_by_column[column] != continuation_headers.get(column)
                    or header_paths_by_column[column]
                    != continuation_header_paths.get(column)
                )
                for column in main_columns & continuation_columns
            ):
                return None
            headers_by_column.update(continuation_headers)
            header_paths_by_column.update(continuation_header_paths)
            optional_parent_prefix_lengths_by_column.update(
                continuation_evidence.get(
                    "optional_parent_prefix_lengths_by_column",
                    {},
                )
            )
    elif not continuation_precedes and not _continuation_rows_match_proven_axis(
        parser=parser,
        worksheet=worksheet,
        main=main,
        continuation=continuation,
        continuation_rows=continuation_rows,
    ):
        return None

    missing_names = {
        column
        for column in continuation_columns
        if column not in headers_by_column
        or not headers_by_column[column]
        or headers_by_column[column].startswith("Column_")
    }
    is_unnamed_superset = bool(continuation_columns - main_columns) and bool(missing_names)
    if missing_names and not is_unnamed_superset:
        return None
    if (
        not continuation_precedes
        and not named_continuation
        and len(continuation_rows) > 1
        and not _members_prove_repeated_axis(
        continuation["members"]
        )
    ):
        return None

    table_ref = _table_ref(
        source_sha256,
        main["table"]["sheet_ordinal"],
        main["table"]["table_ordinal"],
        _region_membership_sha256(
            main["table"]["sheet_ordinal"],
            main["members"] | continuation["members"],
        ),
    )
    context = _canonical_union_context(
        main["rows"],
        headers_by_column,
        entry_limit=table_context_entry_limit,
        value_bytes=table_context_value_bytes,
    )
    rows = [dict(row) for row in main["rows"]]
    for source_row in continuation_rows:
        row = dict(source_row)
        if not is_unnamed_superset:
            row["row_role_kwd"] = "data"
        rows.append(row)

    members = main["members"] | continuation["members"]
    merged_ranges = list(worksheet.merged_cells.ranges)
    union_columns = sorted(headers_by_column)
    ordinal_by_column = {
        column: column_ordinal
        for column_ordinal, column in enumerate(union_columns, start=1)
    }
    for row in rows:
        columns = [
            column
            for column in union_columns
            if (row["row_ordinal_int"], column) in members
        ]
        values = [
            _cell_value(
                parser,
                worksheet,
                row["row_ordinal_int"],
                column,
                merged_ranges,
            )
            for column in columns
        ]
        row["ordered_fields_list"] = json.dumps(
            _ordered_fields(
                [headers_by_column[column] for column in columns],
                values,
                note=row["row_role_kwd"] == "note",
                sheet_ordinal=main["table"]["sheet_ordinal"],
                header_paths=[header_paths_by_column[column] for column in columns],
                column_ordinals=[ordinal_by_column[column] for column in columns],
                absolute_column_ordinals=columns,
                source_anchors=[
                    _source_cell_anchor(
                        row["row_ordinal_int"],
                        column,
                        merged_ranges,
                    )
                    for column in columns
                ],
            ),
            ensure_ascii=False,
            separators=(",", ":"),
        )

    rows.sort(key=lambda row: row["row_ordinal_int"])
    data_index = 0
    for row in rows:
        row["table_ref_kwd"] = table_ref
        row["table_context_list"] = context
        if row["row_role_kwd"] == "data":
            data_index += 1
            row["data_row_index_int"] = data_index
        else:
            row["data_row_index_int"] = None

    table = dict(main["table"])
    table.update(
        {
            "table_ref": table_ref,
            "row_count": len(rows),
            "data_row_count": data_index,
            "source_total_count": None if is_unnamed_superset else data_index,
        }
    )
    _apply_enumeration_decision(table, "D1" if is_unnamed_superset else "L1-02")
    for row in rows:
        row["source_total_count_int"] = table["source_total_count"]
        row_ref = f"{table_ref}:{row['row_ordinal_int']}"
        row["row_ref_kwd"] = row_ref
        row["id"] = "tsr_v1_" + _versioned_digest(
            "tabular-row-record/v1",
            producer_generation_ref,
            row_ref,
        )

    structure_evidence = {
        "headers_by_column": headers_by_column,
        "header_paths_by_column": header_paths_by_column,
        "body_row_ordinals": sorted(
            set(main_evidence["body_row_ordinals"])
            | {row["row_ordinal_int"] for row in continuation_rows}
        ),
        "header_depth": main_evidence["header_depth"],
        "optional_parent_prefix_lengths_by_column": (
            optional_parent_prefix_lengths_by_column
        ),
    }
    source_components = {
        **main["source_components"],
        **continuation["source_components"],
    }
    return {
        "table": table,
        "rows": rows,
        "worksheet_name": main["worksheet_name"],
        "bbox": (
            min(main["bbox"][0], continuation["bbox"][0]),
            min(main["bbox"][1], continuation["bbox"][1]),
            max(main["bbox"][2], continuation["bbox"][2]),
            max(main["bbox"][3], continuation["bbox"][3]),
        ),
        "members": members,
        "member_columns": {column for _row, column in members},
        "structure_evidence": structure_evidence,
        "positive_rule": None if is_unnamed_superset else "L1-02",
        "source_components": source_components,
        "emitted_member_events": [
            *main["emitted_member_events"],
            *continuation["emitted_member_events"],
        ],
        "proven_record_slots": (
            []
            if is_unnamed_superset
            else sorted(
                [
                    *main["proven_record_slots"],
                    *(row["row_ordinal_int"] for row in continuation_rows),
                ]
            )
        ),
    }


def _apply_projection_invariants(
    projected: list[dict[str, Any]],
    source_regions: dict[tuple[int, int], dict[str, Any]],
) -> list[dict[str, Any]]:
    assignments = Counter(
        source_region_key
        for item in projected
        for source_region_key in item["source_components"]
    )

    for item in projected:
        expected_members = set().union(*item["source_components"].values())
        expected_events = Counter(expected_members)
        actual_events = Counter(item["emitted_member_events"])
        membership_closed = (
            item["members"] == expected_members
            and actual_events == expected_events
            and all(assignments[key] == 1 for key in item["source_components"])
        )
        if not membership_closed:
            item["positive_rule"] = None
            _clear_complete_decision(item["table"], item["rows"], "D2")
            continue

        if item["positive_rule"] is None:
            continue
        data_rows = [row for row in item["rows"] if row["row_role_kwd"] == "data"]
        proven_slots = item["proven_record_slots"]
        record_count_closed = (
            len(proven_slots) == len(set(proven_slots))
            and len(proven_slots) == len(data_rows)
            and proven_slots == [row["row_ordinal_int"] for row in data_rows]
            and item["table"]["data_row_count"] == len(data_rows)
            and item["table"]["source_total_count"] == len(proven_slots)
            and [row["data_row_index_int"] for row in data_rows]
            == list(range(1, len(data_rows) + 1))
        )
        if record_count_closed:
            continue
        item["positive_rule"] = None
        item["table"]["data_row_count"] = len(data_rows)
        for data_index, row in enumerate(data_rows, start=1):
            row["data_row_index_int"] = data_index
        _clear_complete_decision(item["table"], item["rows"], "D3")

    defects = []
    for source_region_key, region in source_regions.items():
        if assignments[source_region_key] != 0:
            continue
        sheet_ordinal, source_region_ordinal = source_region_key
        status, reason = ENUMERATION_DECISIONS["D4"]
        defects.append(
            {
                "row_kind": "defect_tombstone",
                "table_ref": None,
                "sheet_ordinal": sheet_ordinal,
                "source_region_ordinal": source_region_ordinal,
                "membership_sha256": region["membership_sha256"],
                "enumeration_status": status,
                "enumeration_reason": reason,
                "matched_rule": "D4",
            }
        )
    return defects


def _validate_tabular_structure_producer_audit(audit: dict[str, Any]) -> None:
    expected_fields = {
        "version",
        "producer_generation_ref",
        "enumeration_rule_version",
        "source_sha256",
        "source_regions",
        "output_objects",
        "defects",
    }
    if not isinstance(audit, dict) or set(audit) != expected_fields:
        raise ValueError("producer audit does not match the fixed schema")
    if audit["version"] != "tabular-structure-producer-audit/v1":
        raise ValueError("unsupported producer audit version")
    _validate_generation_ref(audit["producer_generation_ref"])
    if audit["enumeration_rule_version"] != ENUMERATION_RULE_VERSION:
        raise ValueError("unsupported producer audit enumeration rule version")
    if not re.fullmatch(r"[0-9a-f]{64}", str(audit["source_sha256"])):
        raise ValueError("producer audit source SHA-256 is invalid")
    if not isinstance(audit["defects"], list):
        raise ValueError("producer audit defects must be a list")

    source_regions = audit["source_regions"]
    output_objects = audit["output_objects"]
    if not isinstance(source_regions, list) or not isinstance(output_objects, list):
        raise ValueError("producer audit source/output evidence must be lists")
    source_by_ref = {}
    source_ordinals_by_sheet: dict[int, list[int]] = {}
    for source in source_regions:
        expected_source_fields = {
            "source_region_ref",
            "worksheet_ordinal",
            "bbox",
            "row_count",
            "column_count",
            "membership_sha256",
            "member_count",
            "member_coordinate_set",
            "assigned_object_ref",
            "assignment_count",
        }
        if not isinstance(source, dict) or set(source) != expected_source_fields:
            raise ValueError("producer audit source region does not match the fixed schema")
        source_ref = source["source_region_ref"]
        if not isinstance(source_ref, str) or source_ref in source_by_ref:
            raise ValueError("producer audit source region references are not unique")
        source_ref_match = re.fullmatch(r"([1-9][0-9]*):([1-9][0-9]*)", source_ref)
        if (
            source_ref_match is None
            or int(source_ref_match.group(1)) != source["worksheet_ordinal"]
        ):
            raise ValueError("producer audit source region reference is inconsistent")
        source_ordinals_by_sheet.setdefault(source["worksheet_ordinal"], []).append(
            int(source_ref_match.group(2))
        )
        source_by_ref[source_ref] = source
        if (
            not isinstance(source["assignment_count"], int)
            or isinstance(source["assignment_count"], bool)
            or source["assignment_count"] < 0
        ):
            raise ValueError("producer audit source assignment count is invalid")
        if source["assigned_object_ref"] is not None and not isinstance(
            source["assigned_object_ref"], str
        ):
            raise ValueError("producer audit source object reference is invalid")
        if not re.fullmatch(r"[0-9a-f]{64}", str(source["membership_sha256"])):
            raise ValueError("producer audit source membership SHA-256 is invalid")
        member_coordinates = source["member_coordinate_set"]
        if (
            not isinstance(member_coordinates, list)
            or any(not isinstance(value, str) for value in member_coordinates)
            or member_coordinates
            != sorted(set(member_coordinates), key=_audit_coordinate_key)
        ):
            raise ValueError("producer audit source members are not a deterministic set")
        expected_membership = hashlib.sha256(
            "\n".join(member_coordinates).encode("ascii")
        ).hexdigest()
        if (
            source["membership_sha256"] != expected_membership
            or source["member_count"] != len(member_coordinates)
        ):
            raise ValueError("producer audit source membership closure is invalid")
        parsed_members = [_audit_coordinate_key(value) for value in member_coordinates]
        member_rows = {row for sheet, row, _column in parsed_members}
        member_columns = {column for sheet, _row, column in parsed_members}
        if (
            not parsed_members
            or any(sheet != source["worksheet_ordinal"] for sheet, _row, _column in parsed_members)
            or source["bbox"]
            != [min(member_rows), min(member_columns), max(member_rows), max(member_columns)]
            or source["row_count"] != len(member_rows)
            or source["column_count"] != len(member_columns)
        ):
            raise ValueError("producer audit source geometry is inconsistent")
    if any(
        ordinals != list(range(1, len(ordinals) + 1))
        for ordinals in source_ordinals_by_sheet.values()
    ):
        raise ValueError("producer audit source region references are not contiguous")

    output_refs = set()
    real_assignments = Counter()
    object_ordinals_by_sheet: Counter[int] = Counter()
    tombstone_components = []
    for output in output_objects:
        required_output_fields = {
            "row_kind",
            "object_ref",
            "worksheet_ordinal",
            "component_region_refs",
            "component_membership_sha256_list",
            "union_membership_sha256",
            "union_member_count",
            "emitted_member_coordinate_multiset",
            "emitted_cell_multiset_sha256",
            "emitted_member_occurrence_count",
            "member_max_ingest_count",
            "record_slot_sha256",
            "proven_record_slot_count",
            "record_slot_coordinate_sets",
            "emitted_data_row_ordinals",
            "enumeration_rule_version",
            "structure_generation_ref",
            "table_ref",
            "identity_validation_status",
            "matched_rule",
            "decision_chain_stop",
            "enumeration_status",
            "enumeration_reason",
            "covered_count",
            "source_total_count",
        }
        if not isinstance(output, dict) or set(output) != required_output_fields:
            raise ValueError("producer audit output object does not match the fixed schema")
        if output["row_kind"] == "object":
            if not isinstance(output["object_ref"], str) or output["object_ref"] in output_refs:
                raise ValueError("producer audit object references are not unique")
            output_refs.add(output["object_ref"])
            if output["table_ref"] != output["object_ref"]:
                raise ValueError("producer audit object table identity is inconsistent")
            object_ordinals_by_sheet[output["worksheet_ordinal"]] += 1
            object_ordinal = object_ordinals_by_sheet[output["worksheet_ordinal"]]
        elif output["row_kind"] == "defect_tombstone":
            if output["object_ref"] is not None or output["table_ref"] is not None:
                raise ValueError("producer audit tombstone cannot fabricate an object identity")
            object_ordinal = None
        else:
            raise ValueError("producer audit output row kind is invalid")
        if output["structure_generation_ref"] != audit["producer_generation_ref"]:
            raise ValueError("producer audit output generation is inconsistent")
        if output["enumeration_rule_version"] != audit["enumeration_rule_version"]:
            raise ValueError("producer audit output rule version is inconsistent")
        if output["identity_validation_status"] != "pending_independent_validation":
            raise ValueError("producer audit identity validation status is invalid")
        component_refs = output["component_region_refs"]
        if not isinstance(component_refs, list) or not component_refs:
            raise ValueError("producer audit output components are missing")
        if any(ref not in source_by_ref for ref in component_refs):
            raise ValueError("producer audit output references an unknown source region")
        if component_refs != sorted(
            set(component_refs),
            key=lambda ref: tuple(int(part) for part in ref.split(":")),
        ):
            raise ValueError("producer audit component references are not numerically ordered")
        if output["row_kind"] == "defect_tombstone" and output["matched_rule"] != "D4":
            raise ValueError("producer audit tombstone rule is invalid")
        decision = ENUMERATION_DECISIONS.get(output["matched_rule"])
        if decision != (output["enumeration_status"], output["enumeration_reason"]):
            raise ValueError("producer audit output decision is invalid")
        if output["decision_chain_stop"] != output["matched_rule"]:
            raise ValueError("producer audit decision chain stop is inconsistent")
        if (output["enumeration_status"] == "supported_complete") != (
            output["source_total_count"] is not None
        ):
            raise ValueError("producer audit decision conflicts with source total")

        component_sources = [source_by_ref[ref] for ref in component_refs]
        if any(
            source["worksheet_ordinal"] != output["worksheet_ordinal"]
            for source in component_sources
        ):
            raise ValueError("producer audit output worksheet does not match its components")
        if output["component_membership_sha256_list"] != [
            source["membership_sha256"] for source in component_sources
        ]:
            raise ValueError("producer audit component membership closure is invalid")
        component_coordinate_sets = [
            set(source["member_coordinate_set"]) for source in component_sources
        ]
        if output["enumeration_status"] == "supported_complete" and any(
            left & right
            for index, left in enumerate(component_coordinate_sets)
            for right in component_coordinate_sets[index + 1 :]
        ):
            raise ValueError("producer audit complete component memberships are not disjoint")
        union_coordinates = sorted(
            {
                coordinate
                for coordinates in component_coordinate_sets
                for coordinate in coordinates
            },
            key=_audit_coordinate_key,
        )
        union_digest = hashlib.sha256("\n".join(union_coordinates).encode("ascii")).hexdigest()
        if (
            output["union_membership_sha256"] != union_digest
            or output["union_member_count"] != len(union_coordinates)
        ):
            raise ValueError("producer audit union membership closure is invalid")
        if output["row_kind"] == "object" and output["table_ref"] != _table_ref(
            audit["source_sha256"],
            output["worksheet_ordinal"],
            object_ordinal,
            union_digest,
        ):
            raise ValueError("producer audit table identity is inconsistent")

        emitted = output["emitted_member_coordinate_multiset"]
        if (
            not isinstance(emitted, list)
            or any(not isinstance(value, str) for value in emitted)
            or emitted != sorted(emitted, key=_audit_coordinate_key)
        ):
            raise ValueError("producer audit emitted member events are not deterministic")
        emitted_counts = Counter(emitted)
        if (
            output["emitted_cell_multiset_sha256"]
            != _audit_digest("adr039-emitted-cell-multiset/v1", emitted)
            or output["emitted_member_occurrence_count"] != len(emitted)
            or output["member_max_ingest_count"] != max(emitted_counts.values(), default=0)
        ):
            raise ValueError("producer audit emitted member closure is invalid")
        membership_is_closed = (
            output["member_max_ingest_count"] == 1
            and output["emitted_member_occurrence_count"] == output["union_member_count"]
            and set(emitted) == set(union_coordinates)
        )
        component_assignment_is_closed = all(
            sum(
                candidate.get("row_kind") == "object"
                and ref in candidate.get("component_region_refs", [])
                for candidate in output_objects
                if isinstance(candidate, dict)
            )
            == 1
            for ref in component_refs
        )
        d2_evidence = not membership_is_closed or not component_assignment_is_closed
        if output["row_kind"] == "object" and output["matched_rule"] == "D2" and not d2_evidence:
            raise ValueError("producer audit D2 defect decision does not match its evidence")
        if output["row_kind"] == "object" and d2_evidence and output["matched_rule"] != "D2":
            if not membership_is_closed:
                raise ValueError("producer audit unclosed membership must reach D2")
            raise ValueError("producer audit multiple assignments must reach D2")

        record_slots = output["record_slot_coordinate_sets"]
        emitted_data_row_ordinals = output["emitted_data_row_ordinals"]
        if (
            not isinstance(emitted_data_row_ordinals, list)
            or any(
                not isinstance(row_ordinal, int)
                or isinstance(row_ordinal, bool)
                or row_ordinal < 1
                for row_ordinal in emitted_data_row_ordinals
            )
            or emitted_data_row_ordinals != sorted(set(emitted_data_row_ordinals))
        ):
            raise ValueError("producer audit emitted data rows are not deterministic")
        if output["row_kind"] == "defect_tombstone" and emitted_data_row_ordinals:
            raise ValueError("producer audit tombstone cannot emit data rows")
        if (
            not isinstance(record_slots, list)
            or any(
                not isinstance(slot, list)
                or any(not isinstance(value, str) for value in slot)
                or slot != sorted(set(slot), key=_audit_coordinate_key)
                for slot in record_slots
            )
        ):
            raise ValueError("producer audit record slots are not deterministic sets")
        record_slot_lines = ["|".join(slot) for slot in record_slots]
        if (
            output["record_slot_sha256"]
            != _audit_digest("adr039-record-slot/v1", record_slot_lines)
            or output["proven_record_slot_count"] != len(record_slots)
        ):
            raise ValueError("producer audit record slot closure is invalid")

        union_coordinate_set = set(union_coordinates)
        slot_row_ordinals = []
        record_slot_geometry_is_closed = True
        for slot in record_slots:
            parsed_slot = [_audit_coordinate_key(coordinate) for coordinate in slot]
            slot_rows = {row_ordinal for _sheet, row_ordinal, _column in parsed_slot}
            if not slot or not set(slot).issubset(union_coordinate_set) or len(slot_rows) != 1:
                record_slot_geometry_is_closed = False
                continue
            slot_row_ordinal = next(iter(slot_rows))
            expected_slot = {
                coordinate
                for coordinate in union_coordinate_set
                if _audit_coordinate_key(coordinate)[1] == slot_row_ordinal
            }
            if set(slot) != expected_slot:
                record_slot_geometry_is_closed = False
            slot_row_ordinals.append(slot_row_ordinal)
        record_slot_geometry_is_closed = (
            record_slot_geometry_is_closed
            and len(slot_row_ordinals) == len(set(slot_row_ordinals))
            and slot_row_ordinals == sorted(slot_row_ordinals)
        )
        record_slots_match_data_rows = slot_row_ordinals == emitted_data_row_ordinals
        if output["enumeration_status"] == "supported_complete" and not record_slot_geometry_is_closed:
            if any(not set(slot).issubset(union_coordinate_set) for slot in record_slots):
                raise ValueError("producer audit record slot is outside the union membership")
            if len(slot_row_ordinals) != len(set(slot_row_ordinals)):
                raise ValueError("producer audit complete record slots are not unique")
            raise ValueError("producer audit record slot is not the complete source row")
        if output["enumeration_status"] == "supported_complete" and not record_slots_match_data_rows:
            raise ValueError("producer audit record slots do not match emitted data rows")
        record_count_is_closed = (
            record_slot_geometry_is_closed
            and record_slots_match_data_rows
            and output["covered_count"] == output["proven_record_slot_count"]
        )
        d3_evidence = (
            not d2_evidence
            and output["proven_record_slot_count"] > 0
            and not record_count_is_closed
        )
        if output["row_kind"] == "object" and output["matched_rule"] == "D3" and not d3_evidence:
            raise ValueError("producer audit D3 defect decision does not match its evidence")
        if output["row_kind"] == "object" and d3_evidence and output["matched_rule"] != "D3":
            raise ValueError("producer audit record count evidence must reach D3")
        if output["row_kind"] == "object" and output["matched_rule"] == "D4":
            raise ValueError("producer audit D4 defect decision lacks tombstone evidence")

        if output["row_kind"] == "object":
            real_assignments.update(component_refs)
        else:
            tombstone_components.extend(component_refs)

        if output["enumeration_status"] == "supported_complete":
            if (
                not membership_is_closed
                or not record_count_is_closed
                or output["covered_count"] != output["source_total_count"]
            ):
                raise ValueError("producer audit supported object is not closed")

    zero_assignment_refs = {
        source_ref
        for source_ref, source in source_by_ref.items()
        if source["assignment_count"] == 0
    }
    if set(tombstone_components) != zero_assignment_refs or len(tombstone_components) != len(
        zero_assignment_refs
    ):
        raise ValueError("producer audit tombstone assignment closure is inconsistent")
    for source_ref, source in source_by_ref.items():
        assignment_count = real_assignments[source_ref]
        assigned_outputs = [
            output
            for output in output_objects
            if output["row_kind"] == "object"
            and source_ref in output["component_region_refs"]
        ]
        if assignment_count > 1 and any(
            output["matched_rule"] != "D2" for output in assigned_outputs
        ):
            raise ValueError("producer audit multiple assignments must reach D2")
        expected_object_ref = next(
            (
                output["object_ref"]
                for output in assigned_outputs
            ),
            None,
        ) if assignment_count == 1 else None
        if (
            source["assignment_count"] != assignment_count
            or source["assigned_object_ref"] != expected_object_ref
        ):
            raise ValueError("producer audit source/output assignment closure is inconsistent")

    coordinates = []
    for defect in audit["defects"]:
        expected_defect_fields = {
            "row_kind",
            "table_ref",
            "sheet_ordinal",
            "source_region_ordinal",
            "membership_sha256",
            "enumeration_status",
            "enumeration_reason",
            "matched_rule",
        }
        if not isinstance(defect, dict) or set(defect) != expected_defect_fields:
            raise ValueError("producer audit tombstone does not match the fixed schema")
        if defect["row_kind"] != "defect_tombstone" or defect["table_ref"] is not None:
            raise ValueError("producer audit tombstone cannot fabricate an object identity")
        for field_name in ("sheet_ordinal", "source_region_ordinal"):
            value = defect[field_name]
            if not isinstance(value, int) or isinstance(value, bool) or value < 1:
                raise ValueError("producer audit tombstone ordinals must be positive integers")
        if not re.fullmatch(r"[0-9a-f]{64}", str(defect["membership_sha256"])):
            raise ValueError("producer audit tombstone membership SHA-256 is invalid")
        decision = ENUMERATION_DECISIONS.get(defect["matched_rule"])
        if defect["matched_rule"] != "D4" or decision != (
            defect["enumeration_status"],
            defect["enumeration_reason"],
        ):
            raise ValueError("producer audit tombstone decision is invalid")
        source_ref = f"{defect['sheet_ordinal']}:{defect['source_region_ordinal']}"
        if (
            source_ref not in source_by_ref
            or defect["membership_sha256"] != source_by_ref[source_ref]["membership_sha256"]
        ):
            raise ValueError("producer audit D4 defect does not match its source region")
        coordinates.append((defect["sheet_ordinal"], defect["source_region_ordinal"]))
    if coordinates != sorted(set(coordinates)):
        raise ValueError("producer audit tombstones are not unique deterministic source regions")
    defect_coordinates = {
        (defect["sheet_ordinal"], defect["source_region_ordinal"])
        for defect in audit["defects"]
    }
    tombstone_coordinates = {
        tuple(int(part) for part in source_ref.split(":"))
        for source_ref in tombstone_components
    }
    if defect_coordinates != tombstone_coordinates:
        raise ValueError("producer audit D4 defects do not match their tombstones")


def _audit_coordinate_strings(
    sheet_ordinal: int,
    coordinates: set[tuple[int, int]] | list[tuple[int, int]],
) -> list[str]:
    return [
        f"{sheet_ordinal}:{row_ordinal}:{column_ordinal}"
        for row_ordinal, column_ordinal in sorted(coordinates)
    ]


def _audit_coordinate_key(value: str) -> tuple[int, int, int]:
    match = re.fullmatch(r"([1-9][0-9]*):([1-9][0-9]*):([1-9][0-9]*)", value)
    if not match:
        raise ValueError("producer audit coordinate is invalid")
    return tuple(int(part) for part in match.groups())


def _audit_digest(kind: str, values: list[str]) -> str:
    return hashlib.sha256("\n".join([kind, *values]).encode("utf-8")).hexdigest()


def _audit_output_object(
    item: dict[str, Any],
    *,
    producer_generation_ref: str,
    enumeration_rule_version: str,
) -> dict[str, Any]:
    table = item["table"]
    sheet_ordinal = table["sheet_ordinal"]
    member_events = _audit_coordinate_strings(sheet_ordinal, item["emitted_member_events"])
    record_slot_coordinates = [
        _audit_coordinate_strings(
            sheet_ordinal,
            {
                coordinate
                for coordinate in item["members"]
                if coordinate[0] == row_ordinal
            },
        )
        for row_ordinal in item["proven_record_slots"]
    ]
    record_slot_lines = ["|".join(coordinates) for coordinates in record_slot_coordinates]
    component_refs = [
        f"{source_sheet}:{source_ordinal}"
        for source_sheet, source_ordinal in sorted(item["source_components"])
    ]
    component_digests = [
        _region_membership_sha256(source_sheet, item["source_components"][key])
        for key in sorted(item["source_components"])
        for source_sheet in [key[0]]
    ]
    event_counts = Counter(member_events)
    table_ref = table["table_ref"]
    data_row_ordinals = {
        row["row_ordinal_int"]
        for row in item["rows"]
        if row["row_role_kwd"] == "data"
    }
    covered_count = (
        sum(row_ordinal in data_row_ordinals for row_ordinal in item["proven_record_slots"])
        if table["matched_rule"] == "D3"
        else table["data_row_count"]
    )
    return {
        "row_kind": "object",
        "object_ref": table_ref,
        "worksheet_ordinal": sheet_ordinal,
        "component_region_refs": component_refs,
        "component_membership_sha256_list": component_digests,
        "union_membership_sha256": _region_membership_sha256(sheet_ordinal, item["members"]),
        "union_member_count": len(item["members"]),
        "emitted_member_coordinate_multiset": member_events,
        "emitted_cell_multiset_sha256": _audit_digest(
            "adr039-emitted-cell-multiset/v1",
            member_events,
        ),
        "emitted_member_occurrence_count": len(member_events),
        "member_max_ingest_count": max(event_counts.values(), default=0),
        "record_slot_sha256": _audit_digest(
            "adr039-record-slot/v1",
            record_slot_lines,
        ),
        "proven_record_slot_count": len(record_slot_coordinates),
        "record_slot_coordinate_sets": record_slot_coordinates,
        "emitted_data_row_ordinals": sorted(data_row_ordinals),
        "enumeration_rule_version": enumeration_rule_version,
        "structure_generation_ref": producer_generation_ref,
        "table_ref": table_ref,
        "identity_validation_status": "pending_independent_validation",
        "matched_rule": table["matched_rule"],
        "decision_chain_stop": table["matched_rule"],
        "enumeration_status": table["enumeration_status"],
        "enumeration_reason": table["enumeration_reason"],
        "covered_count": covered_count,
        "source_total_count": table["source_total_count"],
    }


def _audit_d4_object(
    defect: dict[str, Any],
    *,
    source_region: dict[str, Any],
    producer_generation_ref: str,
    enumeration_rule_version: str,
) -> dict[str, Any]:
    sheet_ordinal = defect["sheet_ordinal"]
    member_coordinates = set(source_region["members"])
    component_ref = f"{sheet_ordinal}:{defect['source_region_ordinal']}"
    return {
        "row_kind": "defect_tombstone",
        "object_ref": None,
        "worksheet_ordinal": sheet_ordinal,
        "component_region_refs": [component_ref],
        "component_membership_sha256_list": [source_region["membership_sha256"]],
        "union_membership_sha256": source_region["membership_sha256"],
        "union_member_count": len(member_coordinates),
        "emitted_member_coordinate_multiset": [],
        "emitted_cell_multiset_sha256": _audit_digest(
            "adr039-emitted-cell-multiset/v1",
            [],
        ),
        "emitted_member_occurrence_count": 0,
        "member_max_ingest_count": 0,
        "record_slot_sha256": _audit_digest("adr039-record-slot/v1", []),
        "proven_record_slot_count": 0,
        "record_slot_coordinate_sets": [],
        "emitted_data_row_ordinals": [],
        "enumeration_rule_version": enumeration_rule_version,
        "structure_generation_ref": producer_generation_ref,
        "table_ref": None,
        "identity_validation_status": "pending_independent_validation",
        "matched_rule": "D4",
        "decision_chain_stop": "D4",
        "enumeration_status": "defect",
        "enumeration_reason": "missing_projection",
        "covered_count": 0,
        "source_total_count": None,
    }


def _unknown_structure_region(
    *,
    parser,
    worksheet,
    region: dict[str, Any],
    sheet_name: str,
    sheet_ordinal: int,
    table_ordinal: int,
    membership_sha256: str,
    source_sha256: str,
    producer_generation_ref: str,
) -> tuple[dict[str, Any], list[dict[str, Any]]] | None:
    table_ref = _table_ref(source_sha256, sheet_ordinal, table_ordinal, membership_sha256)
    merged_ranges = list(worksheet.merged_cells.ranges)
    rows = []
    for row_ordinal in sorted({row for row, _column in region["members"]}):
        columns = sorted(column for row, column in region["members"] if row == row_ordinal)
        first_column = min(columns)
        fields = _ordered_fields(
            [f"Column_{column_ordinal}" for column_ordinal in columns],
            [
                _cell_value(
                    parser,
                    worksheet,
                    row_ordinal,
                    column_ordinal,
                    merged_ranges,
                )
                for column_ordinal in columns
            ],
            note=False,
            sheet_ordinal=sheet_ordinal,
            header_paths=[[] for _column_ordinal in columns],
            column_ordinals=[
                column_ordinal - first_column + 1 for column_ordinal in columns
            ],
            absolute_column_ordinals=columns,
            source_anchors=[
                _source_cell_anchor(row_ordinal, column_ordinal, merged_ranges)
                for column_ordinal in columns
            ],
        )
        row_ref = f"{table_ref}:{row_ordinal}"
        rows.append(
            {
                "id": "tsr_v1_" + _versioned_digest(
                    "tabular-row-record/v1",
                    producer_generation_ref,
                    row_ref,
                ),
                "tabular_structure_version_kwd": TABULAR_STRUCTURE_VERSION,
                "structure_kind_kwd": "table_row",
                "producer_schema_version_kwd": PRODUCER_SCHEMA_VERSION,
                "producer_generation_ref_kwd": producer_generation_ref,
                "table_ref_kwd": table_ref,
                "table_label_kwd": _sanitize_untrusted_text(sheet_name),
                "table_context_list": "[]",
                "row_ref_kwd": row_ref,
                "row_ordinal_int": row_ordinal,
                "data_row_index_int": None,
                "row_role_kwd": "unknown",
                "source_total_count_int": None,
                "ordered_fields_list": json.dumps(fields, ensure_ascii=False, separators=(",", ":")),
            }
        )
    if not rows:
        return None
    table = {
        "table_ref": table_ref,
        "sheet_ordinal": sheet_ordinal,
        "table_ordinal": table_ordinal,
        "row_count": len(rows),
        "data_row_count": 0,
        "source_total_count": None,
    }
    _apply_enumeration_decision(table, "R8")
    return table, rows


def _build_tabular_structure_projection_with_audit(
    filename: str,
    binary: bytes,
    *,
    producer_generation_ref: str | None = None,
    adr044_conversion_receipt: dict[str, str] | None = None,
    table_context_entry_limit: int = DEFAULT_CONTEXT_ENTRY_LIMIT,
    table_context_value_bytes: int = DEFAULT_CONTEXT_VALUE_BYTES,
    parser=None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Build one immutable generation from complete workbook bytes.

    The returned records are a derived read model. They contain no vectors and
    are never inserted into the ordinary relevance-chunk index.
    """

    if not isinstance(binary, bytes) or not binary:
        raise ValueError("complete immutable workbook bytes are required")
    producer_generation_ref = producer_generation_ref or str(uuid.uuid4())
    _validate_generation_ref(producer_generation_ref)

    # Reuse RAGFlow's mature xls/xlsx loader and header detector. Importing here
    # keeps this fixed-schema producer independent from the normal chunk path.
    if parser is None:
        from rag.app.table import Excel

        parser = Excel()
    converted_source_sha256 = hashlib.sha256(binary).hexdigest()
    has_adr044_receipt = _validate_adr044_conversion_receipt(
        adr044_conversion_receipt,
        converted_source_sha256,
    )
    source_sha256 = (
        adr044_conversion_receipt["original_source_sha256"]
        if has_adr044_receipt
        else converted_source_sha256
    )
    workbook = parser._load_excel_to_workbook(BytesIO(binary))
    formula_coordinates, formula_inventory_proven = _formula_coordinates_by_sheet(binary)
    formula_cached_result_kinds = _formula_cached_result_kinds_by_sheet(binary)
    formula_values = _formula_values_by_sheet(binary)
    records = []
    tables = []
    defects = []
    audit_source_regions = []
    audit_output_objects = []

    for sheet_ordinal, sheet_name in enumerate(workbook.sheetnames, start=1):
        worksheet = workbook[sheet_name]
        sheet_formula_coordinates = (
            formula_coordinates[sheet_ordinal - 1]
            if sheet_ordinal <= len(formula_coordinates)
            else set()
        )
        sheet_formula_values = (
            formula_values[sheet_ordinal - 1]
            if sheet_ordinal <= len(formula_values)
            else {}
        )
        sheet_formula_cached_result_kinds = (
            formula_cached_result_kinds[sheet_ordinal - 1]
            if sheet_ordinal <= len(formula_cached_result_kinds)
            else {}
        )
        sheet_unresolved_formula_coordinates = (
            sheet_formula_coordinates - set(sheet_formula_cached_result_kinds)
        )
        regions = _worksheet_structure_regions(
            parser,
            worksheet,
            sheet_ordinal,
            sheet_unresolved_formula_coordinates,
            sheet_formula_coordinates,
        )
        projected = []
        unprojected = []
        candidates = []
        for region in regions:
            children = region["g1_children"]
            if len(children) <= 1:
                candidates.append((region, False))
                continue
            # Preserve the exact G2 candidate; proven record slots below decide
            # whether the finer G1 split affects the record axis.
            candidates.append((region, True))
        source_regions = {
            (sheet_ordinal, region_index): region
            for region_index, (region, _g1_disagreement) in enumerate(candidates, start=1)
        }

        for region_index, (region, g1_disagreement) in enumerate(candidates, start=1):
            structure_region = region
            region_worksheet, row_offset = _copy_structure_region(parser, worksheet, region)
            result = _project_structure_region(
                parser=parser,
                worksheet=region_worksheet,
                sheet_name=sheet_name,
                sheet_ordinal=sheet_ordinal,
                table_ordinal=region_index,
                membership_sha256=region["membership_sha256"],
                source_sha256=source_sha256,
                producer_generation_ref=producer_generation_ref,
                row_offset=row_offset,
                table_context_entry_limit=table_context_entry_limit,
                table_context_value_bytes=table_context_value_bytes,
                force_unknown_total=not formula_inventory_proven,
            )
            if g1_disagreement and result is not None:
                table, table_rows = result
                record_rows = {
                    row["row_ordinal_int"]
                    for row in table_rows
                    if row["row_role_kwd"] == "data"
                }
                g1_disagreement_is_safe = (
                    table["source_total_count"] is not None
                    and (
                        table["matched_rule"] == "L1-08"
                        or _g1_disagreement_is_outside_record_axis(
                            worksheet,
                            region,
                            record_rows,
                        )
                    )
                )
            else:
                g1_disagreement_is_safe = not g1_disagreement

            if g1_disagreement and not g1_disagreement_is_safe:
                child_results = []
                for child_members in region["g1_children"]:
                    child_region = _g1_child_region(region, child_members)
                    child_worksheet, child_row_offset = _copy_structure_region(
                        parser,
                        worksheet,
                        child_region,
                    )
                    child_result = _project_structure_region(
                        parser=parser,
                        worksheet=child_worksheet,
                        sheet_name=sheet_name,
                        sheet_ordinal=sheet_ordinal,
                        table_ordinal=region_index,
                        membership_sha256=region["membership_sha256"],
                        source_sha256=source_sha256,
                        producer_generation_ref=producer_generation_ref,
                        row_offset=child_row_offset,
                        table_context_entry_limit=table_context_entry_limit,
                        table_context_value_bytes=table_context_value_bytes,
                        force_unknown_total=not formula_inventory_proven,
                    )
                    if child_result is None:
                        continue
                    child_table, child_rows = child_result
                    child_record_rows = {
                        row["row_ordinal_int"]
                        for row in child_rows
                        if row["row_role_kwd"] == "data"
                    }
                    if (
                        child_table["source_total_count"] is not None
                        and _g1_disagreement_is_outside_record_axis(
                            worksheet,
                            region,
                            child_record_rows,
                        )
                    ):
                        child_results.append((child_result, child_region))
                if len(child_results) == 1:
                    result, structure_region = child_results[0]
                    g1_disagreement_is_safe = True

            if g1_disagreement and result is not None and not g1_disagreement_is_safe:
                result = _unknown_structure_region(
                    parser=parser,
                    worksheet=worksheet,
                    region=region,
                    sheet_name=sheet_name,
                    sheet_ordinal=sheet_ordinal,
                    table_ordinal=region_index,
                    membership_sha256=region["membership_sha256"],
                    source_sha256=source_sha256,
                    producer_generation_ref=producer_generation_ref,
                )
            if result is None:
                unprojected.append((region_index, region))
                continue
            table, table_rows = result
            projected.append(
                _new_projected_item(
                    parser=parser,
                    worksheet=worksheet,
                    region=region,
                    source_region_key=(sheet_ordinal, region_index),
                    table=table,
                    rows=table_rows,
                    positive_rule=(
                        table["matched_rule"]
                        if table["source_total_count"] is not None
                        else None
                    ),
                    structure_region=structure_region,
                )
            )

        if unprojected:
            for region_index, region in unprojected:
                result = _unknown_structure_region(
                    parser=parser,
                    worksheet=worksheet,
                    region=region,
                    sheet_name=sheet_name,
                    sheet_ordinal=sheet_ordinal,
                    table_ordinal=region_index,
                    membership_sha256=region["membership_sha256"],
                    source_sha256=source_sha256,
                    producer_generation_ref=producer_generation_ref,
                )
                if result is None:
                    continue
                table, table_rows = result
                projected.append(
                    _new_projected_item(
                        parser=parser,
                        worksheet=worksheet,
                        region=region,
                        source_region_key=(sheet_ordinal, region_index),
                        table=table,
                        rows=table_rows,
                        positive_rule=None,
                    )
                )

        projected = _merge_unique_empty_axis_context(projected, source_regions)

        changed = True
        while changed:
            changed = False
            for continuation in sorted(projected, key=lambda item: item["bbox"]):
                candidates = []
                for main in projected:
                    if main is continuation:
                        continue
                    merged = _merge_continuation_pair(
                        parser=parser,
                        worksheet=worksheet,
                        main=main,
                        continuation=continuation,
                        source_sha256=source_sha256,
                        producer_generation_ref=producer_generation_ref,
                        table_context_entry_limit=table_context_entry_limit,
                        table_context_value_bytes=table_context_value_bytes,
                    )
                    if merged is not None:
                        candidates.append((main, merged))
                if len(candidates) != 1:
                    continue
                main, merged = candidates[0]
                projected = [
                    item
                    for item in projected
                    if item is not main and item is not continuation
                ] + [merged]
                changed = True
                break

        defects.extend(_apply_projection_invariants(projected, source_regions))

        complete_items = [
            item for item in projected if item["table"]["source_total_count"] is not None
        ]
        if len(complete_items) > 1:
            for item in complete_items:
                if item["table"]["matched_rule"] == "L1-08":
                    continue
                _apply_enumeration_decision(item["table"], "L1-03")
                item["positive_rule"] = "L1-03"

        partial_overlap_ids = set()
        nested_unknown_ids = set()
        preliminary_predicates = {
            id(item): _region_negative_predicates(
                worksheet=worksheet,
                sheet_name=sheet_name,
                item=item,
                siblings=projected,
                formula_coordinates=sheet_formula_coordinates,
                formula_values=sheet_formula_values,
                formula_cached_result_kinds=sheet_formula_cached_result_kinds,
                formula_inventory_proven=formula_inventory_proven,
                partial_overlap=False,
            )
            for item in projected
            if not item["table"]["matched_rule"].startswith("D")
        }
        for index, item in enumerate(projected):
            for sibling in projected[index + 1 :]:
                if not _column_sets_intersect(
                    item["member_columns"],
                    sibling["member_columns"],
                ):
                    continue
                if _column_sets_are_nested(
                    item["member_columns"],
                    sibling["member_columns"],
                ):
                    risk_ids = nested_unknown_ids
                else:
                    risk_ids = partial_overlap_ids
                earlier, later = sorted((item, sibling), key=lambda candidate: candidate["bbox"])
                later_predicates = preliminary_predicates.get(id(later), {})
                later_axis_proven = later.get("positive_rule") is not None and not any(
                    later_predicates.values()
                )
                later_rows = _continuation_record_rows(later)
                earlier_proven_record_columns = {
                    column
                    for row, column in earlier["members"]
                    if row in set(earlier["proven_record_slots"])
                }
                later_proven_record_columns = {
                    column
                    for row, column in later["members"]
                    if row in set(later["proven_record_slots"])
                }
                nested_axis_overlap = bool(
                    earlier_proven_record_columns & later["member_columns"]
                    or later_proven_record_columns & earlier["member_columns"]
                )
                nested_axis_compatible = (
                    risk_ids is not nested_unknown_ids
                    or (
                        nested_axis_overlap
                        and (
                            len(later_rows) != 1
                            or (
                                bool(later_rows)
                                and _continuation_rows_match_proven_axis(
                                    parser=parser,
                                    worksheet=worksheet,
                                    main=earlier,
                                    continuation=later,
                                    continuation_rows=later_rows,
                                )
                            )
                        )
                    )
                )
                unknown_is_context_for_following_structure = (
                    later.get("positive_rule") is None
                    and len(later["member_columns"]) == 1
                    and any(
                        following is not earlier
                        and following is not later
                        and following["bbox"][0] > later["bbox"][2]
                        and following.get("structure_evidence") is not None
                        and later["member_columns"].issubset(following["member_columns"])
                        for following in projected
                    )
                )
                if (
                    earlier["bbox"][2] < later["bbox"][0]
                    and not later_axis_proven
                    and nested_axis_compatible
                    and not unknown_is_context_for_following_structure
                    and (
                        risk_ids is partial_overlap_ids
                        or _unknown_region_can_extend_proven_axis(
                            parser=parser,
                            worksheet=worksheet,
                            proven=earlier,
                            unknown=later,
                        )
                        or not _axis_closure_proven(
                            parser=parser,
                            worksheet=worksheet,
                            earlier=earlier,
                            later=later,
                        )
                    )
                ):
                    risk_ids.update((id(earlier), id(later)))

        for item in projected:
            if item["table"]["matched_rule"].startswith("D"):
                continue
            predicates = dict(preliminary_predicates[id(item)])
            predicates["R6"] = id(item) in partial_overlap_ids
            l1_rule = (
                None
                if id(item) in nested_unknown_ids
                else item["positive_rule"]
            )
            matched_rule = _ordered_enumeration_rule(predicates, l1_rule)
            if matched_rule.startswith("L1-"):
                _apply_enumeration_decision(item["table"], matched_rule)
            else:
                _clear_complete_decision(item["table"], item["rows"], matched_rule)

        projected.sort(
            key=lambda item: (
                item["table"]["sheet_ordinal"],
                item["table"]["table_ordinal"],
            )
        )
        for table_ordinal, item in enumerate(projected, start=1):
            _rekey_projected_item(
                item,
                table_ordinal=table_ordinal,
                source_sha256=source_sha256,
                producer_generation_ref=producer_generation_ref,
            )
            if (
                has_adr044_receipt
                and item["table"]["source_total_count"] is not None
                and item["table"]["matched_rule"] != "L1-08"
            ):
                _apply_enumeration_decision(item["table"], "L1-01")
            item["rows"].sort(key=lambda row: row["row_ordinal_int"])
            item["table"]["row_count"] = len(item["rows"])
            _finalize_table_manifest_evidence(item)
            tables.append(item["table"])
            records.extend(item["rows"])

        for source_region_key, region in sorted(source_regions.items()):
            assigned_items = [
                item
                for item in projected
                if source_region_key in item["source_components"]
            ]
            source_sheet, source_ordinal = source_region_key
            source_members = set(region["members"])
            assigned_ref = (
                assigned_items[0]["table"]["table_ref"]
                if len(assigned_items) == 1
                else None
            )
            audit_source_regions.append(
                {
                    "source_region_ref": f"{source_sheet}:{source_ordinal}",
                    "worksheet_ordinal": source_sheet,
                    "bbox": list(region["bbox"]),
                    "row_count": len({row for row, _column in source_members}),
                    "column_count": len({column for _row, column in source_members}),
                    "membership_sha256": region["membership_sha256"],
                    "member_count": len(source_members),
                    "member_coordinate_set": _audit_coordinate_strings(
                        source_sheet,
                        source_members,
                    ),
                    "assigned_object_ref": assigned_ref,
                    "assignment_count": len(assigned_items),
                }
            )
        audit_output_objects.extend(
            _audit_output_object(
                item,
                producer_generation_ref=producer_generation_ref,
                enumeration_rule_version=ENUMERATION_RULE_VERSION,
            )
            for item in projected
        )
        for defect in defects:
            if defect["sheet_ordinal"] != sheet_ordinal:
                continue
            source_region = source_regions[
                (sheet_ordinal, defect["source_region_ordinal"])
            ]
            audit_output_objects.append(
                _audit_d4_object(
                    defect,
                    source_region=source_region,
                    producer_generation_ref=producer_generation_ref,
                    enumeration_rule_version=ENUMERATION_RULE_VERSION,
                )
            )

    projection = {
        "version": PROJECTION_VERSION,
        "producer_schema_version": PRODUCER_SCHEMA_VERSION,
        "producer_generation_ref": producer_generation_ref,
        "structure_algorithm_version": STRUCTURE_PRODUCER_ALGORITHM_VERSION,
        "enumeration_rule_version": ENUMERATION_RULE_VERSION,
        "source_sha256": source_sha256,
        "tables": tables,
        "rows": records,
    }
    validate_tabular_structure_projection(projection)
    audit = {
        "version": "tabular-structure-producer-audit/v1",
        "producer_generation_ref": producer_generation_ref,
        "enumeration_rule_version": ENUMERATION_RULE_VERSION,
        "source_sha256": source_sha256,
        "source_regions": audit_source_regions,
        "output_objects": audit_output_objects,
        "defects": defects,
    }
    _validate_tabular_structure_producer_audit(audit)
    return projection, audit


def build_tabular_structure_projection(
    filename: str,
    binary: bytes,
    *,
    producer_generation_ref: str | None = None,
    adr044_conversion_receipt: dict[str, str] | None = None,
    table_context_entry_limit: int = DEFAULT_CONTEXT_ENTRY_LIMIT,
    table_context_value_bytes: int = DEFAULT_CONTEXT_VALUE_BYTES,
    parser=None,
) -> dict[str, Any]:
    projection, _audit = _build_tabular_structure_projection_with_audit(
        filename,
        binary,
        producer_generation_ref=producer_generation_ref,
        adr044_conversion_receipt=adr044_conversion_receipt,
        table_context_entry_limit=table_context_entry_limit,
        table_context_value_bytes=table_context_value_bytes,
        parser=parser,
    )
    return projection


def _validate_tabular_structure_projection_for_contract(
    projection: dict[str, Any],
    contract: tuple[str, str, str, str],
) -> None:
    (
        producer_schema_version,
        projection_version,
        structure_algorithm_version,
        enumeration_rule_version,
    ) = contract
    if not isinstance(projection, dict) or set(projection) != PROJECTION_FIELDS:
        raise ValueError("structure projection does not match the fixed top-level schema")
    if projection.get("version") != projection_version:
        raise ValueError("unsupported tabular structure projection version")
    if projection.get("producer_schema_version") != producer_schema_version:
        raise ValueError("unsupported table producer schema version")
    if projection.get("structure_algorithm_version") != structure_algorithm_version:
        raise ValueError("unsupported structure algorithm version")
    if projection.get("enumeration_rule_version") != enumeration_rule_version:
        raise ValueError("unsupported enumeration rule version")
    generation_ref = projection.get("producer_generation_ref")
    _validate_generation_ref(generation_ref)
    source_sha256 = projection.get("source_sha256")
    if not isinstance(source_sha256, str) or not re.fullmatch(r"[0-9a-f]{64}", source_sha256):
        raise ValueError("source SHA-256 is invalid")
    rows = projection.get("rows")
    if not isinstance(rows, list):
        raise ValueError("projection rows must be a list")

    ids = set()
    row_refs = set()
    by_table: dict[str, list[dict[str, Any]]] = {}
    ordered_fields_by_row_ref: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        if not isinstance(row, dict) or set(row) != PROJECTION_ROW_FIELDS:
            raise ValueError("structure projection row fields do not match the fixed schema")
        if row["tabular_structure_version_kwd"] != TABULAR_STRUCTURE_VERSION:
            raise ValueError("unsupported structure row version")
        if row["structure_kind_kwd"] != "table_row":
            raise ValueError("invalid structure kind")
        if row["producer_schema_version_kwd"] != producer_schema_version:
            raise ValueError("unsupported row producer schema version")
        if row["producer_generation_ref_kwd"] != generation_ref:
            raise ValueError("mixed producer generations are not allowed")
        for field_name, label in (
            ("id", "record ID"),
            ("table_ref_kwd", "table reference"),
            ("table_label_kwd", "table label"),
            ("row_ref_kwd", "row reference"),
        ):
            if not isinstance(row[field_name], str):
                raise ValueError(f"{label} must be a string")
        if len(row["table_label_kwd"].encode("utf-8")) > DEFAULT_TABLE_LABEL_BYTES:
            raise ValueError("table label exceeds the UTF-8 byte limit")
        if _sanitize_untrusted_text(row["table_label_kwd"]) != row["table_label_kwd"]:
            raise ValueError("table label contains unsafe controls")
        if row["row_role_kwd"] not in {"data", "note", "unknown"}:
            raise ValueError("invalid structure row role")
        if not isinstance(row["row_ordinal_int"], int) or isinstance(row["row_ordinal_int"], bool) or row["row_ordinal_int"] < 1:
            raise ValueError("row ordinal must be a positive integer")
        data_row_index = row["data_row_index_int"]
        if data_row_index is not None and (
            not isinstance(data_row_index, int) or isinstance(data_row_index, bool) or data_row_index < 1
        ):
            raise ValueError("data row index must be a positive integer or null")
        source_total = row["source_total_count_int"]
        if source_total is not None and (
            not isinstance(source_total, int) or isinstance(source_total, bool) or source_total < 0
        ):
            raise ValueError("source total must be a non-negative integer or null")
        expected_id = "tsr_v1_" + _versioned_digest(
            "tabular-row-record/v1",
            generation_ref,
            row["row_ref_kwd"],
        )
        if row["id"] != expected_id:
            raise ValueError("structure record identity does not match its generation and row reference")
        if row["row_ref_kwd"] != f"{row['table_ref_kwd']}:{row['row_ordinal_int']}":
            raise ValueError("row reference does not match table and physical row identity")
        parsed_lists = {}
        for field_name, label in (
            ("table_context_list", "table context"),
            ("ordered_fields_list", "ordered fields"),
        ):
            if not isinstance(row[field_name], str):
                raise ValueError(f"{label} must be JSON text")
            try:
                values = json.loads(row[field_name])
            except (TypeError, json.JSONDecodeError) as exc:
                raise ValueError(f"{label} must be valid JSON") from exc
            if not isinstance(values, list):
                raise ValueError(f"{label} must use the fixed field schema")
            if field_name == "table_context_list":
                invalid = any(
                    not isinstance(item, dict)
                    or set(item) != {"name", "value"}
                    or not isinstance(item["name"], str)
                    or not isinstance(item["value"], str)
                    for item in values
                )
            else:
                invalid = any(
                    not isinstance(item, dict)
                    or set(item)
                    != {
                        "column_id",
                        "column_ordinal",
                        "header_path",
                        "name",
                        "value",
                    }
                    or not isinstance(item["column_id"], str)
                    or re.fullmatch(r"col_v1:[1-9][0-9]*:[1-9][0-9]*", item["column_id"])
                    is None
                    or not isinstance(item["column_ordinal"], int)
                    or isinstance(item["column_ordinal"], bool)
                    or item["column_ordinal"] < 1
                    or not isinstance(item["header_path"], list)
                    or any(
                        not isinstance(segment, str) or not segment.strip()
                        for segment in item["header_path"]
                    )
                    or not isinstance(item["name"], str)
                    or not isinstance(item["value"], str)
                    for item in values
                )
            if invalid:
                raise ValueError(f"{label} must use the fixed field schema")
            parsed_lists[field_name] = values
        ordered_fields_by_row_ref[row["row_ref_kwd"]] = parsed_lists["ordered_fields_list"]
        context = parsed_lists["table_context_list"]
        if len(context) > DEFAULT_CONTEXT_ENTRY_LIMIT:
            raise ValueError("table context exceeds the entry limit")
        for item in context:
            if any(len(item[field].encode("utf-8")) > DEFAULT_CONTEXT_VALUE_BYTES for field in ("name", "value")):
                raise ValueError("table context exceeds the UTF-8 byte limit")
            if any(_sanitize_untrusted_text(item[field]) != item[field] for field in ("name", "value")):
                raise ValueError("table context contains unsafe controls")
        if row["id"] in ids or row["row_ref_kwd"] in row_refs:
            raise ValueError("duplicate structure row identity")
        ids.add(row["id"])
        row_refs.add(row["row_ref_kwd"])
        by_table.setdefault(row["table_ref_kwd"], []).append(row)

    for table_rows in by_table.values():
        if len({row["table_label_kwd"] for row in table_rows}) != 1:
            raise ValueError("table label is not generation-wide")
        if len({row["table_context_list"] for row in table_rows}) != 1:
            raise ValueError("table context is not generation-wide")
        ordinals = [row["row_ordinal_int"] for row in table_rows]
        if ordinals != sorted(ordinals) or len(ordinals) != len(set(ordinals)):
            raise ValueError("structure rows are not in deterministic physical order")
        data_indices = [row["data_row_index_int"] for row in table_rows if row["row_role_kwd"] == "data"]
        if data_indices != list(range(1, len(data_indices) + 1)):
            raise ValueError("data row indices are not globally contiguous")
        if any(row["row_role_kwd"] != "data" and row["data_row_index_int"] is not None for row in table_rows):
            raise ValueError("non-data rows cannot have a data row index")
        totals = {row["source_total_count_int"] for row in table_rows}
        if totals == {None}:
            continue
        if (
            any(row["row_role_kwd"] == "unknown" for row in table_rows)
            or not data_indices
            or totals != {len(data_indices)}
        ):
            raise ValueError("source total is not generation-wide or fail-closed")

    tables = projection.get("tables")
    if not isinstance(tables, list):
        raise ValueError("table manifest must be a list")
    manifest_refs = set()
    manifest_coordinates = []
    for table in tables:
        expected_keys = {
            "table_ref",
            "sheet_ordinal",
            "table_ordinal",
            "row_count",
            "data_row_count",
            "source_total_count",
            "enumeration_status",
            "enumeration_reason",
            "matched_rule",
            "table_label",
            "table_context",
            "ordered_columns",
        }
        if not isinstance(table, dict) or set(table) != expected_keys:
            raise ValueError("table manifest does not match the fixed schema")
        if not isinstance(table["table_ref"], str):
            raise ValueError("table manifest reference must be a string")
        for field_name in ("sheet_ordinal", "table_ordinal"):
            value = table[field_name]
            if not isinstance(value, int) or isinstance(value, bool) or value < 1:
                raise ValueError("table manifest ordinals must be positive integers")
        for field_name in ("row_count", "data_row_count"):
            value = table[field_name]
            if not isinstance(value, int) or isinstance(value, bool) or value < 0:
                raise ValueError("table manifest counts must be non-negative integers")
        source_total = table["source_total_count"]
        if source_total is not None and (
            not isinstance(source_total, int) or isinstance(source_total, bool) or source_total < 0
        ):
            raise ValueError("table manifest source total must be a non-negative integer or null")
        matched_rule = table["matched_rule"]
        decision = ENUMERATION_DECISIONS.get(matched_rule) if isinstance(matched_rule, str) else None
        if decision != (table["enumeration_status"], table["enumeration_reason"]):
            raise ValueError("table manifest enumeration decision is invalid")
        if (table["enumeration_status"] == "supported_complete") != (source_total is not None):
            raise ValueError("table manifest enumeration decision conflicts with source total")
        table_label = table["table_label"]
        if (
            not isinstance(table_label, str)
            or not table_label
            or len(table_label.encode("utf-8")) > DEFAULT_TABLE_LABEL_BYTES
            or _sanitize_untrusted_text(table_label) != table_label
        ):
            raise ValueError("table manifest label is invalid")
        table_context = table["table_context"]
        if not isinstance(table_context, list) or len(table_context) > DEFAULT_CONTEXT_ENTRY_LIMIT:
            raise ValueError("table manifest context is invalid")
        if any(
            not isinstance(item, dict)
            or set(item) != {"name", "value"}
            or any(
                not isinstance(item[field], str)
                or not item[field]
                or len(item[field].encode("utf-8")) > DEFAULT_CONTEXT_VALUE_BYTES
                or _sanitize_untrusted_text(item[field]) != item[field]
                for field in ("name", "value")
            )
            for item in table_context
        ):
            raise ValueError("table manifest context is invalid")
        ordered_columns = table["ordered_columns"]
        if not isinstance(ordered_columns, list):
            raise ValueError("table manifest ordered columns are invalid")
        if table["enumeration_status"] == "supported_complete" and not ordered_columns:
            raise ValueError("supported complete table requires ordered columns")
        column_ids = set()
        column_ordinals = []
        columns_by_id = {}
        for column in ordered_columns:
            if (
                not isinstance(column, dict)
                or set(column) != {"column_id", "column_ordinal", "header_path", "name"}
                or not isinstance(column["column_id"], str)
                or re.fullmatch(r"col_v1:[1-9][0-9]*:[1-9][0-9]*", column["column_id"]) is None
                or not isinstance(column["column_ordinal"], int)
                or isinstance(column["column_ordinal"], bool)
                or column["column_ordinal"] < 1
                or not isinstance(column["header_path"], list)
                or not column["header_path"]
                or any(not isinstance(segment, str) or not segment.strip() for segment in column["header_path"])
                or not isinstance(column["name"], str)
                or not column["name"]
                or column["column_id"] in column_ids
            ):
                raise ValueError("table manifest ordered columns are invalid")
            column_ids.add(column["column_id"])
            column_ordinals.append(column["column_ordinal"])
            columns_by_id[column["column_id"]] = column
        if column_ordinals != list(range(1, len(column_ordinals) + 1)):
            raise ValueError("table manifest ordered columns are invalid")
        table_ref = table["table_ref"]
        table_ref_match = re.fullmatch(r"tbl_v2_([0-9a-f]{64})_[0-9a-f]{64}", table_ref)
        expected_table_ref = (
            _table_ref_for_contract(
                source_sha256,
                table["sheet_ordinal"],
                table["table_ordinal"],
                table_ref_match.group(1),
                producer_schema_version=producer_schema_version,
                projection_version=projection_version,
                structure_algorithm_version=structure_algorithm_version,
                enumeration_rule_version=enumeration_rule_version,
            )
            if table_ref_match
            else None
        )
        if table_ref != expected_table_ref:
            raise ValueError("table reference does not match source and physical table identity")
        if table_ref in manifest_refs:
            raise ValueError("table manifest does not match projected records")
        manifest_refs.add(table_ref)
        manifest_coordinates.append((table["sheet_ordinal"], table["table_ordinal"]))
        table_rows = by_table.get(table_ref, [])
        data_row_count = sum(row["row_role_kwd"] == "data" for row in table_rows)
        source_totals = {row["source_total_count_int"] for row in table_rows}
        manifest_only_empty = (
            not table_rows
            and table["matched_rule"] == "L1-08"
            and table["enumeration_reason"] == "empty_record_axis_proven"
            and table["row_count"] == 0
            and table["data_row_count"] == 0
            and table["source_total_count"] == 0
            and bool(ordered_columns)
        )
        if not table_rows and not manifest_only_empty:
            raise ValueError("table manifest does not match projected records")
        for row in table_rows:
            row_fields = ordered_fields_by_row_ref[row["row_ref_kwd"]]
            if not ordered_columns:
                continue
            for field in row_fields:
                column = columns_by_id.get(field["column_id"])
                if column is None or any(
                    field[key] != column[key]
                    for key in ("column_id", "column_ordinal", "header_path", "name")
                ):
                    raise ValueError("table manifest ordered columns do not match projected records")
        if (
            table["row_count"] != len(table_rows)
            or table["data_row_count"] != data_row_count
            or (table_rows and source_totals != {table["source_total_count"]})
        ):
            raise ValueError("table manifest counts do not match projected records")
    if not set(by_table).issubset(manifest_refs):
        raise ValueError("table manifest is missing projected tables")
    if manifest_coordinates != sorted(manifest_coordinates):
        raise ValueError("table manifest is not in deterministic physical order")
    ordinals_by_sheet: dict[int, list[int]] = {}
    for sheet_ordinal, table_ordinal in manifest_coordinates:
        ordinals_by_sheet.setdefault(sheet_ordinal, []).append(table_ordinal)
    if any(
        ordinals != list(range(1, len(ordinals) + 1))
        for ordinals in ordinals_by_sheet.values()
    ):
        raise ValueError("table manifest ordinals are not contiguous within a worksheet")


def validate_tabular_structure_projection(projection: dict[str, Any]) -> None:
    _validate_tabular_structure_projection_for_contract(
        projection,
        _CURRENT_PROJECTION_CONTRACT,
    )


def partition_tabular_structure_projection(
    projection: dict[str, Any],
    *,
    rows_per_part: int = DEFAULT_ROWS_PER_PART,
) -> list[dict[str, Any]]:
    validate_tabular_structure_projection(projection)
    if rows_per_part < 1:
        raise ValueError("rows_per_part must be positive")
    rows = projection["rows"]
    parts = []
    for offset in range(0, len(rows), rows_per_part):
        part_rows = rows[offset : offset + rows_per_part]
        parts.append(
            {
                "version": PROJECTION_PART_VERSION,
                "producer_generation_ref": projection["producer_generation_ref"],
                "part_number": len(parts) + 1,
                "row_offset": offset,
                "row_count": len(part_rows),
                "rows": part_rows,
            }
        )
    return parts


def _storage_call(method, *args, tenant_id: str | None = None):
    parameters = inspect.signature(method).parameters.values()
    accepts_tenant = any(parameter.name == "tenant_id" or parameter.kind == inspect.Parameter.VAR_KEYWORD for parameter in parameters)
    if tenant_id and accepts_tenant:
        return method(*args, tenant_id=tenant_id)
    return method(*args)


def _put_verified_immutable(storage, bucket: str, object_name: str, payload: bytes, tenant_id: str | None) -> None:
    if _storage_call(storage.obj_exist, bucket, object_name, tenant_id=tenant_id):
        if _storage_call(storage.get, bucket, object_name, tenant_id=tenant_id) != payload:
            raise IOError("refusing to overwrite immutable structure projection object")
        return
    _storage_call(storage.put, bucket, object_name, payload, tenant_id=tenant_id)
    if not _storage_call(storage.obj_exist, bucket, object_name, tenant_id=tenant_id) or _storage_call(storage.get, bucket, object_name, tenant_id=tenant_id) != payload:
        raise IOError("failed to verify structure projection object")


def store_tabular_structure_projection(
    storage,
    *,
    bucket: str,
    document_id: str,
    projection: dict[str, Any],
    rows_per_part: int = DEFAULT_ROWS_PER_PART,
    tenant_id: str | None = None,
) -> dict[str, Any]:
    """Write immutable parts and publish their generation manifest last.

    The manifest only proves that the generation was fully persisted. It is not
    an active-generation pointer; activation remains a separate API concern.
    """

    if not bucket or not document_id:
        raise ValueError("bucket and document_id are required")
    parts = partition_tabular_structure_projection(projection, rows_per_part=rows_per_part)
    generation_ref = projection["producer_generation_ref"]
    prefix = tabular_structure_projection_prefix(
        document_id,
        generation_ref,
    ).removesuffix("/")
    manifest_parts = []

    for part in parts:
        payload = _canonical_json(part)
        payload_sha256 = hashlib.sha256(payload).hexdigest()
        object_name = f"{prefix}/part-{part['part_number']:06d}-{payload_sha256}.json"
        _put_verified_immutable(storage, bucket, object_name, payload, tenant_id)
        manifest_parts.append(
            {
                "part_number": part["part_number"],
                "object_name": object_name,
                "row_offset": part["row_offset"],
                "row_count": part["row_count"],
                "sha256": payload_sha256,
            }
        )

    manifest = {
        "version": PROJECTION_VERSION,
        "producer_schema_version": projection["producer_schema_version"],
        "producer_generation_ref": generation_ref,
        "structure_algorithm_version": projection["structure_algorithm_version"],
        "enumeration_rule_version": projection["enumeration_rule_version"],
        "source_sha256": projection["source_sha256"],
        "row_count": len(projection["rows"]),
        "tables": projection["tables"],
        "parts": manifest_parts,
    }
    manifest_payload = _canonical_json(manifest)
    manifest_sha256 = hashlib.sha256(manifest_payload).hexdigest()
    manifest_object_name = f"{prefix}/manifest-{manifest_sha256}.json"
    _put_verified_immutable(storage, bucket, manifest_object_name, manifest_payload, tenant_id)
    return {
        "producer_generation_ref": generation_ref,
        "manifest_object_name": manifest_object_name,
        "manifest_sha256": manifest_sha256,
        "part_count": len(parts),
        "row_count": len(projection["rows"]),
    }


def _get_immutable_object(storage, bucket: str, object_name: str, tenant_id: str | None) -> bytes:
    payload = _storage_call(storage.get, bucket, object_name, tenant_id=tenant_id)
    if not isinstance(payload, bytes) or not payload:
        raise StructureSnapshotMissing("structure snapshot object is missing")
    return payload


def _decode_snapshot_json(payload: bytes, label: str) -> dict[str, Any]:
    try:
        value = json.loads(payload)
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise StructureSnapshotChanged(f"{label} payload is invalid") from exc
    if not isinstance(value, dict):
        raise StructureSnapshotChanged(f"{label} payload is invalid")
    return value


def list_tabular_structure_projection_objects(
    storage,
    *,
    bucket: str,
    document_id: str,
    producer_generation_ref: str,
    manifest_object_name: str,
    manifest_sha256: str,
    expected_part_count: int | None = None,
    tenant_id: str | None = None,
) -> dict[str, Any]:
    """Return the exact immutable object names recorded by one generation.

    Deletion cannot use a dataset-wide prefix scan because buckets may contain
    multiple documents and generations. The manifest is the authoritative,
    digest-bound inventory; only names that pass the document/generation scope
    checks are returned.
    """

    if not bucket or not document_id or not manifest_object_name:
        raise StructureSnapshotMissing("structure snapshot identity is missing")
    _validate_generation_ref(producer_generation_ref)
    if not re.fullmatch(r"[0-9a-f]{64}", manifest_sha256 or ""):
        raise StructureSnapshotChanged("manifest digest is invalid")
    document_ref = _versioned_digest("tabular-structure-document/v1", document_id)
    expected_prefix = f"_fuxi/tabular-structure/v1/{document_ref}/{producer_generation_ref}/"
    expected_manifest_name = f"{expected_prefix}manifest-{manifest_sha256}.json"
    if manifest_object_name != expected_manifest_name:
        raise StructureSnapshotChanged("manifest document scope changed")

    manifest_payload = _get_immutable_object(storage, bucket, manifest_object_name, tenant_id)
    if hashlib.sha256(manifest_payload).hexdigest() != manifest_sha256:
        raise StructureSnapshotChanged("manifest digest changed")
    manifest = _decode_snapshot_json(manifest_payload, "manifest")
    expected_manifest_fields = {
        "version",
        "producer_schema_version",
        "producer_generation_ref",
        "structure_algorithm_version",
        "enumeration_rule_version",
        "source_sha256",
        "row_count",
        "tables",
        "parts",
    }
    if set(manifest) != expected_manifest_fields:
        raise StructureSnapshotChanged("manifest schema changed")
    if manifest["producer_generation_ref"] != producer_generation_ref:
        raise StructureSnapshotChanged("manifest generation changed")
    if not isinstance(manifest["parts"], list):
        raise StructureSnapshotChanged("manifest parts changed")
    if expected_part_count is not None and (
        not isinstance(expected_part_count, int)
        or isinstance(expected_part_count, bool)
        or expected_part_count < 0
        or len(manifest["parts"]) != expected_part_count
    ):
        raise StructureSnapshotChanged("generation part count changed")

    part_object_names: list[str] = []
    expected_offset = 0
    for expected_part_number, part_manifest in enumerate(manifest["parts"], start=1):
        expected_part_fields = {"part_number", "object_name", "row_offset", "row_count", "sha256"}
        if not isinstance(part_manifest, dict) or set(part_manifest) != expected_part_fields:
            raise StructureSnapshotChanged("part manifest changed")
        part_sha256 = part_manifest["sha256"]
        if (
            part_manifest["part_number"] != expected_part_number
            or part_manifest["row_offset"] != expected_offset
            or not isinstance(part_manifest["row_count"], int)
            or isinstance(part_manifest["row_count"], bool)
            or part_manifest["row_count"] < 0
            or not isinstance(part_manifest["object_name"], str)
            or not isinstance(part_sha256, str)
            or not re.fullmatch(r"[0-9a-f]{64}", part_sha256)
        ):
            raise StructureSnapshotChanged("part manifest changed")
        expected_part_name = f"{expected_prefix}part-{expected_part_number:06d}-{part_sha256}.json"
        if part_manifest["object_name"] != expected_part_name:
            raise StructureSnapshotChanged("part document scope changed")
        part_object_names.append(expected_part_name)
        expected_offset += part_manifest["row_count"]

    if (
        not isinstance(manifest["row_count"], int)
        or isinstance(manifest["row_count"], bool)
        or manifest["row_count"] < 0
        or expected_offset != manifest["row_count"]
    ):
        raise StructureSnapshotChanged("manifest row count changed")
    return {
        "manifest_object_name": manifest_object_name,
        "part_object_names": part_object_names,
        "object_names": [*part_object_names, manifest_object_name],
    }


def _strict_delete_projection_object(
    storage,
    *,
    bucket: str,
    object_name: str,
    tenant_id: str | None,
) -> None:
    strict_delete = getattr(storage, "rm_strict", None)
    if not callable(strict_delete):
        raise RuntimeError("storage backend does not support strict object deletion")
    _storage_call(strict_delete, bucket, object_name, tenant_id=tenant_id)


def delete_tabular_structure_projection_prefix(
    storage,
    *,
    bucket: str,
    document_id: str,
    producer_generation_ref: str,
    tenant_id: str | None = None,
) -> int:
    """Strictly remove one interrupted generation using its exact scope."""

    strict_delete_prefix = getattr(storage, "rm_prefix_strict", None)
    if not callable(strict_delete_prefix):
        raise RuntimeError(
            "storage backend does not support strict prefix deletion"
        )
    prefix = tabular_structure_projection_prefix(
        document_id,
        producer_generation_ref,
    )
    result = _storage_call(
        strict_delete_prefix,
        bucket,
        prefix,
        tenant_id=tenant_id,
    )
    if not isinstance(result, int) or isinstance(result, bool) or result < 0:
        raise RuntimeError("strict prefix deletion returned an invalid result")
    return result


def delete_tabular_structure_projection_parts(
    storage,
    *,
    bucket: str,
    document_id: str,
    producer_generation_ref: str,
    manifest_object_name: str,
    manifest_sha256: str,
    expected_part_count: int | None = None,
    tenant_id: str | None = None,
) -> int:
    """Delete the digest-bound parts while retaining the manifest retry ledger."""

    inventory = list_tabular_structure_projection_objects(
        storage,
        bucket=bucket,
        document_id=document_id,
        producer_generation_ref=producer_generation_ref,
        manifest_object_name=manifest_object_name,
        manifest_sha256=manifest_sha256,
        expected_part_count=expected_part_count,
        tenant_id=tenant_id,
    )
    for object_name in inventory["part_object_names"]:
        _strict_delete_projection_object(
            storage,
            bucket=bucket,
            object_name=object_name,
            tenant_id=tenant_id,
        )
    return len(inventory["part_object_names"])


def delete_tabular_structure_projection_manifest(
    storage,
    *,
    bucket: str,
    manifest_object_name: str,
    tenant_id: str | None = None,
) -> int:
    """Idempotently delete the manifest after parts deletion is durable."""

    _strict_delete_projection_object(
        storage,
        bucket=bucket,
        object_name=manifest_object_name,
        tenant_id=tenant_id,
    )
    return 1


def delete_tabular_structure_projection_objects(
    storage,
    *,
    bucket: str,
    document_id: str,
    producer_generation_ref: str,
    manifest_object_name: str,
    manifest_sha256: str,
    expected_part_count: int | None = None,
    tenant_id: str | None = None,
) -> int:
    """Delete one exact generation projection, parts before its manifest."""

    deleted = delete_tabular_structure_projection_parts(
        storage,
        bucket=bucket,
        document_id=document_id,
        producer_generation_ref=producer_generation_ref,
        manifest_object_name=manifest_object_name,
        manifest_sha256=manifest_sha256,
        expected_part_count=expected_part_count,
        tenant_id=tenant_id,
    )
    return deleted + delete_tabular_structure_projection_manifest(
        storage,
        bucket=bucket,
        manifest_object_name=manifest_object_name,
        tenant_id=tenant_id,
    )


def _load_tabular_structure_projection_for_contracts(
    storage,
    *,
    bucket: str,
    document_id: str,
    producer_generation_ref: str,
    manifest_object_name: str,
    manifest_sha256: str,
    expected_part_count: int | None = None,
    tenant_id: str | None = None,
    accepted_contracts: frozenset[tuple[str, str, str, str]],
) -> dict[str, Any]:
    """Read and verify one exact immutable projection generation."""

    if not bucket or not document_id or not manifest_object_name:
        raise StructureSnapshotMissing("structure snapshot identity is missing")
    _validate_generation_ref(producer_generation_ref)
    if not re.fullmatch(r"[0-9a-f]{64}", manifest_sha256 or ""):
        raise StructureSnapshotChanged("manifest digest is invalid")
    document_ref = _versioned_digest("tabular-structure-document/v1", document_id)
    expected_prefix = f"_fuxi/tabular-structure/v1/{document_ref}/{producer_generation_ref}/"
    expected_manifest_name = f"{expected_prefix}manifest-{manifest_sha256}.json"
    if manifest_object_name != expected_manifest_name:
        raise StructureSnapshotChanged("manifest document scope changed")

    manifest_payload = _get_immutable_object(storage, bucket, manifest_object_name, tenant_id)
    if hashlib.sha256(manifest_payload).hexdigest() != manifest_sha256:
        raise StructureSnapshotChanged("manifest digest changed")
    manifest = _decode_snapshot_json(manifest_payload, "manifest")
    expected_manifest_fields = {
        "version",
        "producer_schema_version",
        "producer_generation_ref",
        "structure_algorithm_version",
        "enumeration_rule_version",
        "source_sha256",
        "row_count",
        "tables",
        "parts",
    }
    if set(manifest) != expected_manifest_fields:
        raise StructureSnapshotChanged("manifest schema changed")
    contract = (
        manifest["producer_schema_version"],
        manifest["version"],
        manifest["structure_algorithm_version"],
        manifest["enumeration_rule_version"],
    )
    if contract not in accepted_contracts:
        raise StructureSnapshotChanged("manifest version changed")
    if manifest["producer_generation_ref"] != producer_generation_ref:
        raise StructureSnapshotChanged("manifest generation changed")
    if not isinstance(manifest["parts"], list):
        raise StructureSnapshotChanged("manifest parts changed")
    if expected_part_count is not None and (
        not isinstance(expected_part_count, int)
        or isinstance(expected_part_count, bool)
        or expected_part_count < 0
        or len(manifest["parts"]) != expected_part_count
    ):
        raise StructureSnapshotChanged("generation part count changed")

    rows: list[dict[str, Any]] = []
    expected_offset = 0
    for expected_part_number, part_manifest in enumerate(manifest["parts"], start=1):
        expected_part_fields = {"part_number", "object_name", "row_offset", "row_count", "sha256"}
        if not isinstance(part_manifest, dict) or set(part_manifest) != expected_part_fields:
            raise StructureSnapshotChanged("part manifest changed")
        if (
            part_manifest["part_number"] != expected_part_number
            or part_manifest["row_offset"] != expected_offset
            or not isinstance(part_manifest["row_count"], int)
            or isinstance(part_manifest["row_count"], bool)
            or part_manifest["row_count"] < 0
            or not isinstance(part_manifest["object_name"], str)
            or not re.fullmatch(r"[0-9a-f]{64}", str(part_manifest["sha256"]))
        ):
            raise StructureSnapshotChanged("part manifest changed")
        expected_part_name = f"{expected_prefix}part-{expected_part_number:06d}-{part_manifest['sha256']}.json"
        if part_manifest["object_name"] != expected_part_name:
            raise StructureSnapshotChanged("part document scope changed")

        part_payload = _get_immutable_object(storage, bucket, part_manifest["object_name"], tenant_id)
        if hashlib.sha256(part_payload).hexdigest() != part_manifest["sha256"]:
            raise StructureSnapshotChanged("part digest changed")
        part = _decode_snapshot_json(part_payload, "part")
        if set(part) != {"version", "producer_generation_ref", "part_number", "row_offset", "row_count", "rows"}:
            raise StructureSnapshotChanged("part schema changed")
        if (
            part["version"] != PROJECTION_PART_VERSION
            or part["producer_generation_ref"] != producer_generation_ref
            or part["part_number"] != expected_part_number
            or part["row_offset"] != expected_offset
            or part["row_count"] != part_manifest["row_count"]
            or not isinstance(part["rows"], list)
            or len(part["rows"]) != part["row_count"]
        ):
            raise StructureSnapshotChanged("part generation changed")
        rows.extend(part["rows"])
        expected_offset += part["row_count"]

    if expected_offset != manifest["row_count"]:
        raise StructureSnapshotChanged("manifest row count changed")
    projection = {
        "version": manifest["version"],
        "producer_schema_version": manifest["producer_schema_version"],
        "producer_generation_ref": producer_generation_ref,
        "structure_algorithm_version": manifest["structure_algorithm_version"],
        "enumeration_rule_version": manifest["enumeration_rule_version"],
        "source_sha256": manifest["source_sha256"],
        "tables": manifest["tables"],
        "rows": rows,
    }
    try:
        _validate_tabular_structure_projection_for_contract(projection, contract)
    except ValueError as exc:
        raise StructureSnapshotChanged("structure projection validation changed") from exc
    return projection


def load_tabular_structure_projection(
    storage,
    *,
    bucket: str,
    document_id: str,
    producer_generation_ref: str,
    manifest_object_name: str,
    manifest_sha256: str,
    expected_part_count: int | None = None,
    tenant_id: str | None = None,
) -> dict[str, Any]:
    """Read one projection using only the current runtime contract."""

    return _load_tabular_structure_projection_for_contracts(
        storage,
        bucket=bucket,
        document_id=document_id,
        producer_generation_ref=producer_generation_ref,
        manifest_object_name=manifest_object_name,
        manifest_sha256=manifest_sha256,
        expected_part_count=expected_part_count,
        tenant_id=tenant_id,
        accepted_contracts=frozenset({_CURRENT_PROJECTION_CONTRACT}),
    )


def load_tabular_structure_projection_for_backfill(
    storage,
    *,
    bucket: str,
    document_id: str,
    producer_generation_ref: str,
    manifest_object_name: str,
    manifest_sha256: str,
    expected_part_count: int | None = None,
    tenant_id: str | None = None,
) -> dict[str, Any]:
    """Validate a known immutable generation before deciding index eligibility."""

    return _load_tabular_structure_projection_for_contracts(
        storage,
        bucket=bucket,
        document_id=document_id,
        producer_generation_ref=producer_generation_ref,
        manifest_object_name=manifest_object_name,
        manifest_sha256=manifest_sha256,
        expected_part_count=expected_part_count,
        tenant_id=tenant_id,
        accepted_contracts=_KNOWN_BACKFILL_PROJECTION_CONTRACTS,
    )


def page_tabular_structure_rows(
    projection: dict[str, Any],
    *,
    table_ref: str,
    cursor: int = 0,
    page_size: int = 30,
    row_transport_version: str | None = None,
) -> dict[str, Any]:
    """Return one stable offset page from an already verified generation."""

    validate_tabular_structure_projection(projection)
    if not isinstance(table_ref, str) or not table_ref:
        raise ValueError("table_ref is required")
    if not isinstance(cursor, int) or isinstance(cursor, bool) or cursor < 0:
        raise ValueError("cursor must be a non-negative integer")
    if not isinstance(page_size, int) or isinstance(page_size, bool) or page_size < 1:
        raise ValueError("page_size must be a positive integer")
    if row_transport_version not in {None, ROW_PAGE_TRANSPORT_VERSION}:
        raise ValueError("unsupported row transport version")

    table = next((item for item in projection["tables"] if item["table_ref"] == table_ref), None)
    if table is None:
        raise StructureSnapshotMissing("table snapshot is missing")
    rows = [row for row in projection["rows"] if row["table_ref_kwd"] == table_ref]
    rows.sort(
        key=lambda row: (
            row["data_row_index_int"] if row["data_row_index_int"] is not None else float("inf"),
            row["row_ordinal_int"],
            row["row_ref_kwd"],
        )
    )
    page_rows = rows[cursor : cursor + page_size]
    if row_transport_version == ROW_PAGE_TRANSPORT_VERSION:
        page_rows = [
            [
                row["row_ordinal_int"],
                row["data_row_index_int"],
                row["row_role_kwd"],
                [
                    [field["column_ordinal"], field["value"]]
                    for field in json.loads(row["ordered_fields_list"])
                ],
            ]
            for row in page_rows
        ]
    next_cursor = cursor + len(page_rows)
    has_more = next_cursor < len(rows)
    result = {
        "producer_generation_ref": projection["producer_generation_ref"],
        "table_ref": table_ref,
        "producer_schema_version": projection["producer_schema_version"],
        "projection_version": projection["version"],
        "structure_algorithm_version": projection["structure_algorithm_version"],
        "enumeration_rule_version": projection["enumeration_rule_version"],
        "rows": page_rows,
        "total": len(rows),
        "source_total_count": table["source_total_count"],
        "has_more": has_more,
        "next_cursor": next_cursor if has_more else None,
    }
    if row_transport_version is not None:
        result["row_transport_version"] = row_transport_version
    return result
